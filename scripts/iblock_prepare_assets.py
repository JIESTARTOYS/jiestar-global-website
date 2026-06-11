#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


QUOTE_FILE = Path("/Volumes/ORICO/iblock/积趣IBLOCK-全品报价单 - 分类（男-女-常规品）.xlsx")
SOURCE_ROOT = Path("/Volumes/ORICO/iblock/iBlock积趣_电商素材")
TARGET_ROOT = Path("/Volumes/ORICO/iblock/iblock-上架前整理")
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
IGNORE_NAMES = {".DS_Store", "Thumbs.db", ".WeDrive"}
DETAIL_SHEETS = {"常规品", "女生品", "男生品"}
SKU_ALIASES = {
    "IB1301-1": ("T25", "T-25"),
    "IB1301-2": ("J20", "J-20"),
    "IB1301-3": ("J05", "J-05"),
    "IB1301-4": ("99A",),
    "IB1301-5": ("Z10", "Z-10"),
    "IB1301-6": ("DF15", "DF-15"),
}
TERM_SKU_OVERRIDES = {
    "警察": "IB1102-1",
    "宇航员": "IB1106",
    "ASTRONAUT": "IB1106",
    "火箭": "IB1107",
    "ROCKET": "IB1107",
    "探测车": "IB1108",
    "PROBE CAR": "IB1108",
    "卫星": "IB1109",
    "SATELLITE": "IB1109",
    "医护员": "IB1116",
    "MEDICAL WORKER": "IB1116",
    "救护车": "IB1117",
    "AMBULANCE": "IB1117",
    "献血车": "IB1118",
    "BLOOD DONATION": "IB1118",
    "救护站": "IB1119",
    "FIRST AID": "IB1119",
}
NUMBERED_MAIN_SKU_OVERRIDES = (
    ("治愈小队_头图800X800_SKU", {11: "IB1116", 12: "IB1117", 13: "IB1118", 14: "IB1119"}),
    ("望宇小队_电商_主图", {2: "IB1106", 3: "IB1106", 10: "IB1106", 4: "IB1107", 5: "IB1107", 11: "IB1107", 6: "IB1108", 7: "IB1108", 12: "IB1108", 8: "IB1109", 9: "IB1109", 13: "IB1109"}),
    ("治愈小队_头图", {2: "IB1116", 6: "IB1116", 10: "IB1116", 3: "IB1117", 7: "IB1117", 8: "IB1117", 4: "IB1118", 5: "IB1119", 9: "IB1119"}),
)
SHARED_PATH_SKU_OVERRIDES = {
    "治愈小队_详情页800": ("IB1116", "IB1117", "IB1118", "IB1119"),
}
PROMOTE_SKU_AS_WHITE = {"IB1116", "IB1117", "IB1118", "IB1119"}


@dataclass(frozen=True)
class ProductRow:
    sku: str
    source_sheets: tuple[str, ...]
    brand_series: str
    product_series: str
    name_cn: str
    barcode: str
    recommended_age: str
    carton_qty: str
    carton_spec: str
    dealer_price_level_1_cny: str
    dealer_price_level_2_cny: str
    retail_price_cny: str
    ecommerce_control_price_cny: str
    product_size_cm: str
    box_size_cm: str
    tax_included_price_cny: str

    @property
    def parent_sku(self) -> str:
        return self.sku.split("-", 1)[0]


@dataclass(frozen=True)
class ImageEntry:
    source: Path
    role: str
    skus: tuple[str, ...]
    ignored_reason: str = ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    return re.sub(r"\s+", " ", text)


def normalize_sku(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).upper()
    text = re.sub(r"IBLOCK[_ -]*IB", "IB", text)
    text = re.sub(r"[^A-Z0-9-]", "", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text


def parent_sku(sku: str) -> str:
    return sku.split("-", 1)[0]


def expand_short_end(start_digits: str, end_digits: str) -> int:
    if len(end_digits) < len(start_digits):
        end_digits = start_digits[: len(start_digits) - len(end_digits)] + end_digits
    return int(end_digits)


def expand_sku_codes(text: str) -> list[str]:
    normalized = unicodedata.normalize("NFKC", text).upper()
    normalized = re.sub(r"IBLOCK[_ -]*IB", "IB", normalized)
    output: list[str] = []
    pattern = re.compile(r"(?<![A-Z0-9])(IB|LL|CM)\s*[-_ ]?\s*(\d{3,5})(?!\d)", re.I)

    full_code_range = re.compile(
        r"(?<![A-Z0-9])(IB|LL|CM)\s*[-_ ]?\s*(\d{3,5})\s*[-_~至]\s*(IB|LL|CM)\s*[-_ ]?\s*(\d{3,5})(?!\d)",
        re.I,
    )
    for hit in full_code_range.finditer(normalized):
        prefix = hit.group(1).upper()
        end_prefix = hit.group(3).upper()
        start_number = int(hit.group(2))
        end_number = int(hit.group(4))
        if prefix != end_prefix or end_number <= start_number or end_number - start_number > 100:
            continue
        width = len(hit.group(2))
        for number in range(start_number, end_number + 1):
            code = f"{prefix}{number:0{width}d}"
            if code not in output:
                output.append(code)

    for hit in pattern.finditer(normalized):
        prefix = hit.group(1).upper()
        digits = hit.group(2)
        first = f"{prefix}{digits}"
        if first not in output:
            output.append(first)

        tail = normalized[hit.end() : hit.end() + 32]
        range_match = re.match(r"\s*[-~至]\s*(?:(IB|LL|CM)\s*[-_ ]?)?(\d{1,5})(?![A-Z0-9])", tail, re.I)
        if not range_match:
            continue
        end_prefix = (range_match.group(1) or prefix).upper()
        end_number = expand_short_end(digits, range_match.group(2))
        start_number = int(digits)
        if end_prefix != prefix or end_number <= start_number or end_number - start_number > 100:
            continue
        for number in range(start_number + 1, end_number + 1):
            code = f"{prefix}{number:0{len(digits)}d}"
            if code not in output:
                output.append(code)

    return output


def name_keywords(name: str) -> list[str]:
    normalized = normalize_text(name)
    pieces = re.split(r"[()（）·.\-—_/、,，\s]+", normalized)
    output: list[str] = []
    for piece in pieces:
        piece = piece.strip()
        if not piece or len(piece) > 12:
            continue
        if re.search(r"[\u4e00-\u9fff]", piece) and piece not in output:
            output.append(piece)
    return output


def sibling_specific_keywords(row: ProductRow, siblings: list[ProductRow]) -> list[str]:
    keywords = name_keywords(row.name_cn)
    if len(siblings) <= 1:
        return keywords
    sibling_keyword_counts = Counter(keyword for sibling in siblings for keyword in name_keywords(sibling.name_cn))
    specific = [keyword for keyword in keywords if sibling_keyword_counts[keyword] == 1]
    return specific or keywords


def row_value(row: tuple[Any, ...], index: int) -> str:
    return normalize_text(row[index]) if len(row) > index else ""


def detail_row_from_sheet(sheet_name: str, row: tuple[Any, ...]) -> ProductRow | None:
    sku = normalize_sku(row_value(row, 4))
    if not sku or not re.search(r"\d", sku) or sku == "产品货号":
        return None
    return ProductRow(
        sku=sku,
        source_sheets=(sheet_name,),
        brand_series=row_value(row, 2),
        product_series=row_value(row, 3),
        name_cn=row_value(row, 5),
        barcode=row_value(row, 8),
        recommended_age=row_value(row, 9),
        carton_qty=row_value(row, 10),
        carton_spec=row_value(row, 11),
        dealer_price_level_1_cny=row_value(row, 12),
        dealer_price_level_2_cny=row_value(row, 13),
        retail_price_cny=row_value(row, 14),
        ecommerce_control_price_cny=row_value(row, 15),
        product_size_cm=row_value(row, 16),
        box_size_cm=row_value(row, 17),
        tax_included_price_cny="",
    )


def simple_row_from_sheet(sheet_name: str, row: tuple[Any, ...]) -> ProductRow | None:
    sku = normalize_sku(row_value(row, 3))
    if not sku or not re.search(r"\d", sku) or sku == "产品货号":
        return None
    return ProductRow(
        sku=sku,
        source_sheets=(sheet_name,),
        brand_series=row_value(row, 1),
        product_series=row_value(row, 2),
        name_cn=row_value(row, 4),
        barcode="",
        recommended_age="",
        carton_qty="",
        carton_spec="",
        dealer_price_level_1_cny="",
        dealer_price_level_2_cny="",
        retail_price_cny=row_value(row, 6),
        ecommerce_control_price_cny="",
        product_size_cm="",
        box_size_cm="",
        tax_included_price_cny=row_value(row, 5),
    )


def merge_product(existing: ProductRow, candidate: ProductRow) -> ProductRow:
    fields = {
        field: getattr(existing, field) or getattr(candidate, field)
        for field in (
            "brand_series",
            "product_series",
            "name_cn",
            "barcode",
            "recommended_age",
            "carton_qty",
            "carton_spec",
            "dealer_price_level_1_cny",
            "dealer_price_level_2_cny",
            "retail_price_cny",
            "ecommerce_control_price_cny",
            "product_size_cm",
            "box_size_cm",
            "tax_included_price_cny",
        )
    }
    source_sheets = tuple(dict.fromkeys([*existing.source_sheets, *candidate.source_sheets]))
    return ProductRow(sku=existing.sku, source_sheets=source_sheets, **fields)


def parse_quote_rows(path: Path) -> tuple[list[ProductRow], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    by_sku: dict[str, ProductRow] = {}
    raw_count = 0

    for sheet in workbook.worksheets:
        if sheet.title in DETAIL_SHEETS:
            for row in sheet.iter_rows(min_row=5, values_only=True):
                product = detail_row_from_sheet(sheet.title, row)
                if not product:
                    continue
                raw_count += 1
                by_sku[product.sku] = merge_product(by_sku[product.sku], product) if product.sku in by_sku else product
        elif sheet.title == "Sheet1":
            for row in sheet.iter_rows(min_row=5, values_only=True):
                product = simple_row_from_sheet(sheet.title, row)
                if not product:
                    continue
                raw_count += 1
                by_sku[product.sku] = merge_product(by_sku[product.sku], product) if product.sku in by_sku else product

    return sorted(by_sku.values(), key=lambda product: natural_string_key(product.sku)), raw_count - len(by_sku)


def classify_role(path: Path) -> tuple[str, str]:
    text = unicodedata.normalize("NFKC", path.as_posix())
    lower = text.lower()
    if "微信" in text or "九宫格" in text:
        return "ignored", "wechat grid"
    if "视频" in text:
        return "ignored", "video still"
    if "设计源文件" in text or "设计策划" in text:
        return "ignored", "design source"
    if "六合一" in text and "详情" not in text and "分切" not in text:
        return "other", ""
    if "白底" in text:
        return "white", ""
    if "sku" in lower or "SKU图" in text or "sku图" in text:
        return "sku", ""
    if "详情" in text or "分切" in text:
        return "detail", ""
    if re.search(r"(头图|主图)[_ -]?(800|800X800|800x800)", text, re.I) or "800-800" in text:
        return "main", ""
    if "实拍" in text or "拍摄" in text or "转盘" in text:
        return "photo", ""
    return "other", ""


def image_files(root: Path) -> list[Path]:
    return sorted(
        [
            path
            for path in root.rglob("*")
            if path.is_file()
            and path.suffix.lower() in IMAGE_EXTS
            and not path.name.startswith("._")
            and path.name not in IGNORE_NAMES
        ],
        key=lambda path: natural_string_key(path.as_posix()),
    )


def natural_string_key(value: Any) -> list[Any]:
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", str(value))]


def candidate_skus_for_path(path: Path, rows: list[ProductRow]) -> list[str]:
    path_text = unicodedata.normalize("NFKC", path.as_posix()).upper()
    path_text_cn = unicodedata.normalize("NFKC", path.as_posix())
    expanded = set(expand_sku_codes(path_text))
    filename_expanded = set(expand_sku_codes(path.name))
    known_skus = {row.sku for row in rows}
    by_parent: dict[str, list[ProductRow]] = defaultdict(list)
    for row in rows:
        by_parent[row.parent_sku].append(row)

    output: list[str] = []
    special_skus = special_skus_for_path(path_text_cn, known_skus)
    if special_skus is not None:
        return special_skus

    expanded_known_skus = expanded & known_skus
    if len(expanded_known_skus) > 1:
        expanded_rows = [row for row in rows if row.sku in expanded_known_skus]
        specific_range_matches = [
            row.sku
            for row in expanded_rows
            if any(keyword in path_text_cn for keyword in sibling_specific_keywords(row, expanded_rows))
        ]
        if specific_range_matches:
            return sorted(dict.fromkeys(specific_range_matches), key=natural_string_key)

        filename_known_skus = filename_expanded & expanded_known_skus
        if filename_known_skus and filename_known_skus != expanded_known_skus:
            return sorted(filename_known_skus, key=natural_string_key)

    for sku in sorted(expanded & known_skus, key=natural_string_key):
        output.append(sku)

    for code in sorted(expanded, key=natural_string_key):
        siblings = by_parent.get(code, [])
        specific_matches: list[str] = []
        fallback_matches: list[str] = []
        for row in siblings:
            if row.sku in output:
                continue
            if row.sku == row.parent_sku:
                if not any(keyword in path_text_cn for keyword in ("组合", "套装", "混装")):
                    fallback_matches.append(row.sku)
                continue
            specific_keywords = sibling_specific_keywords(row, siblings)
            if specific_keywords and any(keyword in path_text_cn for keyword in specific_keywords):
                specific_matches.append(row.sku)
                continue
            keywords = name_keywords(row.name_cn)
            if keywords and any(keyword in path_text_cn for keyword in keywords):
                fallback_matches.append(row.sku)
        if specific_matches:
            output = [sku for sku in output if sku != code]
        for sku in specific_matches or fallback_matches:
            if sku not in output:
                output.append(sku)

    for row in rows:
        if row.sku in output:
            continue
        if expanded_known_skus and row.sku not in expanded_known_skus:
            continue
        aliases = SKU_ALIASES.get(row.sku, ())
        if aliases and any(alias.upper() in path_text for alias in aliases):
            output.append(row.sku)

    return output


def special_skus_for_path(path_text: str, known_skus: set[str]) -> list[str] | None:
    normalized = unicodedata.normalize("NFKC", path_text)
    upper = normalized.upper()
    for marker, skus in SHARED_PATH_SKU_OVERRIDES.items():
        if marker.upper() in upper:
            return [sku for sku in skus if sku in known_skus]
    for marker, number_map in NUMBERED_MAIN_SKU_OVERRIDES:
        if marker.upper() not in upper:
            continue
        match = re.search(r"\((\d+)\)|(?:头图|素材)_(\d+)_|头图(\d+)_|副本\s*(\d+)-", normalized, re.I)
        if not match:
            continue
        number = int(next(group for group in match.groups() if group))
        sku = number_map.get(number)
        return [sku] if sku in known_skus else []

    matched = [
        sku
        for term, sku in TERM_SKU_OVERRIDES.items()
        if sku in known_skus and (term in normalized or term in upper)
    ]
    if matched:
        return list(dict.fromkeys(matched))
    return None


def collect_images(root: Path, rows: list[ProductRow]) -> tuple[list[ImageEntry], list[ImageEntry], list[Path]]:
    entries: list[ImageEntry] = []
    ignored: list[ImageEntry] = []
    unmatched: list[Path] = []

    for source in image_files(root):
        relative = source.relative_to(root)
        role, ignored_reason = classify_role(relative)
        skus = tuple(candidate_skus_for_path(relative, rows))
        if role == "sku" and len(skus) == 1 and skus[0] in PROMOTE_SKU_AS_WHITE:
            role = "white"
        if role == "ignored":
            ignored.append(ImageEntry(source=source, role=role, skus=skus, ignored_reason=ignored_reason))
            continue
        if not skus:
            unmatched.append(source)
            continue
        entries.append(ImageEntry(source=source, role=role, skus=skus))

    return entries, ignored, unmatched


def paths_for_role(entries: list[ImageEntry], sku: str, role: str) -> list[Path]:
    paths = [entry.source for entry in entries if entry.role == role and sku in entry.skus]
    return sorted(dict.fromkeys(paths), key=lambda path: natural_string_key(path.as_posix()))


def copy_targets_for_sku(sku: str, entries: list[ImageEntry]) -> list[tuple[Path, str, str]]:
    copies: list[tuple[Path, str, str]] = []
    white = paths_for_role(entries, sku, "white")
    main = paths_for_role(entries, sku, "main")
    sku_images = paths_for_role(entries, sku, "sku")
    detail = paths_for_role(entries, sku, "detail")

    for index, source in enumerate(white, start=1):
        name = f"{sku}-白底{source.suffix.lower()}" if index == 1 else f"{sku}-白底-{index:02d}{source.suffix.lower()}"
        copies.append((source, name, "white"))
    for index, source in enumerate(main, start=1):
        copies.append((source, f"{sku}-{index}{source.suffix.lower()}", "main"))
    for index, source in enumerate(sku_images, start=1):
        name = f"{sku}-sku{source.suffix.lower()}" if index == 1 else f"{sku}-sku-{index:02d}{source.suffix.lower()}"
        copies.append((source, name, "sku"))
    for index, source in enumerate(detail, start=1):
        copies.append((source, f"{sku}-详情-{index:02d}{source.suffix.lower()}", "detail"))

    return copies


def readiness_for_counts(counts: dict[str, int], ambiguous: bool) -> tuple[str, str]:
    reasons: list[str] = []
    if ambiguous:
        reasons.append("possible combination or shared-series assets need manual review")
    if counts["white_image_count"] == 0 and counts["sku_image_count"] == 0:
        reasons.append("missing white/SKU image")
    if counts["main_image_count"] == 0:
        reasons.append("missing main image")
    if counts["detail_image_count"] == 0:
        reasons.append("missing detail image")
    if ambiguous:
        return "MANUAL_REVIEW", "; ".join(reasons)
    if reasons:
        return "IMAGE_REVIEW", "; ".join(reasons)
    return "READY_TO_CREATE", ""


def build_catalog_row(product: ProductRow, entries: list[ImageEntry], target_root: Path) -> dict[str, Any]:
    matched_entries = [entry for entry in entries if product.sku in entry.skus]
    counts = {
        "white_image_count": len(paths_for_role(entries, product.sku, "white")),
        "main_image_count": len(paths_for_role(entries, product.sku, "main")),
        "sku_image_count": len(paths_for_role(entries, product.sku, "sku")),
        "detail_image_count": len(paths_for_role(entries, product.sku, "detail")),
    }
    ambiguous = any(len(entry.skus) > 1 and product.sku in entry.skus for entry in matched_entries)
    readiness, reason = readiness_for_counts(counts, ambiguous)
    custom_series = product.product_series or product.brand_series or "Building Block Sets"

    return {
        "sku": product.sku,
        "vendor": "iBlock",
        "category": "Interlocking Blocks",
        "brand_series": product.brand_series,
        "product_series": product.product_series,
        "name_cn": product.name_cn,
        "source_sheets": "; ".join(product.source_sheets),
        "barcode": product.barcode,
        "recommended_age": product.recommended_age,
        "carton_qty": product.carton_qty,
        "carton_spec": product.carton_spec,
        "dealer_price_level_1_cny": product.dealer_price_level_1_cny,
        "dealer_price_level_2_cny": product.dealer_price_level_2_cny,
        "retail_price_cny": product.retail_price_cny,
        "ecommerce_control_price_cny": product.ecommerce_control_price_cny,
        "product_size_cm": product.product_size_cm,
        "white_image_count": counts["white_image_count"],
        "main_image_count": counts["main_image_count"],
        "sku_image_count": counts["sku_image_count"],
        "detail_image_count": counts["detail_image_count"],
        "readiness": readiness,
        "review_reason": reason,
        "box_size_cm": product.box_size_cm,
        "tax_included_price_cny": product.tax_included_price_cny,
        "product_type": custom_series,
        "custom_series": custom_series,
        "local_image_count": len(matched_entries),
        "image_folder": (target_root / "images" / product.sku).as_posix(),
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name, rows in sheets.items():
        sheet = workbook.create_sheet(sheet_name[:31])
        if not rows:
            continue
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)

    workbook.save(path)


def remove_appledouble_files(root: Path) -> int:
    removed = 0
    if not root.exists():
        return removed
    for path in sorted(root.rglob("._*"), key=lambda value: len(value.parts), reverse=True):
        if path.is_dir():
            shutil.rmtree(path)
        else:
            path.unlink()
        removed += 1
    return removed


def build_copy_plan(rows: list[ProductRow], entries: list[ImageEntry], target_root: Path, dry_run: bool) -> list[dict[str, Any]]:
    copy_rows: list[dict[str, Any]] = []
    for product in rows:
        for source, target_name, role in copy_targets_for_sku(product.sku, entries):
            target = target_root / "images" / product.sku / target_name
            copy_rows.append(
                {
                    "sku": product.sku,
                    "role": role,
                    "source": source.as_posix(),
                    "target": target.as_posix(),
                    "applied": "no" if dry_run else "yes",
                }
            )
            if not dry_run:
                target.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source, target)
    return copy_rows


def build_reports(quote_file: Path, source_root: Path, target_root: Path, dry_run: bool) -> dict[str, Any]:
    rows, duplicate_count = parse_quote_rows(quote_file)
    entries, ignored, unmatched = collect_images(source_root, rows)
    catalog = [build_catalog_row(product, entries, target_root) for product in rows]
    manual_review = [row for row in catalog if row["readiness"] != "READY_TO_CREATE"]
    copy_plan = build_copy_plan(rows, entries, target_root, dry_run=dry_run)
    ignored_rows = [
        {
            "role": entry.role,
            "ignored_reason": entry.ignored_reason,
            "matched_skus": "; ".join(entry.skus),
            "source": entry.source.as_posix(),
        }
        for entry in ignored
    ]
    unmatched_rows = [{"source": path.as_posix()} for path in unmatched]
    reports_dir = target_root / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    catalog_csv = reports_dir / "iblock-catalog-ready.csv"
    manual_csv = reports_dir / "iblock-manual-review.csv"
    copy_csv = reports_dir / "iblock-asset-copy-plan.csv"
    ignored_csv = reports_dir / "iblock-ignored-assets.csv"
    unmatched_csv = reports_dir / "iblock-unmatched-assets.csv"
    workbook_path = target_root / "iblock-catalog-ready.xlsx"
    summary_path = reports_dir / "iblock-summary.json"

    write_csv(catalog_csv, catalog)
    write_csv(manual_csv, manual_review)
    write_csv(copy_csv, copy_plan)
    write_csv(ignored_csv, ignored_rows)
    write_csv(unmatched_csv, unmatched_rows)
    write_workbook(
        workbook_path,
        {
            "Catalog": catalog,
            "ManualReview": manual_review,
            "AssetCopyPlan": copy_plan,
            "IgnoredAssets": ignored_rows,
            "UnmatchedAssets": unmatched_rows,
        },
    )

    role_counts = Counter(entry.role for entry in entries)
    summary = {
        "quote_file": quote_file.as_posix(),
        "source_root": source_root.as_posix(),
        "target_root": target_root.as_posix(),
        "dry_run": dry_run,
        "unique_quote_sku_count": len(rows),
        "duplicate_quote_row_count": duplicate_count,
        "source_image_count": len(entries) + len(ignored) + len(unmatched),
        "matched_image_entry_count": len(entries),
        "matched_image_sku_count": len({sku for entry in entries for sku in entry.skus}),
        "ignored_image_count": len(ignored),
        "unmatched_image_count": len(unmatched),
        "image_role_counts": dict(sorted(role_counts.items())),
        "planned_copy_count": len(copy_plan),
        "ready_to_create_count": sum(1 for row in catalog if row["readiness"] == "READY_TO_CREATE"),
        "image_review_count": sum(1 for row in catalog if row["readiness"] == "IMAGE_REVIEW"),
        "manual_review_count": sum(1 for row in catalog if row["readiness"] == "MANUAL_REVIEW"),
        "missing_white_or_sku_count": sum(
            1 for row in catalog if row["white_image_count"] == 0 and row["sku_image_count"] == 0
        ),
        "missing_main_count": sum(1 for row in catalog if row["main_image_count"] == 0),
        "missing_detail_count": sum(1 for row in catalog if row["detail_image_count"] == 0),
        "outputs": {
            "workbook": workbook_path.as_posix(),
            "catalog_csv": catalog_csv.as_posix(),
            "manual_review_csv": manual_csv.as_posix(),
            "copy_plan_csv": copy_csv.as_posix(),
            "ignored_assets_csv": ignored_csv.as_posix(),
            "unmatched_assets_csv": unmatched_csv.as_posix(),
            "summary_json": summary_path.as_posix(),
        },
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    removed = remove_appledouble_files(target_root)
    if removed:
        summary["appledouble_removed_count"] = removed
        summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        remove_appledouble_files(target_root)
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare iBlock quote data and staged Shopify image assets.")
    parser.add_argument("--quote-file", type=Path, default=QUOTE_FILE)
    parser.add_argument("--source-root", type=Path, default=SOURCE_ROOT)
    parser.add_argument("--target-root", type=Path, default=TARGET_ROOT)
    parser.add_argument("--dry-run", action="store_true", help="Write reports but do not copy images.")
    parser.add_argument("--apply", action="store_true", help="Copy normalized images into the target root.")
    args = parser.parse_args()

    if args.dry_run and args.apply:
        parser.error("--dry-run and --apply cannot be used together")
    dry_run = not args.apply
    summary = build_reports(args.quote_file, args.source_root, args.target_root, dry_run=dry_run)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
