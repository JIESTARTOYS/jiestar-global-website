#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image, UnidentifiedImageError


TARGET_ROOT = Path("/Volumes/ORICO/iblock/iblock-上架前整理")
PRODUCT_ROOT = TARGET_ROOT / "shopify-products"
REPORTS_ROOT = TARGET_ROOT / "reports"
WORKBOOK_PATH = TARGET_ROOT / "iblock-shopify-readiness.xlsx"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}


def classify_role(filename: str) -> str:
    lower = filename.lower()
    if "-白底" in filename:
        return "white"
    if "-sku" in lower:
        return "sku"
    if "-详情" in filename:
        return "detail"
    return "main"


def sha1_file(path: Path) -> str:
    digest = hashlib.sha1()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def image_dimensions(path: Path) -> tuple[int, int, str]:
    try:
        with Image.open(path) as image:
            return image.width, image.height, image.format or ""
    except UnidentifiedImageError:
        return 0, 0, "UNREADABLE"


def sku_from_group_filename(filename: str) -> str:
    return filename.split("__", 1)[0] if "__" in filename else filename.split("-", 1)[0]


def issue(
    group: str,
    issue_type: str,
    severity: str,
    file_path: Path | None = None,
    sku: str = "",
    details: str = "",
) -> dict[str, Any]:
    return {
        "upload_group": group,
        "sku": sku,
        "issue_type": issue_type,
        "severity": severity,
        "file": file_path.as_posix() if file_path else "",
        "details": details,
    }


def audit_detail_shape(group: str, path: Path, sku: str, width: int, height: int) -> dict[str, Any] | None:
    if not width or not height:
        return issue(group, "unreadable_image", "blocker", path, sku, "Pillow could not read this image file.")
    if width < 600 or height < 600:
        return issue(group, "small_image_review", "warning", path, sku, f"{width}x{height}")
    if width > height:
        return issue(group, "detail_image_shape_review", "warning", path, sku, f"landscape detail image {width}x{height}")
    if height < width * 1.15:
        return issue(group, "detail_image_shape_review", "warning", path, sku, f"near-square detail image {width}x{height}")
    return None


def audit_group(group_folder: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    group = group_folder.name
    image_folder = group_folder / "images"
    issues: list[dict[str, Any]] = []
    files = sorted(
        [
            path
            for path in image_folder.iterdir()
            if path.is_file() and path.suffix.lower() in IMAGE_EXTS and not path.name.startswith("._")
        ]
        if image_folder.exists()
        else []
    )
    role_counts = Counter()
    sku_counts: dict[str, Counter[str]] = defaultdict(Counter)
    hash_to_files: dict[str, list[Path]] = defaultdict(list)
    dimension_counts = Counter()
    unreadable_count = 0

    for path in files:
        sku = sku_from_group_filename(path.name)
        role = classify_role(path.name)
        role_counts[role] += 1
        sku_counts[sku][role] += 1
        width, height, fmt = image_dimensions(path)
        if fmt == "UNREADABLE":
            unreadable_count += 1
        else:
            dimension_counts[f"{width}x{height}"] += 1
        if role == "detail":
            shape_issue = audit_detail_shape(group, path, sku, width, height)
            if shape_issue:
                issues.append(shape_issue)
        if width and height and (width < 500 or height < 500):
            issues.append(issue(group, "small_image_review", "warning", path, sku, f"{width}x{height}"))
        hash_to_files[sha1_file(path)].append(path)

    duplicate_file_count = 0
    duplicate_sets = 0
    for digest, duplicate_files in hash_to_files.items():
        if len(duplicate_files) <= 1:
            continue
        duplicate_sets += 1
        duplicate_file_count += len(duplicate_files)
        issues.append(
            issue(
                group,
                "exact_duplicate_in_product_group",
                "warning",
                duplicate_files[0],
                "",
                f"{len(duplicate_files)} identical files: "
                + "; ".join(path.name for path in duplicate_files[:12])
                + ("" if len(duplicate_files) <= 12 else f"; +{len(duplicate_files) - 12} more")
                + f" | sha1={digest}",
            )
        )

    for sku, counts in sorted(sku_counts.items()):
        if counts["white"] == 0 and counts["sku"] == 0:
            issues.append(issue(group, "missing_white_or_sku_in_group_folder", "blocker", sku=sku))
        if counts["main"] == 0:
            issues.append(issue(group, "missing_main_in_group_folder", "blocker", sku=sku))
        if counts["detail"] == 0:
            issues.append(issue(group, "missing_detail_in_group_folder", "blocker", sku=sku))

    summary = {
        "upload_group": group,
        "image_count": len(files),
        "sku_count": len(sku_counts),
        "white_count": role_counts["white"],
        "main_count": role_counts["main"],
        "sku_image_count": role_counts["sku"],
        "detail_count": role_counts["detail"],
        "exact_duplicate_set_count": duplicate_sets,
        "exact_duplicate_file_count": duplicate_file_count,
        "unreadable_image_count": unreadable_count,
        "top_dimensions": "; ".join(f"{size}:{count}" for size, count in dimension_counts.most_common(8)),
        "folder": group_folder.as_posix(),
    }
    return issues, summary


def audit_product_root(product_root: Path = PRODUCT_ROOT) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []
    for group_folder in sorted([path for path in product_root.iterdir() if path.is_dir()]):
        group_issues, summary = audit_group(group_folder)
        issues.extend(group_issues)
        summaries.append(summary)
    return issues, summaries


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def update_workbook(issues: list[dict[str, Any]], summaries: list[dict[str, Any]]) -> None:
    if not WORKBOOK_PATH.exists():
        return
    workbook = load_workbook(WORKBOOK_PATH)
    for sheet_name in ("ImageQualitySummary", "ImageQualityIssues"):
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_name, rows in (("ImageQualitySummary", summaries), ("ImageQualityIssues", issues)):
        sheet = workbook.create_sheet(sheet_name)
        if not rows:
            continue
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 70)
    workbook.save(WORKBOOK_PATH)


def run(update_xlsx: bool = True) -> dict[str, Any]:
    issues, summaries = audit_product_root(PRODUCT_ROOT)
    issue_csv = REPORTS_ROOT / "iblock-image-quality-issues.csv"
    summary_csv = REPORTS_ROOT / "iblock-image-quality-summary.csv"
    summary_json = REPORTS_ROOT / "iblock-image-quality-summary.json"
    write_csv(issue_csv, issues)
    write_csv(summary_csv, summaries)
    if update_xlsx:
        update_workbook(issues, summaries)
    blocker_count = sum(1 for row in issues if row["severity"] == "blocker")
    warning_count = sum(1 for row in issues if row["severity"] == "warning")
    duplicate_groups = sum(1 for row in summaries if row["exact_duplicate_set_count"])
    result = {
        "product_group_count": len(summaries),
        "image_count": sum(int(row["image_count"]) for row in summaries),
        "blocker_issue_count": blocker_count,
        "warning_issue_count": warning_count,
        "groups_with_exact_duplicates": duplicate_groups,
        "issue_csv": issue_csv.as_posix(),
        "summary_csv": summary_csv.as_posix(),
    }
    summary_json.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit iBlock product-group image quality before Shopify upload.")
    parser.add_argument("--no-xlsx", action="store_true", help="Do not update the readiness workbook.")
    args = parser.parse_args()
    run(update_xlsx=not args.no_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
