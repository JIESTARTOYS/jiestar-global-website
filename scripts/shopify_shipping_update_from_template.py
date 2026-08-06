#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import subprocess
import sys
import time
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


OUT_DIR = Path("/private/tmp/jiestar-shopify-shipping-update")
PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK_DIR = PROJECT_ROOT / "outputs" / "019fb5f3-8f25-7f02-9403-727aefd39ceb"
DEFAULT_AIR_RATE_WORKBOOK = DEFAULT_WORKBOOK_DIR / "Shopify普货空运结算模板_20260731.xlsx"
DEFAULT_SEA_RATE_WORKBOOK = DEFAULT_WORKBOOK_DIR / "Shopify普货海运结算模板_20260731.xlsx"
DEFAULT_WEIGHT_WORKBOOK = DEFAULT_WORKBOOK_DIR / "Shopify商品计费重量主表_20260731.xlsx"
API_VERSION_FALLBACK = "2026-01"

STANDARD_GOODS_PROFILE = "JIESTAR Standard goods"
MANUAL_REVIEW_PROFILE = "JIESTAR Manual Shipping Review"
LEGACY_BATTERY_PROFILE = "JIESTAR Battery/electric goods"
CHENGHAI_LOCATION_NAME = "JIESTAR Chenghai Warehouse"
TARGET_PROFILE_NAMES = [STANDARD_GOODS_PROFILE, MANUAL_REVIEW_PROFILE]
MANAGED_PROFILE_NAMES = [*TARGET_PROFILE_NAMES, LEGACY_BATTERY_PROFILE]
TARGET_COUNTRY_CODES = {"AU", "BE", "CA", "DE", "ES", "FR", "GB", "IT", "NL", "PL", "SE", "US"}
REQUIRED_SCOPES = {"read_products", "write_products", "read_locations", "read_shipping", "write_shipping"}
WEIGHTS_ONLY_REQUIRED_SCOPES = {"read_products", "write_products", "read_shipping", "write_shipping"}
VARIANT_ASSOCIATE_BATCH_SIZE = 250
EXPECTED_AIR_RATE_COUNT = 240
EXPECTED_SEA_RATE_COUNT = 31
EXPECTED_TOTAL_RATE_COUNT = EXPECTED_AIR_RATE_COUNT + EXPECTED_SEA_RATE_COUNT
APPROVAL_SCHEMA_VERSION = 2

SUMMARY_JSON = "shipping-update-summary.json"
APPLY_RESULT_JSON = "shipping-apply-result.json"
AIR_RATE_PLAN_CSV = "shipping-air-rate-plan.csv"
SEA_RATE_PLAN_CSV = "shipping-sea-rate-plan.csv"
WEIGHT_UPDATES_CSV = "shipping-weight-updates.csv"
PROFILE_ASSIGNMENTS_CSV = "shipping-profile-assignments.csv"
HEAVY_REVIEW_CSV = "shipping-blocked-heavy.csv"
ACTIVE_TO_DRAFT_CSV = "shipping-active-to-draft.csv"
DRAFT_BACKLOG_CSV = "shipping-draft-backlog.csv"
LEGACY_MIGRATION_CSV = "shipping-legacy-battery-migration.csv"
UNMATCHED_ACTIVE_CSV = "shipping-unmatched-active.csv"
API_DIFF_CSV = "shipping-api-diff.csv"


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_hash(value: Any) -> str:
    data = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(data).hexdigest()


def clean(value: Any) -> str:
    return str(value or "").strip()


def sku_key(value: Any) -> str:
    return clean(value).upper()


def dec(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def money_str(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP):.2f}"


def decimal_str(value: Decimal | None, places: str = "0.###") -> str:
    if value is None:
        return ""
    normalized = value.normalize()
    return format(normalized, "f")


def grams_from_weight(weight: dict[str, Any] | None) -> int | None:
    if not weight:
        return None
    value = dec(weight.get("value"))
    unit = clean(weight.get("unit")).upper()
    if value is None:
        return None
    multipliers = {
        "GRAMS": Decimal("1"),
        "KILOGRAMS": Decimal("1000"),
        "POUNDS": Decimal("453.59237"),
        "OUNCES": Decimal("28.349523125"),
    }
    multiplier = multipliers.get(unit)
    if multiplier is None:
        return None
    return int((value * multiplier).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def chunks(items: list[Any], size: int) -> Iterable[list[Any]]:
    for index in range(0, len(items), size):
        yield items[index : index + size]


def filter_variants_by_vendor(variants: list[ShopifyVariant], vendor: str | None) -> list[ShopifyVariant]:
    vendor_key = clean(vendor).casefold()
    if not vendor_key:
        return variants
    return [variant for variant in variants if clean(variant.vendor).casefold() == vendor_key]


def row_dicts(sheet: Any) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [clean(cell) for cell in next(rows)]
    except StopIteration:
        return []
    output: list[dict[str, Any]] = []
    for row in rows:
        if not row or not any(cell not in (None, "") for cell in row):
            continue
        output.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers)) if headers[idx]})
    return output


@dataclass(frozen=True)
class RateRow:
    source_profile: str
    target_profile: str
    country: str
    country_code: str
    service_type: str
    rate_name: str
    transit_time: str
    min_weight_kg: Decimal
    max_weight_kg: Decimal
    price_usd: Decimal
    freight_cost_rmb: Decimal | None


@dataclass(frozen=True)
class WeightTarget:
    sku: str
    sku_key: str
    handle: str
    title: str
    vendor: str
    target_weight_g: int | None
    source_profile: str
    target_profile: str
    weight_import_status: str
    dimension_verification: str
    shipping_status: str
    listing_status: str
    notes: str


@dataclass(frozen=True)
class ShopifyVariant:
    product_id: str
    product_handle: str
    product_title: str
    product_status: str
    vendor: str
    variant_id: str
    variant_title: str
    sku: str
    sku_key: str
    inventory_item_id: str
    inventory_tracked: bool | None
    requires_shipping: bool | None
    current_weight_g: int | None


class ShopifyAdmin:
    def __init__(self) -> None:
        load_dotenv(Path(".env.local"))
        self.domain = os.environ.get("SHOPIFY_STORE_DOMAIN", "").strip()
        self.version = os.environ.get("SHOPIFY_API_VERSION", API_VERSION_FALLBACK).strip() or API_VERSION_FALLBACK
        self.token = os.environ.get("SHOPIFY_ADMIN_ACCESS_TOKEN", "").strip()
        if not self.domain:
            raise RuntimeError("Missing SHOPIFY_STORE_DOMAIN in .env.local")
        if not self.token:
            raise RuntimeError("Missing SHOPIFY_ADMIN_ACCESS_TOKEN in .env.local")
        self.endpoint = f"https://{self.domain}/admin/api/{self.version}/graphql.json"

    def graphql(self, query: str, variables: dict[str, Any] | None = None) -> dict[str, Any]:
        body = json.dumps({"query": query, "variables": variables or {}}).encode("utf-8")
        for attempt in range(6):
            request = urllib.request.Request(
                self.endpoint,
                data=body,
                method="POST",
                headers={"Content-Type": "application/json", "X-Shopify-Access-Token": self.token},
            )
            try:
                with urllib.request.urlopen(request, timeout=120) as response:
                    payload = json.loads(response.read().decode("utf-8"))
            except urllib.error.HTTPError as error:
                error_body = error.read().decode("utf-8", errors="ignore")
                if error.code in {429, 500, 502, 503, 504} and attempt < 5:
                    time.sleep(min(30, 2**attempt))
                    continue
                if error.code == 404:
                    payload = self._graphql_via_curl(body)
                else:
                    raise RuntimeError(f"Shopify HTTP {error.code}: {error_body[:1200]}") from error
            except (urllib.error.URLError, TimeoutError, OSError):
                payload = self._graphql_via_curl(body)

            errors = payload.get("errors")
            if errors:
                error_text = json.dumps(errors, ensure_ascii=False)
                retryable = any(
                    str(error.get("extensions", {}).get("code", "")).upper() == "THROTTLED"
                    or "throttl" in str(error).lower()
                    or "timeout" in str(error).lower()
                    for error in errors
                    if isinstance(error, dict)
                )
                if retryable and attempt < 5:
                    time.sleep(min(30, 2**attempt))
                    continue
                raise RuntimeError(f"Shopify GraphQL errors: {error_text}")
            return payload["data"]
        raise RuntimeError("Shopify GraphQL retry limit exceeded")

    def _graphql_via_curl(self, body: bytes) -> dict[str, Any]:
        result = subprocess.run(
            [
                "curl",
                "--silent",
                "--show-error",
                "--fail-with-body",
                "--http1.1",
                "--connect-timeout",
                "20",
                "--max-time",
                "120",
                "--retry",
                "6",
                "--retry-all-errors",
                "--request",
                "POST",
                "--header",
                "Content-Type: application/json",
                "--header",
                f"X-Shopify-Access-Token: {self.token}",
                "--data-binary",
                "@-",
                self.endpoint,
            ],
            input=body,
            capture_output=True,
            timeout=180,
        )
        response_body = result.stdout.decode("utf-8", errors="ignore")
        if result.returncode:
            message = result.stderr.decode("utf-8", errors="ignore")
            raise RuntimeError(f"Shopify curl request failed: {(message or response_body)[:1200]}")
        return json.loads(response_body)

    def access_scopes(self) -> set[str]:
        data = self.graphql(
            """
            query AppScopes {
              currentAppInstallation {
                accessScopes {
                  handle
                }
              }
            }
            """
        )
        return {scope["handle"] for scope in data["currentAppInstallation"]["accessScopes"]}

    def variants_by_product_status(self, status: str) -> list[ShopifyVariant]:
        normalized_status = clean(status).lower()
        if normalized_status not in {"active", "draft"}:
            raise ValueError(f"Unsupported Shopify product status: {status}")
        query = """
        query VariantsForShippingAudit($cursor: String, $query: String!) {
          products(first: 250, after: $cursor, query: $query, sortKey: TITLE) {
            pageInfo {
              hasNextPage
              endCursor
            }
            nodes {
              id
              handle
              title
              status
              vendor
              variants(first: 250) {
                nodes {
                  id
                  title
                  sku
                  inventoryItem {
                    id
                    tracked
                    requiresShipping
                    measurement {
                      weight {
                        value
                        unit
                      }
                    }
                  }
                }
              }
            }
          }
        }
        """
        variants: list[ShopifyVariant] = []
        cursor = None
        while True:
            data = self.graphql(query, {"cursor": cursor, "query": f"status:{normalized_status}"})
            page = data["products"]
            for product in page["nodes"]:
                for variant in product.get("variants", {}).get("nodes", []):
                    inventory_item = variant.get("inventoryItem") or {}
                    sku = clean(variant.get("sku"))
                    variants.append(
                        ShopifyVariant(
                            product_id=product["id"],
                            product_handle=clean(product.get("handle")),
                            product_title=clean(product.get("title")),
                            product_status=clean(product.get("status")),
                            vendor=clean(product.get("vendor")),
                            variant_id=variant["id"],
                            variant_title=clean(variant.get("title")),
                            sku=sku,
                            sku_key=sku_key(sku),
                            inventory_item_id=inventory_item.get("id") or "",
                            inventory_tracked=inventory_item.get("tracked"),
                            requires_shipping=inventory_item.get("requiresShipping"),
                            current_weight_g=grams_from_weight((inventory_item.get("measurement") or {}).get("weight")),
                        )
                    )
            if not page["pageInfo"]["hasNextPage"]:
                break
            cursor = page["pageInfo"]["endCursor"]
        return variants

    def active_variants(self) -> list[ShopifyVariant]:
        return self.variants_by_product_status("active")

    def draft_variants(self) -> list[ShopifyVariant]:
        return self.variants_by_product_status("draft")

    def locations(self) -> list[dict[str, Any]]:
        data = self.graphql(
            """
            query LocationsForDeliveryProfile {
              locations(first: 50) {
                nodes {
                  id
                  name
                  isActive
                  fulfillsOnlineOrders
                }
              }
            }
            """
        )
        return data["locations"]["nodes"]

    def delivery_profiles(self, include_profile_items: bool = False) -> list[dict[str, Any]]:
        query = """
        query DeliveryProfileSnapshot($cursor: String) {
          deliveryProfiles(first: 5, after: $cursor) {
            pageInfo {
              hasNextPage
              endCursor
            }
            nodes {
              id
              name
              default
              legacyMode
              productVariantsCount {
                count
              }
              profileLocationGroups {
                locationGroup {
                  id
                  locations(first: 5) {
                    nodes {
                      id
                      name
                    }
                  }
                }
                locationGroupZones(first: 30) {
                  nodes {
                    zone {
                      id
                      name
                      countries {
                        code {
                          countryCode
                        }
                      }
                    }
                    methodDefinitions(first: 50) {
                      nodes {
                        id
                        name
                        active
                        description
                        rateProvider {
                          __typename
                          ... on DeliveryRateDefinition {
                            id
                            price {
                              amount
                              currencyCode
                            }
                          }
                        }
                        methodConditions {
                          id
                          field
                          operator
                          conditionCriteria {
                            __typename
                            ... on Weight {
                              value
                              unit
                            }
                            ... on MoneyV2 {
                              amount
                              currencyCode
                            }
                          }
                        }
                      }
                    }
                  }
                }
              }
            }
          }
        }
        """
        profiles: list[dict[str, Any]] = []
        cursor = None
        while True:
            data = self.graphql(query, {"cursor": cursor})
            page = data["deliveryProfiles"]
            profiles.extend(page["nodes"])
            if not page["pageInfo"]["hasNextPage"]:
                break
            cursor = page["pageInfo"]["endCursor"]
        if include_profile_items:
            for profile in profiles:
                if clean(profile.get("name")) in MANAGED_PROFILE_NAMES:
                    profile["profileItemVariantIds"] = self.delivery_profile_variant_ids(profile["id"])
        return profiles

    def delivery_profile_variant_ids(self, profile_id: str) -> list[str]:
        query = """
        query DeliveryProfileItems($id: ID!, $cursor: String) {
          node(id: $id) {
            ... on DeliveryProfile {
              profileItems(first: 250, after: $cursor) {
                pageInfo {
                  hasNextPage
                  endCursor
                }
                nodes {
                  variants(first: 250) {
                    nodes {
                      id
                    }
                  }
                }
              }
            }
          }
        }
        """
        variant_ids: list[str] = []
        cursor = None
        while True:
            data = self.graphql(query, {"id": profile_id, "cursor": cursor})
            page = data["node"]["profileItems"]
            for item in page["nodes"]:
                variant_ids.extend(variant["id"] for variant in item.get("variants", {}).get("nodes", []))
            if not page["pageInfo"]["hasNextPage"]:
                break
            cursor = page["pageInfo"]["endCursor"]
        return sorted(set(variant_ids))

    def variants_bulk_update(self, product_id: str, variants: list[dict[str, Any]]) -> list[dict[str, Any]]:
        if not variants:
            return []
        data = self.graphql(
            """
            mutation ProductVariantsBulkUpdate($productId: ID!, $variants: [ProductVariantsBulkInput!]!) {
              productVariantsBulkUpdate(productId: $productId, variants: $variants, allowPartialUpdates: true) {
                product {
                  id
                }
                productVariants {
                  id
                  sku
                  inventoryItem {
                    id
                    requiresShipping
                    measurement {
                      weight {
                        value
                        unit
                      }
                    }
                  }
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"productId": product_id, "variants": variants},
        )
        return data["productVariantsBulkUpdate"]["userErrors"]

    def delivery_profile_create(self, profile: dict[str, Any]) -> dict[str, Any]:
        data = self.graphql(
            """
            mutation DeliveryProfileCreate($profile: DeliveryProfileInput!) {
              deliveryProfileCreate(profile: $profile) {
                profile {
                  id
                  name
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"profile": profile},
        )
        return data["deliveryProfileCreate"]

    def delivery_profile_update(self, profile_id: str, profile: dict[str, Any]) -> dict[str, Any]:
        data = self.graphql(
            """
            mutation DeliveryProfileUpdate($id: ID!, $profile: DeliveryProfileInput!) {
              deliveryProfileUpdate(id: $id, profile: $profile) {
                profile {
                  id
                  name
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"id": profile_id, "profile": profile},
        )
        return data["deliveryProfileUpdate"]

    def product_update_status(self, product_id: str, status: str) -> dict[str, Any]:
        data = self.graphql(
            """
            mutation UpdateProductStatus($product: ProductUpdateInput!) {
              productUpdate(product: $product) {
                product {
                  id
                  status
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"product": {"id": product_id, "status": status}},
        )
        return data["productUpdate"]

    def delivery_profile_remove(self, profile_id: str) -> dict[str, Any]:
        data = self.graphql(
            """
            mutation DeliveryProfileRemove($id: ID!) {
              deliveryProfileRemove(id: $id) {
                job {
                  id
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"id": profile_id},
        )
        return data["deliveryProfileRemove"]


def load_rate_workbook(path: Path, mode: str) -> list[RateRow]:
    normalized_mode = clean(mode).lower()
    if normalized_mode not in {"air", "sea"}:
        raise ValueError(f"Unsupported shipping mode: {mode}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    required_sheets = {"Inputs", "Rate_Source", "Shopify_Rates", "QA"}
    missing = required_sheets - set(workbook.sheetnames)
    if missing:
        raise RuntimeError(f"{normalized_mode.title()} rate workbook missing sheets: {sorted(missing)}")

    expected_count = EXPECTED_AIR_RATE_COUNT if normalized_mode == "air" else EXPECTED_SEA_RATE_COUNT
    expected_rate_name = "Air Shipping" if normalized_mode == "air" else "Sea Shipping"
    expected_countries = TARGET_COUNTRY_CODES if normalized_mode == "air" else {"AU", "US"}
    rate_rows: list[RateRow] = []
    for row in row_dicts(workbook["Shopify_Rates"]):
        if clean(row.get("Active")).lower() != "active":
            continue
        row_text = " ".join(clean(value) for value in row.values()).lower()
        if "battery/electric" in row_text or "battery-capable" in row_text:
            raise RuntimeError(f"Battery/electric rate text is prohibited in {normalized_mode} workbook: {row}")
        source_profile = clean(row.get("Shipping Profile"))
        if source_profile != "Standard goods":
            raise RuntimeError(f"Only Standard goods rates are accepted; found {source_profile!r}")
        country_code = clean(row.get("Country Code")).upper()
        if country_code not in expected_countries:
            raise RuntimeError(f"Unexpected {normalized_mode} country code: {country_code}")
        rate_name = clean(row.get("Rate Name"))
        if rate_name != expected_rate_name:
            raise RuntimeError(f"{normalized_mode.title()} checkout rate name must be {expected_rate_name!r}; found {rate_name!r}")
        service_type = clean(row.get("Service Type"))
        if normalized_mode not in service_type.lower():
            raise RuntimeError(f"{normalized_mode.title()} rate row has mismatched service type: {service_type!r}")
        min_weight = dec(row.get("Min Weight kg"))
        max_weight = dec(row.get("Max Weight kg"))
        price = dec(row.get("Price USD"))
        freight_cost_rmb = dec(row.get("Freight Cost RMB"))
        if min_weight is None or max_weight is None or price is None or freight_cost_rmb is None:
            raise RuntimeError(f"Invalid {normalized_mode} rate row for {country_code}: {row}")
        rate_rows.append(
            RateRow(
                source_profile=normalized_mode.title(),
                target_profile=STANDARD_GOODS_PROFILE,
                country=clean(row.get("Zone Country")),
                country_code=country_code,
                service_type=service_type,
                rate_name=rate_name,
                transit_time=clean(row.get("Transit Time")),
                min_weight_kg=min_weight,
                max_weight_kg=max_weight,
                price_usd=price,
                freight_cost_rmb=freight_cost_rmb,
            )
        )

    if len(rate_rows) != expected_count:
        raise RuntimeError(f"Expected exactly {expected_count} {normalized_mode} rates, found {len(rate_rows)}")
    country_counts = Counter(row.country_code for row in rate_rows)
    expected_country_counts = (
        {country_code: 20 for country_code in TARGET_COUNTRY_CODES}
        if normalized_mode == "air"
        else {"US": 20, "AU": 11}
    )
    if country_counts != Counter(expected_country_counts):
        raise RuntimeError(
            f"{normalized_mode.title()} country/tier counts do not match the approved plan: "
            f"expected={dict(sorted(expected_country_counts.items()))}; actual={dict(sorted(country_counts.items()))}"
        )
    if any(row.max_weight_kg > Decimal("10") for row in rate_rows):
        raise RuntimeError(f"{normalized_mode.title()} workbook contains a rate over 10kg")
    return rate_rows


def load_weight_workbook(path: Path) -> dict[str, WeightTarget]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    required_sheets = {"Inputs", "Product_Chargeable_Weight", "Shopify商品重量导入", "QA"}
    missing = required_sheets - set(workbook.sheetnames)
    if missing:
        raise RuntimeError(f"Weight workbook missing sheets: {sorted(missing)}")

    target_by_sku: dict[str, WeightTarget] = {}
    duplicate_conflicts: list[str] = []
    for row in row_dicts(workbook["Shopify商品重量导入"]):
        sku = clean(row.get("Variant SKU"))
        key = sku_key(sku)
        if not key:
            continue
        source_profile = clean(row.get("Shipping Profile Suggestion"))
        if "battery" in source_profile.lower() or "electric" in source_profile.lower():
            raise RuntimeError(f"Battery/electric profile routing is prohibited for SKU {sku}")
        weight_status = clean(row.get("Weight Import Status"))
        if weight_status not in {"Yes", "Review", "No"}:
            raise RuntimeError(f"Unsupported Weight Import Status for SKU {sku}: {weight_status!r}")
        target_profile = ""
        if weight_status == "Yes":
            if source_profile != "Standard goods":
                raise RuntimeError(f"Eligible SKU {sku} must use Standard goods; found {source_profile!r}")
            target_profile = STANDARD_GOODS_PROFILE
        elif weight_status == "Review":
            if source_profile not in {"Manual review", ""}:
                raise RuntimeError(f"Review SKU {sku} must use Manual review; found {source_profile!r}")
            target_profile = MANUAL_REVIEW_PROFILE

        target_weight = row.get("Variant Weight")
        target_weight_g: int | None
        try:
            target_weight_g = int(Decimal(str(target_weight)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        except (InvalidOperation, TypeError, ValueError):
            target_weight_g = None
        dimension_verification = clean(row.get("Dimension Verification"))
        if weight_status in {"Yes", "Review"}:
            if target_weight_g is None or target_weight_g <= 0:
                raise RuntimeError(f"Eligible/review SKU {sku} has no positive Shopify weight")
            if "no +2cm buffer" not in dimension_verification.lower():
                raise RuntimeError(f"SKU {sku} is not verified against the no-+2cm dimension rule")

        target = WeightTarget(
            sku=sku,
            sku_key=key,
            handle=clean(row.get("Handle")),
            title=clean(row.get("Title")),
            vendor=clean(row.get("Vendor")),
            target_weight_g=target_weight_g,
            source_profile=source_profile,
            target_profile=target_profile,
            weight_import_status=weight_status,
            dimension_verification=dimension_verification,
            shipping_status=clean(row.get("Shipping Status")),
            listing_status=clean(row.get("Listing Status")),
            notes=clean(row.get("Notes")),
        )
        if key in target_by_sku:
            if target_by_sku[key] != target:
                duplicate_conflicts.append(key)
            continue
        target_by_sku[key] = target
    if duplicate_conflicts:
        raise RuntimeError(f"Conflicting duplicate SKU rows in weight workbook: {sorted(set(duplicate_conflicts))[:20]}")
    return target_by_sku


def load_shipping_workbooks(
    air_rate_workbook: Path,
    sea_rate_workbook: Path,
    weight_workbook: Path,
) -> tuple[list[RateRow], list[RateRow], dict[str, WeightTarget]]:
    air_rates = load_rate_workbook(air_rate_workbook, "air")
    sea_rates = load_rate_workbook(sea_rate_workbook, "sea")
    if len(air_rates) + len(sea_rates) != EXPECTED_TOTAL_RATE_COUNT:
        raise RuntimeError("Combined standard-goods rate count must be exactly 271")
    return air_rates, sea_rates, load_weight_workbook(weight_workbook)


def rate_csv_rows(rate_rows: list[RateRow]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in rate_rows:
        rows.append(
            {
                "target_profile": row.target_profile,
                "source_profile": row.source_profile,
                "zone_country": row.country,
                "country_code": row.country_code,
                "service_type": row.service_type,
                "rate_name": row.rate_name,
                "transit_time": row.transit_time,
                "min_weight_kg": decimal_str(row.min_weight_kg),
                "max_weight_kg": decimal_str(row.max_weight_kg),
                "price_usd": money_str(row.price_usd),
                "freight_cost_rmb": decimal_str(row.freight_cost_rmb),
                "active": "Active",
            }
        )
    return rows


def build_shipping_report(
    variants: list[ShopifyVariant],
    targets: dict[str, WeightTarget],
) -> tuple[
    list[dict[str, str]],
    list[dict[str, str]],
    list[dict[str, str]],
    list[dict[str, str]],
    list[dict[str, str]],
]:
    weight_updates: list[dict[str, str]] = []
    assignments: list[dict[str, str]] = []
    heavy_review: list[dict[str, str]] = []
    draft_actions: list[dict[str, str]] = []
    unmatched: list[dict[str, str]] = []

    invalid_by_variant: dict[str, tuple[str, WeightTarget | None]] = {}
    for variant in variants:
        if not variant.sku_key:
            invalid_by_variant[variant.variant_id] = ("missing_variant_sku", None)
            continue
        target = targets.get(variant.sku_key)
        if not target:
            invalid_by_variant[variant.variant_id] = ("unmatched_weight_workbook", None)
            continue
        if target.weight_import_status == "No":
            invalid_by_variant[variant.variant_id] = ("unverified_weight_or_dimensions", target)
            continue
        target_weight_g = target.target_weight_g
        if target_weight_g is None or target_weight_g <= 0:
            invalid_by_variant[variant.variant_id] = ("invalid_weight_workbook_value", target)
            continue

    draft_product_ids = {
        variant.product_id
        for variant in variants
        if variant.variant_id in invalid_by_variant
    }
    variants_by_product: dict[str, list[ShopifyVariant]] = defaultdict(list)
    for variant in variants:
        variants_by_product[variant.product_id].append(variant)

    for product_id in sorted(draft_product_ids):
        product_variants = variants_by_product[product_id]
        invalid_variants = [
            variant
            for variant in product_variants
            if variant.variant_id in invalid_by_variant
        ]
        reasons = sorted({invalid_by_variant[variant.variant_id][0] for variant in invalid_variants})
        draft_actions.append(
            {
                "action": "change_product_status_to_draft",
                "reason": ";".join(reasons),
                "product_id": product_id,
                "handle": product_variants[0].product_handle,
                "product_title": product_variants[0].product_title,
                "vendor": product_variants[0].vendor,
                "product_variant_count": str(len(product_variants)),
                "invalid_variant_count": str(len(invalid_variants)),
                "invalid_skus": ";".join(sorted({variant.sku for variant in invalid_variants if variant.sku})),
            }
        )
        for variant in product_variants:
            reason, target = invalid_by_variant.get(
                variant.variant_id,
                ("sibling_variant_requires_product_draft", targets.get(variant.sku_key)),
            )
            unmatched.append(
                unmatched_row(
                    variant,
                    reason,
                    target,
                    planned_action="change_product_status_to_draft",
                )
            )

    for variant in variants:
        if variant.product_id in draft_product_ids:
            continue
        target = targets[variant.sku_key]
        target_weight_g = target.target_weight_g
        if target_weight_g is None:
            raise RuntimeError(f"Internal error: validated target weight missing for {variant.sku}")

        manual_review = (
            target.weight_import_status == "Review"
            or target.target_profile == MANUAL_REVIEW_PROFILE
            or target_weight_g > 10000
        )
        target_profile = MANUAL_REVIEW_PROFILE if manual_review else STANDARD_GOODS_PROFILE
        weight_action = "noop" if variant.current_weight_g == target_weight_g and variant.requires_shipping is True else "update"
        reason = "manual_review_no_checkout" if manual_review else "eligible_weight_and_profile_update"
        update_row = {
            "action": weight_action,
            "reason": reason,
            "product_id": variant.product_id,
            "variant_id": variant.variant_id,
            "inventory_item_id": variant.inventory_item_id,
            "handle": variant.product_handle,
            "product_title": variant.product_title,
            "vendor": variant.vendor,
            "variant_title": variant.variant_title,
            "sku": variant.sku,
            "current_weight_g": "" if variant.current_weight_g is None else str(variant.current_weight_g),
            "target_weight_g": str(target_weight_g),
            "current_requires_shipping": "" if variant.requires_shipping is None else str(variant.requires_shipping),
            "target_requires_shipping": "True",
            "source_profile": target.source_profile,
            "target_profile": target_profile,
            "weight_import_status": target.weight_import_status,
            "dimension_verification": target.dimension_verification,
            "shipping_status": target.shipping_status,
            "template_handle": target.handle,
            "template_title": target.title,
        }
        weight_updates.append(update_row)
        assignments.append(
            {
                "action": "assign",
                "reason": reason,
                "product_id": variant.product_id,
                "variant_id": variant.variant_id,
                "handle": variant.product_handle,
                "product_title": variant.product_title,
                "vendor": variant.vendor,
                "sku": variant.sku,
                "target_profile": target_profile,
                "target_weight_g": str(target_weight_g),
                "shipping_status": target.shipping_status,
            }
        )
        if manual_review:
            heavy_review.append(update_row)

    return weight_updates, assignments, heavy_review, draft_actions, unmatched


def build_scoped_shipping_report(
    variants: list[ShopifyVariant],
    targets: dict[str, WeightTarget],
    *,
    scope_to_targets: bool,
) -> tuple[
    list[dict[str, str]],
    list[dict[str, str]],
    list[dict[str, str]],
    list[dict[str, str]],
    list[dict[str, str]],
]:
    if not scope_to_targets:
        return build_shipping_report(variants, targets)

    scoped_variants = [variant for variant in variants if variant.sku_key in targets]
    outside_scope = [variant for variant in variants if variant.sku_key not in targets]
    weight_updates, assignments, heavy_review, draft_actions, unmatched = build_shipping_report(scoped_variants, targets)
    unmatched.extend(
        unmatched_row(
            variant,
            "outside_scoped_weight_workbook",
            planned_action="report_only_no_write",
        )
        for variant in outside_scope
    )
    return weight_updates, assignments, heavy_review, draft_actions, unmatched


def unmatched_row(
    variant: ShopifyVariant,
    reason: str,
    target: WeightTarget | None = None,
    planned_action: str = "",
) -> dict[str, str]:
    return {
        "skip_reason": reason,
        "planned_action": planned_action,
        "product_id": variant.product_id,
        "variant_id": variant.variant_id,
        "inventory_item_id": variant.inventory_item_id,
        "handle": variant.product_handle,
        "product_title": variant.product_title,
        "vendor": variant.vendor,
        "variant_title": variant.variant_title,
        "sku": variant.sku,
        "current_weight_g": "" if variant.current_weight_g is None else str(variant.current_weight_g),
        "current_requires_shipping": "" if variant.requires_shipping is None else str(variant.requires_shipping),
        "template_weight_status": target.weight_import_status if target else "",
        "template_profile": target.target_profile if target else "",
        "shipping_status": target.shipping_status if target else "",
    }


def build_draft_backlog(
    variants: list[ShopifyVariant],
    targets: dict[str, WeightTarget],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for variant in variants:
        target = targets.get(variant.sku_key)
        rows.append(
            {
                "action": "report_only_no_write",
                "product_id": variant.product_id,
                "variant_id": variant.variant_id,
                "handle": variant.product_handle,
                "product_title": variant.product_title,
                "vendor": variant.vendor,
                "variant_title": variant.variant_title,
                "sku": variant.sku,
                "current_weight_g": "" if variant.current_weight_g is None else str(variant.current_weight_g),
                "source_match": "matched" if target else "unmatched",
                "source_weight_status": target.weight_import_status if target else "",
                "source_target_weight_g": "" if not target or target.target_weight_g is None else str(target.target_weight_g),
                "note": "Draft variants are never updated by this workflow.",
            }
        )
    return rows


def build_legacy_migration_report(
    active_variants: list[ShopifyVariant],
    draft_variants: list[ShopifyVariant],
    assignments: list[dict[str, str]],
    draft_actions: list[dict[str, str]],
    delivery_snapshot: dict[str, Any],
) -> list[dict[str, str]]:
    legacy_ids = set(delivery_snapshot.get("legacy_battery_variant_ids") or [])
    variants_by_id = {variant.variant_id: variant for variant in [*active_variants, *draft_variants]}
    assignments_by_id = {row["variant_id"]: row for row in assignments}
    draft_product_ids = {row["product_id"] for row in draft_actions}
    rows: list[dict[str, str]] = []
    for variant_id in sorted(legacy_ids):
        variant = variants_by_id.get(variant_id)
        assignment = assignments_by_id.get(variant_id)
        target_profile = assignment["target_profile"] if assignment else ""
        if assignment:
            action = "migrate_to_standard_profile" if target_profile == STANDARD_GOODS_PROFILE else "migrate_to_manual_profile"
        elif variant and variant.product_id in draft_product_ids:
            action = "change_product_to_draft"
        elif variant and clean(variant.product_status).upper() == "DRAFT":
            action = "already_draft_report_only"
        else:
            action = "blocked_unmatched_legacy_variant"
        rows.append(
            {
                "action": action,
                "variant_id": variant_id,
                "product_id": variant.product_id if variant else "",
                "handle": variant.product_handle if variant else "",
                "product_title": variant.product_title if variant else "",
                "vendor": variant.vendor if variant else "",
                "sku": variant.sku if variant else "",
                "current_profile": LEGACY_BATTERY_PROFILE,
                "target_profile": target_profile,
            }
        )
    return rows


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str] | None = None) -> None:
    if fieldnames is None:
        keys: list[str] = []
        seen = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    keys.append(key)
                    seen.add(key)
        fieldnames = keys
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def expected_rate_key(row: RateRow) -> tuple[str, str, str, str, str]:
    return (
        row.country_code,
        row.rate_name,
        decimal_str(row.min_weight_kg),
        decimal_str(row.max_weight_kg),
        money_str(row.price_usd),
    )


def location_group_has_location(
    profile_location_group: dict[str, Any],
    *,
    location_id: str | None = None,
    location_name: str | None = None,
) -> bool:
    locations = (((profile_location_group.get("locationGroup") or {}).get("locations") or {}).get("nodes") or [])
    return any(
        (not location_id or location.get("id") == location_id)
        and (not location_name or clean(location.get("name")) == location_name)
        for location in locations
    )


def profile_rate_keys(
    profile: dict[str, Any] | None,
    *,
    location_id: str | None = None,
    location_name: str | None = None,
) -> set[tuple[str, str, str, str, str]]:
    keys: set[tuple[str, str, str, str, str]] = set()
    if not profile:
        return keys
    for location_group in profile.get("profileLocationGroups") or []:
        if (location_id or location_name) and not location_group_has_location(
            location_group,
            location_id=location_id,
            location_name=location_name,
        ):
            continue
        zones = ((location_group.get("locationGroupZones") or {}).get("nodes") or [])
        for zone_node in zones:
            country_codes = [
                clean(((country.get("code") or {}).get("countryCode"))).upper()
                for country in ((zone_node.get("zone") or {}).get("countries") or [])
            ]
            for method in ((zone_node.get("methodDefinitions") or {}).get("nodes") or []):
                provider = method.get("rateProvider") or {}
                price = dec(((provider.get("price") or {}).get("amount")))
                if price is None:
                    continue
                min_weight: Decimal | None = None
                max_weight: Decimal | None = None
                for condition in method.get("methodConditions") or []:
                    criteria = condition.get("conditionCriteria") or {}
                    if clean(criteria.get("__typename")) != "Weight":
                        continue
                    value = dec(criteria.get("value"))
                    operator = clean(condition.get("operator")).upper()
                    if "GREATER" in operator:
                        min_weight = value
                    elif "LESS" in operator:
                        max_weight = value
                if min_weight is None or max_weight is None:
                    continue
                for country_code in country_codes:
                    keys.add(
                        (
                            country_code,
                            clean(method.get("name")),
                            decimal_str(min_weight),
                            decimal_str(max_weight),
                            money_str(price),
                        )
                    )
    return keys


def delivery_access_snapshot(
    admin: ShopifyAdmin,
    rate_rows: list[RateRow],
) -> tuple[list[dict[str, str]], dict[str, Any]]:
    rows: list[dict[str, str]] = []
    snapshot: dict[str, Any] = {"profiles_accessible": False, "locations_accessible": False, "errors": []}

    try:
        locations = admin.locations()
        snapshot["locations_accessible"] = True
        snapshot["location_count"] = len(locations)
        snapshot["locations"] = [
            {
                "id": location.get("id"),
                "name": location.get("name"),
                "isActive": location.get("isActive"),
                "fulfillsOnlineOrders": location.get("fulfillsOnlineOrders"),
            }
            for location in locations
        ]
        rows.append({"area": "locations", "status": "read_ok", "name": "", "id": "", "detail": str(len(locations))})
    except Exception as error:
        snapshot["errors"].append({"area": "locations", "error": str(error)})
        rows.append({"area": "locations", "status": "read_error", "name": "", "id": "", "detail": str(error)})

    try:
        profiles = admin.delivery_profiles(include_profile_items=True)
        snapshot["profiles_accessible"] = True
        snapshot["profile_count"] = len(profiles)
        target_profiles = []
        profiles_by_name: dict[str, dict[str, Any]] = {}
        for profile in profiles:
            name = clean(profile.get("name"))
            profiles_by_name[name] = profile
            method_count = 0
            zone_count = 0
            for location_group in profile.get("profileLocationGroups") or []:
                zones = ((location_group.get("locationGroupZones") or {}).get("nodes") or [])
                zone_count += len(zones)
                for zone in zones:
                    method_count += len(((zone.get("methodDefinitions") or {}).get("nodes") or []))
            if name in MANAGED_PROFILE_NAMES:
                target_profiles.append(
                    {
                        "id": profile.get("id"),
                        "name": name,
                        "zone_count": zone_count,
                        "method_count": method_count,
                    }
                )
            rows.append(
                {
                    "area": "delivery_profile",
                    "status": "read_ok",
                    "name": name,
                    "id": profile.get("id") or "",
                    "detail": f"zones={zone_count}; methods={method_count}; variants={((profile.get('productVariantsCount') or {}).get('count'))}",
                }
            )
        snapshot["target_profiles"] = target_profiles
        legacy_profile = profiles_by_name.get(LEGACY_BATTERY_PROFILE)
        snapshot["legacy_battery_profile_exists"] = legacy_profile is not None
        snapshot["legacy_battery_profile_id"] = legacy_profile.get("id") if legacy_profile else None
        snapshot["legacy_battery_variant_ids"] = (
            legacy_profile.get("profileItemVariantIds") or []
            if legacy_profile
            else []
        )
        snapshot["legacy_battery_variant_count"] = len(snapshot["legacy_battery_variant_ids"])

        expected_keys = {expected_rate_key(row) for row in rate_rows}
        standard_profile = profiles_by_name.get(STANDARD_GOODS_PROFILE)
        standard_actual_keys = profile_rate_keys(
            standard_profile,
            location_name=CHENGHAI_LOCATION_NAME,
        )
        missing_rate_keys = sorted(expected_keys - standard_actual_keys)
        unexpected_rate_keys = sorted(standard_actual_keys - expected_keys)
        for key in missing_rate_keys:
            rows.append(
                {
                    "area": "rate_diff",
                    "status": "missing",
                    "name": STANDARD_GOODS_PROFILE,
                    "id": "",
                    "detail": "|".join(key),
                }
            )
        for key in unexpected_rate_keys:
            rows.append(
                {
                    "area": "rate_diff",
                    "status": "unexpected",
                    "name": STANDARD_GOODS_PROFILE,
                    "id": "",
                    "detail": "|".join(key),
                }
            )
        manual_profile = profiles_by_name.get(MANUAL_REVIEW_PROFILE)
        manual_rate_keys = profile_rate_keys(
            manual_profile,
            location_name=CHENGHAI_LOCATION_NAME,
        )
        for key in sorted(manual_rate_keys):
            rows.append(
                {
                    "area": "rate_diff",
                    "status": "manual_profile_must_have_zero_rates",
                    "name": MANUAL_REVIEW_PROFILE,
                    "id": "",
                    "detail": "|".join(key),
                }
            )
        snapshot["expected_standard_rate_count"] = len(expected_keys)
        snapshot["actual_standard_rate_count"] = len(standard_actual_keys)
        snapshot["missing_standard_rate_count"] = len(missing_rate_keys)
        snapshot["unexpected_standard_rate_count"] = len(unexpected_rate_keys)
        snapshot["manual_profile_rate_count"] = len(manual_rate_keys)
        snapshot["rate_diff_count"] = len(missing_rate_keys) + len(unexpected_rate_keys) + len(manual_rate_keys)
    except Exception as error:
        snapshot["errors"].append({"area": "deliveryProfiles", "error": str(error)})
        rows.append({"area": "deliveryProfiles", "status": "read_error", "name": "", "id": "", "detail": str(error)})

    return rows, snapshot


def build_summary(
    air_rate_workbook: Path,
    sea_rate_workbook: Path,
    weight_workbook: Path,
    out_dir: Path,
    admin: ShopifyAdmin,
    scopes: set[str],
    air_rate_rows: list[RateRow],
    sea_rate_rows: list[RateRow],
    active_variants: list[ShopifyVariant],
    draft_variants: list[ShopifyVariant],
    targets: dict[str, WeightTarget],
    weight_updates: list[dict[str, str]],
    assignments: list[dict[str, str]],
    heavy_review: list[dict[str, str]],
    draft_actions: list[dict[str, str]],
    draft_backlog: list[dict[str, str]],
    legacy_migration: list[dict[str, str]],
    unmatched: list[dict[str, str]],
    delivery_snapshot: dict[str, Any],
    applied: bool,
    default_package_zero_verified: bool,
    vendor_filter: str | None = None,
    skip_rate_sync: bool = False,
) -> dict[str, Any]:
    rate_rows = [*air_rate_rows, *sea_rate_rows]
    rate_profile_counts = Counter(row.target_profile for row in rate_rows)
    rate_country_counts = Counter(row.country_code for row in rate_rows)
    target_weight_counts = Counter(target.weight_import_status for target in targets.values())
    assignment_profile_counts = Counter(row["target_profile"] for row in assignments)
    unmatched_reason_counts = Counter(row["skip_reason"] for row in unmatched)
    weight_action_counts = Counter(row["action"] for row in weight_updates)
    required_scopes = WEIGHTS_ONLY_REQUIRED_SCOPES if skip_rate_sync else REQUIRED_SCOPES
    missing_scopes = sorted(required_scopes - scopes)
    workbook_paths = {
        "air_rate": str(air_rate_workbook),
        "sea_rate": str(sea_rate_workbook),
        "weight": str(weight_workbook),
    }
    workbook_sha256s = {
        "air_rate": sha256_file(air_rate_workbook),
        "sea_rate": sha256_file(sea_rate_workbook),
        "weight": sha256_file(weight_workbook),
    }
    blocked_legacy_rows = [row for row in legacy_migration if row["action"] == "blocked_unmatched_legacy_variant"]
    approval_payload = {
        "approval_schema_version": APPROVAL_SCHEMA_VERSION,
        "workbooks": workbook_paths,
        "workbook_sha256s": workbook_sha256s,
        "rate_plan": [row.__dict__ | {"min_weight_kg": str(row.min_weight_kg), "max_weight_kg": str(row.max_weight_kg), "price_usd": str(row.price_usd), "freight_cost_rmb": str(row.freight_cost_rmb) if row.freight_cost_rmb is not None else None} for row in rate_rows],
        "weight_updates": sorted(
            [
                {
                    "variant_id": row["variant_id"],
                    "inventory_item_id": row["inventory_item_id"],
                    "sku": row["sku"],
                    "target_weight_g": row["target_weight_g"],
                    "target_profile": row["target_profile"],
                    "reason": row["reason"],
                }
                for row in weight_updates
            ],
            key=lambda item: (item["variant_id"], item["target_profile"]),
        ),
        "draft_actions": sorted(
            [
                {
                    "product_id": row["product_id"],
                    "reason": row["reason"],
                    "invalid_skus": row["invalid_skus"],
                }
                for row in draft_actions
            ],
            key=lambda item: item["product_id"],
        ),
        "legacy_battery_variant_ids": sorted(delivery_snapshot.get("legacy_battery_variant_ids") or []),
        "active_variant_ids": sorted(row.variant_id for row in active_variants),
        "draft_variant_ids": sorted(row.variant_id for row in draft_variants),
        "default_package_zero_verified": bool(default_package_zero_verified),
        "vendor_filter": clean(vendor_filter),
        "skip_rate_sync": skip_rate_sync,
    }
    approval_signature = canonical_hash(approval_payload)
    delivery_ready = delivery_snapshot.get("profiles_accessible") and (
        skip_rate_sync or delivery_snapshot.get("locations_accessible")
    )
    package_ready = skip_rate_sync or default_package_zero_verified
    legacy_ready = skip_rate_sync or not blocked_legacy_rows
    ready_for_apply = not missing_scopes and delivery_ready and package_ready and legacy_ready
    blockers: list[str] = []
    if missing_scopes:
        blockers.append("missing_shopify_scopes")
    if not delivery_ready:
        blockers.append("delivery_profile_or_location_read_failed")
    if not package_ready:
        blockers.append("default_shopify_package_zero_not_verified")
    if not legacy_ready:
        blockers.append("legacy_battery_profile_contains_unmatched_variants")
    return {
        "approval_schema_version": APPROVAL_SCHEMA_VERSION,
        "applied": applied,
        "ready_for_apply": bool(ready_for_apply),
        "apply_blockers": blockers,
        "approval_signature": approval_signature,
        "shopify_store_domain": admin.domain,
        "shopify_api_version": admin.version,
        "vendor_filter": clean(vendor_filter),
        "skip_rate_sync": skip_rate_sync,
        "default_package_zero_verified": bool(default_package_zero_verified),
        "workbooks": workbook_paths,
        "workbook_sha256s": workbook_sha256s,
        "template_air_rate_row_count": len(air_rate_rows),
        "template_sea_rate_row_count": len(sea_rate_rows),
        "template_rate_row_count": len(rate_rows),
        "template_rate_profile_counts": dict(sorted(rate_profile_counts.items())),
        "template_rate_country_count": len(rate_country_counts),
        "template_rate_country_counts": dict(sorted(rate_country_counts.items())),
        "template_weight_sku_count": len(targets),
        "template_weight_status_counts": dict(sorted(target_weight_counts.items())),
        "shopify_active_variant_count": len(active_variants),
        "shopify_draft_variant_count": len(draft_variants),
        "matched_active_variant_count": len(weight_updates),
        "normal_checkout_variant_count": sum(1 for row in assignments if row["target_profile"] != MANUAL_REVIEW_PROFILE),
        "manual_review_no_checkout_variant_count": len(heavy_review),
        "active_to_draft_product_count": len(draft_actions),
        "draft_backlog_variant_count": len(draft_backlog),
        "legacy_battery_profile_migration_variant_count": len(legacy_migration),
        "legacy_battery_profile_blocked_variant_count": len(blocked_legacy_rows),
        "unmatched_active_variant_count": len(unmatched),
        "weight_action_counts": dict(sorted(weight_action_counts.items())),
        "assignment_profile_counts": dict(sorted(assignment_profile_counts.items())),
        "unmatched_reason_counts": dict(sorted(unmatched_reason_counts.items())),
        "access_scope_count": len(scopes),
        "access_scopes": sorted(scopes),
        "required_scopes": sorted(required_scopes),
        "missing_required_scopes": missing_scopes,
        "delivery_access": delivery_snapshot,
        "reports": {
            "air_rate_plan": str(out_dir / AIR_RATE_PLAN_CSV),
            "sea_rate_plan": str(out_dir / SEA_RATE_PLAN_CSV),
            "weight_updates": str(out_dir / WEIGHT_UPDATES_CSV),
            "profile_assignments": str(out_dir / PROFILE_ASSIGNMENTS_CSV),
            "blocked_heavy": str(out_dir / HEAVY_REVIEW_CSV),
            "active_to_draft": str(out_dir / ACTIVE_TO_DRAFT_CSV),
            "draft_backlog": str(out_dir / DRAFT_BACKLOG_CSV),
            "legacy_battery_migration": str(out_dir / LEGACY_MIGRATION_CSV),
            "unmatched_active": str(out_dir / UNMATCHED_ACTIVE_CSV),
            "api_diff": str(out_dir / API_DIFF_CSV),
        },
    }


def condition_inputs(min_weight_kg: Decimal, max_weight_kg: Decimal) -> list[dict[str, Any]]:
    return [
        {
            "criteria": {"value": float(min_weight_kg), "unit": "KILOGRAMS"},
            "operator": "GREATER_THAN_OR_EQUAL_TO",
        },
        {
            "criteria": {"value": float(max_weight_kg), "unit": "KILOGRAMS"},
            "operator": "LESS_THAN_OR_EQUAL_TO",
        },
    ]


def method_definition_input(row: RateRow) -> dict[str, Any]:
    return {
        "name": row.rate_name,
        "description": f"{row.service_type}; {row.transit_time}".strip("; "),
        "active": True,
        "rateDefinition": {
            "price": {
                "amount": money_str(row.price_usd),
                "currencyCode": "USD",
            }
        },
        "weightConditionsToCreate": condition_inputs(row.min_weight_kg, row.max_weight_kg),
    }


def zones_to_create(rate_rows: list[RateRow]) -> list[dict[str, Any]]:
    by_country: dict[str, list[RateRow]] = defaultdict(list)
    for row in rate_rows:
        by_country[row.country_code].append(row)
    zones: list[dict[str, Any]] = []
    for country_code in sorted(by_country):
        country_rows = sorted(by_country[country_code], key=lambda item: (item.min_weight_kg, item.max_weight_kg, item.rate_name))
        country_name = country_rows[0].country or country_code
        zones.append(
            {
                "name": f"JIESTAR {country_code} {country_name}",
                "countries": [{"code": country_code, "includeAllProvinces": True}],
                "methodDefinitionsToCreate": [method_definition_input(row) for row in country_rows],
            }
        )
    return zones


def target_location_id(locations: list[dict[str, Any]]) -> str:
    matches = [location for location in locations if clean(location.get("name")) == CHENGHAI_LOCATION_NAME]
    if len(matches) != 1:
        raise RuntimeError(
            f"Expected exactly one Shopify location named {CHENGHAI_LOCATION_NAME!r}; found {len(matches)}"
        )
    location = matches[0]
    if location.get("isActive") is not True:
        raise RuntimeError(f"Shopify location {CHENGHAI_LOCATION_NAME!r} is not active")
    return location["id"]


def profile_existing_ids(
    profile: dict[str, Any],
    location_id: str,
) -> tuple[list[str], list[str], str | None]:
    zone_ids: list[str] = []
    method_ids: list[str] = []
    location_group_id: str | None = None
    for location_group in profile.get("profileLocationGroups") or []:
        if not location_group_has_location(location_group, location_id=location_id):
            continue
        location_group_id = location_group.get("locationGroup", {}).get("id") or location_group_id
        for zone in ((location_group.get("locationGroupZones") or {}).get("nodes") or []):
            zone_id = zone.get("zone", {}).get("id")
            if zone_id:
                zone_ids.append(zone_id)
            for method in ((zone.get("methodDefinitions") or {}).get("nodes") or []):
                if method.get("id"):
                    method_ids.append(method["id"])
    return zone_ids, method_ids, location_group_id


def create_or_update_profile(
    admin: ShopifyAdmin,
    existing_profile: dict[str, Any] | None,
    profile_name: str,
    variant_ids: list[str],
    rate_rows: list[RateRow],
    location_id: str,
) -> dict[str, Any]:
    first_variant_ids = variant_ids[:VARIANT_ASSOCIATE_BATCH_SIZE]
    remaining_variant_batches = list(chunks(variant_ids[VARIANT_ASSOCIATE_BATCH_SIZE:], VARIANT_ASSOCIATE_BATCH_SIZE))
    if existing_profile is None:
        location_group: dict[str, Any] = {"locationsToAdd": [location_id]}
        if rate_rows:
            location_group["zonesToCreate"] = zones_to_create(rate_rows)
        profile_input: dict[str, Any] = {
            "name": profile_name,
            "variantsToAssociate": first_variant_ids,
            "locationGroupsToCreate": [location_group],
        }
        result = admin.delivery_profile_create(profile_input)
        associate_results = []
        created_profile_id = (result.get("profile") or {}).get("id")
        if created_profile_id and not result.get("userErrors"):
            for batch in remaining_variant_batches:
                associate_results.append(admin.delivery_profile_update(created_profile_id, {"variantsToAssociate": batch}))
        return {
            "operation": "create",
            "profile_name": profile_name,
            "variant_count": len(variant_ids),
            "associate_batch_count": 1 + len(remaining_variant_batches) if variant_ids else 0,
            "result": result,
            "associate_results": associate_results,
        }

    zone_ids, method_ids, location_group_id = profile_existing_ids(existing_profile, location_id)
    current_variant_ids = set(existing_profile.get("profileItemVariantIds") or [])
    target_variant_ids = set(variant_ids)
    stale_variant_ids = sorted(current_variant_ids - target_variant_ids)

    delete_input: dict[str, Any] = {"name": profile_name}
    if method_ids:
        delete_input["methodDefinitionsToDelete"] = method_ids
    if zone_ids:
        delete_input["zonesToDelete"] = zone_ids
    if stale_variant_ids:
        delete_input["variantsToDissociate"] = stale_variant_ids
    delete_result = admin.delivery_profile_update(existing_profile["id"], delete_input)
    if delete_result.get("userErrors"):
        return {"operation": "delete_existing", "profile_name": profile_name, "result": delete_result}

    update_input: dict[str, Any] = {"name": profile_name, "variantsToAssociate": first_variant_ids}
    if rate_rows:
        if location_group_id:
            update_input["locationGroupsToUpdate"] = [
                {
                    "id": location_group_id,
                    "zonesToCreate": zones_to_create(rate_rows),
                }
            ]
        else:
            update_input["locationGroupsToCreate"] = [
                {
                    "locationsToAdd": [location_id],
                    "zonesToCreate": zones_to_create(rate_rows),
                }
            ]
    result = admin.delivery_profile_update(existing_profile["id"], update_input)
    associate_results = []
    if not result.get("userErrors"):
        for batch in remaining_variant_batches:
            associate_results.append(admin.delivery_profile_update(existing_profile["id"], {"variantsToAssociate": batch}))
    return {
        "operation": "update",
        "profile_name": profile_name,
        "variant_count": len(variant_ids),
        "associate_batch_count": 1 + len(remaining_variant_batches) if variant_ids else 0,
        "stale_variant_dissociate_count": len(stale_variant_ids),
        "deleted_zone_count": len(zone_ids),
        "deleted_method_count": len(method_ids),
        "delete_result": delete_result,
        "result": result,
        "associate_results": associate_results,
    }


def associate_variants_without_rate_sync(
    admin: ShopifyAdmin,
    assignments: list[dict[str, str]],
) -> list[dict[str, Any]]:
    """Associate scoped variants without deleting rates, zones, or unrelated variants."""
    profiles = admin.delivery_profiles(include_profile_items=False)
    existing_by_name = {
        clean(profile.get("name")): profile
        for profile in profiles
        if clean(profile.get("name")) in TARGET_PROFILE_NAMES
    }
    variants_by_profile: dict[str, list[str]] = defaultdict(list)
    for row in assignments:
        variants_by_profile[row["target_profile"]].append(row["variant_id"])

    results: list[dict[str, Any]] = []
    for profile_name, variant_ids in sorted(variants_by_profile.items()):
        profile = existing_by_name.get(profile_name)
        if not profile:
            results.append(
                {
                    "type": "delivery_profile_association",
                    "profile_name": profile_name,
                    "variant_count": len(set(variant_ids)),
                    "ok": False,
                    "errors": [{"message": "Target delivery profile does not exist; rate sync is disabled."}],
                }
            )
            continue
        for batch in chunks(sorted(set(variant_ids)), VARIANT_ASSOCIATE_BATCH_SIZE):
            result = admin.delivery_profile_update(profile["id"], {"variantsToAssociate": batch})
            errors = result.get("userErrors") or []
            results.append(
                {
                    "type": "delivery_profile_association",
                    "profile_name": profile_name,
                    "profile_id": profile["id"],
                    "variant_count": len(batch),
                    "ok": not errors,
                    "errors": errors,
                }
            )
    return results


def apply_shipping(
    admin: ShopifyAdmin,
    rate_rows: list[RateRow],
    weight_updates: list[dict[str, str]],
    assignments: list[dict[str, str]],
    draft_actions: list[dict[str, str]],
    skip_rate_sync: bool = False,
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []

    grouped_weight_updates: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in weight_updates:
        if row["action"] != "update":
            continue
        grouped_weight_updates[row["product_id"]].append(
            {
                "id": row["variant_id"],
                "inventoryItem": {
                    "requiresShipping": True,
                    "measurement": {
                        "weight": {
                            "value": float(row["target_weight_g"]),
                            "unit": "GRAMS",
                        }
                    },
                },
            }
        )

    for product_id, updates in grouped_weight_updates.items():
        for batch in chunks(updates, 100):
            errors = admin.variants_bulk_update(product_id, batch)
            results.append(
                {
                    "type": "variant_weight_update",
                    "product_id": product_id,
                    "variant_count": len(batch),
                    "ok": not errors,
                    "errors": errors,
                }
            )

    for row in draft_actions:
        result = admin.product_update_status(row["product_id"], "DRAFT")
        errors = result.get("userErrors") or []
        results.append(
            {
                "type": "product_status_update",
                "product_id": row["product_id"],
                "handle": row["handle"],
                "target_status": "DRAFT",
                "ok": not errors,
                "errors": errors,
                "result": result,
            }
        )

    if skip_rate_sync:
        results.extend(associate_variants_without_rate_sync(admin, assignments))
        return results

    locations = admin.locations()
    profiles = admin.delivery_profiles(include_profile_items=True)
    existing_by_name = {
        clean(profile.get("name")): profile
        for profile in profiles
        if clean(profile.get("name")) in MANAGED_PROFILE_NAMES
    }
    location_id = target_location_id(locations)

    rates_by_profile: dict[str, list[RateRow]] = defaultdict(list)
    for row in rate_rows:
        rates_by_profile[row.target_profile].append(row)
    variants_by_profile: dict[str, list[str]] = defaultdict(list)
    for row in assignments:
        variants_by_profile[row["target_profile"]].append(row["variant_id"])

    for profile_name in TARGET_PROFILE_NAMES:
        profile_result = create_or_update_profile(
            admin=admin,
            existing_profile=existing_by_name.get(profile_name),
            profile_name=profile_name,
            variant_ids=sorted(set(variants_by_profile.get(profile_name, []))),
            rate_rows=sorted(rates_by_profile.get(profile_name, []), key=lambda item: (item.country_code, item.min_weight_kg, item.rate_name)),
            location_id=location_id,
        )
        profile_errors = []
        for key in ("delete_result", "result"):
            value = profile_result.get(key)
            if isinstance(value, dict):
                profile_errors.extend(value.get("userErrors") or [])
        for value in profile_result.get("associate_results") or []:
            if isinstance(value, dict):
                profile_errors.extend(value.get("userErrors") or [])
        results.append({"type": "delivery_profile", "ok": not profile_errors, "errors": profile_errors, **profile_result})

    legacy_profile = existing_by_name.get(LEGACY_BATTERY_PROFILE)
    if legacy_profile:
        remaining_ids = admin.delivery_profile_variant_ids(legacy_profile["id"])
        for batch in chunks(remaining_ids, VARIANT_ASSOCIATE_BATCH_SIZE):
            result = admin.delivery_profile_update(legacy_profile["id"], {"variantsToDissociate": batch})
            errors = result.get("userErrors") or []
            results.append(
                {
                    "type": "legacy_profile_dissociation",
                    "profile_name": LEGACY_BATTERY_PROFILE,
                    "profile_id": legacy_profile["id"],
                    "variant_count": len(batch),
                    "ok": not errors,
                    "errors": errors,
                }
            )
        final_remaining_ids = admin.delivery_profile_variant_ids(legacy_profile["id"])
        if final_remaining_ids:
            results.append(
                {
                    "type": "legacy_profile_remove",
                    "profile_name": LEGACY_BATTERY_PROFILE,
                    "profile_id": legacy_profile["id"],
                    "remaining_variant_count": len(final_remaining_ids),
                    "ok": False,
                    "errors": [{"message": "Legacy profile still contains variants after migration/dissociation."}],
                }
            )
        else:
            remove_result = admin.delivery_profile_remove(legacy_profile["id"])
            remove_errors = remove_result.get("userErrors") or []
            results.append(
                {
                    "type": "legacy_profile_remove",
                    "profile_name": LEGACY_BATTERY_PROFILE,
                    "profile_id": legacy_profile["id"],
                    "remaining_variant_count": 0,
                    "ok": not remove_errors,
                    "errors": remove_errors,
                    "result": remove_result,
                }
            )
    else:
        results.append(
            {
                "type": "legacy_profile_remove",
                "profile_name": LEGACY_BATTERY_PROFILE,
                "operation": "noop_already_absent",
                "ok": True,
                "errors": [],
            }
        )

    return results


def read_approved_summary(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise RuntimeError(f"Approved summary does not exist: {path}")
    with path.open(encoding="utf-8") as file:
        summary = json.load(file)
    required = {"approval_schema_version", "approval_signature", "ready_for_apply", "workbook_sha256s"}
    missing = required - set(summary)
    if missing:
        raise RuntimeError(f"Approved summary missing required fields: {sorted(missing)}")
    if summary.get("approval_schema_version") != APPROVAL_SCHEMA_VERSION:
        raise RuntimeError("Old shipping approval reports are prohibited; generate and review a fresh version-2 dry-run.")
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Update Shopify weights and the standard-goods air/sea shipping profiles from three reviewed workbooks."
    )
    parser.add_argument("--air-rate-workbook", type=Path, default=DEFAULT_AIR_RATE_WORKBOOK)
    parser.add_argument("--sea-rate-workbook", type=Path, default=DEFAULT_SEA_RATE_WORKBOOK)
    parser.add_argument("--weight-workbook", type=Path, default=DEFAULT_WEIGHT_WORKBOOK)
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    parser.add_argument("--vendor", help="Limit Shopify variants to an exact vendor match, case-insensitive.")
    parser.add_argument(
        "--skip-rate-sync",
        action="store_true",
        help="Update scoped weights and associate variants to existing profiles without changing rates, zones, or unrelated variants.",
    )
    parser.add_argument(
        "--default-package-zero-verified",
        action="store_true",
        help="Assert that Shopify's default checkout package has been set and read back as 0g. Required for a full apply.",
    )
    parser.add_argument("--apply", action="store_true", help="Apply reviewed shipping updates to Shopify.")
    parser.add_argument("--yes", action="store_true", help="Required with --apply.")
    parser.add_argument("--input-approved-report", type=Path, help="Required with --apply. Use shipping-update-summary.json from reviewed dry-run.")
    args = parser.parse_args()

    if args.apply and (not args.yes or not args.input_approved_report):
        raise RuntimeError("--apply requires --yes and --input-approved-report")
    if args.vendor and not args.skip_rate_sync:
        raise RuntimeError("--vendor is only supported with --skip-rate-sync; a full 271-rate sync must audit all active variants.")

    args.out_dir.mkdir(parents=True, exist_ok=True)
    air_rate_rows, sea_rate_rows, targets = load_shipping_workbooks(
        args.air_rate_workbook,
        args.sea_rate_workbook,
        args.weight_workbook,
    )
    rate_rows = [*air_rate_rows, *sea_rate_rows]
    admin = ShopifyAdmin()
    scopes = admin.access_scopes()
    active_variants = filter_variants_by_vendor(admin.active_variants(), args.vendor)
    draft_variants = filter_variants_by_vendor(admin.draft_variants(), args.vendor)
    weight_updates, assignments, heavy_review, draft_actions, unmatched = build_scoped_shipping_report(
        active_variants,
        targets,
        scope_to_targets=args.skip_rate_sync,
    )
    draft_backlog = build_draft_backlog(draft_variants, targets)
    api_diff_rows, delivery_snapshot = delivery_access_snapshot(admin, rate_rows)
    legacy_migration = build_legacy_migration_report(
        active_variants,
        draft_variants,
        assignments,
        draft_actions,
        delivery_snapshot,
    )
    summary = build_summary(
        air_rate_workbook=args.air_rate_workbook,
        sea_rate_workbook=args.sea_rate_workbook,
        weight_workbook=args.weight_workbook,
        out_dir=args.out_dir,
        admin=admin,
        scopes=scopes,
        air_rate_rows=air_rate_rows,
        sea_rate_rows=sea_rate_rows,
        active_variants=active_variants,
        draft_variants=draft_variants,
        targets=targets,
        weight_updates=weight_updates,
        assignments=assignments,
        heavy_review=heavy_review,
        draft_actions=draft_actions,
        draft_backlog=draft_backlog,
        legacy_migration=legacy_migration,
        unmatched=unmatched,
        delivery_snapshot=delivery_snapshot,
        applied=args.apply,
        default_package_zero_verified=args.default_package_zero_verified,
        vendor_filter=args.vendor,
        skip_rate_sync=args.skip_rate_sync,
    )

    # The approved summary path intentionally contains the signature. CSV files are
    # review aids; the JSON signature is the apply guard.
    if args.apply:
        approved = read_approved_summary(args.input_approved_report)
        if not summary.get("ready_for_apply"):
            raise RuntimeError("Current Shopify access is not ready_for_apply; fix Shopify app scopes and rerun dry-run.")
        if not approved.get("ready_for_apply"):
            raise RuntimeError("Approved dry-run summary was not ready_for_apply; fix Shopify app permissions first.")
        if approved.get("approval_signature") != summary["approval_signature"]:
            raise RuntimeError("Current shipping dry-run signature differs from the approved report; rerun dry-run and review again.")
        apply_results = apply_shipping(
            admin,
            rate_rows,
            weight_updates,
            assignments,
            draft_actions,
            skip_rate_sync=args.skip_rate_sync,
        )
        (args.out_dir / APPLY_RESULT_JSON).write_text(json.dumps(apply_results, ensure_ascii=False, indent=2), encoding="utf-8")
        summary["apply_result_path"] = str(args.out_dir / APPLY_RESULT_JSON)
        summary["apply_ok_count"] = sum(1 for row in apply_results if row.get("ok"))
        summary["apply_error_count"] = sum(1 for row in apply_results if not row.get("ok"))
        summary["applied_weight_update_count"] = sum(row.get("variant_count", 0) for row in apply_results if row.get("type") == "variant_weight_update")
        summary["applied_product_to_draft_count"] = sum(1 for row in apply_results if row.get("type") == "product_status_update" and row.get("ok"))

    write_csv(args.out_dir / AIR_RATE_PLAN_CSV, rate_csv_rows(air_rate_rows))
    write_csv(args.out_dir / SEA_RATE_PLAN_CSV, rate_csv_rows(sea_rate_rows))
    write_csv(args.out_dir / WEIGHT_UPDATES_CSV, weight_updates)
    write_csv(args.out_dir / PROFILE_ASSIGNMENTS_CSV, assignments)
    write_csv(args.out_dir / HEAVY_REVIEW_CSV, heavy_review)
    write_csv(args.out_dir / ACTIVE_TO_DRAFT_CSV, draft_actions)
    write_csv(args.out_dir / DRAFT_BACKLOG_CSV, draft_backlog)
    write_csv(args.out_dir / LEGACY_MIGRATION_CSV, legacy_migration)
    write_csv(args.out_dir / UNMATCHED_ACTIVE_CSV, unmatched)
    write_csv(args.out_dir / API_DIFF_CSV, api_diff_rows, ["area", "status", "name", "id", "detail"])
    (args.out_dir / SUMMARY_JSON).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(str(error), file=sys.stderr)
        raise SystemExit(1)
