#!/usr/bin/env python3
"""Generate English B2B product catalog PDFs from local ORICO quotation sheets."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import textwrap
import unicodedata
from collections import defaultdict
from dataclasses import asdict, dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Iterable

import openpyxl
from PIL import Image, ImageOps
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import (
    Flowable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


SOURCE_ROOT = Path("/Volumes/ORICO/各品牌报价表")
OUTPUT_ROOT = Path("output/pdf/product-catalogs")
TMP_ROOT = Path("tmp/pdfs/product-catalogs")

JIESTAR_BOOK = SOURCE_ROOT / "杰星-外销报价表（含图片.内附6个文档）202606.xlsx"
TK_BOOK = SOURCE_ROOT / "TK TWO积木报价表26.05.25(1).xlsx"
SMALL_ANGLE_BOOK = SOURCE_ROOT / "小角度BLOCK报价 新).xlsx"
GULY_BOOK = SOURCE_ROOT / "锦童宝玩具厂积木报价表(1).xlsx"
ZOIN_BOOK = SOURCE_ROOT / "集域产品报价2026.3.17(1).xlsx"
IBLOCK_BOOK = SOURCE_ROOT / "积趣IBLOCK-全品报价单 - 分类（男-女-常规品）.xlsx"
JIQI_BOOK = SOURCE_ROOT / "积琪报价表-26.6.3 32款(3).xlsx"

JIESTAR_EN_BOOK = Path("/Volumes/ORICO/jiestar电商图/杰星整理表.xlsx")
XBERT_EN_BOOK = Path("/Volumes/ORICO/Xbert/砖悦电商图/砖悦整理表.xlsx")
TK_META_BOOK = Path("/Volumes/ORICO/TK two/TK Two产品元字段资料表.xlsx")
GULY_META_BOOK = Path("/Volumes/ORICO/GULY/GULY产品元字段资料表.xlsx")
ZOIN_READY_BOOK = Path("/Volumes/ORICO/积域资料/Zoin-上架前整理/reports/zoin-catalog-ready.xlsx")
IBLOCK_READY_BOOK = Path("/Volumes/ORICO/iblock/iblock-上架前整理/iblock-catalog-ready.xlsx")
JIQI_META_BOOK = Path("/Volumes/ORICO/积琪积木/JIQI产品元字段资料表.xlsx")
IBLOCK_IMAGE_ROOT = Path("/Volumes/ORICO/iblock/iblock-上架前整理/images")


PRICE_WORDS = (
    "price",
    "factory",
    "dealer",
    "retail",
    "control",
    "cost",
    "出厂",
    "单价",
    "控价",
    "经销",
    "零售价",
    "含税",
)


SERIES_EN_MAP = {
    "坦克": "Tank",
    "车模": "Car Model",
    "飞机": "Aircraft",
    "警察": "Police",
    "花类": "Flowers",
    "街景": "Street View",
    "冰雪": "Winter Fantasy",
    "军事": "Military",
    "消防": "Fire Rescue",
    "动物": "Animals",
    "昆虫": "Insects",
    "躲豆豆": "Hide-And-Seek Scenes",
    "创意摩载具摆件": "Creative Motorcycle Display",
    "科技摩托车": "Technic Motorcycles",
    "古励积木报价": "Technic Vehicles",
    "双高积木报价": "Technic Motorcycles",
    "展鹏": "Space Models",
    "常规品": "Core Collection",
    "女生品": "Lifestyle Collection",
    "男生品": "Technic Collection",
    "十二生肖": "Zodiac",
    "国魂·重器崛起": "Military Engineering",
    "花愿祈": "Floral Wishes",
    "虫界漫游": "Insect Explorer",
}

BRAND_ACCENTS = {
    "JIESTAR": "#D71920",
    "Xbert": "#111827",
    "TK Two": "#D7A923",
    "Small Angle": "#2B8FFF",
    "GULY": "#E85D04",
    "Zoin": "#10A37F",
    "iBlock": "#7C3AED",
    "JIQI": "#DC2626",
}

BRAND_LOGOS = {
    "JIESTAR": Path("public/images/brand/jiestar-logo-color.png"),
    "Xbert": Path("public/images/sub-brands/zhuanyue-xbert-logo.png"),
    "TK Two": Path("public/images/sub-brands/tktwo-logo.png"),
    "Small Angle": Path("public/images/sub-brands/xiaojiaodu-logo.png"),
    "GULY": Path("public/images/sub-brands/guly-logo.png"),
    "Zoin": Path("public/images/sub-brands/zoin-logo.png"),
    "iBlock": Path("public/images/sub-brands/iblock-logo.png"),
    "JIQI": Path("public/images/sub-brands/jiqi-logo.png"),
}

INK = "#101828"
MUTED = "#667085"
LINE = "#D9E1EA"
PAPER = "#FFFFFF"
SOFT = "#F6F8FB"
DEEP = "#111827"
JIESTAR_INDEX_ROWS_PER_COLUMN = 30
JIESTAR_INDEX_COLUMNS = 3
JIESTAR_MOBILE_PAGE_SIZE = (108 * mm, 192 * mm)
JIESTAR_MOBILE_INDEX_ROWS = 22
JIESTAR_MOBILE_EXCLUDED_SERIES = {"Baby and Toddler"}
MOBILE_BRAND_ORDER = ["JIESTAR", "Xbert", "TK Two", "Small Angle", "GULY", "Zoin", "iBlock", "JIQI"]


@dataclass
class Product:
    brand: str
    source_brand: str
    series: str
    source_series: str
    sku: str
    name_en: str
    name_source: str
    name_confidence: str
    source_name_cn: str
    piece_count: str = "-"
    carton_qty: str = "-"
    color_box_size: str = "-"
    outer_carton_size: str = "-"
    gross_net_weight_kg: str = "-"
    model_size_cm: str = "-"
    recommended_age: str = "-"
    image_path: str = ""
    image_source: str = ""
    source_file: str = ""
    source_sheet: str = ""
    source_row: int = 0


def clean_text(value: Any) -> str:
    if value is None:
        return "-"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\r", "\n")
    text = re.sub(r"\s+", " ", text).strip()
    return text if text else "-"


def clean_sku(value: Any) -> str:
    text = clean_text(value)
    if text == "-":
        return ""
    text = text.splitlines()[0]
    text = re.sub(r"（.*?）|\(.*?\)", "", text)
    text = re.sub(r"[\u4e00-\u9fff]", "", text)
    text = text.strip().upper()
    text = re.sub(r"[^A-Z0-9._/-]", "", text)
    if text in {"SKU", "货号", "产品货号", "ITEM NO.", "ITEM NO"}:
        return ""
    return text


def pdf_safe_text(value: Any) -> str:
    text = clean_text(value)
    if text == "-":
        return "-"
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\u200b", "").replace("\u200c", "").replace("\u200d", "")
    text = re.sub(r"[\u4e00-\u9fff]", " ", text)
    text = re.sub(r"[^\x20-\x7E]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text if text else "-"


def normalize_dimension(value: Any) -> str:
    text = clean_text(value)
    if text == "-":
        return "-"
    text = unicodedata.normalize("NFKC", text)
    if "外" in text and ("内" in text or "：" in text or ":" in text):
        parts = re.split(r"外[:：]?", text, maxsplit=1)
        if len(parts) == 2 and parts[1].strip():
            text = parts[1].strip()
    text = text.replace("×", "*").replace("x", "*").replace("X", "*")
    text = re.sub(r"[\u4e00-\u9fff：:]+", " ", text)
    text = re.sub(r"[^0-9A-Za-z*./+ -]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text if text else "-"


def normalize_carton_qty(value: Any) -> str:
    text = clean_text(value)
    if text == "-":
        return "-"
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("×", "*").replace("x", "*").replace("X", "*")
    text = text.replace("无内", " no inner")
    text = re.sub(r"(\d+)\s*[盒只]", r"\1 pcs", text)
    text = re.sub(r"(\d+)\s*内", r"\1 inner", text)
    text = re.sub(r"pcs\s*(?=\d)", "pcs / ", text)
    text = re.sub(r"[\u4e00-\u9fff：:]+", " ", text)
    text = re.sub(r"[^0-9A-Za-z*./+ -]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text if text else "-"


def piece_count(value: Any) -> str:
    text = clean_text(value)
    if text == "-":
        return "-"
    match = re.search(r"(\d{2,5})", text)
    return match.group(1) if match else text


def english_series(value: Any, fallback: str = "Building Sets") -> str:
    text = clean_text(value)
    if text == "-":
        return fallback
    if text in SERIES_EN_MAP:
        return SERIES_EN_MAP[text]
    if re.search(r"[A-Za-z]", text) and not re.search(r"[\u4e00-\u9fff]", text):
        return text
    return fallback


def has_chinese(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", text or ""))


def safe_title(text: str) -> str:
    text = pdf_safe_text(text)
    text = text.replace("&", "and")
    text = re.sub(r"\s+", " ", text).strip(" -")
    return text or "-"


def choose_title(
    brand: str,
    sku: str,
    source_name: str,
    series: str,
    title_maps: dict[str, dict[str, str]],
) -> tuple[str, str, str]:
    brand_map = title_maps.get(brand, {})
    if sku in brand_map:
        title = safe_title(brand_map[sku])
        if title != "-" and not has_chinese(title):
            return title, "metadata", "high"

    source = clean_text(source_name)
    if source != "-" and re.search(r"[A-Za-z]", source) and not has_chinese(source):
        return safe_title(f"{brand} {source}"), "source", "medium"

    series_en = english_series(series)
    return safe_title(f"{brand} {series_en} Building Set {sku}"), "generated", "low"


def load_workbook(path: Path, *, data_only: bool = True) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=data_only, read_only=False)


def load_title_maps() -> dict[str, dict[str, str]]:
    maps: dict[str, dict[str, str]] = defaultdict(dict)

    if JIESTAR_EN_BOOK.exists():
        wb = load_workbook(JIESTAR_EN_BOOK)
        for ws in wb.worksheets:
            headers = [clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
            sku_col = find_header(headers, "货号")
            title_col = find_header(headers, "Product Name (EN)")
            if sku_col and title_col:
                for row in range(2, ws.max_row + 1):
                    sku = clean_sku(ws.cell(row, sku_col).value)
                    title = clean_text(ws.cell(row, title_col).value)
                    if sku and title != "-" and not has_chinese(title):
                        maps["JIESTAR"][sku] = title

    if XBERT_EN_BOOK.exists():
        wb = load_workbook(XBERT_EN_BOOK)
        for ws_name in ("砖悦积木",):
            if ws_name not in wb.sheetnames:
                continue
            ws = wb[ws_name]
            headers = [clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
            sku_col = find_header(headers, "货号")
            title_col = find_header(headers, "Product Name (EN)")
            if sku_col and title_col:
                for row in range(2, ws.max_row + 1):
                    sku = clean_sku(ws.cell(row, sku_col).value)
                    title = clean_text(ws.cell(row, title_col).value)
                    if sku and title != "-" and not has_chinese(title):
                        maps["Xbert"][sku] = f"Xbert {title} Building Set"

    for brand, path, sheet in (
        ("TK Two", TK_META_BOOK, "TK Two Metafields"),
        ("GULY", GULY_META_BOOK, "GULY Metafields"),
        ("JIQI", JIQI_META_BOOK, "JIQI Metafields"),
    ):
        if path.exists():
            load_metadata_titles(path, sheet, brand, maps)

    if ZOIN_READY_BOOK.exists():
        load_metadata_titles(ZOIN_READY_BOOK, "CatalogReady", "Zoin", maps)
    if IBLOCK_READY_BOOK.exists():
        load_iblock_titles(maps)

    return maps


def load_metadata_titles(
    path: Path, sheet_name: str, brand: str, maps: dict[str, dict[str, str]]
) -> None:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    headers = [clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    sku_col = find_header(headers, "sku")
    title_col = find_header(headers, "shopify_title")
    if not sku_col or not title_col:
        return
    for row in range(2, ws.max_row + 1):
        sku = clean_sku(ws.cell(row, sku_col).value)
        title = clean_text(ws.cell(row, title_col).value)
        if sku and title != "-" and not has_chinese(title):
            maps[brand][sku] = title


def load_iblock_titles(maps: dict[str, dict[str, str]]) -> None:
    wb = load_workbook(IBLOCK_READY_BOOK)
    if "Catalog" not in wb.sheetnames:
        return
    ws = wb["Catalog"]
    headers = [clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    sku_col = find_header(headers, "sku")
    series_col = find_header(headers, "product_series")
    name_col = find_header(headers, "name_cn")
    if not sku_col or not name_col:
        return
    for row in range(2, ws.max_row + 1):
        sku = clean_sku(ws.cell(row, sku_col).value)
        series = english_series(ws.cell(row, series_col).value if series_col else "-")
        if sku:
            maps["iBlock"][sku] = f"iBlock {series} Building Set {sku}"


def find_header(headers: list[str], needle: str) -> int | None:
    normalized = [h.replace("\n", "").strip().lower() for h in headers]
    target = needle.replace("\n", "").strip().lower()
    for index, header in enumerate(normalized, start=1):
        if header == target:
            return index
    for index, header in enumerate(normalized, start=1):
        if target in header:
            return index
    return None


def row_image_map(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    image_dir: Path,
    *,
    preferred_cols: set[int] | None = None,
) -> dict[int, Path]:
    mapping: dict[int, Path] = {}
    image_dir.mkdir(parents=True, exist_ok=True)
    for index, image in enumerate(getattr(ws, "_images", []), start=1):
        try:
            marker = image.anchor._from
            row = marker.row + 1
            col = marker.col + 1
        except Exception:
            continue
        if preferred_cols and col not in preferred_cols:
            continue
        if row in mapping:
            continue
        ext = "jpg" if (getattr(image, "format", "") or "").lower() in {"jpeg", "jpg"} else "png"
        raw_path = image_dir / f"{ws.title}-{row}-{index}.{ext}"
        try:
            raw_path.write_bytes(image._data())
            optimized = optimize_image(raw_path, image_dir / f"{ws.title}-{row}-{index}-optimized.jpg")
            mapping[row] = optimized
        except Exception:
            continue
    return mapping


def optimize_image(source: Path, target: Path, max_px: int = 1100) -> Path:
    with Image.open(source) as image:
        image = ImageOps.exif_transpose(image)
        image.thumbnail((max_px, max_px), Image.Resampling.LANCZOS)
        canvas = Image.new("RGB", image.size, "white")
        if image.mode == "RGBA":
            canvas.paste(image, mask=image.split()[-1])
        else:
            canvas.paste(image.convert("RGB"))
        canvas.save(target, "JPEG", quality=82, optimize=True)
    return target


def build_products() -> list[Product]:
    title_maps = load_title_maps()
    products: list[Product] = []

    products.extend(extract_jiestar_and_xbert(title_maps))
    products.extend(extract_tk_two(title_maps))
    products.extend(extract_small_angle(title_maps))
    products.extend(extract_guly(title_maps))
    products.extend(extract_zoin(title_maps))
    products.extend(extract_iblock(title_maps))
    products.extend(extract_jiqi(title_maps))

    return dedupe_products(products)


def dedupe_products(products: list[Product]) -> list[Product]:
    best: dict[tuple[str, str], Product] = {}
    order: list[tuple[str, str]] = []
    for product in products:
        key = (product.brand, product.sku)
        if key not in best:
            best[key] = product
            order.append(key)
            continue
        if product_quality_score(product) > product_quality_score(best[key]):
            best[key] = product
    return [best[key] for key in order]


def product_quality_score(product: Product) -> int:
    fields = (
        product.image_path,
        product.piece_count,
        product.carton_qty,
        product.color_box_size,
        product.outer_carton_size,
        product.gross_net_weight_kg,
        product.model_size_cm,
        product.recommended_age,
    )
    score = sum(1 for field in fields if field and field != "-")
    if product.name_confidence == "high":
        score += 3
    elif product.name_confidence == "medium":
        score += 1
    return score


def make_product(
    *,
    brand: str,
    source_brand: str | None = None,
    series: Any,
    sku: Any,
    source_name: Any,
    title_maps: dict[str, dict[str, str]],
    source_file: Path,
    source_sheet: str,
    source_row: int,
    piece_count_value: Any = None,
    carton_qty: Any = None,
    color_box_size: Any = None,
    outer_carton_size: Any = None,
    gross_net_weight_kg: Any = None,
    model_size_cm: Any = None,
    recommended_age: Any = None,
    image_path: Path | None = None,
    image_source: str = "",
) -> Product | None:
    sku_text = clean_sku(sku)
    if not sku_text:
        return None
    source_series = clean_text(series)
    series_en = english_series(series)
    name_en, name_source, confidence = choose_title(
        brand, sku_text, clean_text(source_name), source_series, title_maps
    )
    return Product(
        brand=brand,
        source_brand=source_brand or brand,
        series=series_en,
        source_series=source_series,
        sku=sku_text,
        name_en=name_en,
        name_source=name_source,
        name_confidence=confidence,
        source_name_cn=clean_text(source_name),
        piece_count=piece_count(piece_count_value),
        carton_qty=normalize_carton_qty(carton_qty),
        color_box_size=normalize_dimension(color_box_size),
        outer_carton_size=normalize_dimension(outer_carton_size),
        gross_net_weight_kg=pdf_safe_text(gross_net_weight_kg),
        model_size_cm=normalize_dimension(model_size_cm),
        recommended_age=pdf_safe_text(recommended_age),
        image_path=str(image_path or ""),
        image_source=image_source if image_path else "",
        source_file=str(source_file),
        source_sheet=source_sheet,
        source_row=source_row,
    )


def extract_jiestar_and_xbert(title_maps: dict[str, dict[str, str]]) -> list[Product]:
    wb = load_workbook(JIESTAR_BOOK)
    image_base = TMP_ROOT / "images" / "jiestar-source"
    products: list[Product] = []
    configs = {
        "杰星积木": ("JIESTAR", 3, 4, 5, 7, 8, 9, 10, 12, 13, 15),
        "杰星X系列": ("JIESTAR", 3, 4, 5, 7, 8, 9, 10, 12, 13, 15),
        "杰星FF系列": ("JIESTAR", 3, 4, 5, 7, 8, 9, 10, 12, 13, 15),
        "杰星JJ【积木】": ("JIESTAR", 3, 4, 5, 7, 8, 9, 10, 12, 13, 15),
        "杰星婴童": ("JIESTAR", "Baby and Toddler", 3, 4, 6, 7, 8, 9, 11, 12, None),
        "砖悦": ("Xbert", 3, 4, 5, 7, 8, 9, 10, 12, 13, 15),
    }
    for sheet_name, config in configs.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        image_map = row_image_map(ws, image_base / sheet_name, preferred_cols={2})
        brand = config[0]
        series_ref = config[1]
        for row in range(4, ws.max_row + 1):
            series = series_ref if isinstance(series_ref, str) else ws.cell(row, series_ref).value
            product = make_product(
                brand=brand,
                series=series,
                sku=ws.cell(row, config[2]).value,
                source_name=ws.cell(row, config[3]).value,
                title_maps=title_maps,
                source_file=JIESTAR_BOOK,
                source_sheet=sheet_name,
                source_row=row,
                color_box_size=ws.cell(row, config[4]).value,
                model_size_cm=ws.cell(row, config[5]).value,
                outer_carton_size=ws.cell(row, config[6]).value,
                carton_qty=ws.cell(row, config[7]).value,
                gross_net_weight_kg=ws.cell(row, config[8]).value,
                recommended_age=ws.cell(row, config[9]).value,
                piece_count_value=ws.cell(row, config[10]).value if config[10] else None,
                image_path=image_map.get(row),
                image_source="workbook",
            )
            if product:
                products.append(product)
    return products


def extract_tk_two(title_maps: dict[str, dict[str, str]]) -> list[Product]:
    wb = load_workbook(TK_BOOK)
    ws = wb["Sheet1"]
    image_map = row_image_map(ws, TMP_ROOT / "images" / "tk-two", preferred_cols={2})
    products = []
    for row in range(3, ws.max_row + 1):
        product = make_product(
            brand="TK Two",
            series="Tank",
            sku=ws.cell(row, 1).value,
            source_name=ws.cell(row, 3).value,
            title_maps=title_maps,
            source_file=TK_BOOK,
            source_sheet=ws.title,
            source_row=row,
            carton_qty=ws.cell(row, 4).value,
            model_size_cm=ws.cell(row, 9).value,
            color_box_size=ws.cell(row, 10).value,
            outer_carton_size=ws.cell(row, 11).value,
            gross_net_weight_kg=ws.cell(row, 13).value,
            piece_count_value=ws.cell(row, 14).value,
            image_path=image_map.get(row),
            image_source="workbook",
        )
        if product:
            products.append(product)
    return products


def extract_small_angle(title_maps: dict[str, dict[str, str]]) -> list[Product]:
    wb = load_workbook(SMALL_ANGLE_BOOK)
    ws = wb["Sheet1"]
    image_map = row_image_map(ws, TMP_ROOT / "images" / "small-angle", preferred_cols={1})
    products = []
    for row in range(5, ws.max_row + 1):
        product = make_product(
            brand="Small Angle",
            series=ws.cell(row, 2).value,
            sku=ws.cell(row, 3).value,
            source_name=ws.cell(row, 5).value,
            title_maps=title_maps,
            source_file=SMALL_ANGLE_BOOK,
            source_sheet=ws.title,
            source_row=row,
            piece_count_value=ws.cell(row, 4).value,
            carton_qty=ws.cell(row, 7).value,
            color_box_size=ws.cell(row, 8).value,
            outer_carton_size=ws.cell(row, 9).value,
            gross_net_weight_kg=format_gross_net(ws.cell(row, 10).value, ws.cell(row, 11).value),
            image_path=image_map.get(row),
            image_source="workbook",
        )
        if product:
            products.append(product)
    return products


def extract_guly(title_maps: dict[str, dict[str, str]]) -> list[Product]:
    wb = load_workbook(GULY_BOOK)
    products = []
    configs = {
        "古励积木报价": (1, 2, 6, 7, 8, 9, 10),
        "双高积木报价": (1, 2, 6, 7, 8, 9, 10),
        "展鹏": (1, 2, 5, 6, 7, 8, 9),
    }
    for sheet_name, config in configs.items():
        ws = wb[sheet_name]
        image_map = row_image_map(ws, TMP_ROOT / "images" / "guly" / sheet_name, preferred_cols={3})
        start_row = 5 if sheet_name != "双高积木报价" else 3
        for row in range(start_row, ws.max_row + 1):
            product = make_product(
                brand="GULY",
                series=sheet_name,
                sku=ws.cell(row, config[0]).value,
                source_name=ws.cell(row, config[1]).value,
                title_maps=title_maps,
                source_file=GULY_BOOK,
                source_sheet=sheet_name,
                source_row=row,
                carton_qty=ws.cell(row, config[2]).value,
                color_box_size=ws.cell(row, config[3]).value,
                outer_carton_size=ws.cell(row, config[4]).value,
                gross_net_weight_kg=ws.cell(row, config[5]).value,
                piece_count_value=ws.cell(row, config[6]).value,
                image_path=image_map.get(row),
                image_source="workbook",
            )
            if product:
                products.append(product)
    return products


def extract_zoin(title_maps: dict[str, dict[str, str]]) -> list[Product]:
    wb = load_workbook(ZOIN_BOOK)
    ws = wb["Sheet1"]
    image_map = row_image_map(ws, TMP_ROOT / "images" / "zoin", preferred_cols={2})
    products = []
    for row in range(3, ws.max_row + 1):
        product = make_product(
            brand="Zoin",
            series=ws.cell(row, 4).value,
            sku=ws.cell(row, 5).value,
            source_name=ws.cell(row, 6).value,
            title_maps=title_maps,
            source_file=ZOIN_BOOK,
            source_sheet=ws.title,
            source_row=row,
            color_box_size=ws.cell(row, 8).value,
            model_size_cm=ws.cell(row, 9).value,
            outer_carton_size=ws.cell(row, 10).value,
            carton_qty=ws.cell(row, 11).value,
            gross_net_weight_kg=ws.cell(row, 13).value,
            recommended_age=ws.cell(row, 14).value,
            image_path=image_map.get(row),
            image_source="workbook",
        )
        if product:
            products.append(product)
    return products


def extract_iblock(title_maps: dict[str, dict[str, str]]) -> list[Product]:
    wb = load_workbook(IBLOCK_BOOK, data_only=False)
    products: list[Product] = []
    image_index = build_iblock_image_index()

    for sheet_name in ("常规品", "女生品", "男生品"):
        ws = wb[sheet_name]
        for row in range(5, ws.max_row + 1):
            sku = clean_sku(ws.cell(row, 5).value)
            image_path = image_index.get(sku)
            raw_series = clean_text(ws.cell(row, 4).value)
            product = make_product(
                brand="iBlock",
                series=sheet_name if raw_series == "-" else raw_series,
                sku=sku,
                source_name=ws.cell(row, 6).value,
                title_maps=title_maps,
                source_file=IBLOCK_BOOK,
                source_sheet=sheet_name,
                source_row=row,
                carton_qty=ws.cell(row, 11).value,
                model_size_cm=ws.cell(row, 17).value,
                color_box_size=ws.cell(row, 18).value,
                recommended_age=ws.cell(row, 10).value,
                image_path=image_path,
                image_source="iblock-prepared-folder" if image_path else "",
            )
            if product:
                products.append(product)

    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        for row in range(5, ws.max_row + 1):
            sku = clean_sku(ws.cell(row, 4).value)
            image_path = image_index.get(sku)
            product = make_product(
                brand="iBlock",
                series=ws.cell(row, 3).value or "Zodiac",
                sku=sku,
                source_name=ws.cell(row, 5).value,
                title_maps=title_maps,
                source_file=IBLOCK_BOOK,
                source_sheet=ws.title,
                source_row=row,
                image_path=image_path,
                image_source="iblock-prepared-folder" if image_path else "",
            )
            if product:
                products.append(product)
    return products


def build_iblock_image_index() -> dict[str, Path]:
    if not IBLOCK_IMAGE_ROOT.exists():
        return {}
    folder_map = {p.name.upper(): p for p in IBLOCK_IMAGE_ROOT.iterdir() if p.is_dir()}
    result: dict[str, Path] = {}
    for sku, folder in folder_map.items():
        files = [
            f
            for f in folder.iterdir()
            if f.is_file()
            and f.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}
            and not f.name.startswith("._")
        ]
        if not files:
            continue
        preferred = sorted(
            files,
            key=lambda p: (
                0 if "白底" in p.name else 1,
                0 if re.search(r"-1\.(jpg|jpeg|png|webp)$", p.name, re.I) else 1,
                1 if "详情" in p.name else 0,
                len(p.name),
            ),
        )[0]
        out_dir = TMP_ROOT / "images" / "iblock"
        out_dir.mkdir(parents=True, exist_ok=True)
        target = out_dir / f"{sku}.jpg"
        try:
            result[sku] = optimize_image(preferred, target)
        except Exception:
            result[sku] = preferred
    return result


def extract_jiqi(title_maps: dict[str, dict[str, str]]) -> list[Product]:
    wb = load_workbook(JIQI_BOOK)
    ws = wb["Sheet1"]
    image_map = row_image_map(ws, TMP_ROOT / "images" / "jiqi", preferred_cols={2})
    products = []
    for row in range(2, ws.max_row + 1):
        product = make_product(
            brand="JIQI",
            series="Display Model",
            sku=ws.cell(row, 3).value,
            source_name=ws.cell(row, 5).value,
            title_maps=title_maps,
            source_file=JIQI_BOOK,
            source_sheet=ws.title,
            source_row=row,
            piece_count_value=ws.cell(row, 4).value,
            carton_qty=ws.cell(row, 7).value,
            color_box_size=ws.cell(row, 8).value,
            outer_carton_size=ws.cell(row, 9).value,
            gross_net_weight_kg=ws.cell(row, 10).value,
            model_size_cm=ws.cell(row, 11).value,
            recommended_age=ws.cell(row, 12).value,
            image_path=image_map.get(row),
            image_source="workbook",
        )
        if product:
            products.append(product)
    return products


def format_gross_net(gross: Any, net: Any) -> str:
    gross_text = clean_text(gross)
    net_text = clean_text(net)
    if gross_text == "-" and net_text == "-":
        return "-"
    if net_text == "-":
        return gross_text
    if gross_text == "-":
        return net_text
    return f"{gross_text}/{net_text}"


class ProductCard(Flowable):
    def __init__(
        self,
        product: Product,
        width: float,
        height: float,
        styles: dict[str, ParagraphStyle],
        accent: str,
    ):
        super().__init__()
        self.product = product
        self.width = width
        self.height = height
        self.styles = styles
        self.accent = accent

    def wrap(self, avail_width: float, avail_height: float) -> tuple[float, float]:
        return self.width, self.height

    def draw(self) -> None:
        c = self.canv
        p = self.product
        accent = colors.HexColor(self.accent)
        c.saveState()
        c.setStrokeColor(colors.HexColor(LINE))
        c.setFillColor(colors.white)
        c.roundRect(0, 0, self.width, self.height, 7, stroke=1, fill=1)
        c.setFillColor(accent)
        c.roundRect(0, 0, 4 * mm, self.height, 7, stroke=0, fill=1)
        c.setFillColor(colors.HexColor("#F7F9FC"))
        c.rect(3.2 * mm, 0.6 * mm, self.width - 3.8 * mm, self.height - 1.2 * mm, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.roundRect(6 * mm, 5 * mm, self.width - 12 * mm, self.height - 10 * mm, 6, stroke=0, fill=1)

        pad = 8 * mm
        image_w = self.width * 0.42
        image_h = self.height - 2 * pad
        image_x = pad + 2 * mm
        image_y = pad
        c.setFillColor(colors.HexColor(SOFT))
        c.roundRect(image_x, image_y, image_w, image_h, 6, stroke=0, fill=1)
        c.setStrokeColor(colors.HexColor("#E8EEF5"))
        c.roundRect(image_x + 1.2 * mm, image_y + 1.2 * mm, image_w - 2.4 * mm, image_h - 2.4 * mm, 5, stroke=1, fill=0)

        if p.image_path and Path(p.image_path).exists():
            try:
                with Image.open(p.image_path) as img:
                    iw, ih = img.size
                scale = min((image_w - 8 * mm) / iw, (image_h - 8 * mm) / ih)
                draw_w = iw * scale
                draw_h = ih * scale
                c.drawImage(
                    p.image_path,
                    image_x + (image_w - draw_w) / 2,
                    image_y + (image_h - draw_h) / 2,
                    width=draw_w,
                    height=draw_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                self.draw_placeholder(c, image_x, image_y, image_w, image_h, "Image unavailable")
        else:
            self.draw_placeholder(c, image_x, image_y, image_w, image_h, "Image pending")

        text_x = image_x + image_w + 8 * mm
        text_w = self.width - text_x - pad
        top_y = self.height - pad

        c.setFillColor(accent)
        c.roundRect(text_x, top_y - 6 * mm, 30 * mm, 5 * mm, 2.5, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 5.7)
        c.drawCentredString(text_x + 15 * mm, top_y - 4.1 * mm, "B2B CATALOG")

        c.setFillColor(colors.HexColor("#E8EEF5"))
        c.roundRect(text_x + 34 * mm, top_y - 6 * mm, text_w - 34 * mm, 5 * mm, 2.5, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica", 5.7)
        c.drawString(text_x + 36 * mm, top_y - 4.1 * mm, fit_single_line(p.series.upper(), "Helvetica", 5.7, text_w - 38 * mm))

        title_top = top_y - 12 * mm
        title = fit_text(p.name_en, "Helvetica-Bold", 12.2, text_w, max_lines=2)
        title_para = Paragraph(title, self.styles["card_title"])
        _, title_h = title_para.wrap(text_w, 26 * mm)
        title_para.drawOn(c, text_x, title_top - title_h)

        y = title_top - title_h - 4 * mm
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica", 7.5)
        subtitle = f"Item No. {p.sku}  /  {p.brand}"
        c.drawString(text_x, y, fit_single_line(subtitle, "Helvetica", 7.5, text_w))
        y -= 5 * mm
        c.setStrokeColor(colors.HexColor("#E8EEF5"))
        c.line(text_x, y, text_x + text_w, y)
        y -= 4.2 * mm

        fields = [
            ("Pieces", p.piece_count),
            ("Carton Qty", p.carton_qty),
            ("Color Box", p.color_box_size),
            ("Outer Carton", p.outer_carton_size),
            ("Weight KG", p.gross_net_weight_kg),
            ("Model Size", p.model_size_cm),
            ("Age", p.recommended_age),
        ]
        grid_col_w = (text_w - 4 * mm) / 2
        cell_h = 9.7 * mm
        for index, (label, value) in enumerate(fields):
            col = index % 2
            row = index // 2
            x = text_x + col * (grid_col_w + 4 * mm)
            yy = y - row * cell_h
            c.setFillColor(colors.HexColor("#F7F9FC"))
            c.roundRect(x, yy - cell_h + 1 * mm, grid_col_w, cell_h - 1.5 * mm, 2.5, stroke=0, fill=1)
            c.setFillColor(accent if index in (0, 1) else colors.HexColor(MUTED))
            c.setFont("Helvetica-Bold", 5.8)
            c.drawString(x + 2 * mm, yy - 3 * mm, label.upper())
            c.setFillColor(colors.HexColor(INK))
            c.setFont("Helvetica-Bold" if index in (0, 1) else "Helvetica", 7.4)
            c.drawString(x + 2 * mm, yy - 6.7 * mm, fit_single_line(value, "Helvetica", 7.4, grid_col_w - 4 * mm))

        c.restoreState()

    def draw_placeholder(self, c: Any, x: float, y: float, w: float, h: float, label: str) -> None:
        c.setStrokeColor(colors.HexColor("#CBD5E1"))
        c.setDash(3, 3)
        c.roundRect(x + 4 * mm, y + 4 * mm, w - 8 * mm, h - 8 * mm, 4, stroke=1, fill=0)
        c.setDash()
        c.setFillColor(colors.HexColor("#94A3B8"))
        c.setFont("Helvetica", 8)
        c.drawCentredString(x + w / 2, y + h / 2, label)


class CoverPage(Flowable):
    def __init__(self, brand: str, products: list[Product], width: float, height: float, accent: str):
        super().__init__()
        self.brand = brand
        self.products = products
        self.width = width
        self.height = height
        self.accent = accent

    def wrap(self, avail_width: float, avail_height: float) -> tuple[float, float]:
        return self.width, self.height

    def draw(self) -> None:
        c = self.canv
        accent = colors.HexColor(self.accent)
        series_count = len({product.series for product in self.products})
        c.saveState()
        c.setFillColor(colors.HexColor(DEEP))
        c.roundRect(0, 0, self.width, self.height, 10, stroke=0, fill=1)
        c.setFillColor(accent)
        c.rect(0, 0, 8 * mm, self.height, stroke=0, fill=1)

        c.setStrokeColor(colors.HexColor("#263244"))
        for i in range(10):
            x = 24 * mm + i * 17 * mm
            c.line(x, 16 * mm, x + 50 * mm, self.height - 18 * mm)

        c.setFillColor(colors.white)
        logo_path = BRAND_LOGOS.get(self.brand)
        if logo_path and logo_path.exists():
            c.setFillColor(colors.white)
            c.roundRect(24 * mm, self.height - 46 * mm, 34 * mm, 22 * mm, 5, stroke=0, fill=1)
            self.draw_image(c, str(logo_path), 28 * mm, self.height - 42 * mm, 26 * mm, 14 * mm)
            text_x = 64 * mm
        else:
            text_x = 24 * mm
        c.setFont("Helvetica-Bold", 30)
        c.drawString(text_x, self.height - 38 * mm, fit_single_line(self.brand, "Helvetica-Bold", 30, 105 * mm))
        c.setFont("Helvetica", 12)
        c.setFillColor(colors.HexColor("#CAD5E2"))
        c.drawString(text_x, self.height - 49 * mm, "Global B2B Product Catalog")

        c.setFillColor(accent)
        c.roundRect(24 * mm, self.height - 68 * mm, 58 * mm, 7 * mm, 3.5, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 6.7)
        c.drawCentredString(53 * mm, self.height - 65.4 * mm, "WHOLESALE / OEM / ODM")

        stat_y = 36 * mm
        stats = [("PRODUCTS", str(len(self.products))), ("SERIES", str(series_count)), ("TERMS", "BY REQUEST")]
        for index, (label, value) in enumerate(stats):
            x = 24 * mm + index * 48 * mm
            c.setFillColor(colors.HexColor("#1C2533"))
            c.roundRect(x, stat_y, 40 * mm, 22 * mm, 4, stroke=0, fill=1)
            c.setFillColor(colors.HexColor("#91A1B5"))
            c.setFont("Helvetica-Bold", 6)
            c.drawString(x + 4 * mm, stat_y + 14 * mm, label)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 12 if len(value) < 8 else 8)
            c.drawString(x + 4 * mm, stat_y + 6 * mm, fit_single_line(value, "Helvetica-Bold", 12, 32 * mm))

        hero_products = [p for p in self.products if p.image_path and Path(p.image_path).exists()][:3]
        base_x = self.width - 116 * mm
        base_y = 36 * mm
        for index, product in enumerate(hero_products):
            x = base_x + (index % 2) * 36 * mm
            y = base_y + index * 31 * mm
            w = 52 * mm
            h = 40 * mm
            if index == 1:
                x += 16 * mm
            c.setFillColor(colors.HexColor("#F8FAFC"))
            c.roundRect(x, y, w, h, 5, stroke=0, fill=1)
            c.setStrokeColor(colors.HexColor("#39465A"))
            c.roundRect(x, y, w, h, 5, stroke=1, fill=0)
            self.draw_image(c, product.image_path, x + 4 * mm, y + 4 * mm, w - 8 * mm, h - 8 * mm)

        c.setFont("Helvetica", 7.2)
        c.setFillColor(colors.HexColor("#91A1B5"))
        c.drawString(24 * mm, 16 * mm, "English-first product selection reference for overseas business buyers.")
        c.restoreState()

    def draw_image(self, c: Any, path: str, x: float, y: float, w: float, h: float) -> None:
        with Image.open(path) as img:
            iw, ih = img.size
        scale = min(w / iw, h / ih)
        draw_w = iw * scale
        draw_h = ih * scale
        c.drawImage(path, x + (w - draw_w) / 2, y + (h - draw_h) / 2, draw_w, draw_h, preserveAspectRatio=True, mask="auto")


class SeriesHeader(Flowable):
    def __init__(self, title: str, count: int, width: float, accent: str):
        super().__init__()
        self.title = title
        self.count = count
        self.width = width
        self.height = 19 * mm
        self.accent = accent

    def wrap(self, avail_width: float, avail_height: float) -> tuple[float, float]:
        return self.width, self.height

    def draw(self) -> None:
        c = self.canv
        c.saveState()
        c.setFillColor(colors.HexColor(self.accent))
        c.roundRect(0, 4 * mm, 5 * mm, 11 * mm, 2.5, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 19)
        c.drawString(8 * mm, 8.2 * mm, fit_single_line(self.title, "Helvetica-Bold", 19, self.width - 55 * mm))
        c.setFillColor(colors.HexColor("#EEF2F6"))
        c.roundRect(self.width - 38 * mm, 8 * mm, 38 * mm, 7 * mm, 3.5, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica-Bold", 6.8)
        c.drawCentredString(self.width - 19 * mm, 10.4 * mm, f"{self.count} PRODUCTS")
        c.setStrokeColor(colors.HexColor("#E6ECF3"))
        c.line(8 * mm, 4 * mm, self.width, 4 * mm)
        c.restoreState()


def fit_single_line(text: Any, font: str, size: float, max_width: float) -> str:
    value = clean_text(text)
    if stringWidth(value, font, size) <= max_width:
        return value
    ellipsis = "..."
    while value and stringWidth(value + ellipsis, font, size) > max_width:
        value = value[:-1]
    return value + ellipsis if value else "-"


def fit_text(text: str, font: str, size: float, max_width: float, max_lines: int) -> str:
    words = clean_text(text).split()
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if stringWidth(candidate, font, size) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
        if len(lines) == max_lines:
            break
    if current and len(lines) < max_lines:
        lines.append(current)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
    if len(lines) == max_lines and len(" ".join(lines).split()) < len(words):
        lines[-1] = fit_single_line(lines[-1] + "...", font, size, max_width)
    return "<br/>".join(lines) if lines else "-"


def make_styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "cover_title": ParagraphStyle(
            "CoverTitle",
            parent=base["Title"],
            fontName="Helvetica-Bold",
            fontSize=30,
            leading=36,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#111827"),
        ),
        "cover_subtitle": ParagraphStyle(
            "CoverSubtitle",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=12,
            leading=18,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#475569"),
        ),
        "series_title": ParagraphStyle(
            "SeriesTitle",
            parent=base["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=28,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#111827"),
        ),
        "series_meta": ParagraphStyle(
            "SeriesMeta",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#64748B"),
        ),
        "card_title": ParagraphStyle(
            "CardTitle",
            parent=base["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=13,
            textColor=colors.HexColor("#111827"),
        ),
    }


def group_by_brand(products: Iterable[Product]) -> dict[str, list[Product]]:
    grouped: dict[str, list[Product]] = defaultdict(list)
    for product in products:
        grouped[product.brand].append(product)
    for brand_products in grouped.values():
        brand_products.sort(key=lambda p: (p.series, natural_key(p.sku)))
    return dict(sorted(grouped.items()))


def natural_key(text: str) -> list[Any]:
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", text)]


def piece_sort_value(product: Product) -> int:
    match = re.search(r"\d+", clean_text(product.piece_count).replace(",", ""))
    return int(match.group(0)) if match else 0


def high_piece_products(products: Iterable[Product], limit: int) -> list[Product]:
    candidates = [p for p in products if p.image_path and Path(p.image_path).exists()]
    candidates.sort(key=lambda p: (-piece_sort_value(p), natural_key(p.sku)))
    return candidates[:limit]


def has_piece_counts(products: Iterable[Product]) -> bool:
    return any(piece_sort_value(product) > 0 for product in products)


def piece_showcase_text(product: Product) -> str:
    pieces = piece_sort_value(product)
    return f"{pieces} pcs" if pieces > 0 else "-"


def contrast_text_color(hex_color: str) -> colors.Color:
    value = hex_color.lstrip("#")
    if len(value) != 6:
        return colors.white
    red, green, blue = (int(value[index : index + 2], 16) for index in (0, 2, 4))
    luminance = (0.299 * red + 0.587 * green + 0.114 * blue) / 255
    return colors.HexColor(INK) if luminance > 0.62 else colors.white


def render_catalogs(
    products: list[Product],
    only_brands: set[str] | None = None,
    *,
    mobile: bool = False,
) -> dict[str, Path]:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    styles = make_styles()
    grouped = group_by_brand(products)
    output_paths: dict[str, Path] = {}
    page_size = landscape(A4)
    page_w, page_h = page_size
    margin = 14 * mm
    card_w = page_w - 2 * margin - 4 * mm
    card_h = (page_h - 2 * margin - 20 * mm) / 2

    for brand, brand_products in grouped.items():
        if only_brands and brand not in only_brands:
            continue
        accent = BRAND_ACCENTS.get(brand, "#10A37F")
        slug = brand_slug(brand)
        output_path = OUTPUT_ROOT / f"{slug}-product-catalog.pdf"
        if mobile:
            output_path = OUTPUT_ROOT / f"{slug}-product-catalog-mobile.pdf"
            render_mobile_catalog(brand, brand_products, output_path, accent)
            output_paths[brand] = output_path
            continue
        if brand == "JIESTAR":
            render_jiestar_catalog(brand_products, output_path, page_size, accent)
            output_paths[brand] = output_path
            continue
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=page_size,
            rightMargin=margin,
            leftMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
            title=f"{brand} Product Catalog",
            author="JIESTAR",
        )
        story: list[Any] = []
        story.extend(make_cover(brand, brand_products, styles, doc.width - 12, doc.height - 12, accent))
        by_series: dict[str, list[Product]] = defaultdict(list)
        for product in brand_products:
            by_series[product.series].append(product)
        for series_index, (series, series_products) in enumerate(sorted(by_series.items())):
            if series_index:
                story.append(PageBreak())
            story.append(SeriesHeader(series, len(series_products), card_w, accent))
            story.append(Spacer(1, 3.5 * mm))
            for index, product in enumerate(series_products):
                story.append(ProductCard(product, card_w, card_h, styles, accent))
                if index % 2 == 0 and index != len(series_products) - 1:
                    story.append(Spacer(1, 8 * mm))
                elif index != len(series_products) - 1:
                    story.append(PageBreak())
            story.append(Spacer(1, 2 * mm))
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        output_paths[brand] = output_path
    return output_paths


def render_jiestar_catalog(
    products: list[Product],
    output_path: Path,
    page_size: tuple[float, float],
    accent: str,
) -> None:
    page_w, page_h = page_size
    margin = 11 * mm
    header_h = 18 * mm
    footer_h = 8 * mm
    gap = 6 * mm
    content_w = page_w - 2 * margin
    card_h = (page_h - 2 * margin - header_h - footer_h - gap) / 2
    plan = build_jiestar_page_plan(products)
    catalog = pdf_canvas.Canvas(str(output_path), pagesize=page_size)
    catalog.setTitle("JIESTAR Product Catalog")
    catalog.setAuthor("JIESTAR")

    bookmark(catalog, "Cover", "cover", 0)
    CoverPage("JIESTAR", products, page_w - 2 * margin, page_h - 2 * margin, accent).drawOn(
        catalog, margin, margin
    )
    catalog.showPage()

    bookmark(catalog, "Product Finder", "product-finder", 0)
    draw_jiestar_finder_page(catalog, products, plan, page_w, page_h, margin, accent)
    draw_jiestar_footer(catalog, 2, plan["total_pages"], page_w, margin)
    catalog.showPage()

    bookmark(catalog, "Specification Field Guide", "specification-field-guide", 0)
    draw_jiestar_field_guide_page(catalog, plan, page_w, page_h, margin, accent)
    draw_jiestar_footer(catalog, 3, plan["total_pages"], page_w, margin)
    catalog.showPage()

    series_bookmarked: set[str] = set()
    for page in plan["product_pages"]:
        series = page["series"]
        if series not in series_bookmarked:
            bookmark(catalog, series, f"series-{slugify_key(series)}", 0)
            series_bookmarked.add(series)
        draw_jiestar_product_page(
            catalog,
            page["number"],
            plan["total_pages"],
            page["products"],
            series,
            plan["series_ranges"][series],
            page_w,
            page_h,
            margin,
            header_h,
            footer_h,
            content_w,
            card_h,
            gap,
            accent,
        )
        catalog.showPage()

    bookmark(catalog, "SKU / Name Index", "sku-name-index", 0)
    for index_page, entries in enumerate(plan["index_pages"], start=plan["index_start_page"]):
        draw_jiestar_index_page(
            catalog,
            entries,
            index_page,
            plan["total_pages"],
            page_w,
            page_h,
            margin,
            accent,
        )
        if index_page != plan["total_pages"]:
            catalog.showPage()
    catalog.save()
    write_jiestar_page_index(plan)


def build_jiestar_page_plan(products: list[Product]) -> dict[str, Any]:
    sorted_products = sorted(products, key=lambda p: (p.series, natural_key(p.sku)))
    by_series: dict[str, list[Product]] = defaultdict(list)
    for product in sorted_products:
        by_series[product.series].append(product)

    page_number = 4
    product_pages = []
    product_page_by_sku: dict[str, int] = {}
    series_ranges: dict[str, dict[str, int]] = {}
    for series, series_products in sorted(by_series.items()):
        start_page = page_number
        for index in range(0, len(series_products), 2):
            page_products = series_products[index : index + 2]
            product_pages.append({"number": page_number, "series": series, "products": page_products})
            for product in page_products:
                product_page_by_sku[product.sku] = page_number
            page_number += 1
        series_ranges[series] = {
            "start": start_page,
            "end": page_number - 1,
            "count": len(series_products),
        }

    index_start_page = page_number
    index_entries = [
        {
            "sku": product.sku,
            "name": product.name_en,
            "series": product.series,
            "page": product_page_by_sku[product.sku],
        }
        for product in sorted(products, key=lambda p: natural_key(p.sku))
    ]
    per_index_page = JIESTAR_INDEX_ROWS_PER_COLUMN * JIESTAR_INDEX_COLUMNS
    index_pages = [
        index_entries[index : index + per_index_page]
        for index in range(0, len(index_entries), per_index_page)
    ]
    total_pages = index_start_page + len(index_pages) - 1
    return {
        "product_pages": product_pages,
        "product_page_by_sku": product_page_by_sku,
        "series_ranges": series_ranges,
        "index_start_page": index_start_page,
        "index_pages": index_pages,
        "total_pages": total_pages,
    }


def write_jiestar_page_index(plan: dict[str, Any]) -> None:
    path = OUTPUT_ROOT / "jiestar-catalog-page-index.csv"
    fields = ["sku", "name", "series", "page"]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for entries in plan["index_pages"]:
            writer.writerows(entries)


def mobile_products_for_brand(brand: str, products: Iterable[Product]) -> list[Product]:
    filtered = list(products)
    if brand == "JIESTAR":
        filtered = [product for product in filtered if product.series not in JIESTAR_MOBILE_EXCLUDED_SERIES]
    return filtered


def render_mobile_catalog(brand: str, products: list[Product], output_path: Path, accent: str) -> None:
    products = mobile_products_for_brand(brand, products)
    page_w, page_h = JIESTAR_MOBILE_PAGE_SIZE
    margin = 7 * mm
    plan = build_mobile_page_plan(products)
    catalog = pdf_canvas.Canvas(str(output_path), pagesize=JIESTAR_MOBILE_PAGE_SIZE)
    catalog.setTitle(f"{brand} 2026 Product Catalog - Mobile")
    catalog.setAuthor("JIESTAR")

    bookmark(catalog, "Cover", "mobile-cover", 0)
    draw_mobile_cover(catalog, brand, products, plan, page_w, page_h, margin, accent)
    catalog.showPage()

    bookmark(catalog, "Product Finder", "mobile-product-finder", 0)
    draw_mobile_finder(catalog, brand, plan, page_w, page_h, margin, accent)
    draw_mobile_footer(catalog, brand, 2, plan["total_pages"], page_w, margin)
    catalog.showPage()

    bookmark(catalog, "Specification Guide", "mobile-specification-guide", 0)
    draw_mobile_spec_guide(catalog, brand, plan, page_w, page_h, margin, accent)
    draw_mobile_footer(catalog, brand, 3, plan["total_pages"], page_w, margin)
    catalog.showPage()

    bookmark(catalog, "Brand / Cooperation", "mobile-brand-cooperation", 0)
    draw_mobile_cooperation(catalog, brand, page_w, page_h, margin, accent)
    draw_mobile_footer(catalog, brand, 4, plan["total_pages"], page_w, margin)
    catalog.showPage()

    for series in plan["series_order"]:
        meta = plan["series_ranges"][series]
        bookmark(catalog, series, f"mobile-series-{slugify_key(series)}", 0)
        draw_mobile_series_divider(catalog, brand, series, meta, page_w, page_h, margin, accent)
        draw_mobile_footer(catalog, brand, meta["divider"], plan["total_pages"], page_w, margin)
        catalog.showPage()
        for product in meta["products"]:
            draw_mobile_product_page(
                catalog,
                brand,
                product,
                plan["product_page_by_sku"][product.sku],
                plan["total_pages"],
                page_w,
                page_h,
                margin,
                accent,
            )
            catalog.showPage()

    bookmark(catalog, "SKU / Name Index", "mobile-sku-name-index", 0)
    for page_number, entries in enumerate(plan["index_pages"], start=plan["index_start_page"]):
        draw_mobile_index_page(catalog, brand, entries, page_number, plan["total_pages"], page_w, page_h, margin, accent)
        catalog.showPage()

    bookmark(catalog, "Contact", "mobile-contact", 0)
    draw_mobile_contact(catalog, brand, plan["contact_page"], plan["total_pages"], page_w, page_h, margin, accent)
    catalog.save()
    write_mobile_page_index(brand, plan)


def build_mobile_series_plan(products: list[Product], start_page: int) -> dict[str, Any]:
    sorted_products = sorted(products, key=lambda p: (p.series, -piece_sort_value(p), natural_key(p.sku)))
    by_series: dict[str, list[Product]] = defaultdict(list)
    for product in sorted_products:
        by_series[product.series].append(product)

    page_number = start_page
    product_page_by_sku: dict[str, int] = {}
    series_ranges: dict[str, dict[str, Any]] = {}
    series_order: list[str] = []
    for series, series_products in sorted(by_series.items()):
        series_products = sorted(series_products, key=lambda p: (-piece_sort_value(p), natural_key(p.sku)))
        series_order.append(series)
        divider_page = page_number
        page_number += 1
        product_start = page_number
        for product in series_products:
            product_page_by_sku[product.sku] = page_number
            page_number += 1
        series_ranges[series] = {
            "divider": divider_page,
            "start": product_start,
            "end": page_number - 1,
            "count": len(series_products),
            "products": series_products,
        }

    return {
        "series_order": series_order,
        "series_ranges": series_ranges,
        "product_page_by_sku": product_page_by_sku,
        "next_page": page_number,
    }


def build_mobile_page_plan(products: list[Product]) -> dict[str, Any]:
    plan = build_mobile_series_plan(products, 5)
    index_start_page = plan["next_page"]
    index_entries = [
        {
            "sku": product.sku,
            "name": product.name_en,
            "series": product.series,
            "page": plan["product_page_by_sku"][product.sku],
        }
        for product in sorted(products, key=lambda p: natural_key(p.sku))
    ]
    index_pages = [
        index_entries[index : index + JIESTAR_MOBILE_INDEX_ROWS]
        for index in range(0, len(index_entries), JIESTAR_MOBILE_INDEX_ROWS)
    ]
    contact_page = index_start_page + len(index_pages)
    return {
        "series_order": plan["series_order"],
        "series_ranges": plan["series_ranges"],
        "product_page_by_sku": plan["product_page_by_sku"],
        "index_start_page": index_start_page,
        "index_pages": index_pages,
        "contact_page": contact_page,
        "total_pages": contact_page,
    }


def write_mobile_page_index(brand: str, plan: dict[str, Any]) -> None:
    path = OUTPUT_ROOT / f"{brand_slug(brand)}-catalog-mobile-page-index.csv"
    fields = ["sku", "name", "series", "page"]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for entries in plan["index_pages"]:
            writer.writerows(entries)


def ordered_mobile_brand_groups(products: list[Product]) -> list[tuple[str, list[Product]]]:
    grouped = group_by_brand(products)
    ordered_brands = [brand for brand in MOBILE_BRAND_ORDER if brand in grouped]
    ordered_brands.extend(brand for brand in grouped if brand not in ordered_brands)
    result = []
    for brand in ordered_brands:
        brand_products = mobile_products_for_brand(brand, grouped[brand])
        if brand_products:
            result.append((brand, brand_products))
    return result


def build_combined_mobile_page_plan(products: list[Product]) -> dict[str, Any]:
    brand_sections = []
    page_number = 4
    for brand, brand_products in ordered_mobile_brand_groups(products):
        cover_page = page_number
        finder_page = page_number + 1
        section_plan = build_mobile_series_plan(brand_products, page_number + 2)
        brand_sections.append(
            {
                "brand": brand,
                "products": brand_products,
                "cover": cover_page,
                "finder": finder_page,
                "start": cover_page,
                "end": section_plan["next_page"] - 1,
                "plan": section_plan,
            }
        )
        page_number = section_plan["next_page"]

    index_start_page = page_number
    index_entries = []
    for section in brand_sections:
        brand = section["brand"]
        section_plan = section["plan"]
        for product in sorted(section["products"], key=lambda p: natural_key(p.sku)):
            index_entries.append(
                {
                    "brand": brand,
                    "sku": product.sku,
                    "name": product.name_en,
                    "series": product.series,
                    "page": section_plan["product_page_by_sku"][product.sku],
                }
            )

    index_pages = [
        index_entries[index : index + JIESTAR_MOBILE_INDEX_ROWS]
        for index in range(0, len(index_entries), JIESTAR_MOBILE_INDEX_ROWS)
    ]
    contact_page = index_start_page + len(index_pages)
    total_pages = contact_page
    for section in brand_sections:
        section["plan"]["index_start_page"] = index_start_page
        section["plan"]["total_pages"] = total_pages

    return {
        "brand_sections": brand_sections,
        "index_start_page": index_start_page,
        "index_pages": index_pages,
        "contact_page": contact_page,
        "total_pages": total_pages,
        "product_count": len(index_entries),
    }


def render_combined_mobile_catalog(products: list[Product], output_path: Path) -> None:
    page_w, page_h = JIESTAR_MOBILE_PAGE_SIZE
    margin = 7 * mm
    plan = build_combined_mobile_page_plan(products)
    catalog = pdf_canvas.Canvas(str(output_path), pagesize=JIESTAR_MOBILE_PAGE_SIZE)
    catalog.setTitle("All Brands 2026 Product Catalog - Mobile")
    catalog.setAuthor("JIESTAR")

    bookmark(catalog, "Cover", "combined-mobile-cover", 0)
    draw_combined_mobile_cover(catalog, plan, page_w, page_h, margin, BRAND_ACCENTS["JIESTAR"])
    catalog.showPage()

    bookmark(catalog, "Brand Finder", "combined-mobile-brand-finder", 0)
    draw_combined_mobile_brand_finder(catalog, plan, page_w, page_h, margin, BRAND_ACCENTS["JIESTAR"])
    draw_mobile_footer(catalog, "All Brands", 2, plan["total_pages"], page_w, margin)
    catalog.showPage()

    bookmark(catalog, "Specification Guide", "combined-mobile-specification-guide", 0)
    draw_mobile_spec_guide(
        catalog,
        "All Brands",
        {"index_start_page": plan["index_start_page"]},
        page_w,
        page_h,
        margin,
        BRAND_ACCENTS["JIESTAR"],
    )
    draw_mobile_footer(catalog, "All Brands", 3, plan["total_pages"], page_w, margin)
    catalog.showPage()

    for section in plan["brand_sections"]:
        brand = section["brand"]
        accent = BRAND_ACCENTS.get(brand, "#10A37F")
        section_plan = section["plan"]
        bookmark(catalog, brand, f"combined-mobile-brand-{slugify_key(brand)}", 0)
        brand_display_plan = {
            "series_order": section_plan["series_order"],
            "series_ranges": section_plan["series_ranges"],
            "total_pages": section["end"] - section["start"] + 1,
        }
        draw_mobile_cover(catalog, brand, section["products"], brand_display_plan, page_w, page_h, margin, accent)
        catalog.showPage()

        draw_mobile_finder(catalog, brand, section_plan, page_w, page_h, margin, accent)
        draw_mobile_footer(catalog, brand, section["finder"], plan["total_pages"], page_w, margin)
        catalog.showPage()

        for series in section_plan["series_order"]:
            meta = section_plan["series_ranges"][series]
            bookmark(
                catalog,
                f"{brand} / {series}",
                f"combined-mobile-series-{slugify_key(brand)}-{slugify_key(series)}",
                1,
            )
            draw_mobile_series_divider(catalog, brand, series, meta, page_w, page_h, margin, accent)
            draw_mobile_footer(catalog, brand, meta["divider"], plan["total_pages"], page_w, margin)
            catalog.showPage()
            for product in meta["products"]:
                draw_mobile_product_page(
                    catalog,
                    brand,
                    product,
                    section_plan["product_page_by_sku"][product.sku],
                    plan["total_pages"],
                    page_w,
                    page_h,
                    margin,
                    accent,
                )
                catalog.showPage()

    bookmark(catalog, "SKU / Name Index", "combined-mobile-sku-name-index", 0)
    for page_number, entries in enumerate(plan["index_pages"], start=plan["index_start_page"]):
        draw_combined_mobile_index_page(
            catalog,
            entries,
            page_number,
            plan["total_pages"],
            page_w,
            page_h,
            margin,
            BRAND_ACCENTS["JIESTAR"],
        )
        catalog.showPage()

    bookmark(catalog, "Contact", "combined-mobile-contact", 0)
    draw_mobile_contact(catalog, "JIESTAR", plan["contact_page"], plan["total_pages"], page_w, page_h, margin, BRAND_ACCENTS["JIESTAR"])
    catalog.save()
    write_combined_mobile_page_index(plan)


def write_combined_mobile_page_index(plan: dict[str, Any]) -> None:
    path = OUTPUT_ROOT / "all-brands-catalog-mobile-page-index.csv"
    fields = ["brand", "sku", "name", "series", "page"]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for entries in plan["index_pages"]:
            writer.writerows(entries)


def draw_combined_mobile_cover(
    c: pdf_canvas.Canvas,
    plan: dict[str, Any],
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    all_products = [product for section in plan["brand_sections"] for product in section["products"]]
    c.setFillColor(colors.white)
    c.rect(0, 0, page_w, page_h, stroke=0, fill=1)
    c.setFillColor(colors.HexColor("#F4F6F8"))
    c.rect(0, 0, page_w, 52 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor(accent))
    c.rect(0, page_h - 42 * mm, page_w, 42 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor("#B51217"))
    c.rect(0, page_h - 42 * mm, page_w, 5 * mm, stroke=0, fill=1)
    draw_logo(c, "JIESTAR", margin, page_h - 22 * mm, 22 * mm, 13 * mm)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 5.2)
    c.drawRightString(page_w - margin, page_h - 13 * mm, "ALL BRAND MOBILE GUIDE")
    c.setFont("Helvetica-Bold", 17.5)
    c.drawString(margin, page_h - 31 * mm, "2026 Product Catalog")
    c.setFont("Helvetica", 6.2)
    c.drawString(margin, page_h - 36.5 * mm, "All brand building block selections for global buyers")

    hero_products = high_piece_products(all_products, 3)
    if hero_products:
        draw_mobile_accent_label(c, "ALL BRAND SHOWCASE", margin, page_h - 50 * mm, accent)
        draw_mobile_showcase_card(c, hero_products[0], margin, 91 * mm, page_w - 2 * margin, 61 * mm, accent, main=True)

    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(margin, 81 * mm, "Featured high-piece sets")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 5.2)
    c.drawRightString(page_w - margin, 81 * mm, "Across all brands")
    for index, product in enumerate(hero_products[1:3]):
        card_w = (page_w - 2 * margin - 4 * mm) / 2
        x = margin + index * (card_w + 4 * mm)
        draw_mobile_showcase_card(c, product, x, 50 * mm, card_w, 27 * mm, accent)

    stats = [
        ("Products", str(plan["product_count"])),
        ("Brands", str(len(plan["brand_sections"]))),
        ("Pages", str(plan["total_pages"])),
    ]
    stat_y = 25 * mm
    stat_w = (page_w - 2 * margin - 6 * mm) / 3
    for index, (label, value) in enumerate(stats):
        x = margin + index * (stat_w + 3 * mm)
        draw_mobile_metric(c, label, value, x, stat_y, stat_w, 17 * mm, accent, inverted=(index == 0))
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 5.6)
    c.drawString(margin, 13 * mm, "Organized by brand, series, SKU, and product specifications.")
    c.setFillColor(colors.HexColor(accent))
    c.rect(margin, 9 * mm, page_w - 2 * margin, 1.2 * mm, stroke=0, fill=1)
    c.restoreState()


def draw_combined_mobile_brand_finder(
    c: pdf_canvas.Canvas,
    plan: dict[str, Any],
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    draw_mobile_header(c, "JIESTAR", "All Brands / Brand Finder", page_w, page_h, margin, accent)
    y = page_h - 28 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 18)
    c.drawString(margin, y, "Brand Finder")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 6.2)
    c.drawString(margin, y - 6 * mm, "Start by brand, then use each brand Product Finder for series navigation.")

    y -= 18 * mm
    stats = [
        ("Products", plan["product_count"]),
        ("Brands", len(plan["brand_sections"])),
        ("Index", f"Page {plan['index_start_page']}"),
    ]
    card_w = (page_w - 2 * margin - 6 * mm) / 3
    for index, (label, value) in enumerate(stats):
        x = margin + index * (card_w + 3 * mm)
        draw_mobile_metric(c, label, str(value), x, y, card_w, 15 * mm, accent, inverted=(index == 0))

    y -= 11 * mm
    c.setFillColor(colors.HexColor(DEEP))
    c.roundRect(margin, y - 7 * mm, page_w - 2 * margin, 7 * mm, 2.5, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 5.3)
    c.drawString(margin + 2 * mm, y - 4.5 * mm, "BRAND")
    c.drawString(page_w - margin - 43 * mm, y - 4.5 * mm, "PRODUCTS")
    c.drawString(page_w - margin - 27 * mm, y - 4.5 * mm, "SERIES")
    c.drawRightString(page_w - margin - 2 * mm, y - 4.5 * mm, "PAGES")

    row_h = 8.5 * mm
    y -= 7 * mm
    for index, section in enumerate(plan["brand_sections"]):
        y -= row_h
        brand = section["brand"]
        c.setFillColor(colors.HexColor("#FFFFFF" if index % 2 == 0 else "#F8FAFC"))
        c.rect(margin, y, page_w - 2 * margin, row_h, stroke=0, fill=1)
        c.setStrokeColor(colors.HexColor("#E5EAF0"))
        c.line(margin, y, page_w - margin, y)
        c.setFillColor(colors.HexColor(BRAND_ACCENTS.get(brand, accent)))
        c.roundRect(margin + 2 * mm, y + 2.5 * mm, 2.2 * mm, 2.2 * mm, 1.1, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 6.2)
        c.drawString(margin + 6 * mm, y + 3 * mm, fit_single_line(brand, "Helvetica-Bold", 6.2, 34 * mm))
        c.setFont("Helvetica", 6)
        c.drawString(page_w - margin - 40 * mm, y + 3 * mm, str(len(section["products"])))
        c.drawString(page_w - margin - 25 * mm, y + 3 * mm, str(len(section["plan"]["series_order"])))
        c.drawRightString(page_w - margin - 2 * mm, y + 3 * mm, f"{section['cover']}-{section['end']}")
    c.restoreState()


def draw_combined_mobile_index_page(
    c: pdf_canvas.Canvas,
    entries: list[dict[str, Any]],
    page_number: int,
    total_pages: int,
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    draw_mobile_header(c, "JIESTAR", "All Brands / SKU - Name Index", page_w, page_h, margin, accent)
    y = page_h - 28 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 15)
    c.drawString(margin, y, "All Brand SKU Index")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 5.8)
    c.drawString(margin, y - 5.5 * mm, "Sorted by brand and Item No. Use page numbers to find product detail pages.")
    y -= 14 * mm
    c.setFillColor(colors.HexColor(DEEP))
    c.roundRect(margin, y - 6 * mm, page_w - 2 * margin, 6 * mm, 2.5, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 5.4)
    c.drawString(margin + 2 * mm, y - 4 * mm, "BRAND / ITEM NO. / PRODUCT / PAGE")
    y -= 8 * mm
    row_h = 6.1 * mm
    for index, entry in enumerate(entries):
        yy = y - index * row_h
        c.setFillColor(colors.HexColor("#FFFFFF" if index % 2 == 0 else "#F8FAFC"))
        c.rect(margin, yy - row_h + 0.2 * mm, page_w - 2 * margin, row_h, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(BRAND_ACCENTS.get(entry["brand"], accent)))
        c.setFont("Helvetica-Bold", 4.7)
        c.drawString(margin + 1.5 * mm, yy - 2 * mm, fit_single_line(entry["brand"], "Helvetica-Bold", 4.7, 18 * mm))
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 5.7)
        c.drawString(margin + 21 * mm, yy - 2 * mm, fit_single_line(entry["sku"], "Helvetica-Bold", 5.7, 23 * mm))
        c.setFillColor(colors.HexColor(accent))
        c.drawRightString(page_w - margin - 1.5 * mm, yy - 2 * mm, f"P{entry['page']}")
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica", 4.8)
        name = f"{entry['name']} - {entry['series']}"
        c.drawString(margin + 1.5 * mm, yy - 4.5 * mm, fit_single_line(name, "Helvetica", 4.8, page_w - 2 * margin - 3 * mm))
    c.restoreState()
    draw_mobile_footer(c, "All Brands", page_number, total_pages, page_w, margin)


def draw_mobile_accent_label(
    c: pdf_canvas.Canvas,
    text: str,
    x: float,
    y: float,
    accent: str,
    *,
    fill: str = "#FFF1F2",
    text_color: str | None = None,
) -> None:
    c.setFillColor(colors.HexColor(fill))
    width = min(max(stringWidth(text, "Helvetica-Bold", 4.8) + 7 * mm, 18 * mm), 48 * mm)
    c.roundRect(x, y, width, 6 * mm, 3, stroke=0, fill=1)
    c.setFillColor(colors.HexColor(text_color or accent))
    c.setFont("Helvetica-Bold", 4.8)
    c.drawCentredString(x + width / 2, y + 2.1 * mm, text)


def draw_mobile_metric(
    c: pdf_canvas.Canvas,
    label: str,
    value: str,
    x: float,
    y: float,
    w: float,
    h: float,
    accent: str,
    *,
    inverted: bool = False,
) -> None:
    c.setFillColor(colors.HexColor(accent if inverted else "#F7F9FC"))
    c.roundRect(x, y, w, h, 4, stroke=0, fill=1)
    inverted_text = contrast_text_color(accent)
    c.setFillColor(inverted_text if inverted else colors.HexColor(MUTED))
    c.setFont("Helvetica-Bold", 4.5)
    c.drawString(x + 2.4 * mm, y + h - 5 * mm, label.upper())
    c.setFillColor(inverted_text if inverted else colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 8.4)
    c.drawString(x + 2.4 * mm, y + 3.2 * mm, fit_single_line(value, "Helvetica-Bold", 8.4, w - 4.8 * mm))


def draw_mobile_showcase_card(
    c: pdf_canvas.Canvas,
    product: Product,
    x: float,
    y: float,
    w: float,
    h: float,
    accent: str,
    *,
    main: bool = False,
) -> None:
    c.setFillColor(colors.HexColor("#DDE4EC"))
    c.roundRect(x + 0.8 * mm, y - 0.8 * mm, w, h, 6, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.roundRect(x, y, w, h, 6, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor("#E7ECF2"))
    c.roundRect(x, y, w, h, 6, stroke=1, fill=0)

    image_h = h - (16 * mm if main else 11 * mm)
    if product.image_path and Path(product.image_path).exists():
        draw_image_fit(c, product.image_path, x + 4 * mm, y + h - image_h - 4 * mm, w - 8 * mm, image_h - 4 * mm)
    else:
        draw_placeholder(c, x + 2 * mm, y + h - image_h - 3 * mm, w - 4 * mm, image_h - 4 * mm, "[Product Image]")

    if main:
        c.setFillColor(colors.HexColor(accent))
        c.roundRect(x + 4 * mm, y + 5 * mm, 25 * mm, 7 * mm, 3.5, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 5.2)
        c.drawCentredString(x + 16.5 * mm, y + 7.6 * mm, fit_single_line(product.sku, "Helvetica-Bold", 5.2, 21 * mm))
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 7.2)
        c.drawRightString(x + w - 4 * mm, y + 8 * mm, piece_showcase_text(product))
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica", 4.8)
        c.drawString(x + 4 * mm, y + 2.6 * mm, fit_single_line(product.name_en, "Helvetica", 4.8, w - 8 * mm))
        return

    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 5.2)
    c.drawString(x + 3 * mm, y + 4.7 * mm, fit_single_line(product.sku, "Helvetica-Bold", 5.2, w * 0.42))
    c.setFillColor(colors.HexColor(accent))
    c.drawRightString(x + w - 3 * mm, y + 4.7 * mm, piece_showcase_text(product))
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 4.4)
    c.drawString(x + 3 * mm, y + 2 * mm, fit_single_line(product.name_en, "Helvetica", 4.4, w - 6 * mm))


def draw_mobile_cover(
    c: pdf_canvas.Canvas,
    brand: str,
    products: list[Product],
    plan: dict[str, Any],
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    c.setFillColor(colors.white)
    c.rect(0, 0, page_w, page_h, stroke=0, fill=1)
    c.setFillColor(colors.HexColor("#F4F6F8"))
    c.rect(0, 0, page_w, 52 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor(accent))
    c.rect(0, page_h - 42 * mm, page_w, 42 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor("#B51217"))
    c.rect(0, page_h - 42 * mm, page_w, 5 * mm, stroke=0, fill=1)
    draw_logo(c, brand, margin, page_h - 22 * mm, 22 * mm, 13 * mm)
    c.setFillColor(contrast_text_color(accent))
    c.setFont("Helvetica-Bold", 5.2)
    c.drawRightString(page_w - margin, page_h - 13 * mm, "MOBILE B2B GUIDE")
    c.setFont("Helvetica-Bold", 17.5)
    c.drawString(margin, page_h - 31 * mm, "2026 Product Catalog")
    c.setFont("Helvetica", 6.2)
    c.drawString(margin, page_h - 36.5 * mm, "Building block selections for global buyers")

    hero_products = high_piece_products(products, 3)
    has_pieces = has_piece_counts(products)
    if hero_products:
        cover_label = "HIGH PIECE COUNT SHOWCASE" if has_pieces else "FEATURED PRODUCT SHOWCASE"
        draw_mobile_accent_label(c, cover_label, margin, page_h - 50 * mm, accent)
        draw_mobile_showcase_card(c, hero_products[0], margin, 91 * mm, page_w - 2 * margin, 61 * mm, accent, main=True)

    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(margin, 81 * mm, "Featured high-piece sets" if has_pieces else "Featured products")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 5.2)
    c.drawRightString(page_w - margin, 81 * mm, "Sorted by piece count" if has_pieces else "Selected from catalog")
    for index, product in enumerate(hero_products[1:3]):
        card_w = (page_w - 2 * margin - 4 * mm) / 2
        x = margin + index * (card_w + 4 * mm)
        draw_mobile_showcase_card(c, product, x, 50 * mm, card_w, 27 * mm, accent)

    stats = [
        ("Products", str(len(products))),
        ("Series", str(len(plan["series_order"]))),
        ("Pages", str(plan["total_pages"])),
    ]
    stat_y = 25 * mm
    stat_w = (page_w - 2 * margin - 6 * mm) / 3
    for index, (label, value) in enumerate(stats):
        x = margin + index * (stat_w + 3 * mm)
        draw_mobile_metric(c, label, value, x, stat_y, stat_w, 17 * mm, accent, inverted=(index == 0))
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 5.6)
    c.drawString(margin, 13 * mm, "For distributors, wholesale buyers, OEM / ODM, and product cooperation.")
    c.setFillColor(colors.HexColor(accent))
    c.rect(margin, 9 * mm, page_w - 2 * margin, 1.2 * mm, stroke=0, fill=1)
    c.restoreState()


def draw_mobile_finder(
    c: pdf_canvas.Canvas,
    brand: str,
    plan: dict[str, Any],
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    draw_mobile_header(c, brand, f"{brand} / Product Finder", page_w, page_h, margin, accent)
    y = page_h - 28 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 18)
    c.drawString(margin, y, "Product Finder")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 6.5)
    c.drawString(margin, y - 6 * mm, "Choose a series, then use Item No. for inquiries and page search.")

    y -= 18 * mm
    stats = [
        ("Products", sum(meta["count"] for meta in plan["series_ranges"].values())),
        ("Series", len(plan["series_order"])),
        ("Index", f"Page {plan['index_start_page']}"),
    ]
    card_w = (page_w - 2 * margin - 6 * mm) / 3
    for index, (label, value) in enumerate(stats):
        x = margin + index * (card_w + 3 * mm)
        c.setFillColor(colors.HexColor(SOFT))
        c.roundRect(x, y, card_w, 15 * mm, 3, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica-Bold", 4.8)
        c.drawString(x + 2 * mm, y + 9 * mm, label.upper())
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 8)
        c.drawString(x + 2 * mm, y + 4 * mm, str(value))

    y -= 11 * mm
    c.setFillColor(colors.HexColor(DEEP))
    c.roundRect(margin, y - 7 * mm, page_w - 2 * margin, 7 * mm, 2.5, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 5.3)
    c.drawString(margin + 2 * mm, y - 4.5 * mm, "SERIES")
    c.drawString(page_w - margin - 34 * mm, y - 4.5 * mm, "PRODUCTS")
    c.drawRightString(page_w - margin - 2 * mm, y - 4.5 * mm, "PAGES")

    row_h = 8 * mm
    y -= 7 * mm
    for index, series in enumerate(plan["series_order"]):
        meta = plan["series_ranges"][series]
        y -= row_h
        c.setFillColor(colors.HexColor("#FFFFFF" if index % 2 == 0 else "#F8FAFC"))
        c.rect(margin, y, page_w - 2 * margin, row_h, stroke=0, fill=1)
        c.setStrokeColor(colors.HexColor("#E5EAF0"))
        c.line(margin, y, page_w - margin, y)
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 6.2)
        c.drawString(margin + 2 * mm, y + 3 * mm, fit_single_line(series, "Helvetica-Bold", 6.2, 46 * mm))
        c.setFont("Helvetica", 6)
        c.drawString(page_w - margin - 31 * mm, y + 3 * mm, str(meta["count"]))
        c.drawRightString(page_w - margin - 2 * mm, y + 3 * mm, f"{meta['divider']}-{meta['end']}")
    c.restoreState()


def draw_mobile_spec_guide(
    c: pdf_canvas.Canvas,
    brand: str,
    plan: dict[str, Any],
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    draw_mobile_header(c, brand, f"{brand} / Specification Guide", page_w, page_h, margin, accent)
    y = page_h - 28 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, y, "Specification Guide")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 6.2)
    c.drawString(margin, y - 6 * mm, "Use these fields to compare products before confirming samples and orders.")

    items = [
        ("Item No.", "Primary reference for inquiry and ordering."),
        ("Pieces", "Piece count for model scale comparison."),
        ("Carton Qty", "Units packed in one master carton."),
        ("Color Box", "Individual product package size."),
        ("Outer Carton", "Master carton dimensions."),
        ("Weight KG", "Gross / net carton weight."),
        ("Model Size", "Finished model dimensions."),
        ("Age", "Recommended age when supplied."),
    ]
    y -= 20 * mm
    card_h = 16 * mm
    for index, (label, detail) in enumerate(items):
        x = margin + (index % 2) * ((page_w - 2 * margin - 4 * mm) / 2 + 4 * mm)
        yy = y - (index // 2) * (card_h + 5 * mm)
        w = (page_w - 2 * margin - 4 * mm) / 2
        c.setFillColor(colors.HexColor(SOFT))
        c.roundRect(x, yy - card_h, w, card_h, 3.5, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(accent))
        c.setFont("Helvetica-Bold", 5.5)
        c.drawString(x + 2.5 * mm, yy - 5 * mm, label.upper())
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica", 5.5)
        for line_index, line in enumerate(fit_canvas_lines(detail, "Helvetica", 5.5, w - 5 * mm, 2)):
            c.drawString(x + 2.5 * mm, yy - 10 * mm - line_index * 2.8 * mm, line)

    y = 34 * mm
    c.setFillColor(colors.HexColor(DEEP))
    c.roundRect(margin, y, page_w - 2 * margin, 24 * mm, 4, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(margin + 4 * mm, y + 15 * mm, "Fast lookup")
    c.setFont("Helvetica", 5.8)
    c.setFillColor(colors.HexColor("#CBD5E1"))
    c.drawString(margin + 4 * mm, y + 9 * mm, "Use PDF bookmarks for series navigation.")
    c.drawString(margin + 4 * mm, y + 4.5 * mm, f"SKU / Name Index starts on page {plan['index_start_page']}.")
    c.restoreState()


def draw_mobile_cooperation(
    c: pdf_canvas.Canvas,
    brand: str,
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    draw_mobile_header(c, brand, f"{brand} / Brand and Cooperation", page_w, page_h, margin, accent)
    y = page_h - 30 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, y, "Build Product Lines")
    c.drawString(margin, y - 8 * mm, f"With {fit_single_line(brand, 'Helvetica-Bold', 16, 50 * mm)}")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 6.3)
    copy = f"{brand} supports global partners with product selection, wholesale supply, OEM / ODM customization, product co-development, and long-term sub-brand cooperation."
    for index, line in enumerate(fit_canvas_lines(copy, "Helvetica", 6.3, page_w - 2 * margin, 4)):
        c.drawString(margin, y - 20 * mm - index * 4 * mm, line)

    items = [
        ("Wholesale Supply", "Existing product selections for channel and ecommerce buyers."),
        ("OEM / ODM", "Customization discussions for product direction and packaging."),
        ("Co-Development", "Product planning for differentiated market needs."),
        ("Sub-Brand Cooperation", "Long-term product line and brand direction support."),
    ]
    card_y = y - 45 * mm
    for index, (title, detail) in enumerate(items):
        yy = card_y - index * 22 * mm
        c.setFillColor(colors.HexColor(SOFT))
        c.roundRect(margin, yy - 16 * mm, page_w - 2 * margin, 16 * mm, 4, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(accent))
        c.setFont("Helvetica-Bold", 6.2)
        c.drawString(margin + 3 * mm, yy - 6 * mm, title.upper())
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica", 5.6)
        c.drawString(margin + 3 * mm, yy - 11 * mm, fit_single_line(detail, "Helvetica", 5.6, page_w - 2 * margin - 6 * mm))
    c.restoreState()


def draw_mobile_series_divider(
    c: pdf_canvas.Canvas,
    brand: str,
    series: str,
    meta: dict[str, Any],
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    c.setFillColor(colors.white)
    c.rect(0, 0, page_w, page_h, stroke=0, fill=1)
    c.setFillColor(colors.HexColor("#F4F6F8"))
    c.rect(0, 0, page_w, 62 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor(accent))
    c.rect(0, page_h - 18 * mm, page_w, 18 * mm, stroke=0, fill=1)
    draw_logo(c, brand, margin, page_h - 13.8 * mm, 14 * mm, 8 * mm)
    c.setFillColor(contrast_text_color(accent))
    c.setFont("Helvetica-Bold", 5.2)
    c.drawRightString(page_w - margin, page_h - 10 * mm, "SERIES FOCUS")

    has_pieces = has_piece_counts(meta["products"])
    series_label = "TOP PIECE COUNT PRODUCTS" if has_pieces else "FEATURED PRODUCTS"
    draw_mobile_accent_label(c, series_label, margin, page_h - 31 * mm, accent)
    c.setFillColor(colors.HexColor(INK))
    title_y = page_h - 44 * mm
    title_lines = fit_canvas_lines(series, "Helvetica-Bold", 20, page_w - 2 * margin, 2)
    c.setFont("Helvetica-Bold", 20)
    for index, line in enumerate(title_lines):
        c.drawString(margin, title_y - index * 8.8 * mm, line)

    stats_y = page_h - 67 * mm if len(title_lines) == 1 else page_h - 75 * mm
    stat_w = (page_w - 2 * margin - 4 * mm) / 2
    draw_mobile_metric(c, "Products", str(meta["count"]), margin, stats_y, stat_w, 15 * mm, accent, inverted=True)
    draw_mobile_metric(
        c,
        "Product Pages",
        f"{meta['start']}-{meta['end']}",
        margin + stat_w + 4 * mm,
        stats_y,
        stat_w,
        15 * mm,
        accent,
    )

    sample_products = high_piece_products(meta["products"], 3)
    if sample_products:
        draw_mobile_showcase_card(c, sample_products[0], margin, 67 * mm, page_w - 2 * margin, 59 * mm, accent, main=True)

    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 7)
    c.drawString(margin, 56 * mm, "More high-piece options" if has_pieces else "More featured options")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 4.9)
    c.drawRightString(page_w - margin, 56 * mm, "Use SKU for inquiry")
    for index, product in enumerate(sample_products[1:3]):
        card_w = (page_w - 2 * margin - 4 * mm) / 2
        x = margin + index * (card_w + 4 * mm)
        draw_mobile_showcase_card(c, product, x, 26 * mm, card_w, 26 * mm, accent)
    c.setFillColor(colors.HexColor(accent))
    c.rect(margin, 14 * mm, page_w - 2 * margin, 1.2 * mm, stroke=0, fill=1)
    c.restoreState()


def draw_mobile_product_page(
    c: pdf_canvas.Canvas,
    brand: str,
    product: Product,
    page_number: int,
    total_pages: int,
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    draw_mobile_header(c, brand, f"{brand} / {product.series}", page_w, page_h, margin, accent)
    image_y = page_h - 86 * mm
    image_h = 56 * mm
    c.setFillColor(colors.HexColor(SOFT))
    c.roundRect(margin, image_y, page_w - 2 * margin, image_h, 6, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor("#E8EEF5"))
    c.roundRect(margin + 1 * mm, image_y + 1 * mm, page_w - 2 * margin - 2 * mm, image_h - 2 * mm, 5, stroke=1, fill=0)
    if product.image_path and Path(product.image_path).exists():
        draw_image_fit(c, product.image_path, margin + 4 * mm, image_y + 4 * mm, page_w - 2 * margin - 8 * mm, image_h - 8 * mm)
    else:
        draw_placeholder(c, margin, image_y, page_w - 2 * margin, image_h, "[Product Image Placeholder]")

    title_y = image_y - 7 * mm
    c.setFillColor(colors.HexColor(accent))
    c.roundRect(margin, title_y - 6 * mm, 21 * mm, 5.5 * mm, 2.7, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 5.4)
    c.drawCentredString(margin + 10.5 * mm, title_y - 3.9 * mm, "ITEM NO.")
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 9.2)
    c.drawString(margin + 25 * mm, title_y - 4.2 * mm, fit_single_line(product.sku, "Helvetica-Bold", 9.2, page_w - 2 * margin - 25 * mm))

    name_y = title_y - 12 * mm
    c.setFont("Helvetica-Bold", 11)
    for index, line in enumerate(fit_canvas_lines(product.name_en, "Helvetica-Bold", 11, page_w - 2 * margin, 2)):
        c.drawString(margin, name_y - index * 5 * mm, line)

    fields = [
        ("Pieces", product.piece_count),
        ("Carton Qty", product.carton_qty),
        ("Color Box", product.color_box_size),
        ("Outer Carton", product.outer_carton_size),
        ("Weight KG", product.gross_net_weight_kg),
        ("Model Size", product.model_size_cm),
        ("Age", product.recommended_age),
        ("Series", product.series),
    ]
    grid_top = name_y - 16 * mm
    col_gap = 3 * mm
    row_gap = 2.2 * mm
    cell_w = (page_w - 2 * margin - col_gap) / 2
    cell_h = 11.5 * mm
    for index, (label, value) in enumerate(fields):
        col = index % 2
        row = index // 2
        x = margin + col * (cell_w + col_gap)
        y = grid_top - row * (cell_h + row_gap)
        c.setFillColor(colors.HexColor("#F7F9FC"))
        c.roundRect(x, y - cell_h, cell_w, cell_h, 3, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(accent) if index in (0, 1) else colors.HexColor(MUTED))
        c.setFont("Helvetica-Bold", 5.2)
        c.drawString(x + 2.5 * mm, y - 5 * mm, label.upper())
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold" if index in (0, 1) else "Helvetica", 6.2)
        c.drawString(x + 2.5 * mm, y - 9 * mm, fit_single_line(value, "Helvetica", 6.2, cell_w - 5 * mm))
    c.restoreState()
    draw_mobile_footer(c, brand, page_number, total_pages, page_w, margin)


def draw_mobile_index_page(
    c: pdf_canvas.Canvas,
    brand: str,
    entries: list[dict[str, Any]],
    page_number: int,
    total_pages: int,
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    draw_mobile_header(c, brand, f"{brand} / SKU - Name Index", page_w, page_h, margin, accent)
    y = page_h - 28 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 15)
    c.drawString(margin, y, "SKU / Name Index")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 5.8)
    c.drawString(margin, y - 5.5 * mm, "Sorted by Item No. Use page numbers to find product detail pages.")
    y -= 14 * mm
    c.setFillColor(colors.HexColor(DEEP))
    c.roundRect(margin, y - 6 * mm, page_w - 2 * margin, 6 * mm, 2.5, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 5.4)
    c.drawString(margin + 2 * mm, y - 4 * mm, "ITEM NO. / PRODUCT / PAGE")
    y -= 8 * mm
    row_h = 6.1 * mm
    for index, entry in enumerate(entries):
        yy = y - index * row_h
        c.setFillColor(colors.HexColor("#FFFFFF" if index % 2 == 0 else "#F8FAFC"))
        c.rect(margin, yy - row_h + 0.2 * mm, page_w - 2 * margin, row_h, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 5.7)
        c.drawString(margin + 1.5 * mm, yy - 2 * mm, fit_single_line(entry["sku"], "Helvetica-Bold", 5.7, 23 * mm))
        c.setFillColor(colors.HexColor(accent))
        c.drawRightString(page_w - margin - 1.5 * mm, yy - 2 * mm, f"P{entry['page']}")
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica", 4.9)
        name = f"{entry['name']} - {entry['series']}"
        c.drawString(margin + 1.5 * mm, yy - 4.5 * mm, fit_single_line(name, "Helvetica", 4.9, page_w - 2 * margin - 3 * mm))
    c.restoreState()
    draw_mobile_footer(c, brand, page_number, total_pages, page_w, margin)


def draw_mobile_contact(
    c: pdf_canvas.Canvas,
    brand: str,
    page_number: int,
    total_pages: int,
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    c.setFillColor(colors.HexColor(DEEP))
    c.rect(0, 0, page_w, page_h, stroke=0, fill=1)
    draw_logo(c, brand, margin, page_h - 28 * mm, 24 * mm, 14 * mm)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(margin, page_h - 50 * mm, "Contact JIESTAR")
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor("#CBD5E1"))
    c.drawString(margin, page_h - 59 * mm, f"For {brand} wholesale supply, OEM / ODM, and product cooperation.")
    contacts = [
        ("Website", "www.jiestartoys.com"),
        ("Business Email", "info@jiestartoys.com"),
        ("Support Email", "support@jiestartoys.com"),
        ("WhatsApp", "[WhatsApp]"),
        ("Phone", "[Phone]"),
        ("Address", "[Address]"),
    ]
    y = page_h - 82 * mm
    for label, value in contacts:
        c.setFillColor(colors.HexColor("#1F2937"))
        c.roundRect(margin, y - 14 * mm, page_w - 2 * margin, 12 * mm, 4, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(accent))
        c.setFont("Helvetica-Bold", 5.5)
        c.drawString(margin + 3 * mm, y - 6 * mm, label.upper())
        c.setFillColor(colors.white)
        c.setFont("Helvetica", 7)
        c.drawString(margin + 3 * mm, y - 10.5 * mm, fit_single_line(value, "Helvetica", 7, page_w - 2 * margin - 6 * mm))
        y -= 17 * mm
    c.restoreState()
    draw_mobile_footer(c, brand, page_number, total_pages, page_w, margin)


def draw_mobile_header(c: pdf_canvas.Canvas, brand: str, label: str, page_w: float, page_h: float, margin: float, accent: str) -> None:
    c.saveState()
    y = page_h - 12 * mm
    draw_logo(c, brand, margin, y - 3.7 * mm, 12 * mm, 7.3 * mm)
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 5.5)
    c.drawString(
        margin + 15 * mm,
        y - 0.4 * mm,
        fit_single_line(label, "Helvetica-Bold", 5.5, page_w - 2 * margin - 35 * mm),
    )
    c.setStrokeColor(colors.HexColor("#E8EEF5"))
    c.line(margin, y - 7 * mm, page_w - margin, y - 7 * mm)
    c.setFillColor(colors.HexColor(accent))
    c.roundRect(page_w - margin - 16 * mm, y - 4.7 * mm, 16 * mm, 4 * mm, 2, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 3.7)
    c.drawCentredString(page_w - margin - 8 * mm, y - 3.2 * mm, "9:16")
    c.restoreState()


def draw_mobile_footer(c: pdf_canvas.Canvas, brand: str, page_number: int, total_pages: int, page_w: float, margin: float) -> None:
    c.saveState()
    y = 7 * mm
    c.setStrokeColor(colors.HexColor("#E8EEF5"))
    c.line(margin, y + 3.5 * mm, page_w - margin, y + 3.5 * mm)
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.setFont("Helvetica", 4.8)
    c.drawString(margin, y, f"{brand} 2026 Product Catalog")
    c.drawRightString(page_w - margin, y, f"Page {page_number} / {total_pages}")
    c.restoreState()


def bookmark(c: pdf_canvas.Canvas, title: str, key: str, level: int) -> None:
    c.bookmarkPage(key)
    c.addOutlineEntry(pdf_safe_text(title), key, level=level, closed=False)


def slugify_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-") or "section"


def draw_jiestar_finder_page(
    c: pdf_canvas.Canvas,
    products: list[Product],
    plan: dict[str, Any],
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    draw_jiestar_header(c, "JIESTAR / Product Finder", page_w, page_h, margin, accent)
    title_y = page_h - margin - 27 * mm
    draw_logo(c, "JIESTAR", margin, title_y + 3 * mm, 30 * mm, 14 * mm)
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 22)
    c.drawString(margin + 36 * mm, title_y + 10 * mm, "Product Finder")
    c.setFont("Helvetica", 8.5)
    c.setFillColor(colors.HexColor(MUTED))
    c.drawString(
        margin + 36 * mm,
        title_y + 3 * mm,
        "Start by series, then confirm specifications by Item No. Use the PDF bookmarks for quick jumps.",
    )

    stats = [
        ("Products", str(len(products))),
        ("Series", str(len(plan["series_ranges"]))),
        ("Product Pages", str(len(plan["product_pages"]))),
        ("Index Starts", f"Page {plan['index_start_page']}"),
    ]
    stat_w = (page_w - 2 * margin - 18 * mm) / 4
    stat_y = title_y - 19 * mm
    for index, (label, value) in enumerate(stats):
        x = margin + index * (stat_w + 6 * mm)
        c.setFillColor(colors.HexColor("#F7F9FC"))
        c.roundRect(x, stat_y, stat_w, 14 * mm, 4, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica-Bold", 5.8)
        c.drawString(x + 3 * mm, stat_y + 9 * mm, label.upper())
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x + 3 * mm, stat_y + 4 * mm, value)

    table_y = stat_y - 13 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 12)
    c.drawString(margin, table_y, "Series Contents")
    table_top = table_y - 6 * mm
    col_x = [margin, margin + 82 * mm, margin + 110 * mm, margin + 139 * mm]
    headers = ("Series", "Products", "Pages", "Buyer Use")
    c.setFillColor(colors.HexColor(DEEP))
    c.roundRect(margin, table_top - 7 * mm, page_w - 2 * margin, 7 * mm, 2.5, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 6.5)
    for x, header in zip(col_x, headers):
        c.drawString(x + 2 * mm, table_top - 4.8 * mm, header.upper())

    row_h = 7.6 * mm
    y = table_top - 7 * mm
    for idx, (series, meta) in enumerate(plan["series_ranges"].items()):
        y -= row_h
        c.setFillColor(colors.HexColor("#FFFFFF" if idx % 2 == 0 else "#F8FAFC"))
        c.rect(margin, y, page_w - 2 * margin, row_h, stroke=0, fill=1)
        c.setStrokeColor(colors.HexColor("#E5EAF0"))
        c.line(margin, y, page_w - margin, y)
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 7.1)
        c.drawString(col_x[0] + 2 * mm, y + 2.6 * mm, fit_single_line(series, "Helvetica-Bold", 7.1, 76 * mm))
        c.setFont("Helvetica", 7)
        c.drawString(col_x[1] + 2 * mm, y + 2.6 * mm, str(meta["count"]))
        page_label = f"{meta['start']}-{meta['end']}" if meta["start"] != meta["end"] else str(meta["start"])
        c.drawString(col_x[2] + 2 * mm, y + 2.6 * mm, page_label)
        c.setFillColor(colors.HexColor(MUTED))
        c.drawString(
            col_x[3] + 2 * mm,
            y + 2.6 * mm,
            fit_single_line(series_buyer_hint(series), "Helvetica", 7, page_w - col_x[3] - margin - 4 * mm),
        )

    c.restoreState()


def draw_jiestar_field_guide_page(
    c: pdf_canvas.Canvas,
    plan: dict[str, Any],
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    draw_jiestar_header(c, "JIESTAR / Specification Field Guide", page_w, page_h, margin, accent)
    title_y = page_h - margin - 25 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 20)
    c.drawString(margin, title_y, "Specification Field Guide")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 8)
    c.drawString(
        margin,
        title_y - 6 * mm,
        "Use these fields to shortlist products before confirming samples, packing, and shipment details.",
    )

    steps = [
        ("1", "Start With Series", "Jump from bookmarks or the series contents page."),
        ("2", "Confirm Item No.", "Use Item No. in inquiries to avoid product-name ambiguity."),
        ("3", "Check Logistics", "Compare carton quantity, package size, carton size, and weight."),
    ]
    step_y = title_y - 44 * mm
    step_w = (page_w - 2 * margin - 12 * mm) / 3
    for index, (num, heading, detail) in enumerate(steps):
        x = margin + index * (step_w + 6 * mm)
        c.setFillColor(colors.HexColor("#F7F9FC"))
        c.roundRect(x, step_y, step_w, 28 * mm, 5, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(accent))
        c.circle(x + 8 * mm, step_y + 19 * mm, 4 * mm, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8)
        c.drawCentredString(x + 8 * mm, step_y + 16.7 * mm, num)
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 9)
        c.drawString(x + 16 * mm, step_y + 18 * mm, heading)
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica", 6.7)
        for line_index, line in enumerate(fit_canvas_lines(detail, "Helvetica", 6.7, step_w - 12 * mm, 2)):
            c.drawString(x + 6 * mm, step_y + 9 * mm - line_index * 3.2 * mm, line)

    field_y = step_y - 18 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 12)
    c.drawString(margin, field_y, "Visible Product Fields")
    guide_items = [
        ("Item No.", "Primary ordering and inquiry reference."),
        ("Pieces", "Piece count for scale and value comparison."),
        ("Carton Qty", "Units packed in one master carton."),
        ("Color Box", "Individual product package dimensions."),
        ("Outer Carton", "Master carton dimensions for logistics."),
        ("Weight KG", "Gross / net carton weight when available."),
        ("Model Size", "Finished model dimensions when available."),
        ("Age", "Recommended age when supplied by the source table."),
    ]
    card_w = (page_w - 2 * margin - 12 * mm) / 4
    card_h = 18 * mm
    for index, (label, detail) in enumerate(guide_items):
        x = margin + (index % 4) * (card_w + 4 * mm)
        y = field_y - 8 * mm - (index // 4) * (card_h + 5 * mm) - card_h
        c.setFillColor(colors.HexColor(SOFT))
        c.roundRect(x, y, card_w, card_h, 4, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(accent))
        c.setFont("Helvetica-Bold", 6)
        c.drawString(x + 3 * mm, y + 11 * mm, label.upper())
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica", 6)
        for line_index, line in enumerate(fit_canvas_lines(detail, "Helvetica", 6, card_w - 6 * mm, 2)):
            c.drawString(x + 3 * mm, y + 6.5 * mm - line_index * 3 * mm, line)

    lookup_y = margin + 34 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 12)
    c.drawString(margin, lookup_y, "Fast Lookup Tools")
    lookup_items = [
        ("PDF Bookmarks", "Jump directly to each series in the PDF sidebar."),
        ("Series Contents", f"Product pages begin on page {min(v['start'] for v in plan['series_ranges'].values())}."),
        ("SKU / Name Index", f"Sorted Item No. index starts on page {plan['index_start_page']}."),
    ]
    lookup_y -= 9 * mm
    lookup_w = (page_w - 2 * margin - 10 * mm) / 3
    for index, (label, detail) in enumerate(lookup_items):
        x = margin + index * (lookup_w + 5 * mm)
        c.setFillColor(colors.white)
        c.setStrokeColor(colors.HexColor("#E5EAF0"))
        c.roundRect(x, lookup_y - 18 * mm, lookup_w, 18 * mm, 4, stroke=1, fill=1)
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold", 7)
        c.drawString(x + 3 * mm, lookup_y - 6 * mm, label)
        c.setFillColor(colors.HexColor(MUTED))
        c.setFont("Helvetica", 6)
        c.drawString(x + 3 * mm, lookup_y - 11 * mm, fit_single_line(detail, "Helvetica", 6, lookup_w - 6 * mm))
    c.restoreState()


def series_buyer_hint(series: str) -> str:
    hints = {
        "Aircraft": "Aviation display sets, military aircraft, helicopters.",
        "Animals": "Small collectible animals and animal-themed sets.",
        "Baby and Toddler": "Younger-age sets and early building items.",
        "Building Sets": "General assortment and broad category selection.",
        "Car Model": "Cars, racing, vehicles, and display models.",
        "Fire Rescue": "Emergency rescue and fire-themed sets.",
        "Flowers": "Floral display models and decorative gifts.",
        "Military": "Military vehicles, defense themes, and display sets.",
        "Police": "Police vehicles and city security themes.",
        "Street View": "Architecture, shops, city scenes, and dioramas.",
        "Tank": "Tank and armored vehicle product lines.",
        "Winter Fantasy": "Seasonal winter and fantasy display theme.",
    }
    return hints.get(series, "Browse by theme, then confirm specs by Item No.")


def draw_jiestar_product_page(
    c: pdf_canvas.Canvas,
    page_number: int,
    total_pages: int,
    products: list[Product],
    series: str,
    series_meta: dict[str, int],
    page_w: float,
    page_h: float,
    margin: float,
    header_h: float,
    footer_h: float,
    content_w: float,
    card_h: float,
    gap: float,
    accent: str,
) -> None:
    draw_jiestar_header(c, f"JIESTAR / {series}", page_w, page_h, margin, accent)
    header_bottom = page_h - margin - header_h
    c.saveState()
    c.setFillColor(colors.HexColor(accent))
    c.roundRect(margin, header_bottom + 2 * mm, 36 * mm, 7 * mm, 3.5, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 6.5)
    c.drawCentredString(margin + 18 * mm, header_bottom + 4.4 * mm, series.upper())
    c.setFillColor(colors.HexColor("#EEF2F6"))
    c.roundRect(margin + 40 * mm, header_bottom + 2 * mm, 38 * mm, 7 * mm, 3.5, stroke=0, fill=1)
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica-Bold", 6.3)
    c.drawCentredString(margin + 59 * mm, header_bottom + 4.4 * mm, f"{series_meta['count']} PRODUCTS")
    page_range = f"PAGES {series_meta['start']}-{series_meta['end']}"
    c.drawRightString(page_w - margin, header_bottom + 4.4 * mm, page_range)
    c.restoreState()

    top_card_y = margin + footer_h + card_h + gap
    bottom_card_y = margin + footer_h
    draw_jiestar_product_card(c, products[0], margin, top_card_y, content_w, card_h, accent)
    if len(products) > 1:
        draw_jiestar_product_card(c, products[1], margin, bottom_card_y, content_w, card_h, accent)
    draw_jiestar_footer(c, page_number, total_pages, page_w, margin)


def draw_jiestar_product_card(
    c: pdf_canvas.Canvas,
    product: Product,
    x: float,
    y: float,
    w: float,
    h: float,
    accent: str,
) -> None:
    c.saveState()
    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor(LINE))
    c.roundRect(x, y, w, h, 6, stroke=1, fill=1)
    c.setFillColor(colors.HexColor(accent))
    c.roundRect(x, y, 3.2 * mm, h, 6, stroke=0, fill=1)

    pad = 6 * mm
    image_w = w * 0.39
    image_x = x + pad
    image_y = y + pad
    image_h = h - 2 * pad
    c.setFillColor(colors.HexColor(SOFT))
    c.roundRect(image_x, image_y, image_w, image_h, 5, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor("#E8EEF5"))
    c.roundRect(image_x + 1 * mm, image_y + 1 * mm, image_w - 2 * mm, image_h - 2 * mm, 4, stroke=1, fill=0)
    if product.image_path and Path(product.image_path).exists():
        draw_image_fit(c, product.image_path, image_x + 4 * mm, image_y + 4 * mm, image_w - 8 * mm, image_h - 8 * mm)
    else:
        draw_placeholder(c, image_x, image_y, image_w, image_h, "Image pending")

    text_x = image_x + image_w + 8 * mm
    text_w = x + w - text_x - pad
    top_y = y + h - pad
    c.setFillColor(colors.HexColor(accent))
    c.roundRect(text_x, top_y - 5.5 * mm, 23 * mm, 5 * mm, 2.5, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 5.6)
    c.drawCentredString(text_x + 11.5 * mm, top_y - 3.6 * mm, "JIESTAR")
    c.setFillColor(colors.HexColor("#EEF2F6"))
    c.roundRect(text_x + 26 * mm, top_y - 5.5 * mm, text_w - 26 * mm, 5 * mm, 2.5, stroke=0, fill=1)
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 5.6)
    c.drawString(text_x + 28 * mm, top_y - 3.6 * mm, fit_single_line(product.series.upper(), "Helvetica", 5.6, text_w - 30 * mm))

    title_y = top_y - 11 * mm
    title_lines = fit_canvas_lines(product.name_en, "Helvetica-Bold", 10.4, text_w, 2)
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 10.4)
    for line in title_lines:
        c.drawString(text_x, title_y, line)
        title_y -= 4.6 * mm
    sku_y = title_y - 1 * mm
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor(MUTED))
    c.drawString(text_x, sku_y, fit_single_line(f"Item No. {product.sku}", "Helvetica", 7, text_w))
    c.setStrokeColor(colors.HexColor("#E8EEF5"))
    c.line(text_x, sku_y - 3.2 * mm, text_x + text_w, sku_y - 3.2 * mm)

    fields = [
        ("Pieces", product.piece_count),
        ("Carton Qty", product.carton_qty),
        ("Color Box", product.color_box_size),
        ("Outer Carton", product.outer_carton_size),
        ("Weight KG", product.gross_net_weight_kg),
        ("Model Size", product.model_size_cm),
        ("Age", product.recommended_age),
    ]
    grid_y = sku_y - 8 * mm
    cell_gap = 3 * mm
    col_w = (text_w - cell_gap) / 2
    cell_h = 7.5 * mm
    for index, (label, value) in enumerate(fields):
        col = index % 2
        row = index // 2
        cell_x = text_x + col * (col_w + cell_gap)
        cell_y = grid_y - row * cell_h
        c.setFillColor(colors.HexColor("#F7F9FC"))
        c.roundRect(cell_x, cell_y - cell_h + 0.6 * mm, col_w, cell_h - 1 * mm, 2.3, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(accent) if index in (0, 1) else colors.HexColor(MUTED))
        c.setFont("Helvetica-Bold", 5.1)
        c.drawString(cell_x + 2 * mm, cell_y - 2.4 * mm, label.upper())
        c.setFillColor(colors.HexColor(INK))
        c.setFont("Helvetica-Bold" if index in (0, 1) else "Helvetica", 6.5)
        c.drawString(cell_x + 2 * mm, cell_y - 5.4 * mm, fit_single_line(value, "Helvetica", 6.5, col_w - 4 * mm))
    c.restoreState()


def draw_jiestar_index_page(
    c: pdf_canvas.Canvas,
    entries: list[dict[str, Any]],
    page_number: int,
    total_pages: int,
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    draw_jiestar_header(c, "JIESTAR / SKU - Name Index", page_w, page_h, margin, accent)
    title_y = page_h - margin - 22 * mm
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, title_y, "SKU / Name Index")
    c.setFillColor(colors.HexColor(MUTED))
    c.setFont("Helvetica", 7)
    c.drawString(margin, title_y - 5 * mm, "Sorted by Item No. Page numbers jump to the product card section.")

    col_gap = 6 * mm
    col_w = (page_w - 2 * margin - col_gap * (JIESTAR_INDEX_COLUMNS - 1)) / JIESTAR_INDEX_COLUMNS
    start_y = title_y - 12 * mm
    row_h = 5 * mm
    for col in range(JIESTAR_INDEX_COLUMNS):
        x = margin + col * (col_w + col_gap)
        c.setFillColor(colors.HexColor(DEEP))
        c.roundRect(x, start_y - 5.8 * mm, col_w, 5.8 * mm, 2.5, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 5.8)
        c.drawString(x + 2 * mm, start_y - 3.8 * mm, "ITEM NO. / PRODUCT NAME / PAGE")
        col_entries = entries[col * JIESTAR_INDEX_ROWS_PER_COLUMN : (col + 1) * JIESTAR_INDEX_ROWS_PER_COLUMN]
        y = start_y - 6 * mm
        for row_index, entry in enumerate(col_entries):
            y -= row_h
            c.setFillColor(colors.HexColor("#FFFFFF" if row_index % 2 == 0 else "#F8FAFC"))
            c.rect(x, y, col_w, row_h, stroke=0, fill=1)
            c.setFillColor(colors.HexColor(INK))
            c.setFont("Helvetica-Bold", 5.8)
            c.drawString(x + 1.8 * mm, y + 2.7 * mm, fit_single_line(entry["sku"], "Helvetica-Bold", 5.8, 20 * mm))
            c.setFillColor(colors.HexColor(accent))
            c.drawRightString(x + col_w - 2 * mm, y + 2.7 * mm, f"P{entry['page']}")
            c.setFillColor(colors.HexColor(MUTED))
            c.setFont("Helvetica", 5.1)
            name = f"{entry['name']} - {entry['series']}"
            c.drawString(x + 1.8 * mm, y + 0.7 * mm, fit_single_line(name, "Helvetica", 5.1, col_w - 4 * mm))
    draw_jiestar_footer(c, page_number, total_pages, page_w, margin)


def draw_jiestar_header(
    c: pdf_canvas.Canvas,
    label: str,
    page_w: float,
    page_h: float,
    margin: float,
    accent: str,
) -> None:
    c.saveState()
    y = page_h - margin - 8 * mm
    draw_logo(c, "JIESTAR", margin, y - 1.5 * mm, 19 * mm, 8 * mm)
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(margin + 24 * mm, y + 1.2 * mm, fit_single_line(label, "Helvetica-Bold", 8, 120 * mm))
    c.setFillColor(colors.HexColor(accent))
    c.roundRect(page_w - margin - 38 * mm, y - 0.8 * mm, 38 * mm, 5.5 * mm, 2.7, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 5.5)
    c.drawCentredString(page_w - margin - 19 * mm, y + 1.2 * mm, "GLOBAL B2B CATALOG")
    c.setStrokeColor(colors.HexColor("#E8EEF5"))
    c.line(margin, y - 4 * mm, page_w - margin, y - 4 * mm)
    c.restoreState()


def draw_jiestar_footer(c: pdf_canvas.Canvas, page_number: int, total_pages: int, page_w: float, margin: float) -> None:
    c.saveState()
    y = 7.5 * mm
    c.setStrokeColor(colors.HexColor("#E8EEF5"))
    c.line(margin, y + 4.2 * mm, page_w - margin, y + 4.2 * mm)
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.setFont("Helvetica", 6.7)
    c.drawString(margin, y, "JIESTAR Global B2B Product Catalog")
    c.drawRightString(page_w - margin, y, f"Page {page_number} / {total_pages}")
    c.restoreState()


def draw_logo(c: pdf_canvas.Canvas, brand: str, x: float, y: float, w: float, h: float) -> None:
    logo_path = BRAND_LOGOS.get(brand)
    if logo_path and logo_path.exists():
        try:
            draw_image_fit(c, str(logo_path), x, y, w, h)
            return
        except Exception:
            pass
    c.setFillColor(colors.HexColor(INK))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(x, y + h / 2, brand)


def draw_image_fit(c: pdf_canvas.Canvas, path: str, x: float, y: float, w: float, h: float) -> None:
    with Image.open(path) as img:
        iw, ih = img.size
    scale = min(w / iw, h / ih)
    draw_w = iw * scale
    draw_h = ih * scale
    c.drawImage(path, x + (w - draw_w) / 2, y + (h - draw_h) / 2, draw_w, draw_h, preserveAspectRatio=True, mask="auto")


def draw_placeholder(c: pdf_canvas.Canvas, x: float, y: float, w: float, h: float, label: str) -> None:
    c.setStrokeColor(colors.HexColor("#CBD5E1"))
    c.setDash(3, 3)
    c.roundRect(x + 4 * mm, y + 4 * mm, w - 8 * mm, h - 8 * mm, 4, stroke=1, fill=0)
    c.setDash()
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.setFont("Helvetica", 8)
    c.drawCentredString(x + w / 2, y + h / 2, label)


def fit_canvas_lines(text: Any, font: str, size: float, max_width: float, max_lines: int) -> list[str]:
    words = clean_text(text).split()
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if stringWidth(candidate, font, size) <= max_width:
            current = candidate
            continue
        if current:
            lines.append(current)
        current = word
        if len(lines) == max_lines:
            break
    if current and len(lines) < max_lines:
        lines.append(current)
    if len(lines) == max_lines and len(" ".join(lines).split()) < len(words):
        lines[-1] = fit_single_line(f"{lines[-1]}...", font, size, max_width)
    return lines or ["-"]


def make_cover(
    brand: str,
    products: list[Product],
    styles: dict[str, ParagraphStyle],
    cover_w: float,
    cover_h: float,
    accent: str,
) -> list[Any]:
    return [CoverPage(brand, products, cover_w, cover_h, accent), PageBreak()]


def footer(canvas: Any, doc: SimpleDocTemplate) -> None:
    canvas.saveState()
    logo_path = BRAND_LOGOS.get("JIESTAR")
    if logo_path and logo_path.exists():
        try:
            canvas.drawImage(
                str(logo_path),
                doc.leftMargin,
                5.8 * mm,
                width=7.5 * mm,
                height=7.5 * mm,
                preserveAspectRatio=True,
                mask="auto",
            )
            text_x = doc.leftMargin + 9 * mm
        except Exception:
            text_x = doc.leftMargin
    else:
        text_x = doc.leftMargin
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#94A3B8"))
    canvas.drawString(text_x, 8 * mm, "JIESTAR Global B2B Product Catalog")
    canvas.drawRightString(A4[1] - doc.rightMargin, 8 * mm, f"Page {doc.page}")
    canvas.restoreState()


def brand_slug(brand: str) -> str:
    return {
        "JIESTAR": "jiestar",
        "Xbert": "xbert",
        "TK Two": "tk-two",
        "Small Angle": "small-angle",
        "GULY": "guly",
        "Zoin": "zoin",
        "iBlock": "iblock",
        "JIQI": "jiqi",
    }[brand]


def write_reports(products: list[Product], pdf_paths: dict[str, Path]) -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    product_csv = OUTPUT_ROOT / "catalog-products.csv"
    with product_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(asdict(products[0]).keys()))
        writer.writeheader()
        for product in products:
            writer.writerow(asdict(product))

    qa_rows = []
    for brand, brand_products in group_by_brand(products).items():
        missing_images = [p for p in brand_products if not p.image_path]
        missing_pieces = [p for p in brand_products if p.piece_count == "-"]
        missing_size_weight = [
            p
            for p in brand_products
            if p.color_box_size == "-"
            or p.outer_carton_size == "-"
            or p.gross_net_weight_kg == "-"
            or p.model_size_cm == "-"
        ]
        low_confidence = [p for p in brand_products if p.name_confidence == "low"]
        qa_rows.append(
            {
                "brand": brand,
                "products": len(brand_products),
                "series": len({p.series for p in brand_products}),
                "pdf": str(pdf_paths.get(brand, "")),
                "missing_images": len(missing_images),
                "missing_piece_counts": len(missing_pieces),
                "missing_size_or_weight_fields": len(missing_size_weight),
                "low_confidence_english_names": len(low_confidence),
            }
        )

    qa_csv = OUTPUT_ROOT / "catalog-qa-summary.csv"
    with qa_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(qa_rows[0].keys()))
        writer.writeheader()
        writer.writerows(qa_rows)

    issues_csv = OUTPUT_ROOT / "catalog-qa-issues.csv"
    with issues_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        fields = ["brand", "sku", "issue", "source_file", "source_sheet", "source_row", "source_name_cn"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for product in products:
            if not product.image_path:
                writer.writerow(issue_row(product, "missing image"))
            if product.piece_count == "-":
                writer.writerow(issue_row(product, "missing piece count"))
            if product.name_confidence == "low":
                writer.writerow(issue_row(product, "low confidence English name"))
            for field_name in ("color_box_size", "outer_carton_size", "gross_net_weight_kg", "model_size_cm"):
                if getattr(product, field_name) == "-":
                    writer.writerow(issue_row(product, f"missing {field_name}"))

    summary_json = OUTPUT_ROOT / "catalog-generation-summary.json"
    summary_json.write_text(json.dumps(qa_rows, ensure_ascii=False, indent=2), encoding="utf-8")


def issue_row(product: Product, issue: str) -> dict[str, Any]:
    return {
        "brand": product.brand,
        "sku": product.sku,
        "issue": issue,
        "source_file": product.source_file,
        "source_sheet": product.source_sheet,
        "source_row": product.source_row,
        "source_name_cn": product.source_name_cn,
    }


def validate_no_price_words(pdf_paths: dict[str, Path]) -> dict[str, list[str]]:
    from pypdf import PdfReader

    findings: dict[str, list[str]] = {}
    for brand, path in pdf_paths.items():
        reader = PdfReader(str(path))
        text = "\n".join(page.extract_text() or "" for page in reader.pages[: min(10, len(reader.pages))])
        lower = text.lower()
        hits = [word for word in PRICE_WORDS if word.lower() in lower]
        if hits:
            findings[brand] = hits
    return findings


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--skip-pdf", action="store_true")
    parser.add_argument(
        "--brand",
        choices=list(BRAND_ACCENTS.keys()),
        help="Render only one brand PDF. Use this for focused design revisions.",
    )
    parser.add_argument(
        "--mobile",
        action="store_true",
        help="Render 9:16 mobile-first catalogs instead of horizontal catalogs.",
    )
    parser.add_argument(
        "--combined-mobile",
        action="store_true",
        help="Render one combined 9:16 mobile-first catalog containing all brands.",
    )
    args = parser.parse_args()
    if args.combined_mobile and args.brand:
        raise SystemExit("--combined-mobile cannot be used with --brand.")

    if TMP_ROOT.exists():
        shutil.rmtree(TMP_ROOT)
    TMP_ROOT.mkdir(parents=True, exist_ok=True)
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)

    products = build_products()
    if not products:
        raise SystemExit("No products extracted.")

    if args.skip_pdf:
        write_reports(products, {})
        print(f"Extracted {len(products)} products.")
        return

    if args.combined_mobile:
        output_path = OUTPUT_ROOT / "all-brands-product-catalog-mobile.pdf"
        render_combined_mobile_catalog(products, output_path)
        pdf_paths = {"All Brands": output_path}
    else:
        only_brands = {args.brand} if args.brand else None
        pdf_paths = render_catalogs(products, only_brands=only_brands, mobile=args.mobile)
    if not args.brand and not args.combined_mobile:
        write_reports(products, pdf_paths)
    findings = validate_no_price_words(pdf_paths)
    if findings:
        raise SystemExit(f"Price-like words found in generated PDFs: {findings}")

    rendered_products = [product for product in products if not args.brand or product.brand == args.brand]
    if args.mobile or args.combined_mobile:
        rendered_products = [
            product
            for product in rendered_products
            if not (product.brand == "JIESTAR" and product.series in JIESTAR_MOBILE_EXCLUDED_SERIES)
        ]
    rendered_count = len(rendered_products)
    print(f"Generated {len(pdf_paths)} PDFs for {rendered_count} products.")
    for brand, path in pdf_paths.items():
        print(f"{brand}: {path}")


if __name__ == "__main__":
    main()
