#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import http.client
import json
import mimetypes
import os
import re
import subprocess
import sys
import time
import urllib.error
import urllib.request
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path("/Volumes/ORICO/jiestar电商图/英文详情")
WORKBOOK = Path("/Volumes/ORICO/jiestar电商图/杰星整理表.xlsx")
OUT_DIR = Path("/private/tmp/jiestar-shopify-import")
SAMPLE_FOLDERS = [
    "57016英文",
    "51018+51019+51020英文",
    "57017+57018英文",
]
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
API_VERSION_FALLBACK = "2026-01"
PRICE = "999"
DETAIL_SLICE_MAX_HEIGHT = 7000
DETAIL_SLICE_MIN_BYTES = 8 * 1024 * 1024
DETAIL_SLICE_QUALITY = 88

TITLE_OVERRIDES = {
    "57016": "JIESTAR Medieval European Library Building Block Set with LED Lights",
    "51018+51019+51020": "JIESTAR Engineering Vehicle Building Block Set 3-Pack",
    "57017+57018": "JIESTAR Engineering Vehicle Building Block Set 2-Pack",
}

VARIANT_TITLE_OVERRIDES = {
    "51018": "Dump Truck",
    "51019": "Heavy Road Roller",
    "51020": "Bulldozer",
    "57017": "2-in-1 Excavator",
    "57018": "Mining Truck",
}


@dataclass
class WorkbookRow:
    sheet: str
    sku: str
    series_cn: str
    series_en: str
    name_cn: str
    name_en: str
    package_size: str
    finished_size: str
    age: str
    notes: str


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def normalize_sku(value: Any) -> str:
    sku = str(value).strip().upper() if value is not None else ""
    return sku[:-2] if sku.endswith(".0") else sku


def clean_cell(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def load_workbook_rows() -> dict[str, WorkbookRow]:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    rows: dict[str, WorkbookRow] = {}

    for sheet in workbook.worksheets:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        headers = [clean_cell(cell.value) for cell in header_row]
        header_index = {header: index for index, header in enumerate(headers)}
        sku_index = header_index.get("货号")

        if sku_index is None:
            continue

        for row in sheet.iter_rows(min_row=2, values_only=True):
            sku = normalize_sku(row[sku_index] if sku_index < len(row) else "")

            if not sku:
                continue

            def value(*names: str) -> str:
                for name in names:
                    index = header_index.get(name)
                    if index is not None and index < len(row):
                        return clean_cell(row[index])
                return ""

            rows[sku] = WorkbookRow(
                sheet=sheet.title,
                sku=sku,
                series_cn=value("系列"),
                series_en=value("Series (EN)"),
                name_cn=value("品名", "名称"),
                name_en=value("Product Name (EN)", "Name (EN)"),
                package_size=value("包装规格"),
                finished_size=value("组装尺寸", "造型规格", "产品规格"),
                age=value("年龄段", "适合年龄"),
                notes=value("备注"),
            )

    return rows


def strip_language_suffix(folder_name: str) -> str:
    return re.sub(r"(英文|中文)$", "", folder_name).strip()


def expand_codes(folder_name: str) -> list[str]:
    base = strip_language_suffix(folder_name)
    base = re.sub(r"（.*?）", "", base).strip()
    codes: list[str] = []

    for part in base.split("+"):
        part = part.strip()
        range_match = re.fullmatch(r"([A-Z]*)(\d{4,6})-([A-Z]*)(\d{4,6})", part, re.I)

        if range_match:
            prefix_a, start, prefix_b, end = range_match.groups()
            prefix = (prefix_a or prefix_b).upper()
            start_num = int(start)
            end_num = int(end)

            if end_num >= start_num and end_num - start_num <= 200:
                codes.extend(f"{prefix}{number:0{len(start)}d}" for number in range(start_num, end_num + 1))
                continue

        single_match = re.fullmatch(r"([A-Z]*)(\d{4,6})", part, re.I)

        if single_match:
            codes.append((single_match.group(1) + single_match.group(2)).upper())

    return codes


def slugify(value: str) -> str:
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", value.lower())).strip("-")


def parse_piece_count(notes: str) -> str:
    match = re.search(r"(\d+)\s*块", notes)
    return match.group(1) if match else ""


def product_title(base: str, variants: list[WorkbookRow]) -> str:
    if base in TITLE_OVERRIDES:
        return TITLE_OVERRIDES[base]

    if len(variants) == 1:
        name = variants[0].name_en or variants[0].name_cn or variants[0].sku
        name = re.sub(r"\s+", " ", name).strip(" .")
        return f"JIESTAR {name.title()} Building Block Set"

    series = variants[0].series_en or "Building Block"
    return f"JIESTAR {series} Building Block Set {len(variants)}-Pack"


def variant_option_name(row: WorkbookRow) -> str:
    if row.sku in VARIANT_TITLE_OVERRIDES:
        return f"{row.sku} - {VARIANT_TITLE_OVERRIDES[row.sku]}"

    raw_name = row.name_en or row.name_cn or row.sku
    name = re.sub(r"\s+", " ", raw_name).strip()
    return f"{row.sku} - {name}"


def image_files(folder: Path, sku_codes: list[str] | None = None) -> dict[str, list[Path]]:
    sku_code_set = {sku.upper().removesuffix("E") for sku in sku_codes or []}
    files = sorted(
        [
            path
            for path in folder.rglob("*")
            if path.is_file() and path.suffix.lower() in IMAGE_EXTS and not path.name.startswith("._")
        ],
        key=lambda path: str(path),
    )
    buckets = {
        "white": [],
        "numbered": [],
        "sku": [],
        "detail": [],
        "transparent": [],
        "other": [],
    }

    for path in files:
        lower_name = path.name.lower()

        if "白底" in path.name:
            buckets["white"].append(path)
        elif "详情" in path.name:
            buckets["detail"].append(path)
        elif "透明" in path.name:
            buckets["transparent"].append(path)
        elif is_numbered_main_image(path):
            buckets["numbered"].append(path)
        elif "sku" in lower_name or path.stem.upper() in sku_code_set or ("款" in path.name and len(sku_code_set) == 1):
            buckets["sku"].append(path)
        else:
            buckets["other"].append(path)

    buckets["numbered"].sort(key=image_sort_key)
    buckets["sku"].sort(key=lambda path: path.name)
    return buckets


def is_numbered_main_image(path: Path) -> bool:
    match = re.search(r"-(\d+)(?:\.\w+)$", path.name)
    return bool(match and int(match.group(1)) <= 50)


def image_sort_key(path: Path) -> tuple[int, str]:
    match = re.search(r"-(\d+)(?:\.\w+)$", path.name)
    return (int(match.group(1)) if match else 999, path.name)


def list_product_folders(scope: str) -> list[str]:
    if scope == "samples":
        return SAMPLE_FOLDERS

    return sorted(path.name for path in ROOT.iterdir() if path.is_dir() and not path.name.startswith("."))


def build_manifest(folder_names: list[str] | None = None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows_by_sku = load_workbook_rows()
    manifest: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for folder_name in folder_names or SAMPLE_FOLDERS:
        folder = ROOT / folder_name
        base = strip_language_suffix(folder_name)
        codes = expand_codes(folder_name)
        variants = []

        if not folder.exists():
            skipped.append({"folder": folder_name, "reason": "folder_missing"})
            continue

        for code in codes:
            row = rows_by_sku.get(code) or rows_by_sku.get(f"{code}E")

            if row:
                variants.append(row)

        if not variants:
            skipped.append({"folder": folder_name, "reason": "no_workbook_rows", "codes": codes})
            continue

        images = image_files(folder, [row.sku for row in variants])
        fallback_detail = images["other"][0] if not images["detail"] and len(images["other"]) == 1 else None
        main_media = images["white"][:1] + images["numbered"]
        detail_image = images["detail"][0] if images["detail"] else fallback_detail

        if not images["white"]:
            skipped.append({"folder": folder_name, "reason": "missing_white_image", "codes": codes})
            continue

        if not main_media:
            skipped.append({"folder": folder_name, "reason": "missing_main_media", "codes": codes})
            continue

        handle = slugify(f"{base}-{product_title(base, variants)}")
        primary = variants[0]
        piece_counts = [parse_piece_count(row.notes) for row in variants if parse_piece_count(row.notes)]
        piece_count_total = str(sum(int(count) for count in piece_counts)) if piece_counts else ""
        ages = sorted({row.age for row in variants if row.age})
        finished_sizes = [row.finished_size for row in variants if row.finished_size]
        package_sizes = sorted({row.package_size for row in variants if row.package_size})

        manifest.append(
            {
                "folder": folder_name,
                "folder_path": str(folder),
                "base": base,
                "handle": handle,
                "title": product_title(base, variants),
                "vendor": "JieStar",
                "status": "DRAFT",
                "product_type": primary.series_en or "Building Block Sets",
                "price": PRICE,
                "variants": [
                    {
                        "sku": row.sku,
                        "option_name": variant_option_name(row),
                        "title_source": row.name_en,
                        "series": row.series_en,
                        "age": row.age,
                        "piece_count": parse_piece_count(row.notes),
                        "package_size": row.package_size,
                        "finished_size": row.finished_size,
                    }
                    for row in variants
                ],
                "metafields": {
                    "specs.piece_count": piece_count_total,
                    "specs.recommended_age": ", ".join(ages),
                    "specs.finished_model_size": " / ".join(finished_sizes),
                    "specs.package_size": ", ".join(package_sizes),
                    "specs.difficulty_level": "See product package",
                    "custom.series": primary.series_en,
                },
                "main_media": [str(path) for path in main_media],
                "sku_images": [str(path) for path in images["sku"]],
                "detail_image": str(detail_image) if detail_image else "",
                "transparent_images": [str(path) for path in images["transparent"]],
                "missing": {
                    "white": not bool(images["white"]),
                    "detail": detail_image is None,
                    "sku_images": len(images["sku"]) < len(variants),
                },
            }
        )

    return manifest, skipped


def write_manifest(manifest: list[dict[str, Any]], skipped: list[dict[str, Any]], name: str = "sample") -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / f"{name}-manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / f"{name}-skipped.json").write_text(json.dumps(skipped, ensure_ascii=False, indent=2), encoding="utf-8")

    with (OUT_DIR / f"{name}-manifest.csv").open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["folder", "handle", "title", "variant_skus", "main_media_count", "detail_image", "missing"])

        for item in manifest:
            writer.writerow(
                [
                    item["folder"],
                    item["handle"],
                    item["title"],
                    ", ".join(variant["sku"] for variant in item["variants"]),
                    len(item["main_media"]),
                    item["detail_image"],
                    json.dumps(item["missing"], ensure_ascii=False),
                ]
            )


class ShopifyAdmin:
    def __init__(self) -> None:
        load_dotenv(Path(".env.local"))
        self.domain = os.environ.get("SHOPIFY_STORE_DOMAIN", "").strip()
        self.version = os.environ.get("SHOPIFY_API_VERSION", API_VERSION_FALLBACK).strip() or API_VERSION_FALLBACK
        self.token = os.environ.get("SHOPIFY_ADMIN_ACCESS_TOKEN", "").strip()

        if not self.domain:
            raise RuntimeError("Missing SHOPIFY_STORE_DOMAIN in .env.local")

        if not self.token:
            raise RuntimeError("Missing SHOPIFY_ADMIN_ACCESS_TOKEN in .env.local")

        self.endpoint = f"https://{self.domain}/admin/api/{self.version}/graphql.json"

    def graphql(self, query: str, variables: dict[str, Any] | None = None) -> dict[str, Any]:
        payload = json.dumps({"query": query, "variables": variables or {}}).encode("utf-8")
        request = urllib.request.Request(
            self.endpoint,
            data=payload,
            method="POST",
            headers={
                "Content-Type": "application/json",
                "X-Shopify-Access-Token": self.token,
            },
        )

        try:
            with urlopen_with_retries(request, timeout=90) as response:
                body = response.read().decode("utf-8")
        except urllib.error.HTTPError as error:
            body = error.read().decode("utf-8", errors="ignore")
            if error.code not in {404, 429, 500, 502, 503, 504}:
                raise RuntimeError(f"Shopify HTTP {error.code}: {body[:1200]}") from error
            body = self._graphql_via_curl(payload)
        except (urllib.error.URLError, TimeoutError, OSError, http.client.RemoteDisconnected):
            body = self._graphql_via_curl(payload)

        data = json.loads(body)

        if data.get("errors"):
            raise RuntimeError(json.dumps(data["errors"], ensure_ascii=False, indent=2))

        return data["data"]

    def _graphql_via_curl(self, payload: bytes) -> str:
        result = subprocess.run(
            [
                "curl",
                "--silent",
                "--show-error",
                "--fail-with-body",
                "--http1.1",
                "--connect-timeout",
                "20",
                "--max-time",
                "120",
                "--retry",
                "6",
                "--retry-all-errors",
                "--request",
                "POST",
                "--header",
                "Content-Type: application/json",
                "--header",
                f"X-Shopify-Access-Token: {self.token}",
                "--data-binary",
                "@-",
                self.endpoint,
            ],
            input=payload,
            capture_output=True,
            timeout=180,
        )
        body = result.stdout.decode("utf-8", errors="ignore")
        if result.returncode:
            message = result.stderr.decode("utf-8", errors="ignore")
            raise RuntimeError(f"Shopify curl request failed: {(message or body)[:1200]}")
        return body

    def preflight(self) -> dict[str, Any]:
        return self.graphql(
            """
            query Preflight {
              shop {
                name
                myshopifyDomain
              }
            }
            """
        )

    def product_by_handle(self, handle: str) -> dict[str, Any] | None:
        data = self.graphql(
            """
            query ExistingProduct($query: String!) {
              products(first: 1, query: $query) {
                nodes {
                  id
                  title
                  handle
                  status
                }
              }
            }
            """,
            {"query": f"handle:{handle}"},
        )
        nodes = data["products"]["nodes"]
        return nodes[0] if nodes else None

    def stage_upload(self, path: Path) -> str:
        mime_type = mimetypes.guess_type(path.name)[0] or "image/jpeg"
        data = self.graphql(
            """
            mutation StageUpload($input: [StagedUploadInput!]!) {
              stagedUploadsCreate(input: $input) {
                stagedTargets {
                  url
                  resourceUrl
                  parameters {
                    name
                    value
                  }
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {
                "input": [
                    {
                        "resource": "IMAGE",
                        "filename": path.name,
                        "mimeType": mime_type,
                        "httpMethod": "POST",
                    }
                ]
            },
        )
        result = data["stagedUploadsCreate"]
        assert_no_user_errors("stagedUploadsCreate", result["userErrors"])
        target = result["stagedTargets"][0]
        upload_multipart(target["url"], target["parameters"], path, mime_type)
        return target["resourceUrl"]

    def file_create(self, path: Path, alt: str) -> str:
        resource_url = self.stage_upload(path)
        data = self.graphql(
            """
            mutation FileCreate($files: [FileCreateInput!]!) {
              fileCreate(files: $files) {
                files {
                  id
                  fileStatus
                  alt
                  ... on MediaImage {
                    image {
                      url
                    }
                  }
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {
                "files": [
                    {
                        "alt": alt,
                        "contentType": "IMAGE",
                        "originalSource": resource_url,
                    }
                ]
            },
        )
        result = data["fileCreate"]
        assert_no_user_errors("fileCreate", result["userErrors"])
        file_node = result["files"][0]
        file_id = file_node["id"]

        for _ in range(120):
            url = (file_node.get("image") or {}).get("url") if file_node else None

            if url:
                return url

            time.sleep(2)
            data = self.graphql(
                """
                query FileNode($id: ID!) {
                  node(id: $id) {
                    ... on MediaImage {
                      id
                      fileStatus
                      image {
                        url
                      }
                    }
                  }
                }
                """,
                {"id": file_id},
            )
            file_node = data["node"]

        raise RuntimeError(f"Shopify file did not return a CDN URL: {path}")

    def product_set(self, item: dict[str, Any], description_html: str) -> dict[str, Any]:
        variants = [
            {
                "optionValues": [
                    {
                        "optionName": "SKU",
                        "name": variant["option_name"],
                    }
                ],
                "price": item["price"],
                "inventoryItem": {
                    "sku": variant["sku"],
                    "tracked": False,
                },
            }
            for variant in item["variants"]
        ]
        data = self.graphql(
            """
            mutation ProductSet($input: ProductSetInput!, $synchronous: Boolean!) {
              productSet(input: $input, synchronous: $synchronous) {
                product {
                  id
                  title
                  handle
                  status
                  vendor
                  variants(first: 250) {
                    nodes {
                      id
                      title
                      price
                      sku
                      inventoryItem {
                        id
                        tracked
                      }
                    }
                  }
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {
                "synchronous": True,
                "input": {
                    "title": item["title"],
                    "handle": item["handle"],
                    "vendor": item["vendor"],
                    "status": item["status"],
                    "productType": item["product_type"],
                    "descriptionHtml": description_html,
                    "productOptions": [
                        {
                            "name": "SKU",
                            "values": [{"name": variant["option_name"]} for variant in item["variants"]],
                        }
                    ],
                    "variants": variants,
                    "metafields": product_metafields(item),
                },
            },
        )
        result = data["productSet"]
        assert_no_user_errors("productSet", result["userErrors"])
        return result["product"]

    def product_update_media(self, product_id: str, media: list[dict[str, Any]]) -> None:
        if not media:
            return
        data = self.graphql(
            """
            mutation ProductUpdateMedia($product: ProductUpdateInput!, $media: [CreateMediaInput!]) {
              productUpdate(product: $product, media: $media) {
                product {
                  id
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"product": {"id": product_id}, "media": media},
        )
        assert_no_user_errors("productUpdate", data["productUpdate"]["userErrors"])

    def product_update_description(self, product_id: str, description_html: str) -> None:
        data = self.graphql(
            """
            mutation ProductUpdateDescription($product: ProductUpdateInput!) {
              productUpdate(product: $product) {
                product {
                  id
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"product": {"id": product_id, "descriptionHtml": description_html}},
        )
        assert_no_user_errors("productUpdateDescription", data["productUpdate"]["userErrors"])

    def delete_files(self, file_ids: list[str]) -> None:
        if not file_ids:
            return
        data = self.graphql(
            """
            mutation FileDelete($fileIds: [ID!]!) {
              fileDelete(fileIds: $fileIds) {
                deletedFileIds
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"fileIds": file_ids},
        )
        assert_no_user_errors("fileDelete", data["fileDelete"]["userErrors"])

    def detach_variant_media(self, product_id: str, variant_media: list[dict[str, Any]]) -> None:
        if not variant_media:
            return
        data = self.graphql(
            """
            mutation ProductVariantDetachMedia($productId: ID!, $variantMedia: [ProductVariantDetachMediaInput!]!) {
              productVariantDetachMedia(productId: $productId, variantMedia: $variantMedia) {
                product {
                  id
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"productId": product_id, "variantMedia": variant_media},
        )
        assert_no_user_errors("productVariantDetachMedia", data["productVariantDetachMedia"]["userErrors"])

    def append_variant_media(self, product_id: str, variant_media: list[dict[str, Any]]) -> None:
        if not variant_media:
            return
        data = self.graphql(
            """
            mutation ProductVariantAppendMedia($productId: ID!, $variantMedia: [ProductVariantAppendMediaInput!]!) {
              productVariantAppendMedia(productId: $productId, variantMedia: $variantMedia) {
                product {
                  id
                }
                productVariants {
                  id
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"productId": product_id, "variantMedia": variant_media},
        )
        errors = [
            error
            for error in data["productVariantAppendMedia"]["userErrors"]
            if error.get("message") != "The given variant already has attached media."
        ]
        assert_no_user_errors("productVariantAppendMedia", errors)

    def reorder_media(self, product_id: str, expected_alts: list[str]) -> None:
        product = self.fetch_product(product_id)
        media_nodes = product["media"]["nodes"]
        moves = []

        for position, alt in enumerate(expected_alts):
            node = next((item for item in media_nodes if item.get("alt") == alt), None)

            if node:
                moves.append({"id": node["id"], "newPosition": str(position)})

        if not moves:
            return

        data = self.graphql(
            """
            mutation ProductReorderMedia($id: ID!, $moves: [MoveInput!]!) {
              productReorderMedia(id: $id, moves: $moves) {
                job {
                  id
                }
                mediaUserErrors {
                  field
                  message
                }
              }
            }
            """,
            {"id": product_id, "moves": moves},
        )
        assert_no_user_errors("productReorderMedia", data["productReorderMedia"]["mediaUserErrors"])

    def fetch_product(self, product_id: str) -> dict[str, Any]:
        data = self.graphql(
            """
            query ProductVerify($id: ID!) {
              product(id: $id) {
                id
                title
                handle
                status
                vendor
                descriptionHtml
                onlineStorePreviewUrl
                media(first: 250, sortKey: POSITION) {
                  nodes {
                    id
                    alt
                    mediaContentType
                    preview {
                      status
                    }
                    ... on MediaImage {
                      image {
                        url
                      }
                    }
                  }
                }
                metafields(first: 20, namespace: "specs") {
                  nodes {
                    namespace
                    key
                    value
                  }
                }
                variants(first: 250) {
                  nodes {
                    id
                    title
                    price
                    sku
                    image {
                      id
                      url
                      altText
                    }
                    media(first: 10) {
                      nodes {
                        id
                        alt
                        mediaContentType
                        ... on MediaImage {
                          image {
                            url
                          }
                        }
                      }
                    }
                    inventoryItem {
                      tracked
                    }
                  }
                }
              }
            }
            """,
            {"id": product_id},
        )
        return data["product"]


def product_metafields(item: dict[str, Any]) -> list[dict[str, str]]:
    output = []
    metafield_types = {
        "specs.piece_count": "number_integer",
    }

    for full_key, value in item["metafields"].items():
        if not value:
            continue

        namespace, key = full_key.split(".", 1)
        metafield_type = metafield_types.get(full_key, "single_line_text_field")
        clean_value = str(value)

        if metafield_type == "single_line_text_field":
            clean_value = re.sub(r"\s+", " ", clean_value).strip()

        output.append(
            {
                "namespace": namespace,
                "key": key,
                "type": metafield_type,
                "value": clean_value,
            }
        )

    return output


def assert_no_user_errors(operation: str, errors: list[dict[str, Any]]) -> None:
    if errors:
        raise RuntimeError(f"{operation} userErrors: {json.dumps(errors, ensure_ascii=False, indent=2)}")


def upload_multipart(url: str, parameters: list[dict[str, str]], path: Path, mime_type: str) -> None:
    boundary = f"----jiestar-{uuid.uuid4().hex}"
    chunks: list[bytes] = []

    for parameter in parameters:
        chunks.append(f"--{boundary}\r\n".encode())
        chunks.append(f'Content-Disposition: form-data; name="{parameter["name"]}"\r\n\r\n'.encode())
        chunks.append(parameter["value"].encode())
        chunks.append(b"\r\n")

    chunks.append(f"--{boundary}\r\n".encode())
    chunks.append(f'Content-Disposition: form-data; name="file"; filename="{path.name}"\r\n'.encode())
    chunks.append(f"Content-Type: {mime_type}\r\n\r\n".encode())
    chunks.append(path.read_bytes())
    chunks.append(b"\r\n")
    chunks.append(f"--{boundary}--\r\n".encode())
    body = b"".join(chunks)

    request = urllib.request.Request(
        url,
        data=body,
        method="POST",
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Content-Length": str(len(body)),
        },
    )

    retryable_statuses = {404, 429, 500, 502, 503, 504}

    last_error: Exception | None = None
    for attempt in range(1, 4):
        try:
            with urlopen_with_retries(request, timeout=120) as response:
                response.read()
            return
        except urllib.error.HTTPError as error:
            response_body = error.read().decode("utf-8", errors="ignore")

            if error.code == 400:
                last_error = error
                break

            if error.code in retryable_statuses and attempt < 3:
                last_error = error
                time.sleep(2 * attempt)
                continue

            if error.code not in retryable_statuses:
                raise RuntimeError(
                    f"Staged upload failed with HTTP {error.code}: {response_body[:1200]}"
                ) from error
            last_error = error
        except (urllib.error.URLError, TimeoutError, OSError, http.client.RemoteDisconnected) as error:
            last_error = error
            if attempt < 3:
                time.sleep(2 * attempt)
                continue

    command = [
        "curl",
        "--silent",
        "--show-error",
        "--fail-with-body",
        "--http1.1",
        "--connect-timeout",
        "20",
        "--max-time",
        "300",
        "--retry",
        "6",
        "--retry-all-errors",
    ]
    for parameter in parameters:
        command.extend(["--form-string", f'{parameter["name"]}={parameter["value"]}'])
    command.extend(["--form", f"file=@{path};type={mime_type}", url])

    result = subprocess.run(command, capture_output=True, timeout=360)
    if result.returncode:
        message = result.stderr.decode("utf-8", errors="ignore")
        response_body = result.stdout.decode("utf-8", errors="ignore")
        raise RuntimeError(
            f"Staged upload failed via curl: {(message or response_body or str(last_error))[:1200]}"
        )


def urlopen_with_retries(request: urllib.request.Request, timeout: int, attempts: int = 6):
    last_error: Exception | None = None
    retryable_statuses = {404, 429, 500, 502, 503, 504}

    for attempt in range(1, attempts + 1):
        try:
            return urllib.request.urlopen(request, timeout=timeout)
        except urllib.error.HTTPError as error:
            if error.code not in retryable_statuses or attempt == attempts:
                raise

            error.read()
            error.close()
            last_error = error
        except (urllib.error.URLError, TimeoutError, OSError, http.client.RemoteDisconnected) as error:
            last_error = error

        if attempt == attempts:
            break

        time.sleep(2 * attempt)

    assert last_error is not None
    raise last_error


def detail_image_paths(path: Path) -> list[Path]:
    if not path.exists():
        return []

    try:
        from PIL import Image
    except ImportError:
        return [path]

    with Image.open(path) as image:
        width, height = image.size

        if height <= DETAIL_SLICE_MAX_HEIGHT and path.stat().st_size <= DETAIL_SLICE_MIN_BYTES:
            return [path]

        output_dir = OUT_DIR / "detail-slices" / path.stem
        output_dir.mkdir(parents=True, exist_ok=True)
        output_paths = []

        for index, top in enumerate(range(0, height, DETAIL_SLICE_MAX_HEIGHT), start=1):
            bottom = min(top + DETAIL_SLICE_MAX_HEIGHT, height)
            output_path = output_dir / f"{path.stem}-part-{index:02d}.jpg"
            crop = image.crop((0, top, width, bottom))

            if crop.mode != "RGB":
                crop = crop.convert("RGB")

            crop.save(output_path, format="JPEG", quality=DETAIL_SLICE_QUALITY, optimize=True)
            output_paths.append(output_path)

    return output_paths


def upload_detail_images(admin: ShopifyAdmin, item: dict[str, Any]) -> list[str]:
    if not item["detail_image"]:
        return []

    paths = detail_image_paths(Path(item["detail_image"]))
    urls = []

    for index, path in enumerate(paths, start=1):
        suffix = f" part {index}" if len(paths) > 1 else ""
        urls.append(admin.file_create(path, f"{item['title']} details{suffix}"))

    return urls


def build_description_html(item: dict[str, Any], detail_urls: list[str] | None = None) -> str:
    if not detail_urls:
        return ""

    images = "\n".join(
        f'<p><img src="{html.escape(url)}" alt="{html.escape(item["title"])} details part {index}" /></p>'
        for index, url in enumerate(detail_urls, start=1)
    )
    return images


def append_sku_images_to_variants(admin: ShopifyAdmin, product_id: str) -> None:
    product = admin.fetch_product(product_id)
    media_by_sku = {}
    sku_media_ids = []

    for media in product["media"]["nodes"]:
        alt = media.get("alt") or ""
        match = re.search(r"SKU image - ([A-Z]*\d{4,6})(?:-sku)?\.", alt, re.I)

        if "SKU image - " in alt:
            sku_media_ids.append(media["id"])

        if match:
            media_by_sku[match.group(1).upper()] = media["id"]

    variant_media = []

    for variant in product["variants"]["nodes"]:
        if variant.get("image") or variant.get("media", {}).get("nodes"):
            continue

        sku = (variant.get("sku") or "").upper()
        media_id = (
            media_by_sku.get(sku)
            or media_by_sku.get(sku.removesuffix("E"))
            or media_by_sku.get(sku.removeprefix("X"))
        )

        if not media_id and len(product["variants"]["nodes"]) == 1 and len(sku_media_ids) == 1:
            media_id = sku_media_ids[0]

        if media_id:
            variant_media.append({"variantId": variant["id"], "mediaIds": [media_id]})

    admin.append_variant_media(product_id, variant_media)


def media_inputs_for_item(admin: ShopifyAdmin, item: dict[str, Any], existing_alts: set[str]) -> tuple[list[dict[str, Any]], list[str]]:
    media_inputs = []
    expected_alts = []

    for media_path in item["main_media"]:
        path = Path(media_path)
        alt = f"{item['title']} - {path.name}"
        expected_alts.append(alt)

        if alt in existing_alts:
            continue

        media_inputs.append(
            {
                "originalSource": admin.stage_upload(path),
                "alt": alt,
                "mediaContentType": "IMAGE",
            }
        )

    for media_path in item["sku_images"]:
        path = Path(media_path)
        alt = f"{item['title']} SKU image - {path.name}"
        expected_alts.append(alt)

        if alt in existing_alts:
            continue

        media_inputs.append(
            {
                "originalSource": admin.stage_upload(path),
                "alt": alt,
                "mediaContentType": "IMAGE",
            }
        )

    return media_inputs, expected_alts


def sync_product_assets(admin: ShopifyAdmin, item: dict[str, Any], product_id: str, update_description: bool = False) -> dict[str, Any]:
    if update_description and item["detail_image"]:
        detail_urls = upload_detail_images(admin, item)
        admin.product_update_description(product_id, build_description_html(item, detail_urls))

    product = admin.fetch_product(product_id)
    cleanup_unexpected_media(admin, item, product_id, product)
    product = admin.fetch_product(product_id)
    existing_alts = {media.get("alt") or "" for media in product["media"]["nodes"]}
    media_inputs, expected_alts = media_inputs_for_item(admin, item, existing_alts)

    if media_inputs:
        admin.product_update_media(product_id, media_inputs)
        time.sleep(8)

    append_sku_images_to_variants(admin, product_id)
    admin.reorder_media(product_id, expected_alts)
    time.sleep(4)
    return admin.fetch_product(product_id)


def cleanup_unexpected_media(admin: ShopifyAdmin, item: dict[str, Any], product_id: str, product: dict[str, Any]) -> None:
    main_names = {Path(path).name for path in item["main_media"]}
    sku_names = {Path(path).name for path in item["sku_images"]}
    delete_ids = []
    detach_inputs = []

    for media in product["media"]["nodes"]:
        alt = media.get("alt") or ""
        is_sku_media = "SKU image - " in alt
        name = alt.rsplit(" - ", 1)[-1]

        if is_sku_media:
            name = alt.rsplit("SKU image - ", 1)[-1]

        if (is_sku_media and name in sku_names) or (not is_sku_media and name in main_names):
            continue

        delete_ids.append(media["id"])

        for variant in product["variants"]["nodes"]:
            if any(node["id"] == media["id"] for node in variant.get("media", {}).get("nodes", [])):
                detach_inputs.append({"variantId": variant["id"], "mediaIds": [media["id"]]})

    admin.detach_variant_media(product_id, detach_inputs)
    admin.delete_files(delete_ids)


def create_products(manifest: list[dict[str, Any]], report_name: str = "sample-created-products") -> list[dict[str, Any]]:
    admin = ShopifyAdmin()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    created = []

    for item in manifest:
        try:
            existing = admin.product_by_handle(item["handle"])

            if existing:
                verified = sync_product_assets(admin, item, existing["id"], update_description=False)
                created.append({"manifest": item, "product": verified, "existing": True, "ok": True})
                (OUT_DIR / f"{report_name}.json").write_text(json.dumps(created, ensure_ascii=False, indent=2), encoding="utf-8")
                continue

            detail_urls = []

            if item["detail_image"]:
                detail_urls = upload_detail_images(admin, item)

            product = admin.product_set(item, build_description_html(item, detail_urls))
            product_id = product["id"]
            verified = sync_product_assets(admin, item, product_id, update_description=False)
            created.append({"manifest": item, "product": verified, "existing": False, "ok": True})
        except Exception as error:
            created.append({"manifest": item, "ok": False, "error": str(error)})

        (OUT_DIR / f"{report_name}.json").write_text(json.dumps(created, ensure_ascii=False, indent=2), encoding="utf-8")

    return created


def select_batch(manifest: list[dict[str, Any]], offset: int, batch_size: int) -> list[dict[str, Any]]:
    return manifest[offset : offset + batch_size]


def run_preflight() -> None:
    admin = ShopifyAdmin()
    print(json.dumps(admin.preflight(), ensure_ascii=False, indent=2))


def main() -> int:
    parser = argparse.ArgumentParser(description="Create JIESTAR Shopify draft products from the local image folders.")
    parser.add_argument("--dry-run", action="store_true", help="Generate the sample manifest only.")
    parser.add_argument("--preflight", action="store_true", help="Check Shopify Admin API credentials with a read-only query.")
    parser.add_argument("--create-samples", action="store_true", help="Create the three sample products as Shopify drafts.")
    parser.add_argument("--create-batch", action="store_true", help="Create one small batch of products as Shopify drafts.")
    parser.add_argument("--scope", choices=["samples", "all"], default="samples", help="Which folders to include in the manifest.")
    parser.add_argument("--offset", type=int, default=0, help="Manifest offset for --create-batch.")
    parser.add_argument("--batch-size", type=int, default=10, help="Maximum products to create in one batch.")
    args = parser.parse_args()

    if not (args.dry_run or args.preflight or args.create_samples or args.create_batch):
        parser.error("Choose one of --dry-run, --preflight, --create-samples, or --create-batch")

    if args.preflight:
        run_preflight()
        return 0

    folder_names = list_product_folders(args.scope)
    manifest, skipped = build_manifest(folder_names)
    manifest_name = args.scope
    write_manifest(manifest, skipped, manifest_name)

    print(
        json.dumps(
            {
                "manifest_json": str(OUT_DIR / f"{manifest_name}-manifest.json"),
                "skipped_json": str(OUT_DIR / f"{manifest_name}-skipped.json"),
                "products": len(manifest),
                "skipped": len(skipped),
            },
            ensure_ascii=False,
            indent=2,
        )
    )

    if args.create_samples:
        sample_manifest = [item for item in manifest if item["folder"] in SAMPLE_FOLDERS]
        created = create_products(sample_manifest, "sample-created-products")
        print(json.dumps({"processed": len(created), "report": str(OUT_DIR / "sample-created-products.json")}, ensure_ascii=False, indent=2))

    if args.create_batch:
        if args.batch_size < 1 or args.batch_size > 25:
            parser.error("--batch-size must be between 1 and 25")

        batch = select_batch(manifest, args.offset, args.batch_size)
        report_name = f"batch-offset-{args.offset}-size-{args.batch_size}"
        created = create_products(batch, report_name)
        print(
            json.dumps(
                {
                    "processed": len(created),
                    "ok": sum(1 for item in created if item.get("ok")),
                    "failed": sum(1 for item in created if not item.get("ok")),
                    "report": str(OUT_DIR / f"{report_name}.json"),
                    "next_offset": args.offset + args.batch_size,
                },
                ensure_ascii=False,
                indent=2,
            )
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
