#!/usr/bin/env python3
"""Generate and, only after approval, apply the JIESTAR U.S. warehouse setup.

Dry-run is the default. It writes a source-SKU audit, conservative weight-band
rates, Shopify diffs, and an approval hash. Live writes require all of:
--apply, --yes, an unchanged --approved-report, and --publication-id.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from dataclasses import asdict, dataclass
from decimal import Decimal, ROUND_FLOOR
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from shopify_shipping_update_from_template import (
    STANDARD_GOODS_PROFILE,
    ShopifyAdmin,
    ShopifyVariant,
    canonical_hash,
    clean,
    decimal_str,
    money_str,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path(
    "/Users/chensen/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_td8knmekr3je22_1a81/temp/RWTemp/2026-08/"
    "45642a058de49fd708c2b0075922e26f/亚马逊美国海外仓多渠道配送费用报陈森 2026.08.01(1).xlsx"
)
DEFAULT_OUT_DIR = Path("/private/tmp/jiestar-us-warehouse-approval")
US_LOCATION_NAME = "Amazon U.S. Warehouse"
COLLECTION_TITLE = "U.S. Warehouse"
COLLECTION_HANDLE = "us-warehouse"
RATE_NAME = "U.S. Warehouse Shipping"
ZONE_NAME = "JIESTAR U.S. Warehouse"
MAX_WEIGHT_G = 10_000
BAND_SIZE_G = 500
EXPECTED_ACTIVE_SKUS = {
    "59159", "59160", "59161", "66124", "JJ9045", "JJ9083", "JJ9234",
    "JJ9235", "LL001", "LL002", "LL003", "LL004", "LL911", "TK8001",
}
EXPECTED_PRODUCT_COUNT = 12
EXPECTED_VARIANT_COUNT = 14
DEFERRED_SKUS = {"JJ9219", "33001", "33002", "X88052", "X88053"}
REQUIRED_SCOPES = {
    "read_inventory", "write_inventory", "read_locations", "read_products",
    "write_products", "read_publications", "write_publications", "read_shipping", "write_shipping",
}

SUMMARY_JSON = "us-warehouse-approval-summary.json"
SOURCE_AUDIT_CSV = "us-warehouse-sku-audit.csv"
RATE_PLAN_CSV = "us-warehouse-rate-plan.csv"
SHOPIFY_DIFF_CSV = "us-warehouse-shopify-diff.csv"
SNAPSHOT_JSON = "us-warehouse-pre-apply-snapshot.json"
APPLY_RESULT_JSON = "us-warehouse-apply-result.json"


@dataclass(frozen=True)
class EligibleSku:
    sku: str
    variant_id: str
    inventory_item_id: str
    product_id: str
    product_handle: str
    product_title: str
    weight_g: int
    source_cost_usd: Decimal
    inventory_tracked: bool | None


@dataclass(frozen=True)
class RateBand:
    min_weight_kg: Decimal
    max_weight_kg: Decimal
    max_source_cost_usd: Decimal
    target_rate_usd: Decimal
    witness_skus: str


def source_costs(workbook_path: Path) -> dict[str, Decimal]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        rows = list(sheet.iter_rows(min_row=2, max_col=5, values_only=True))
    finally:
        workbook.close()

    output: dict[str, Decimal] = {}
    for row in rows:
        sku = clean(row[0]).upper()
        if not sku:
            continue
        currency = clean(row[2]).lower()
        if currency not in {"美金", "usd", "$"}:
            raise RuntimeError(f"Unexpected source currency for {sku}: {row[2]!r}")
        output[sku] = Decimal(str(row[1]))
    if len(output) != 19:
        raise RuntimeError(f"Expected 19 source SKUs, found {len(output)}")
    return output


def round_up_to_99(value: Decimal) -> Decimal:
    whole = value.to_integral_value(rounding=ROUND_FLOOR)
    candidate = whole + Decimal("0.99")
    if candidate < value:
        candidate += Decimal("1")
    return candidate.quantize(Decimal("0.01"))


def customer_rate(cost: Decimal) -> Decimal:
    return round_up_to_99((cost * Decimal("1.05") + Decimal("1.19")) / Decimal("0.8901"))


def build_rate_bands(items: list[EligibleSku]) -> list[RateBand]:
    """Use unbounded knapsack to price every reachable eligible cart up to 10kg."""
    dp: list[Decimal | None] = [None] * (MAX_WEIGHT_G + 1)
    parents: list[tuple[int, str] | None] = [None] * (MAX_WEIGHT_G + 1)
    dp[0] = Decimal("0")

    for item in items:
        for total_weight in range(item.weight_g, MAX_WEIGHT_G + 1):
            previous = dp[total_weight - item.weight_g]
            if previous is None:
                continue
            candidate = previous + item.source_cost_usd
            if dp[total_weight] is None or candidate > dp[total_weight]:
                dp[total_weight] = candidate
                parents[total_weight] = (total_weight - item.weight_g, item.sku)

    observations: list[tuple[Decimal | None, str]] = []
    for lower_g in range(0, MAX_WEIGHT_G, BAND_SIZE_G):
        upper_g = lower_g + BAND_SIZE_G
        reachable = [(weight, dp[weight]) for weight in range(lower_g + 1, upper_g + 1) if dp[weight] is not None]
        if not reachable:
            observations.append((None, ""))
            continue
        witness_weight, band_cost = max(reachable, key=lambda pair: pair[1] or Decimal("0"))
        assert band_cost is not None
        witness: list[str] = []
        cursor = witness_weight
        while cursor and parents[cursor]:
            cursor, sku = parents[cursor]  # type: ignore[misc]
            witness.append(sku)
        witness_counts = defaultdict(int)
        for sku in witness:
            witness_counts[sku] += 1
        current_witness = " + ".join(
            f"{sku} x{count}" for sku, count in sorted(witness_counts.items())
        )
        observations.append((band_cost, current_witness))

    first_reachable = next(((cost, witness) for cost, witness in observations if cost is not None), None)
    if not first_reachable:
        return []

    bands: list[RateBand] = []
    monotonic_cost, monotonic_witness = first_reachable
    for index, (band_cost, current_witness) in enumerate(observations):
        lower_g = index * BAND_SIZE_G
        upper_g = lower_g + BAND_SIZE_G
        if band_cost is not None and band_cost >= monotonic_cost:
            monotonic_cost = band_cost
            monotonic_witness = current_witness
        bands.append(
            RateBand(
                min_weight_kg=Decimal("0") if lower_g == 0 else Decimal(lower_g + 1) / Decimal("1000"),
                max_weight_kg=Decimal(upper_g) / Decimal("1000"),
                max_source_cost_usd=monotonic_cost,
                target_rate_usd=customer_rate(monotonic_cost),
                witness_skus=monotonic_witness,
            )
        )
    return bands


def match_eligible_variants(
    active_variants: list[ShopifyVariant],
    costs: dict[str, Decimal],
) -> tuple[list[EligibleSku], list[dict[str, str]]]:
    eligible: list[EligibleSku] = []
    audit: list[dict[str, str]] = []
    active_by_sku = {variant.sku_key: variant for variant in active_variants if variant.sku_key}
    for sku, cost in sorted(costs.items()):
        variant = active_by_sku.get(sku)
        is_expected = sku in EXPECTED_ACTIVE_SKUS
        if is_expected and variant and variant.current_weight_g and variant.inventory_item_id:
            eligible.append(
                EligibleSku(
                    sku=sku,
                    variant_id=variant.variant_id,
                    inventory_item_id=variant.inventory_item_id,
                    product_id=variant.product_id,
                    product_handle=variant.product_handle,
                    product_title=variant.product_title,
                    weight_g=variant.current_weight_g,
                    source_cost_usd=cost,
                    inventory_tracked=variant.inventory_tracked,
                )
            )
            status = "eligible_active"
        elif is_expected:
            status = "blocked_missing_active_variant_or_weight"
        elif sku in DEFERRED_SKUS:
            status = "deferred_not_in_first_collection"
        else:
            status = "not_in_first_collection"
        audit.append(
            {
                "sku": sku,
                "source_cost_usd": money_str(cost),
                "shopify_status": status,
                "product_handle": variant.product_handle if variant else "",
                "weight_g": str(variant.current_weight_g or "") if variant else "",
                "inventory_tracked": str(variant.inventory_tracked).lower() if variant else "",
            }
        )
    return eligible, audit


def group_locations(group: dict[str, Any]) -> list[dict[str, Any]]:
    return (((group.get("locationGroup") or {}).get("locations") or {}).get("nodes") or [])


def find_location_group(profile: dict[str, Any] | None, location_id: str | None) -> dict[str, Any] | None:
    if not profile or not location_id:
        return None
    return next(
        (
            group for group in profile.get("profileLocationGroups") or []
            if any(location.get("id") == location_id for location in group_locations(group))
        ),
        None,
    )


def group_rate_fingerprint(group: dict[str, Any] | None) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    if not group:
        return rows
    for zone in ((group.get("locationGroupZones") or {}).get("nodes") or []):
        countries = ",".join(
            clean(((country.get("code") or {}).get("countryCode")))
            for country in ((zone.get("zone") or {}).get("countries") or [])
        )
        for method in ((zone.get("methodDefinitions") or {}).get("nodes") or []):
            rows.append({
                "zone_id": clean((zone.get("zone") or {}).get("id")),
                "zone_name": clean((zone.get("zone") or {}).get("name")),
                "countries": countries,
                "method_id": clean(method.get("id")),
                "method_name": clean(method.get("name")),
            })
    return rows


class UsWarehouseAdmin(ShopifyAdmin):
    def collection_snapshot(self) -> dict[str, Any] | None:
        data = self.graphql(
            """
            query UsWarehouseCollection($identifier: CollectionIdentifierInput!) {
              collectionByIdentifier(identifier: $identifier) {
                id handle title
                products(first: 250) { nodes { id handle title } }
              }
            }
            """,
            {"identifier": {"handle": COLLECTION_HANDLE}},
        )
        return data.get("collectionByIdentifier")

    def publication_snapshot(self) -> list[dict[str, Any]]:
        data = self.graphql(
            """
            query Publications { publications(first: 50) { nodes { id name app { title } } } }
            """
        )
        return data["publications"]["nodes"]

    def create_collection(self, product_ids: list[str]) -> dict[str, Any]:
        data = self.graphql(
            """
            mutation CreateUsWarehouseCollection($input: CollectionInput!) {
              collectionCreate(input: $input) {
                collection { id handle title }
                userErrors { field message }
              }
            }
            """,
            {"input": {
                "title": COLLECTION_TITLE,
                "handle": COLLECTION_HANDLE,
                "descriptionHtml": "<p>Selected JIESTAR products eligible for U.S. warehouse fulfillment.</p>",
                "sortOrder": "MANUAL",
                "products": product_ids,
            }},
        )
        return data["collectionCreate"]

    def add_collection_products(self, collection_id: str, product_ids: list[str]) -> dict[str, Any]:
        data = self.graphql(
            """
            mutation AddUsWarehouseProducts($id: ID!, $productIds: [ID!]!) {
              collectionAddProducts(id: $id, productIds: $productIds) {
                collection { id handle }
                userErrors { field message }
              }
            }
            """,
            {"id": collection_id, "productIds": product_ids},
        )
        return data["collectionAddProducts"]

    def publish_collection(self, collection_id: str, publication_id: str) -> dict[str, Any]:
        data = self.graphql(
            """
            mutation PublishUsWarehouseCollection($id: ID!, $input: [PublicationInput!]!, $publicationId: ID!) {
              publishablePublish(id: $id, input: $input) {
                publishable { publishedOnPublication(publicationId: $publicationId) }
                userErrors { field message }
              }
            }
            """,
            {"id": collection_id, "input": [{"publicationId": publication_id}], "publicationId": publication_id},
        )
        return data["publishablePublish"]

    def activate_inventory_item(self, inventory_item_id: str, location_id: str) -> dict[str, Any]:
        data = self.graphql(
            """
            mutation ActivateUsWarehouseInventory($inventoryItemId: ID!, $locationId: ID!) {
              inventoryActivate(inventoryItemId: $inventoryItemId, locationId: $locationId) {
                inventoryLevel { id }
                userErrors { field message }
              }
            }
            """,
            {"inventoryItemId": inventory_item_id, "locationId": location_id},
        )
        return data["inventoryActivate"]


def method_definition(band: RateBand) -> dict[str, Any]:
    return {
        "name": RATE_NAME,
        "description": "U.S. warehouse fulfillment; final delivery timing depends on destination.",
        "active": True,
        "rateDefinition": {"price": {"amount": money_str(band.target_rate_usd), "currencyCode": "USD"}},
        "weightConditionsToCreate": [
            {"operator": "GREATER_THAN_OR_EQUAL_TO", "criteria": {"value": float(band.min_weight_kg), "unit": "KILOGRAMS"}},
            {"operator": "LESS_THAN_OR_EQUAL_TO", "criteria": {"value": float(band.max_weight_kg), "unit": "KILOGRAMS"}},
        ],
    }


def zone_input(bands: list[RateBand]) -> dict[str, Any]:
    return {
        "name": ZONE_NAME,
        "countries": [{"code": "US", "includeAllProvinces": True}],
        "methodDefinitionsToCreate": [method_definition(band) for band in bands],
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0]) if rows else ["status"]
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows or [{"status": "no_rows"}])


def build_live_plan(admin: UsWarehouseAdmin, workbook: Path) -> dict[str, Any]:
    costs = source_costs(workbook)
    active_variants = admin.active_variants()
    draft_variants = admin.draft_variants()
    eligible, source_audit = match_eligible_variants(active_variants, costs)
    bands = build_rate_bands(eligible)
    locations = admin.locations()
    location_matches = [location for location in locations if clean(location.get("name")) == US_LOCATION_NAME]
    location = location_matches[0] if len(location_matches) == 1 else None
    profiles = admin.delivery_profiles(include_profile_items=True)
    standard_profile = next((profile for profile in profiles if clean(profile.get("name")) == STANDARD_GOODS_PROFILE), None)
    location_group = find_location_group(standard_profile, location.get("id") if location else None)
    collection = admin.collection_snapshot()
    publications = admin.publication_snapshot()
    scopes = admin.access_scopes()

    eligible_product_ids = sorted({item.product_id for item in eligible})
    existing_product_ids = sorted(product["id"] for product in (collection or {}).get("products", {}).get("nodes", []))
    missing_product_ids = sorted(set(eligible_product_ids) - set(existing_product_ids))
    extra_product_ids = sorted(set(existing_product_ids) - set(eligible_product_ids))
    draft_skus = {variant.sku_key for variant in draft_variants}
    blockers: list[str] = []
    if len(eligible) != EXPECTED_VARIANT_COUNT or len(eligible_product_ids) != EXPECTED_PRODUCT_COUNT:
        blockers.append("expected_exactly_14_active_variants_across_12_products")
    if set(item.sku for item in eligible) != EXPECTED_ACTIVE_SKUS:
        blockers.append("expected_active_sku_set_mismatch")
    if any(item.weight_g <= 0 for item in eligible):
        blockers.append("eligible_variant_has_missing_weight")
    if any(item.inventory_tracked is not False for item in eligible):
        blockers.append("eligible_variant_inventory_tracking_is_not_disabled")
    if len(location_matches) != 1:
        blockers.append("create_one_internal_location_named_amazon_us_warehouse")
    elif location.get("isActive") is not True or location.get("fulfillsOnlineOrders") is not True:
        blockers.append("us_warehouse_location_must_be_active_for_online_fulfillment")
    if not standard_profile:
        blockers.append("standard_goods_profile_not_found")
    missing_scopes = sorted(REQUIRED_SCOPES - scopes)
    if missing_scopes:
        blockers.append("missing_required_admin_scopes")
    if "JJ9219" not in draft_skus:
        blockers.append("jj9219_is_not_confirmed_draft")
    if extra_product_ids:
        blockers.append("us_warehouse_collection_contains_unapproved_products")

    rate_rows = [
        {
            "location_group": (location_group or {}).get("locationGroup", {}).get("id", "create_new_group"),
            "zone": ZONE_NAME,
            "country": "US",
            "rate_name": RATE_NAME,
            "min_weight_kg": decimal_str(band.min_weight_kg),
            "max_weight_kg": decimal_str(band.max_weight_kg),
            "max_source_cost_usd": money_str(band.max_source_cost_usd),
            "formula": "(cost * 1.05 + 1.19) / 0.8901; round up to x.99",
            "target_rate_usd": money_str(band.target_rate_usd),
            "witness_skus": band.witness_skus,
        }
        for band in bands
    ]
    shopify_diff = [
        {"area": "location", "status": "ready" if location else "missing", "current": clean((location or {}).get("name")), "target": US_LOCATION_NAME},
        {"area": "location_group", "status": "exists" if location_group else "create", "current": clean((location_group or {}).get("locationGroup", {}).get("id")), "target": STANDARD_GOODS_PROFILE},
        {"area": "collection", "status": "exists" if collection else "create", "current": clean((collection or {}).get("id")), "target": COLLECTION_HANDLE},
        {"area": "collection_products_missing", "status": str(len(missing_product_ids)), "current": ",".join(missing_product_ids), "target": "12 exact products"},
        {"area": "collection_products_extra", "status": str(len(extra_product_ids)), "current": ",".join(extra_product_ids), "target": "0"},
        {"area": "us_group_existing_rates", "status": str(len(group_rate_fingerprint(location_group))), "current": json.dumps(group_rate_fingerprint(location_group), ensure_ascii=False), "target": str(len(bands))},
    ]

    plan = {
        "workbook": str(workbook),
        "location_name": US_LOCATION_NAME,
        "location_id": (location or {}).get("id"),
        "standard_profile_id": (standard_profile or {}).get("id"),
        "location_group_id": (location_group or {}).get("locationGroup", {}).get("id"),
        "eligible": [{**asdict(item), "source_cost_usd": money_str(item.source_cost_usd)} for item in eligible],
        "rates": rate_rows,
        "collection_id": (collection or {}).get("id"),
        "collection_product_ids": eligible_product_ids,
        "collection_missing_product_ids": missing_product_ids,
        "collection_extra_product_ids": extra_product_ids,
        "existing_us_group_rates": group_rate_fingerprint(location_group),
    }
    return {
        "mode": "dry_run_approval_only",
        "ready_for_apply": not blockers,
        "blockers": blockers,
        "missing_required_scopes": missing_scopes,
        "eligible_variant_count": len(eligible),
        "eligible_product_count": len(eligible_product_ids),
        "source_sku_count": len(costs),
        "deferred_skus": sorted(DEFERRED_SKUS),
        "publications": publications,
        "source_audit": source_audit,
        "rate_rows": rate_rows,
        "shopify_diff": shopify_diff,
        "plan": plan,
        "approval_hash": canonical_hash(plan),
    }


def save_report(report: dict[str, Any], out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary_path = out_dir / SUMMARY_JSON
    summary_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(out_dir / SOURCE_AUDIT_CSV, report["source_audit"])
    write_csv(out_dir / RATE_PLAN_CSV, report["rate_rows"])
    write_csv(out_dir / SHOPIFY_DIFF_CSV, report["shopify_diff"])
    return summary_path


def apply_approved_plan(
    admin: UsWarehouseAdmin,
    live_report: dict[str, Any],
    approved_report_path: Path,
    publication_id: str,
    out_dir: Path,
) -> dict[str, Any]:
    approved = json.loads(approved_report_path.read_text(encoding="utf-8"))
    if not approved.get("ready_for_apply"):
        raise RuntimeError("Approved report is not marked ready_for_apply")
    if approved.get("approval_hash") != live_report.get("approval_hash"):
        raise RuntimeError("Live Shopify/source plan differs from the approved report; generate a new approval report")
    if not publication_id.startswith("gid://shopify/Publication/"):
        raise RuntimeError("--publication-id must be a full Shopify Publication GID from the approval report")

    plan = live_report["plan"]
    snapshot = {
        "approval_hash": live_report["approval_hash"],
        "locations": admin.locations(),
        "profiles": admin.delivery_profiles(include_profile_items=True),
        "collection": admin.collection_snapshot(),
    }
    (out_dir / SNAPSHOT_JSON).write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")

    results: list[dict[str, Any]] = []
    location_id = plan["location_id"]
    for item in plan["eligible"]:
        result = admin.activate_inventory_item(item["inventory_item_id"], location_id)
        results.append({"type": "inventory_activate", "sku": item["sku"], "result": result})

    collection = admin.collection_snapshot()
    if not collection:
        created = admin.create_collection(plan["collection_product_ids"])
        results.append({"type": "collection_create", "result": created})
        if created.get("userErrors"):
            raise RuntimeError(f"Collection creation failed: {created['userErrors']}")
        collection_id = created["collection"]["id"]
    else:
        collection_id = collection["id"]
        missing_ids = plan["collection_missing_product_ids"]
        if missing_ids:
            added = admin.add_collection_products(collection_id, missing_ids)
            results.append({"type": "collection_add_products", "result": added})
            if added.get("userErrors"):
                raise RuntimeError(f"Adding collection products failed: {added['userErrors']}")
        if plan["collection_extra_product_ids"]:
            raise RuntimeError("Collection has extra products; remove them manually and generate a new approval report")

    published = admin.publish_collection(collection_id, publication_id)
    results.append({"type": "collection_publish", "result": published})
    if published.get("userErrors"):
        raise RuntimeError(f"Collection publication failed: {published['userErrors']}")

    bands = [
        RateBand(
            min_weight_kg=Decimal(row["min_weight_kg"]),
            max_weight_kg=Decimal(row["max_weight_kg"]),
            max_source_cost_usd=Decimal(row["max_source_cost_usd"]),
            target_rate_usd=Decimal(row["target_rate_usd"]),
            witness_skus=row["witness_skus"],
        )
        for row in plan["rates"]
    ]
    profile_id = plan["standard_profile_id"]
    group_id = plan["location_group_id"]
    existing_rates = plan["existing_us_group_rates"]
    if existing_rates:
        delete_input: dict[str, Any] = {"name": STANDARD_GOODS_PROFILE}
        method_ids = [row["method_id"] for row in existing_rates if row.get("method_id")]
        zone_ids = sorted({row["zone_id"] for row in existing_rates if row.get("zone_id")})
        if method_ids:
            delete_input["methodDefinitionsToDelete"] = method_ids
        if zone_ids:
            delete_input["zonesToDelete"] = zone_ids
        deleted = admin.delivery_profile_update(profile_id, delete_input)
        results.append({"type": "us_group_rate_delete", "result": deleted})
        if deleted.get("userErrors"):
            raise RuntimeError(f"Deleting old U.S. group rates failed: {deleted['userErrors']}")

    if group_id:
        rate_input = {
            "name": STANDARD_GOODS_PROFILE,
            "locationGroupsToUpdate": [{"id": group_id, "zonesToCreate": [zone_input(bands)]}],
        }
    else:
        rate_input = {
            "name": STANDARD_GOODS_PROFILE,
            "locationGroupsToCreate": [{"locationsToAdd": [location_id], "zonesToCreate": [zone_input(bands)]}],
        }
    rate_result = admin.delivery_profile_update(profile_id, rate_input)
    results.append({"type": "us_group_rate_apply", "result": rate_result})
    if rate_result.get("userErrors"):
        raise RuntimeError(f"Applying U.S. group rates failed: {rate_result['userErrors']}")

    verified = build_live_plan(admin, Path(plan["workbook"]))
    output = {"approval_hash": live_report["approval_hash"], "results": results, "verified_readback": verified}
    (out_dir / APPLY_RESULT_JSON).write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--yes", action="store_true")
    parser.add_argument("--approved-report", type=Path)
    parser.add_argument("--publication-id")
    args = parser.parse_args()

    if args.apply and (not args.yes or not args.approved_report or not args.publication_id):
        parser.error("--apply requires --yes, --approved-report, and --publication-id")

    admin = UsWarehouseAdmin()
    report = build_live_plan(admin, args.workbook)
    summary_path = save_report(report, args.out_dir)
    print(f"Approval report: {summary_path}")
    print(f"Ready for apply: {report['ready_for_apply']}")
    print(f"Approval hash: {report['approval_hash']}")
    if report["blockers"]:
        print("Blockers: " + ", ".join(report["blockers"]))

    if not args.apply:
        return 0
    if not report["ready_for_apply"]:
        raise RuntimeError("Live report has blockers; apply is disabled")
    apply_approved_plan(admin, report, args.approved_report, args.publication_id, args.out_dir)
    print(f"Apply and immediate readback complete: {args.out_dir / APPLY_RESULT_JSON}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
