#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import time
import urllib.parse
import urllib.request
from html import unescape
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))

import iblock_prepare_assets as prep


TARGET_ROOT = Path("/Volumes/ORICO/iblock/iblock-上架前整理")
SOURCE_ROOT = Path("/Volumes/ORICO/iblock/iBlock积趣_电商素材")
CATALOG_CSV = TARGET_ROOT / "reports" / "iblock-catalog-ready.csv"
BRICK4_WEB = "https://brick4.com"
BRICK4_CDN = "https://cdn.brick4.com/"
IBLOCK_BRAND_ID = "76"
USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"

MULTI_SKU_RANGES = {
    "IB1065": ("IB1065-IB1068", range(1065, 1069)),
    "IB1076": ("IB1076-IB1081", range(1076, 1082)),
    "IB1085": ("IB1085-IB1090", range(1085, 1091)),
    "IB1111": ("IB1111-IB1114", range(1111, 1115)),
}
CITY_DREAM_CHILD_GROUPS = {"IB1101", "IB1102", "IB1103", "IB1104"}
SINGLE_SKU_GUOHUN = {f"IB{number}" for number in range(1204, 1213)}
EXCLUDED_UPLOAD_SKUS = {
    "IB1101-5": "not a real upload SKU; IB1101 is the parent product with four child products",
    "IB1102-5": "not a real upload SKU; IB1102 is the parent product with four child products",
}
PARENT_MAIN_ALLOWED_FOR_SUBSKU = {"IB2101", "IB2202"}
PIECE_COUNT_FALLBACK = "See product package"
IB1301_ALIASES = {
    "IB1301-1": ["T25", "T-25"],
    "IB1301-2": ["J20", "J-20"],
    "IB1301-3": ["J05", "J-05"],
    "IB1301-4": ["99A"],
    "IB1301-5": ["Z10", "Z-10"],
    "IB1301-6": ["DF15", "DF"],
}
TITLE_OVERRIDES = {
    "IB1061": "iBlock Mythic Fire Warrior Mecha Building Block Set",
    "IB1062": "iBlock Mythic Ice Dragon Mecha Building Block Set",
    "IB1063": "iBlock Mythic Monkey Warrior Mecha Building Block Set",
    "IB1064": "iBlock Mythic Third-Eye Warrior Mecha Building Block Set",
    "IB1065": "iBlock Bottled Fairy Tale Queen Building Block Set",
    "IB1066": "iBlock Bottled Fairy Tale Tea Party Building Block Set",
    "IB1067": "iBlock Bottled Fairy Tale Snow Princess Building Block Set",
    "IB1068": "iBlock Bottled Fairy Tale Mermaid Party Building Block Set",
    "IB1201": "iBlock Aircraft Carrier Display Model Building Block Set",
    "IB1202": "iBlock Nuclear Submarine Display Model Building Block Set",
    "IB1203": "iBlock Stealth Fighter Mecha Building Block Set",
    "IB1204": "iBlock Battle Tank Mecha Building Block Set",
    "IB1205": "iBlock Amphibious Assault Ship Display Model Building Block Set",
    "IB1206": "iBlock Sixth Generation Fighter Display Model Building Block Set",
    "IB1207": "iBlock Military Drone Model Building Block Set",
    "IB1208": "iBlock Space Exploration Vehicle Building Block Set",
    "IB1209": "iBlock Carrier Fighter Display Model Building Block Set",
    "IB1210": "iBlock Military Transport Aircraft Building Block Set",
    "IB1211": "iBlock Aircraft Carrier Display Model Building Block Set",
    "IB1212": "iBlock Destroyer Ship Display Model Building Block Set",
    "IB1301-1": "iBlock Military Tank Mini Building Block Set",
    "IB1301-2": "iBlock Stealth Fighter Mini Military Building Block Set",
    "IB1301-3": "iBlock Stealth Fighter Mini Military Building Block Set",
    "IB1301-4": "iBlock Battle Tank Mini Military Building Block Set",
    "IB1301-5": "iBlock Attack Helicopter Mini Military Building Block Set",
    "IB1301-6": "iBlock Missile Launcher Mini Military Building Block Set",
}
TITLE_REPLACEMENTS = [
    ("深渊掠夺者", "Abyss Raider"),
    ("魔术师·机械之芯", "Mechanical Magician"),
    ("魔术师 机械之芯", "Mechanical Magician"),
    ("怪奇怪的系列 掌运【神抽】盲盒", "Lucky Mystery Mini Figure"),
    ("花屿蝶", "Butterfly Flower Display"),
    ("弗朗花", "Plumeria Flower"),
    ("樱花", "Cherry Blossom"),
    ("绣球花", "Hydrangea"),
    ("向日葵", "Sunflower"),
    ("四叶草", "Four Leaf Clover"),
    ("蓝风铃", "Blue Bellflower"),
    ("破界救援", "Rescue Team"),
    ("救援小队", "Rescue Team"),
    ("消防员", "Firefighter"),
    ("消防车", "Fire Truck"),
    ("救援机", "Rescue Aircraft"),
    ("救援艇", "Rescue Boat"),
    ("飓风行动", "Rapid Response Team"),
    ("飓风小队", "Rapid Response Team"),
    ("特警", "Police Officer"),
    ("警车", "Police Car"),
    ("追击机", "Pursuit Aircraft"),
    ("摩托车", "Motorcycle"),
    ("特种部队", "Special Operations Team"),
    ("特种小队", "Special Operations Team"),
    ("特种兵", "Special Forces Figure"),
    ("主战坦克", "Battle Tank"),
    ("东风导弹", "Missile Launcher"),
    ("无人机", "Drone"),
    ("重工崛起", "Construction Team"),
    ("重工小队", "Construction Team"),
    ("工程师", "Engineer"),
    ("搅拌车", "Concrete Mixer Truck"),
    ("推土机", "Bulldozer"),
    ("挖掘机", "Excavator"),
    ("望宇逐梦", "Space Exploration Team"),
    ("望宇小队", "Space Exploration Team"),
    ("宇航员", "Astronaut"),
    ("火箭", "Rocket"),
    ("探测车", "Rover"),
    ("卫星", "Satellite"),
    ("空战王牌", "Air Mission Team"),
    ("空战小队", "Air Mission Team"),
    ("飞行员", "Pilot"),
    ("运输机", "Transport Aircraft"),
    ("战斗机", "Fighter Jet"),
    ("直升机", "Helicopter"),
    ("治愈天使", "Medical Care Team"),
    ("治愈小队", "Medical Care Team"),
    ("医护员", "Medical Worker"),
    ("救护车", "Ambulance"),
    ("献血车", "Blood Donation Vehicle"),
    ("救护站", "First Aid Station"),
    ("套装", "Display Set"),
    ("展示盒", "Display Box"),
    ("花愿祈", "Floral Wish"),
    ("芍药", "Peony"),
    ("朱顶红", "Amaryllis"),
    ("牡丹菊", "Chrysanthemum"),
    ("大丽花", "Dahlia"),
    ("剑兰", "Gladiolus"),
    ("桃花", "Peach Blossom"),
    ("圣诞风铃", "Christmas Bell"),
    ("金牛座", "Taurus Zodiac"),
    ("巨蟹座", "Cancer Zodiac"),
    ("天蝎座", "Scorpio Zodiac"),
    ("射手座", "Sagittarius Zodiac"),
    ("双鱼座", "Pisces Zodiac"),
    ("双子座", "Gemini Zodiac"),
    ("处女座", "Virgo Zodiac"),
    ("水瓶座", "Aquarius Zodiac"),
    ("狮子座", "Leo Zodiac"),
    ("天秤座", "Libra Zodiac"),
    ("白羊座", "Aries Zodiac"),
    ("摩羯座", "Capricorn Zodiac"),
    ("四时花境", "Seasonal Flower Garden"),
    ("独角仙", "Rhinoceros Beetle"),
    ("蚂蚱", "Grasshopper"),
    ("七星瓢虫", "Ladybug"),
    ("蜻蜓", "Dragonfly"),
    ("蜘蛛", "Spider"),
    ("蝴蝶", "Butterfly"),
    ("蝎子", "Scorpion"),
    ("知了", "Cicada"),
    ("蜜蜂", "Bee"),
]
SERIES_FALLBACKS = [
    ("十二生肖", "Zodiac Character"),
    ("十二星座", "Zodiac Character"),
    ("极速方程", "Formula Race Car"),
    ("花漾玲珑", "Floral Display"),
    ("四时花境", "Seasonal Flower Garden"),
    ("城市梦英雄", "City Heroes"),
    ("国魂", "Military Display Model"),
    ("MINI战线", "Mini Military Model"),
    ("虫界漫游", "Insect World Model"),
]
IP_SENSITIVE_TERMS = [
    "哪吒",
    "悟空",
    "杨戬",
    "敖丙",
    "齐天",
    "封神",
    "白雪公主",
    "红桃皇后",
    "歼-15",
    "福建舰",
    "四川舰",
    "055",
    "99A",
    "T-25",
    "J-20",
    "J35",
    "Z-10",
]


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", unescape(str(value or ""))).strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as file:
        return list(csv.DictReader(file))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for name, rows in sheets.items():
        sheet = workbook.create_sheet(name[:31])
        if not rows:
            continue
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 52)
    workbook.save(path)


def upload_group_for_sku(sku: str) -> tuple[str, str]:
    base = sku.split("-", 1)[0]
    if base in SINGLE_SKU_GUOHUN:
        return sku, "SINGLE_SKU_PRODUCT"
    if sku == base and base in CITY_DREAM_CHILD_GROUPS:
        return sku, "SINGLE_SKU_PRODUCT"
    if base in CITY_DREAM_CHILD_GROUPS and "-" in sku:
        suffix = sku.split("-", 1)[1]
        if suffix in {"1", "2", "3", "4"}:
            return f"{base}-1-{base}-4", "MULTI_SKU_PRODUCT"
    for group_key, numbers in MULTI_SKU_RANGES.values():
        if base.startswith("IB") and int(base[2:]) in numbers:
            return group_key, "MULTI_SKU_PRODUCT"
    if base == "IB1401" and "-" in sku and sku.split("-", 1)[1] in {"1", "2", "3", "4", "5", "6"}:
        return "IB1401-1-IB1401-6", "MULTI_SKU_PRODUCT"
    if base == "IB2001" and "-" in sku:
        try:
            suffix_number = int(sku.split("-", 1)[1])
        except ValueError:
            suffix_number = 0
        if 1 <= suffix_number <= 12:
            return "IB2001-1-IB2001-12", "MULTI_SKU_PRODUCT"
    if base == "IB2202" and (sku == "IB2202" or "-" in sku):
        if sku == "IB2202":
            return "IB2202-1-IB2202-9", "MULTI_SKU_PRODUCT"
        try:
            suffix_number = int(sku.split("-", 1)[1])
        except ValueError:
            suffix_number = 0
        if 1 <= suffix_number <= 9:
            return "IB2202-1-IB2202-9", "MULTI_SKU_PRODUCT"
    return sku, "SINGLE_SKU_PRODUCT"


def is_excluded_sku(sku: str) -> bool:
    return sku.upper() in EXCLUDED_UPLOAD_SKUS


def title_words(value: str) -> list[str]:
    return [piece for piece in re.split(r"[()（）·.\-—_/、,，\s]+", clean_text(value)) if piece]


def local_aliases(sku: str, name_cn: str) -> list[str]:
    aliases = [sku, sku.split("-", 1)[0], *IB1301_ALIASES.get(sku, [])]
    for word in title_words(name_cn):
        compact = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff]", "", word).upper()
        if len(compact) >= 2:
            aliases.append(compact)
    return list(dict.fromkeys(aliases))


def local_match_aliases(sku: str, name_cn: str) -> tuple[list[str], list[str]]:
    aliases = local_aliases(sku, name_cn)
    if "-" not in sku:
        return aliases, []
    broad_parent = sku.split("-", 1)[0]
    strict = [alias for alias in aliases if alias.upper() not in {broad_parent.upper()}]
    return strict, [broad_parent]


def explicit_sku_bases_in_text(text: str) -> set[str]:
    bases: set[str] = set()
    for match in re.finditer(r"(?<![A-Z0-9])(IB|LL|CM)\s*[-_ ]?\s*(\d{3,5})(?:\s*-\s*\d+)?", text.upper(), re.I):
        bases.add(f"{match.group(1).upper()}{match.group(2)}")
    return bases


def safe_shopify_title(sku: str, product_series: str, name_cn: str) -> str:
    if sku in TITLE_OVERRIDES:
        return TITLE_OVERRIDES[sku]
    series = clean_text(product_series).replace("-", "")
    name = clean_text(name_cn)
    replacements = [
        ("麦克拉仑", "Orange Race Car"),
        ("默席迪斯", "Silver Race Car"),
        ("斗牛", "White Race Car"),
        ("法兰", "Red Race Car"),
        ("白雪公主", "Snow Princess"),
        ("红桃皇后", "Queen"),
        ("福建舰", "Aircraft Carrier"),
        ("四川舰", "Amphibious Assault Ship"),
        ("歼-15", "Carrier Fighter"),
        ("J35", "Stealth Fighter"),
        ("99A", "Battle Tank"),
        ("055", "Destroyer Ship"),
        ("哪吒", "Mythic Fire Warrior"),
        ("悟空", "Mythic Monkey Warrior"),
        ("杨戬", "Mythic Third-Eye Warrior"),
        ("敖丙", "Mythic Ice Dragon"),
    ]
    for source, target in replacements:
        name = name.replace(source, target)
    for source, target in TITLE_REPLACEMENTS:
        name = name.replace(source, f" {target} ")
    ascii_words = re.findall(r"[A-Za-z0-9]+(?:-[A-Za-z0-9]+)?", name)
    if ascii_words:
        core = " ".join(ascii_words)
    elif "十二生肖" in series:
        core = f"Zodiac Character {sku.split('-')[-1]}"
    elif "花" in series:
        core = ""
        for source, target in SERIES_FALLBACKS:
            if source in series:
                core = f"{target} {sku.split('-')[-1] if '-' in sku else sku[-2:]}"
                break
        core = core or "Floral Display"
    else:
        core = ""
        for source, target in SERIES_FALLBACKS:
            if source in series:
                core = f"{target} {sku.split('-')[-1] if '-' in sku else sku[-2:]}"
                break
        core = core or sku
    return f"iBlock {core} Building Block Set"


def ip_sensitive_terms_found(*values: str) -> str:
    text = " ".join(clean_text(value) for value in values)
    found = [term for term in IP_SENSITIVE_TERMS if term in text]
    return "; ".join(dict.fromkeys(found))


def request_text(url: str, retries: int = 2) -> str:
    last: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(request, timeout=35) as response:
                return response.read().decode("utf-8", errors="ignore")
        except Exception as error:  # noqa: BLE001 - network evidence is kept in report.
            last = error
            time.sleep(1.0 * attempt)
    raise RuntimeError(f"request failed: {url}: {last}")


def request_json(url: str) -> dict[str, Any]:
    return json.loads(request_text(url))


def exact_iblock_match(node: dict[str, Any], sku: str) -> bool:
    normalized = sku.upper()
    for setnumber in node.get("setnumber") or []:
        number = clean_text(setnumber.get("setnumber")).upper()
        brand = clean_text(setnumber.get("brand"))
        if brand == IBLOCK_BRAND_ID and number == normalized:
            return True
    return False


def clean_image_url(value: str) -> str:
    value = clean_text(value).strip('"').strip("'")
    value = re.sub(r"!(?:richtext|bigpic|setcover|setlist|large|thumb).*$", "", value)
    if value.startswith("//"):
        value = "https:" + value
    if value.startswith("http://"):
        value = "https://" + value[len("http://") :]
    if value.startswith("https://"):
        return value
    return BRICK4_CDN + value.lstrip("/")


def image_extension(url: str) -> str:
    suffix = Path(urllib.parse.urlparse(clean_image_url(url)).path).suffix.lower()
    return suffix if suffix in {".jpg", ".jpeg", ".png", ".webp"} else ".jpg"


def fetch_brick4(sku: str) -> dict[str, Any]:
    params = urllib.parse.urlencode({"s": sku, "page": "1"})
    data = request_json(f"{BRICK4_WEB}/get/set?{params}")
    for node in data.get("data") or []:
        if exact_iblock_match(node, sku):
            set_id = clean_text(node.get("id"))
            url = f"{BRICK4_WEB}/set/{set_id}/"
            if clean_text(node.get("title2url")):
                url += urllib.parse.quote(clean_text(node.get("title2url")).strip("/"))
            album: list[str] = []
            richtext: list[str] = []
            html = ""
            if set_id:
                try:
                    html = request_text(url)
                    album = [clean_image_url(match) for match in re.findall(r'data-imgurl="([^"]+)"', html)]
                    richtext = [clean_image_url(match) for match in re.findall(r'<img[^>]+src="([^"]+)"', html)]
                except RuntimeError:
                    pass
            return {
                "exact": "yes",
                "brick4_id": set_id,
                "brick4_url": url,
                "brick4_title": clean_text(node.get("title")),
                "brick4_subtitle": clean_text(node.get("subtitle")),
                "brick4_pcs": clean_text(node.get("pcs")),
                "brick4_theme": clean_text(node.get("theme")),
                "cover": clean_image_url(clean_text(node.get("cover"))) if clean_text(node.get("cover")) else "",
                "album": album,
                "richtext": richtext,
            }
    return {"exact": "no"}


def has_marker(folder: Path, marker: str) -> bool:
    return folder.exists() and any(path.is_file() and marker in path.name and not path.name.startswith("._") for path in folder.iterdir())


def has_role(folder: Path, role: str) -> bool:
    if not folder.exists():
        return False
    files = [path.name for path in folder.iterdir() if path.is_file() and not path.name.startswith("._")]
    if role == "white":
        return any("-白底" in name for name in files)
    if role == "sku":
        return any("-sku" in name for name in files)
    if role == "detail":
        return any("-详情" in name for name in files)
    if role == "main":
        return any(all(marker not in name for marker in ("-白底", "-sku", "-详情")) for name in files)
    return False


def role_marker(role: str) -> str:
    return {"white": "-白底", "main": "-1", "sku": "-sku", "detail": "-详情"}.get(role, "")


def local_deep_candidates(row: dict[str, str]) -> list[tuple[Path, str]]:
    sku = row["sku"]
    base = sku.split("-", 1)[0]
    strict_aliases, broad_aliases = local_match_aliases(row["sku"], row["name_cn"])
    candidates: list[tuple[Path, str]] = []
    for source in prep.image_files(SOURCE_ROOT):
        relative = source.relative_to(SOURCE_ROOT)
        role, ignored = prep.classify_role(relative)
        if role == "ignored" or ignored:
            continue
        text = relative.as_posix().upper()
        explicit_bases = explicit_sku_bases_in_text(text)
        if explicit_bases and base.upper() not in explicit_bases:
            continue
        if not any(alias.upper() in text for alias in strict_aliases):
            parent_main_match = (
                role == "main"
                and base in PARENT_MAIN_ALLOWED_FOR_SUBSKU
                and broad_aliases
                and any(alias.upper() in text for alias in broad_aliases)
            )
            if parent_main_match:
                candidates.append((source, role))
                continue
            if broad_aliases and any(alias.upper() in text for alias in broad_aliases):
                continue
            continue
        candidates.append((source, role))
    return candidates


def next_copy_name(folder: Path, sku: str, role: str, source: Path, prefix: str) -> Path:
    suffix = source.suffix.lower() if source.suffix.lower() in prep.IMAGE_EXTS else ".jpg"
    existing = [path for path in folder.glob(f"{sku}-*") if path.is_file() and not path.name.startswith("._")]
    if role == "white":
        return folder / f"{sku}-白底-{prefix}{suffix}"
    if role == "sku":
        return folder / f"{sku}-sku-{prefix}{suffix}"
    if role == "detail":
        count = sum(1 for path in existing if "-详情" in path.name) + 1
        return folder / f"{sku}-详情-{prefix}-{count:02d}{suffix}"
    count = sum(1 for path in existing if not any(marker in path.name for marker in ("-白底", "-sku", "-详情"))) + 1
    return folder / f"{sku}-{prefix}-{count:02d}{suffix}"


def copy_local_missing(row: dict[str, str], apply: bool) -> list[dict[str, str]]:
    sku = row["sku"]
    folder = TARGET_ROOT / "images" / sku
    copied: list[dict[str, str]] = []
    for source, role in local_deep_candidates(row):
        if has_role(folder, role):
            continue
        if role not in {"white", "main", "sku", "detail"}:
            continue
        target = next_copy_name(folder, sku, role, source, "local")
        copied.append({"sku": sku, "role": role, "source": source.as_posix(), "target": target.as_posix()})
        if apply:
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
    return copied


def download_brick4_images(row: dict[str, str], spec: dict[str, Any], apply: bool) -> list[dict[str, str]]:
    if spec.get("exact") != "yes":
        return []
    sku = row["sku"]
    folder = TARGET_ROOT / "images" / sku
    downloads: list[dict[str, str]] = []
    if not has_marker(folder, "-白底") and spec.get("cover"):
        url = spec["cover"]
        target = folder / f"{sku}-白底-brick4{image_extension(url)}"
        downloads.append({"sku": sku, "role": "white", "source": url, "target": target.as_posix()})
        if apply:
            target.parent.mkdir(parents=True, exist_ok=True)
            request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(request, timeout=45) as response:
                target.write_bytes(response.read())
    if not has_role(folder, "main") and spec.get("cover"):
        url = spec["cover"]
        target = folder / f"{sku}-brick4-main{image_extension(url)}"
        downloads.append({"sku": sku, "role": "main", "source": url, "target": target.as_posix()})
        if apply:
            target.parent.mkdir(parents=True, exist_ok=True)
            request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(request, timeout=45) as response:
                target.write_bytes(response.read())
    if not has_marker(folder, "-详情"):
        for index, url in enumerate((spec.get("richtext") or spec.get("album") or [])[:10], start=1):
            target = folder / f"{sku}-详情-brick4-{index:02d}{image_extension(url)}"
            downloads.append({"sku": sku, "role": "detail", "source": url, "target": target.as_posix()})
            if apply:
                target.parent.mkdir(parents=True, exist_ok=True)
                request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
                with urllib.request.urlopen(request, timeout=45) as response:
                    target.write_bytes(response.read())
    return downloads


def image_counts(sku: str) -> dict[str, int]:
    folder = TARGET_ROOT / "images" / sku
    files = [path.name for path in folder.iterdir()] if folder.exists() else []
    files = [name for name in files if not name.startswith("._")]
    return {
        "white_image_count_final": sum("-白底" in name for name in files),
        "main_image_count_final": sum(all(marker not in name for marker in ("-白底", "-sku", "-详情")) for name in files),
        "sku_image_count_final": sum("-sku" in name for name in files),
        "detail_image_count_final": sum("-详情" in name for name in files),
    }


def readiness(counts: dict[str, int], upload_mode: str, brick4: dict[str, Any]) -> tuple[str, str]:
    gaps: list[str] = []
    if counts["white_image_count_final"] == 0 and counts["sku_image_count_final"] == 0:
        gaps.append("missing white/SKU image")
    if counts["main_image_count_final"] == 0:
        gaps.append("missing main image")
    if counts["detail_image_count_final"] == 0:
        gaps.append("missing detail image")
    return ("READY_FOR_UPLOAD_PREP" if not gaps else "NEEDS_REVIEW", "; ".join(gaps))


def piece_count_display_value(spec: dict[str, Any]) -> str:
    return clean_text(spec.get("brick4_pcs")) or PIECE_COUNT_FALLBACK


def grouped_image_files(sku: str) -> list[Path]:
    folder = TARGET_ROOT / "images" / sku
    if not folder.exists():
        return []
    return sorted(
        [path for path in folder.iterdir() if path.is_file() and not path.name.startswith("._")],
        key=lambda path: prep.natural_string_key(path.name),
    )


def variant_skus_for_group(rows: list[dict[str, Any]]) -> list[str]:
    skus = [str(row["sku"]) for row in rows]
    child_skus = [sku for sku in skus if "-" in sku]
    if child_skus:
        child_bases = {sku.split("-", 1)[0] for sku in child_skus}
        return [sku for sku in skus if not (sku in child_bases and len(child_bases) == 1)]
    return skus


def write_product_group_assets(rows: list[dict[str, Any]], apply: bool) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    product_root = TARGET_ROOT / "shopify-products"
    if apply and product_root.exists():
        shutil.rmtree(product_root)
    by_group: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_group.setdefault(str(row["upload_group"]), []).append(row)

    group_rows: list[dict[str, Any]] = []
    asset_rows: list[dict[str, str]] = []
    for group, group_members in sorted(by_group.items(), key=lambda item: prep.natural_string_key(item[0])):
        group_members = sorted(group_members, key=lambda row: prep.natural_string_key(str(row["sku"])))
        group_folder = product_root / group
        image_folder = group_folder / "images"
        group_asset_count = 0
        for row in group_members:
            sku = str(row["sku"])
            for source in grouped_image_files(sku):
                target = image_folder / f"{sku}__{source.name}"
                asset_rows.append(
                    {
                        "upload_group": group,
                        "sku": sku,
                        "source": source.as_posix(),
                        "target": target.as_posix(),
                    }
                )
                group_asset_count += 1
                if apply:
                    target.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(source, target)
        variant_skus = variant_skus_for_group(group_members)
        parent_candidates = [str(row["sku"]) for row in group_members if str(row["sku"]) not in variant_skus]
        group_rows.append(
            {
                "upload_group": group,
                "upload_mode": "MULTI_SKU_PRODUCT" if len(variant_skus) > 1 else "SINGLE_SKU_PRODUCT",
                "parent_sku": ", ".join(parent_candidates),
                "variant_skus": ", ".join(variant_skus),
                "variant_count": len(variant_skus),
                "source_row_count": len(group_members),
                "group_image_count": group_asset_count,
                "product_folder": group_folder.as_posix(),
            }
        )
        if apply:
            group_folder.mkdir(parents=True, exist_ok=True)
            (group_folder / "variants.csv").write_text("sku\n" + "\n".join(variant_skus) + "\n", encoding="utf-8")
    return group_rows, asset_rows


def build_rows(
    apply: bool, limit: int | None = None
) -> tuple[list[dict[str, Any]], list[dict[str, str]], list[dict[str, str]], list[dict[str, Any]]]:
    catalog = read_csv(CATALOG_CSV)
    if limit:
        catalog = catalog[:limit]
    specs: dict[str, dict[str, Any]] = {}
    enriched: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    local_copies: list[dict[str, str]] = []
    brick4_downloads: list[dict[str, str]] = []
    for index, row in enumerate(catalog, start=1):
        sku = row["sku"]
        if is_excluded_sku(sku):
            excluded.append(
                {
                    **row,
                    "exclude_reason": EXCLUDED_UPLOAD_SKUS[sku.upper()],
                    "upload_readiness_after_enrichment": "EXCLUDED_NOT_REAL_SKU",
                }
            )
            continue
        local_copies.extend(copy_local_missing(row, apply=apply))
        try:
            spec = fetch_brick4(sku)
        except Exception as error:  # noqa: BLE001 - record network failures per SKU.
            spec = {"exact": "error", "error": str(error)}
        specs[sku] = spec
        brick4_downloads.extend(download_brick4_images(row, spec, apply=apply))
        counts = image_counts(sku)
        group_key, upload_mode = upload_group_for_sku(sku)
        state, reason = readiness(counts, upload_mode, spec)
        safe_title = safe_shopify_title(sku, row["product_series"], row["name_cn"])
        ip_terms = ip_sensitive_terms_found(row["product_series"], row["name_cn"], safe_title)
        enriched.append(
            {
                **row,
                "shopify_price": "999",
                "upload_group": group_key,
                "upload_mode": upload_mode,
                "shopify_title_safe": safe_title,
                "ip_sensitive_terms_found": ip_terms,
                "ip_avoidance_status": "TITLE_REVIEW" if ip_sensitive_terms_found(safe_title) else "TITLE_IP_AVOIDED",
                "specs_piece_count": spec.get("brick4_pcs", ""),
                "specs_piece_count_display": piece_count_display_value(spec),
                "specs_piece_count_note": "" if spec.get("brick4_pcs") else "No manufacturer piece count provided; storefront can display fallback text.",
                "specs_recommended_age": row.get("recommended_age", ""),
                "specs_finished_model_size": row.get("product_size_cm", ""),
                "specs_package_size": row.get("box_size_cm", ""),
                "specs_difficulty_level": "See product package",
                "custom_series": row.get("product_series") or row.get("brand_series") or "",
                "brick4_exact_match": spec.get("exact", ""),
                "brick4_url": spec.get("brick4_url", ""),
                "brick4_title": spec.get("brick4_title", ""),
                **counts,
                "upload_readiness_after_enrichment": state,
                "remaining_gap": reason,
            }
        )
        time.sleep(0.12)
        if index % 25 == 0:
            print(f"[iblock readiness] {index}/{len(catalog)}", flush=True)
    return enriched, local_copies, brick4_downloads, excluded


def run(apply: bool, limit: int | None = None) -> dict[str, Any]:
    rows, local_copies, brick4_downloads, excluded = build_rows(apply=apply, limit=limit)
    product_groups, product_assets = write_product_group_assets(rows, apply=apply)
    reports = TARGET_ROOT / "reports"
    readiness_csv = reports / "iblock-shopify-readiness.csv"
    manual_csv = reports / "iblock-still-missing-for-manual.csv"
    local_csv = reports / "iblock-local-deep-backfill.csv"
    brick4_csv = reports / "iblock-brick4-backfill.csv"
    excluded_csv = reports / "iblock-excluded-not-real-sku.csv"
    product_groups_csv = reports / "iblock-product-groups.csv"
    product_assets_csv = reports / "iblock-product-group-assets.csv"
    web_search_csv = reports / "iblock-web-search-unmatched-brick4.csv"
    workbook = TARGET_ROOT / "iblock-shopify-readiness.xlsx"
    manual_rows = [row for row in rows if row["upload_readiness_after_enrichment"] != "READY_FOR_UPLOAD_PREP"]
    web_search_rows = read_csv(web_search_csv) if web_search_csv.exists() and web_search_csv.stat().st_size else []
    write_csv(readiness_csv, rows)
    write_csv(manual_csv, manual_rows)
    write_csv(local_csv, local_copies)
    write_csv(brick4_csv, brick4_downloads)
    write_csv(excluded_csv, excluded)
    write_csv(product_groups_csv, product_groups)
    write_csv(product_assets_csv, product_assets)
    write_workbook(
        workbook,
        {
            "Readiness": rows,
            "StillMissing": manual_rows,
            "ProductGroups": product_groups,
            "ProductGroupAssets": product_assets,
            "LocalBackfill": local_copies,
            "Brick4Backfill": brick4_downloads,
            "ExcludedNotRealSKU": excluded,
            "WebSearchUnmatched": web_search_rows,
        },
    )
    prep.remove_appledouble_files(TARGET_ROOT)
    summary = {
        "apply": apply,
        "checked_sku_count": len(rows),
        "excluded_not_real_sku_count": len(excluded),
        "ready_for_upload_prep": sum(row["upload_readiness_after_enrichment"] == "READY_FOR_UPLOAD_PREP" for row in rows),
        "still_needs_review": len(manual_rows),
        "product_group_count": len(product_groups),
        "product_group_asset_count": len(product_assets),
        "brick4_exact_match_count": sum(row["brick4_exact_match"] == "yes" for row in rows),
        "piece_count_backfilled_count": sum(bool(row["specs_piece_count"]) for row in rows),
        "local_backfill_file_count": len(local_copies),
        "brick4_backfill_file_count": len(brick4_downloads),
        "outputs": {
            "workbook": workbook.as_posix(),
            "readiness_csv": readiness_csv.as_posix(),
            "manual_csv": manual_csv.as_posix(),
            "local_backfill_csv": local_csv.as_posix(),
            "brick4_backfill_csv": brick4_csv.as_posix(),
            "excluded_csv": excluded_csv.as_posix(),
            "product_groups_csv": product_groups_csv.as_posix(),
            "product_assets_csv": product_assets_csv.as_posix(),
            "web_search_csv": web_search_csv.as_posix(),
        },
    }
    summary_path = reports / "iblock-shopify-readiness-summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    prep.remove_appledouble_files(TARGET_ROOT)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Enrich iBlock staging data for Shopify upload readiness.")
    parser.add_argument("--apply", action="store_true", help="Copy/download missing assets into the staging folder.")
    parser.add_argument("--dry-run", action="store_true", help="Write reports without copying/downloading assets.")
    parser.add_argument("--limit", type=int)
    args = parser.parse_args()
    if args.apply and args.dry_run:
        parser.error("--apply and --dry-run cannot be used together")
    run(apply=args.apply, limit=args.limit)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
