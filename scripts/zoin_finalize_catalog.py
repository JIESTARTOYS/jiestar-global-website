#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


TARGET_ROOT = Path("/Volumes/ORICO/积域资料/Zoin-上架前整理")
REPORT_DIR = Path("/private/tmp/zoin-product-prep")
MANIFEST_CSV = REPORT_DIR / "zoin-products-manifest.csv"
BRICK4_CSV = REPORT_DIR / "zoin-brick4-supplement.csv"
FINAL_AUDIT_JSON = REPORT_DIR / "zoin-final-local-asset-audit.json"

SERIES_EN = {
    "躲豆豆": "Hide-And-Seek Scenes",
    "小火车": "Mini Food Train",
    "花境": "Floral Display",
    "筑梦": "DongDong Dream",
    "植物": "Botanical Display",
    "大国重器机甲": "Heavy Machine Mecha",
    "魂铠": "Mythic Armor Mecha",
    "零限创意": "Creative Anatomy",
    "蔬果动物": "Fruit And Vegetable Animals",
    "种一棵": "Mini Bonsai Trees",
    "节庆": "Festive Display",
    "建筑": "Architecture",
    "DongDong糖": "DongDong Candy",
    "来一口": "Mini Food Display",
    "艺术随想": "Art Inspired Display",
}

NAME_EN = {
    "GT201": "Spring Hide-And-Seek",
    "GT202": "Trash Bin Hide-And-Seek",
    "GT203": "Snack Pack Hide-And-Seek",
    "GT204": "Pillow Hide-And-Seek",
    "GT112": "Popcorn Mini Train",
    "GT113": "Rainbow Candy Box Mini Train",
    "GT114": "Bakery Mini Train",
    "GT115": "Cocoa Ice Cream Mini Train",
    "GT105": "Squirrel Garden",
    "GT106": "DongDong Garden Waltz",
    "GT107": "Galaxy Fishing Garden",
    "GT109": "DongDong Dream Journey",
    "GT110": "DongDong Nutcracker Figure",
    "FT101": "Floral Reverie",
    "FT102": "Blooming Fairy Garden",
    "GK401": "Black Rose",
    "GK402": "Gold Rose",
    "GK405": "Spooky Night Botanical Display",
    "GK508": "Land Battle Mecha",
    "GK509": "Sky Dragon Mecha",
    "GK510": "Cruiser Battle Mecha",
    "GK501": "Monkey Warrior Armor",
    "GK502": "Dark Guardian Armor",
    "GK503": "Light Guardian Armor",
    "GK504": "Ghost Warrior Armor",
    "GK506": "Anubis Armor",
    "GK507": "Horus Armor",
    "GK101": "Brain Model",
    "GK102": "Heart Model",
    "GK103": "Tooth Model",
    "GK104": "Eye Model",
    "GK105": "Raven Model",
    "GK602": "Durian Chicken",
    "GK603": "Bitter Melon Crocodile",
    "GK604": "Lemon Fish",
    "GK605": "Peach Pig",
    "GK403": "Cherry Blossom Bonsai",
    "GK404": "Pine Bonsai",
    "GK406": "Plum Blossom Bonsai",
    "GK407": "Persimmon Bonsai",
    "GK301": "Nine Color Deer",
    "GK303": "Money Catching Snake",
    "GK304": "Fortune Snake",
    "GK201": "Roman Arena Architecture",
    "GK202": "Classical Basilica Throne Architecture",
    "GT111": "DongDong Candy Blind Box",
    "GK606": "Mini Barbecue Blind Box",
    "GK601": "Mini Food Blind Box",
    "GK107": "Macaw Model",
    "GT101": "Masked Opera Figure",
    "GT102": "Winged Victory Figure",
    "GT103": "Moonlit Starry Art Figure",
    "GT104": "Morning Sonata Figure",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as file:
        return list(csv.DictReader(file))


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def handleize(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return re.sub(r"-+", "-", value).strip("-")


def file_counts(sku: str) -> dict[str, Any]:
    folder = TARGET_ROOT / "images" / sku
    files = [path.name for path in folder.iterdir() if path.is_file() and not path.name.startswith("._")]
    white = [name for name in files if "-白底" in name]
    detail = [name for name in files if "-详情" in name]
    main = [name for name in files if "-白底" not in name and "-详情" not in name]
    return {
        "image_folder": folder.as_posix(),
        "image_file_count": len(files),
        "white_image_count_final": len(white),
        "main_image_count_final": len(main),
        "detail_image_count_final": len(detail),
        "missing_white_final": "yes" if not white else "no",
        "missing_main_final": "yes" if not main else "no",
        "missing_detail_final": "yes" if not detail else "no",
    }


def image_status(counts: dict[str, Any]) -> tuple[str, str]:
    reasons: list[str] = []
    if counts["missing_white_final"] == "yes":
        reasons.append("missing white image")
    if counts["missing_main_final"] == "yes":
        reasons.append("missing main image")
    if counts["missing_detail_final"] == "yes":
        reasons.append("missing detail image")
    if reasons:
        return "IMAGE_REVIEW", "; ".join(reasons)
    return "IMAGE_READY", ""


def upload_readiness(counts: dict[str, Any]) -> tuple[str, str]:
    blockers: list[str] = []
    if counts["missing_white_final"] == "yes":
        blockers.append("missing white image")
    if counts["missing_main_final"] == "yes":
        blockers.append("missing main image")

    if blockers:
        return "HOLD_FOR_IMAGE_REVIEW", "; ".join(blockers)

    if counts["missing_detail_final"] == "yes":
        return "READY_BASE_IMAGES_DETAIL_REVIEW", "missing detail image"

    return "READY_FOR_REVIEW", ""


def metafield_gap(row: dict[str, str], brick4: dict[str, str]) -> str:
    reasons: list[str] = []
    if not (row.get("piece_count") or brick4.get("brick4_pcs")):
        reasons.append("missing piece count")
    if brick4.get("exact_match") != "yes":
        reasons.append("no exact Brick4 Zoin match")
    return "; ".join(reasons)


def build_catalog_rows() -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, Any]]:
    manifest = read_csv(MANIFEST_CSV)
    brick4_by_sku = {row["sku"]: row for row in read_csv(BRICK4_CSV)}
    catalog: list[dict[str, str]] = []
    manual: list[dict[str, str]] = []

    for source in manifest:
        sku = source["sku"]
        name_en = NAME_EN.get(sku, source["name_cn"])
        series_en = SERIES_EN.get(source["series"], source["series"] or "Building Block Sets")
        title = f"Zoin {name_en} Building Block Set"
        counts = file_counts(sku)
        brick4 = brick4_by_sku.get(sku, {})
        piece_count = source.get("piece_count") or brick4.get("brick4_pcs", "")
        piece_count_source = "quote_table" if source.get("piece_count") else ("Brick4 exact match" if brick4.get("brick4_pcs") else "")
        img_status, img_reason = image_status(counts)
        readiness, readiness_reason = upload_readiness(counts)
        field_gap = metafield_gap(source, brick4)
        row = {
            "sku": sku,
            "vendor": "Zoin",
            "shopify_title": title,
            "handle": handleize(title),
            "product_type": series_en,
            "category": "Interlocking Blocks",
            "price": "999",
            "status_recommendation": "DRAFT" if readiness == "HOLD_FOR_IMAGE_REVIEW" else "READY_TO_CREATE",
            "upload_readiness": readiness,
            "upload_readiness_reason": readiness_reason,
            "image_prep_status": img_status,
            "image_review_reason": img_reason,
            "metafield_gap": field_gap,
            "metafield_review_needed": "yes" if field_gap else "no",
            "series_cn": source["series"],
            "series_en": series_en,
            "name_cn": source["name_cn"],
            "name_en": name_en,
            "specs_piece_count": piece_count,
            "specs_piece_count_source": piece_count_source,
            "specs_piece_count_source_url": brick4.get("brick4_url") if brick4.get("brick4_pcs") else "",
            "specs_recommended_age": source["recommended_age"],
            "specs_finished_model_size": source["finished_model_size"],
            "specs_package_size": source["package_size"],
            "specs_difficulty_level": "See product package",
            "custom_series": series_en,
            "package_type": source["package_type"],
            "carton_size": source["carton_size"],
            "carton_qty": source["carton_qty"],
            "gross_net_weight": source["gross_net_weight"],
            "volume": source["volume"],
            "factory_price_cny": source["factory_price_cny"],
            "has_minifigure": "yes" if source["has_minifigure"] == "有" else "no",
            "certifications_source": source["certifications"],
            "brick4_exact_match": brick4.get("exact_match", ""),
            "brick4_url": brick4.get("brick4_url", ""),
            "brick4_title": brick4.get("brick4_title", ""),
            "brick4_subtitle": brick4.get("brick4_subtitle", ""),
            **{key: str(value) for key, value in counts.items()},
        }
        catalog.append(row)
        if readiness != "READY_FOR_REVIEW" or field_gap:
            manual.append(row)

    summary = {
        "sku_count": len(catalog),
        "full_image_ready_count": sum(1 for row in catalog if row["image_prep_status"] == "IMAGE_READY"),
        "base_image_ready_count": sum(1 for row in catalog if row["upload_readiness"] != "HOLD_FOR_IMAGE_REVIEW"),
        "ready_for_review_count": sum(1 for row in catalog if row["upload_readiness"] == "READY_FOR_REVIEW"),
        "ready_base_images_detail_review_count": sum(1 for row in catalog if row["upload_readiness"] == "READY_BASE_IMAGES_DETAIL_REVIEW"),
        "hold_for_image_review_count": sum(1 for row in catalog if row["upload_readiness"] == "HOLD_FOR_IMAGE_REVIEW"),
        "ready_to_create_count": sum(1 for row in catalog if row["status_recommendation"] == "READY_TO_CREATE"),
        "manual_review_count": len(manual),
        "image_manual_review_count": sum(1 for row in catalog if row["upload_readiness"] != "READY_FOR_REVIEW"),
        "metafield_manual_review_count": sum(1 for row in catalog if row["metafield_review_needed"] == "yes"),
        "missing_white_count": sum(1 for row in catalog if row["missing_white_final"] == "yes"),
        "missing_main_count": sum(1 for row in catalog if row["missing_main_final"] == "yes"),
        "missing_detail_count": sum(1 for row in catalog if row["missing_detail_final"] == "yes"),
        "missing_piece_count": sum(1 for row in catalog if not row["specs_piece_count"]),
        "exact_brick4_count": sum(1 for row in catalog if row["brick4_exact_match"] == "yes"),
    }
    return catalog, manual, summary


def write_workbook(path: Path, catalog: list[dict[str, str]], manual: list[dict[str, str]], summary: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    warn_fill = PatternFill("solid", fgColor="FCE5CD")

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Metric", "Value"])
    for key, value in summary.items():
        summary_sheet.append([key, value])
    summary_sheet.append([])
    summary_sheet.append(["Remaining image gaps", "GK105 white; GK105/GK602/GK603/GK604/GK605 detail"])
    summary_sheet.append(["Piece count", "Brick4 did not provide pcs values for exact matches; piece_count remains a metafield review gap, not a base image upload blocker."])

    def add_table(sheet_name: str, rows: list[dict[str, str]]) -> None:
        sheet = workbook.create_sheet(sheet_name)
        if not rows:
            return
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        status_col = headers.index("image_prep_status") + 1 if "image_prep_status" in headers else None
        if status_col:
            for row_idx in range(2, sheet.max_row + 1):
                if sheet.cell(row_idx, status_col).value == "IMAGE_REVIEW":
                    sheet.cell(row_idx, status_col).fill = warn_fill
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 46)

    add_table("CatalogReady", catalog)
    add_table("ManualReview", manual)
    workbook.save(path)


def main() -> int:
    catalog, manual, summary = build_catalog_rows()
    catalog_csv = REPORT_DIR / "zoin-catalog-ready.csv"
    manual_csv = REPORT_DIR / "zoin-manual-review.csv"
    summary_json = REPORT_DIR / "zoin-catalog-ready-summary.json"
    workbook_path = REPORT_DIR / "zoin-catalog-ready.xlsx"
    write_csv(catalog_csv, catalog)
    write_csv(manual_csv, manual)
    write_workbook(workbook_path, catalog, manual, summary)
    summary.update(
        {
            "catalog_csv": catalog_csv.as_posix(),
            "manual_csv": manual_csv.as_posix(),
            "workbook_xlsx": workbook_path.as_posix(),
            "summary_json": summary_json.as_posix(),
        }
    )
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    target_reports = TARGET_ROOT / "reports"
    target_reports.mkdir(parents=True, exist_ok=True)
    for source in (catalog_csv, manual_csv, workbook_path, summary_json):
        shutil.copy2(source, target_reports / source.name)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
