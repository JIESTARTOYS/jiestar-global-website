#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import shopify_sample_import as base_import
from shopify_cn_pending_import import ShopifyAdmin as BaseShopifyAdmin


QUOTE_XLSX = Path("/Volumes/ORICO/GULY/锦童宝玩具厂积木报价表(1).xlsx")
IMAGE_ROOT = Path("/Volumes/ORICO/GULY/古励整理")
METADATA_XLSX = Path("/Volumes/ORICO/GULY/GULY产品元字段资料表.xlsx")
OUT_DIR = Path("/private/tmp/jiestar-shopify-guly-import")

VENDOR = "GULY"
STATUS = "ACTIVE"
PRICE = "999"
CATEGORY_ID = "gid://shopify/TaxonomyCategory/tg-5-7-12"
CATEGORY_NAME = "Interlocking Blocks"
OPTION_NAME = "Model"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
REQUIRED_SCOPES = {"read_products", "write_products", "read_files", "write_files"}
PUBLICATION_SCOPES = {"read_publications", "write_publications"}
DETAIL_SLICE_MAX_HEIGHT = 4000
DETAIL_SLICE_MIN_BYTES = 5 * 1024 * 1024
DETAIL_SLICE_QUALITY = 86
MAX_STAGED_UPLOAD_BYTES = 15 * 1024 * 1024
UPLOAD_IMAGE_QUALITY_STEPS = (86, 80, 74, 68)

SOURCE_FIELDS = [
    "sheet",
    "sku",
    "original_sku_cell",
    "original_name_cn",
    "factory_price",
    "motor_price",
    "carton_qty",
    "package_size",
    "outer_carton_size",
    "gross_net_weight",
    "piece_count",
    "barcode",
    "remarks",
]

HIGH_RISK_TITLE_PATTERNS = [
    r"lamborghini",
    r"ferrari",
    r"bugatti",
    r"mclaren",
    r"porsche",
    r"koenigsegg",
    r"mercedes",
    r"tesla",
    r"audi",
    r"ford",
    r"bmw",
    r"ducati",
    r"yamaha",
    r"harley",
    r"aston",
    r"apollo",
    r"star wars",
    r"disney",
    r"marvel",
]


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_header(value: Any) -> str:
    return clean(value).replace(" ", "").replace("\n", "")


def contains_cjk(value: str) -> bool:
    return bool(re.search(r"[\u3400-\u9fff]", value))


def same_money(left: Any, right: Any) -> bool:
    try:
        return Decimal(str(left)) == Decimal(str(right))
    except (InvalidOperation, ValueError):
        return False


def natural_key(value: str | Path) -> tuple[Any, ...]:
    text = str(value).lower()
    parts = re.split(r"(\d+)", text)
    return tuple(int(part) if part.isdigit() else part for part in parts)


def slugify(value: str) -> str:
    return base_import.slugify(value)


def sku_from_cell(value: Any) -> str:
    match = re.search(r"(?<!\d)(\d{5})(?!\d)", str(value or ""))
    return match.group(1) if match else ""


def is_discontinued(value: Any, row_values: tuple[Any, ...]) -> bool:
    text = " ".join(str(item or "") for item in row_values)
    return "停产" in text or "停售" in text or "停做" in text


def read_quote_rows() -> list[dict[str, str]]:
    workbook = load_workbook(QUOTE_XLSX, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    seen: set[str] = set()

    for sheet in workbook.worksheets:
        header_row: int | None = None
        headers: list[str] = []
        for index, cells in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            header_candidates = [normalize_header(cell) for cell in cells]
            if "货号" in header_candidates:
                header_row = index
                headers = header_candidates
                break

        if not header_row:
            continue

        sku_index = headers.index("货号")

        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if sku_index >= len(values):
                continue

            sku = sku_from_cell(values[sku_index])
            if not sku:
                continue

            key = sku.upper()
            if key in seen:
                continue
            seen.add(key)

            row_by_header = {
                header: clean(values[index]) if index < len(values) else ""
                for index, header in enumerate(headers)
                if header
            }
            rows.append(
                {
                    "sheet": sheet.title,
                    "sku": key,
                    "original_sku_cell": clean(values[sku_index]),
                    "original_name_cn": row_by_header.get("产品名称", ""),
                    "factory_price": row_by_header.get("出厂价", ""),
                    "motor_price": row_by_header.get("电机（实收）", ""),
                    "carton_qty": row_by_header.get("装箱量", ""),
                    "package_size": row_by_header.get("彩盒规格（CM)", "")
                    or row_by_header.get("彩盒规格（CM）", ""),
                    "outer_carton_size": row_by_header.get("外箱规格（CM）", "")
                    or row_by_header.get("外箱规格（CM)", ""),
                    "gross_net_weight": row_by_header.get("毛净重", ""),
                    "piece_count": row_by_header.get("颗粒数/pcs", "") or row_by_header.get("颗粒数/PCS", ""),
                    "barcode": row_by_header.get("条码", "") or row_by_header.get("产品条形码", ""),
                    "remarks": row_by_header.get("备注", ""),
                    "discontinued": "yes" if is_discontinued(values[sku_index], values) else "",
                }
            )

    return rows


def all_image_files() -> list[Path]:
    files: list[Path] = []
    for path in IMAGE_ROOT.rglob("*"):
        if not path.is_file():
            continue
        if path.name.startswith("._") or path.name.lower() in {"thumbs.db", ".ds_store"}:
            continue
        if any(part in {"__MACOSX", "源文件"} for part in path.parts):
            continue
        if path.suffix.lower() not in IMAGE_EXTS:
            continue
        files.append(path)
    return sorted(files, key=natural_key)


def image_files_for_sku(sku: str, image_files: list[Path]) -> list[Path]:
    pattern = re.compile(rf"(?<!\d){re.escape(sku)}(?!\d)")
    output = [path for path in image_files if pattern.search(str(path.relative_to(IMAGE_ROOT)))]
    return sorted(output, key=natural_key)


def media_for_sku(sku: str, image_files: list[Path]) -> dict[str, list[Path]]:
    files = image_files_for_sku(sku, image_files)

    def path_text(path: Path) -> str:
        return str(path.relative_to(IMAGE_ROOT))

    def has_part(path: Path, term: str) -> bool:
        return any(term in part for part in path.relative_to(IMAGE_ROOT).parts)

    details = [path for path in files if has_part(path, "详情") and not has_part(path, "白底")]
    details = sorted(details, key=natural_key)

    whites = [path for path in files if has_part(path, "白底") and "详情" not in path.name]
    whites = sorted(
        whites,
        key=lambda path: (
            0 if "尺寸" not in path_text(path) else 1,
            0 if path.suffix.lower() in {".jpg", ".jpeg"} else 1,
            natural_key(path),
        ),
    )

    size_images = [path for path in files if "尺寸" in path_text(path)]
    sku_images = sorted([path for path in files if "sku" in path.name.lower()], key=natural_key)

    main = [
        path
        for path in files
        if path not in details
        and path not in whites
        and path not in sku_images
        and path not in size_images
        and ("主图" in path_text(path) or "800主图" in path.name or re.search(r"主图[-_ ]*\d+", path.name))
    ]
    main = sorted(main, key=natural_key)

    fallback_sku = sku_images[:1] or whites[:1]
    return {
        "white": whites[:1],
        "main": main,
        "sku": fallback_sku,
        "detail": details,
        "size": sorted(size_images, key=natural_key),
        "all": files,
    }


def upload_ready_image_path(path: Path, sku: str, role: str) -> Path:
    if path.stat().st_size <= MAX_STAGED_UPLOAD_BYTES:
        return path

    try:
        from PIL import Image
    except ImportError:
        return path

    output_dir = OUT_DIR / "upload-images" / sku
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{path.stem}-{role}-upload.jpg"

    with Image.open(path) as image:
        if image.mode != "RGB":
            image = image.convert("RGB")

        width, height = image.size
        max_dimension = max(width, height)
        if max_dimension > 2400:
            scale = 2400 / max_dimension
            image = image.resize((round(width * scale), round(height * scale)))

        for quality in UPLOAD_IMAGE_QUALITY_STEPS:
            image.save(output_path, format="JPEG", quality=quality, optimize=True)
            if output_path.stat().st_size <= MAX_STAGED_UPLOAD_BYTES:
                return output_path

        width, height = image.size
        image = image.resize((round(width * 0.85), round(height * 0.85)))
        image.save(output_path, format="JPEG", quality=68, optimize=True)

    return output_path


def scale_from_name(name: str) -> str:
    match = re.search(r"1\s*[:：]\s*(\d+)", name)
    return f"1:{match.group(1)}" if match else ""


def product_type_and_series(name: str) -> tuple[str, str]:
    if "摩托" in name or "机车" in name:
        return "Technic", "Motorcycle"
    if "街景" in name or "咖啡" in name or "花店" in name or "面包屋" in name or "裁缝铺" in name or "汉堡店" in name:
        return "Buildings & Street Scenes", "Buildings & Street Scenes"
    if "招财熊猫" in name or "机械兔" in name or "机械鱼" in name:
        return "Ornament", "Display Model"
    if "机甲" in name or "金刚" in name or "战神" in name or "哪吒" in name or "齐天" in name:
        return "Mecha", "Mecha"
    if "航母" in name or ("战舰" in name and "星球" not in name):
        return "Warship", "Warship"
    if "星球" in name or "太空" in name or "战机" in name or "炮艇" in name:
        return "Space", "Space"
    if "工程" in name or "拖车" in name or "装载" in name or "推土" in name:
        return "Engineering", "Engineering"
    if "军事" in name or "武器" in name or "加特林" in name:
        return "Military", "Military"
    if "发动机" in name:
        return "Technic", "Engine Model"
    if "游戏机" in name or "弹珠" in name or "摇摇乐" in name:
        return "Arcade Game", "Game Machine"
    if "音乐盒" in name:
        if "圣诞" in name:
            return "Christmas", "Music Box"
        return "Ornament", "Music Box"
    if "花鞋" in name:
        return "Ornament", "Display Model"
    if "电视机" in name:
        return "Home Appliance", "Display Model"
    if "圣诞" in name:
        return "Holiday", "Holiday"
    return "Car Model", "Car Model"


def car_descriptor(name: str) -> str:
    if "飞屋" in name:
        return "Flying House Display"
    if "警车" in name:
        return "Police Sports Car"
    if "皮卡" in name:
        return "Electric Pickup Truck"
    if "SUV" in name or "卫士" in name:
        return "Off-Road SUV"
    if "霸王龙" in name:
        return "Armored SUV"
    if "特技" in name or "漂移" in name:
        return "High-Speed Stunt Car"
    if "工程" in name:
        return "Engineering Vehicle"
    if "方程" in name or "赛车" in name:
        return "Race Car"
    if "装甲" in name:
        return "Armored Supercar"
    if "星空" in name:
        return "Starry Supercar"
    if "暗夜" in name:
        return "Dark Hypercar"
    if "飞火流星" in name:
        return "Meteor Supercar"
    if "小牛" in name:
        return "Compact Supercar"
    if "三千年" in name or "三仟年" in name:
        return "Millennium Supercar"
    if "毒药" in name:
        return "Venom Supercar"
    if "one off" in name.lower():
        return "One-Off Supercar"
    if "GT3" in name:
        return "GT3 Race Car"
    if "RS6" in name:
        return "Sport Wagon"
    if "GTR" in name:
        return "Performance Coupe"
    if "963" in name or "勒茫" in name:
        return "Endurance Race Car"
    if "918" in name or "Mission X" in name:
        return "Concept Hypercar"
    if "Tourbillon" in name or "tourbillon" in name:
        return "Grand Touring Hypercar"
    if "SF90" in name or "F80" in name or "Revuelto" in name:
        return "Hybrid Supercar"
    if "W1" in name:
        return "Next-Gen Supercar"
    if "GT Mansory" in name:
        return "GT Supercar"
    if "挑战者" in name:
        return "Classic Muscle Car"
    if "科迩维特" in name:
        return "Muscle Sports Car"
    if "布加" in name or "科妮" in name or "阿波" in name or "啊波" in name or "啊菠" in name or "马丁" in name or "塞纳" in name:
        return "Track Hypercar"
    if "迈凯" in name:
        return "Track Supercar"
    if "兰博" in name or "蘭博" in name or "蓝博" in name:
        return "Angular Supercar"
    if "SP3" in name:
        return "Open-Cockpit Sports Car"
    return "Sports Car"


def motorcycle_descriptor(name: str) -> str:
    if "可变" in name or "变形" in name:
        return "Transforming Technic Motorcycle"
    if "大魔鬼" in name or "夜路德" in name:
        return "Power Cruiser Technic Motorcycle"
    if "哈镭" in name or "哈雷" in name:
        return "Cruiser Technic Motorcycle"
    if "R1300" in name or "泛美" in name or "Desertx" in name:
        return "Adventure Technic Motorcycle"
    if "G450" in name or "H2R" in name:
        return "Track Technic Motorcycle"
    if "幼兽" in name:
        return "Mini Retro Technic Motorcycle"
    if "Tracer" in name or "9GT" in name:
        return "Sport Touring Technic Motorcycle"
    if "XDiavel" in name:
        return "Power Cruiser Technic Motorcycle"
    if "杜卡" in name or "杜咔" in name or "铠旋" in name or "SR-X" in name:
        return "Performance Technic Motorcycle"
    return "Technic Motorcycle"


def mecha_descriptor(name: str) -> str:
    if "百兽" in name or "狮王" in name:
        return "Black Lion Mecha"
    if "超能" in name:
        return "Power Mecha"
    if "齐天暝" in name:
        return "Dark Monkey King Mecha"
    if "齐天" in name:
        return "Monkey King Mecha"
    if "赛博" in name:
        return "Cyber Hero Mecha"
    if "暗黑" in name:
        return "Dark Warrior Mecha"
    if "天焱" in name:
        return "Flame Warrior Mecha"
    return "Mecha"


def space_descriptor(name: str) -> str:
    if "太空猪" in name:
        return "Space Pig"
    if "共和国炮艇" in name:
        return "Republic Space Gunship"
    if "巡洋战舰" in name:
        return "Dark Cruiser Space Battleship"
    if "星河战舰" in name:
        return "Dark Galaxy Space Battleship"
    if "刀锋战舰" in name:
        return "Blade Space Battleship"
    if "拦截战机" in name:
        return "Interceptor Space Fighter"
    if "X战机" in name:
        return "X Space Fighter"
    if "S战斗机" in name:
        return "S Space Fighter"
    if "暗黑仟年" in name:
        return "Dark Millennium Space Model"
    if "战舰" in name:
        return "Space Battleship"
    if "战机" in name:
        return "Space Fighter"
    if "炮艇" in name:
        return "Space Gunship"
    return "Space"


def building_descriptor(name: str) -> str:
    if "汉堡店" in name:
        return "Cat Burger Shop"
    if "咖啡厅" in name:
        return "Pet Cafe"
    if "咖啡屋" in name:
        return "Cafe House"
    if "仙人掌屋" in name:
        return "Cactus House"
    if "章鱼果茶屋" in name:
        return "Octopus Fruit Tea Shop"
    if "云端花店" in name:
        return "Cloud Flower Shop"
    if "面包屋" in name:
        return "Bakery House"
    if "日葵花屋" in name:
        return "Sunflower House"
    if "裁缝铺" in name:
        return "Tailor Shop"
    if "梦屿花房" in name:
        return "Dream Island Flower House"
    return "Street View"


def ornament_descriptor(name: str) -> str:
    if "飞屋" in name:
        return "Flying House Music Box"
    if "圣诞惊喜" in name:
        return "Christmas Surprise Gift Box Music Box"
    if "朝花橙语" in name:
        return "Orange Blossom Floral Sneaker"
    if "蜜落花楹" in name:
        return "Jacaranda Floral Sneaker"
    if "招财熊猫" in name:
        return "Lucky Panda Display"
    if "机械兔" in name:
        return "Mechanical Rabbit Display"
    if "机械鱼" in name:
        return "Mechanical Fish Display"
    return "Display"


def arcade_descriptor(name: str) -> str:
    if "弹珠" in name:
        return "Pinball Table Arcade Game"
    if "水果" in name:
        return "Fruit Rocker Arcade Game"
    return "Arcade Game Machine"


def title_for_row(row: dict[str, str]) -> str:
    sku = row["sku"]
    name = row.get("original_name_cn", "")
    product_type, series = product_type_and_series(name)
    scale = scale_from_name(name)
    scale_prefix = f"{scale} " if scale else ""
    edition = "Electroplated " if "电镀" in name else ""

    if product_type == "Car Model":
        descriptor = car_descriptor(name)
        title = f"GULY {scale_prefix}{edition}{descriptor} Model Kit {sku}"
    elif series == "Motorcycle":
        descriptor = motorcycle_descriptor(name)
        title = f"GULY {scale_prefix}{edition}{descriptor} Model Kit {sku}"
    elif product_type == "Mecha":
        descriptor = mecha_descriptor(name)
        title = f"GULY {edition}{descriptor} Model Kit {sku}"
    elif product_type == "Warship":
        title = f"GULY Aircraft Carrier Display Model Kit {sku}" if "航母" in name else f"GULY Warship Display Model Kit {sku}"
    elif product_type == "Space":
        descriptor = space_descriptor(name)
        title = f"GULY {descriptor} Building Set {sku}"
    elif product_type == "Engineering":
        if "拖车" in name:
            descriptor = "Road Tow Truck"
        elif "装载" in name or "推土" in name:
            descriptor = "Wheel Loader Bulldozer"
        else:
            descriptor = "Engineering Vehicle"
        title = f"GULY {descriptor} Building Set {sku}"
    elif product_type == "Military":
        if "加特林" in name:
            title = f"GULY Rotary Cannon Building Set {sku}"
        elif "直升机" in name:
            title = f"GULY Military Helicopter Building Set {sku}"
        elif "战斗机" in name:
            title = f"GULY Stealth Fighter Building Set {sku}"
        else:
            title = f"GULY Military Building Set {sku}"
    elif series == "Engine Model":
        title = f"GULY V12 Engine Model Kit {sku}"
    elif series == "Game Machine":
        descriptor = arcade_descriptor(name)
        title = f"GULY {descriptor} Building Set {sku}"
    elif series == "Music Box":
        descriptor = ornament_descriptor(name)
        title = f"GULY {descriptor} Building Set {sku}"
    elif product_type == "Holiday":
        descriptor = ornament_descriptor(name)
        title = f"GULY {descriptor} Building Set {sku}"
    elif "花鞋" in name:
        descriptor = ornament_descriptor(name)
        title = f"GULY {descriptor} Building Set {sku}"
    elif "电视机" in name:
        title = f"GULY Retro TV Display Building Set {sku}"
    elif product_type == "Buildings & Street Scenes":
        descriptor = building_descriptor(name)
        title = f"GULY {descriptor} Building Set {sku}"
    elif product_type == "Ornament":
        descriptor = ornament_descriptor(name)
        title = f"GULY {descriptor} Building Set {sku}"
    else:
        title = f"GULY Building Block Set {sku}"

    title = re.sub(r"\s+", " ", title).strip()
    for pattern in HIGH_RISK_TITLE_PATTERNS:
        title = re.sub(pattern, "", title, flags=re.I)
    title = re.sub(r"\s+", " ", title).strip()
    return title


def metafields_for_row(row: dict[str, str], product_type: str, series: str) -> dict[str, str]:
    piece_count = clean(row.get("piece_count", ""))
    if piece_count and not piece_count.isdigit():
        match = re.search(r"\d+", piece_count)
        piece_count = match.group(0) if match else ""

    metafields = {
        "specs.piece_count": piece_count,
        "specs.package_size": clean(row.get("package_size", "")),
        "specs.difficulty_level": "See product package",
        "custom.series": series or product_type,
    }
    return {key: value for key, value in metafields.items() if value}


def manifest_item(row: dict[str, str], image_files: list[Path]) -> tuple[dict[str, Any], list[str]]:
    sku = row["sku"]
    media = media_for_sku(sku, image_files)
    title = title_for_row(row)
    handle = slugify(title)
    product_type, series = product_type_and_series(row.get("original_name_cn", ""))
    issues: list[str] = []

    if row.get("discontinued"):
        issues.append("discontinued")
    if contains_cjk(title):
        issues.append("title_contains_chinese")
    if any(re.search(pattern, title, re.I) for pattern in HIGH_RISK_TITLE_PATTERNS):
        issues.append("title_high_risk_brand_term")
    if not media["white"]:
        issues.append("missing_white_image")
    if not media["main"]:
        issues.append("missing_main_image")
    if not media["detail"]:
        issues.append("missing_detail_image")
    if VENDOR != "GULY":
        issues.append("vendor_mismatch")
    if PRICE != "999":
        issues.append("price_not_999")
    if STATUS != "ACTIVE":
        issues.append("status_not_active")

    main_media = media["white"] + media["main"]
    upload_main_media = [
        str(upload_ready_image_path(path, sku, f"main-{index:02d}"))
        for index, path in enumerate(main_media, start=1)
    ]
    upload_sku_images = [
        str(upload_ready_image_path(path, sku, f"sku-{index:02d}"))
        for index, path in enumerate(media["sku"], start=1)
    ]
    item = {
        "folder": sku,
        "folder_path": str(IMAGE_ROOT),
        "base": sku,
        "handle": handle,
        "title": title,
        "vendor": VENDOR,
        "status": STATUS,
        "product_type": product_type,
        "category": CATEGORY_ID,
        "category_name": CATEGORY_NAME,
        "price": PRICE,
        "option_name": OPTION_NAME,
        "variants": [
            {
                "sku": sku,
                "option_name": f"{sku} - {title.removeprefix(VENDOR).strip()}",
                "title_source": clean(row.get("original_name_cn")),
                "series": series,
                "age": "",
                "piece_count": clean(row.get("piece_count")),
                "package_size": clean(row.get("package_size")),
                "finished_size": "",
            }
        ],
        "metafields": metafields_for_row(row, product_type, series),
        "main_media": upload_main_media,
        "sku_images": upload_sku_images,
        "detail_images": [str(path) for path in media["detail"]],
        "transparent_images": [],
        "source_row": row,
        "media_status": {
            "all_count": len(media["all"]),
            "white_count": len(media["white"]),
            "main_count": len(media["main"]),
            "sku_count": len(media["sku"]),
            "detail_count": len(media["detail"]),
            "size_count": len(media["size"]),
            "first_media": Path(main_media[0]).name if main_media else "",
        },
        "missing": {
            "white": not bool(media["white"]),
            "main": not bool(media["main"]),
            "detail": not bool(media["detail"]),
            "sku_image_fallback_to_white": not any("sku" in Path(path).name.lower() for path in media["sku"]),
        },
    }
    return item, issues


def build_manifest() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, str]]]:
    rows = read_quote_rows()
    image_files = all_image_files()
    manifest: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    rows_by_sku = {row["sku"]: row for row in rows}

    for row in rows:
        item, issues = manifest_item(row, image_files)
        has_any_asset = bool(item["media_status"]["all_count"])
        if issues:
            skipped.append(
                {
                    "sku": row["sku"],
                    "handle": item["handle"],
                    "title": item["title"],
                    "sheet": row["sheet"],
                    "original_name_cn": row["original_name_cn"],
                    "issues": issues,
                    "media_status": item["media_status"],
                    "has_any_asset": has_any_asset,
                }
            )
        else:
            manifest.append(item)

    asset_skus: set[str] = set()
    for path in image_files:
        text = str(path.relative_to(IMAGE_ROOT))
        for match in re.finditer(r"(?<!\d)(\d{5})(?!\d)", text):
            asset_skus.add(match.group(1))

    for sku in sorted(asset_skus - set(rows_by_sku), key=natural_key):
        skipped.append(
            {
                "sku": sku,
                "handle": "",
                "title": "",
                "sheet": "",
                "original_name_cn": "",
                "issues": ["asset_without_quote_row"],
                "media_status": {"all_count": len(image_files_for_sku(sku, image_files))},
                "has_any_asset": True,
            }
        )

    return manifest, skipped, rows


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def source_row_for_skipped(row: dict[str, Any], quote_rows: list[dict[str, str]]) -> dict[str, str]:
    return next((source for source in quote_rows if source["sku"] == row.get("sku")), {})


def metadata_row(item: dict[str, Any], upload_status: str, issues: str, source: dict[str, str]) -> list[Any]:
    metafields = item.get("metafields", {})
    media_status = item.get("media_status", {})
    return [
        item.get("base", ""),
        item.get("vendor", ""),
        item.get("title", ""),
        item.get("handle", ""),
        item.get("status", ""),
        item.get("price", ""),
        item.get("product_type", ""),
        item.get("category_name", ""),
        metafields.get("specs.piece_count", ""),
        metafields.get("specs.finished_model_size", ""),
        metafields.get("specs.package_size", ""),
        metafields.get("specs.difficulty_level", ""),
        metafields.get("custom.series", ""),
        source.get("sheet", ""),
        source.get("original_sku_cell", ""),
        source.get("original_name_cn", ""),
        source.get("factory_price", ""),
        source.get("motor_price", ""),
        source.get("carton_qty", ""),
        source.get("outer_carton_size", ""),
        source.get("gross_net_weight", ""),
        source.get("barcode", ""),
        source.get("remarks", ""),
        json.dumps(media_status, ensure_ascii=False),
        upload_status,
        issues,
    ]


def write_metadata_workbook(manifest: list[dict[str, Any]], skipped: list[dict[str, Any]], quote_rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GULY Metafields"
    headers = [
        "sku",
        "vendor",
        "shopify_title",
        "handle",
        "status",
        "price",
        "product_type",
        "category",
        "specs.piece_count",
        "specs.finished_model_size",
        "specs.package_size",
        "specs.difficulty_level",
        "custom.series",
        "source_sheet",
        "original_sku_cell",
        "original_name_cn",
        "factory_price",
        "motor_price",
        "carton_qty",
        "outer_carton_size",
        "gross_net_weight",
        "barcode",
        "remarks",
        "media_status",
        "upload_status",
        "issues",
    ]
    sheet.append(headers)

    for item in manifest:
        sheet.append(metadata_row(item, "READY", "", item.get("source_row", {})))

    for row in skipped:
        source = source_row_for_skipped(row, quote_rows)
        if source:
            product_type, series = product_type_and_series(source.get("original_name_cn", ""))
            pseudo = {
                "base": row["sku"],
                "vendor": VENDOR,
                "title": row["title"],
                "handle": row["handle"],
                "status": STATUS,
                "price": PRICE,
                "product_type": product_type,
                "category_name": CATEGORY_NAME,
                "metafields": metafields_for_row(source, product_type, series),
                "media_status": row.get("media_status", {}),
            }
        else:
            pseudo = {
                "base": row.get("sku", ""),
                "vendor": VENDOR,
                "title": row.get("title", ""),
                "handle": row.get("handle", ""),
                "status": STATUS,
                "price": PRICE,
                "product_type": "",
                "category_name": CATEGORY_NAME,
                "metafields": {},
                "media_status": row.get("media_status", {}),
            }
        sheet.append(metadata_row(pseudo, "SKIPPED", "; ".join(row.get("issues", [])), source))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_fill = {
        "READY": PatternFill("solid", fgColor="E2F0D9"),
        "SKIPPED": PatternFill("solid", fgColor="FCE4D6"),
    }
    status_col = headers.index("upload_status") + 1
    for row in sheet.iter_rows(min_row=2):
        fill = status_fill.get(str(row[status_col - 1].value or ""))
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    widths = {
        "A": 12,
        "B": 12,
        "C": 48,
        "D": 48,
        "E": 12,
        "F": 10,
        "G": 22,
        "H": 24,
        "I": 16,
        "J": 22,
        "K": 18,
        "L": 24,
        "M": 22,
        "N": 16,
        "O": 18,
        "P": 36,
        "Q": 12,
        "R": 14,
        "S": 12,
        "T": 20,
        "U": 18,
        "V": 18,
        "W": 24,
        "X": 42,
        "Y": 14,
        "Z": 34,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

    METADATA_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(METADATA_XLSX)


def write_reports(manifest: list[dict[str, Any]], skipped: list[dict[str, Any]], quote_rows: list[dict[str, str]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    write_json(OUT_DIR / "guly-manifest.json", manifest)
    write_json(OUT_DIR / "guly-skipped.json", skipped)
    write_metadata_workbook(manifest, skipped, quote_rows)
    write_csv(
        OUT_DIR / "guly-manifest.csv",
        [
            {
                "sku": item["base"],
                "handle": item["handle"],
                "title": item["title"],
                "vendor": item["vendor"],
                "status": item["status"],
                "price": item["price"],
                "product_type": item["product_type"],
                "main_media_count": len(item["main_media"]),
                "detail_count": len(item["detail_images"]),
                "sku_image_count": len(item["sku_images"]),
                "first_media": Path(item["main_media"][0]).name if item["main_media"] else "",
                "source_sheet": item.get("source_row", {}).get("sheet", ""),
                "original_name_cn": item.get("source_row", {}).get("original_name_cn", ""),
            }
            for item in manifest
        ],
        [
            "sku",
            "handle",
            "title",
            "vendor",
            "status",
            "price",
            "product_type",
            "main_media_count",
            "detail_count",
            "sku_image_count",
            "first_media",
            "source_sheet",
            "original_name_cn",
        ],
    )
    write_csv(
        OUT_DIR / "guly-skipped.csv",
        [
            {
                "sku": row["sku"],
                "handle": row.get("handle", ""),
                "title": row.get("title", ""),
                "sheet": row.get("sheet", ""),
                "original_name_cn": row.get("original_name_cn", ""),
                "issues": "; ".join(row.get("issues", [])),
                "has_any_asset": row.get("has_any_asset", ""),
                "media_status": json.dumps(row.get("media_status", {}), ensure_ascii=False),
            }
            for row in skipped
        ],
        ["sku", "handle", "title", "sheet", "original_name_cn", "issues", "has_any_asset", "media_status"],
    )


class ShopifyAdmin(BaseShopifyAdmin):
    def product_set(self, item: dict[str, Any], description_html: str) -> dict[str, Any]:
        variants = [
            {
                "optionValues": [
                    {
                        "optionName": item.get("option_name") or OPTION_NAME,
                        "name": variant["option_name"],
                    }
                ],
                "price": item["price"],
                "inventoryItem": {
                    "sku": variant["sku"],
                    "tracked": False,
                },
            }
            for variant in item["variants"]
        ]
        data = self.graphql(
            """
            mutation ProductSet($input: ProductSetInput!, $synchronous: Boolean!) {
              productSet(input: $input, synchronous: $synchronous) {
                product {
                  id
                  title
                  handle
                  status
                  vendor
                  productType
                  variants(first: 50) {
                    nodes {
                      id
                      title
                      price
                      sku
                      inventoryItem {
                        id
                        tracked
                      }
                    }
                  }
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {
                "synchronous": True,
                "input": {
                    "title": item["title"],
                    "handle": item["handle"],
                    "vendor": item["vendor"],
                    "status": item["status"],
                    "productType": item["product_type"],
                    "descriptionHtml": description_html,
                    "productOptions": [
                        {
                            "name": item.get("option_name") or OPTION_NAME,
                            "values": [{"name": variant["option_name"]} for variant in item["variants"]],
                        }
                    ],
                    "variants": variants,
                    "metafields": base_import.product_metafields(item),
                },
            },
        )
        result = data["productSet"]
        base_import.assert_no_user_errors("productSet", result["userErrors"])
        return result["product"]

    def products_by_sku(self, skus: set[str]) -> list[dict[str, Any]]:
        products: list[dict[str, Any]] = []
        cursor: str | None = None
        query = " OR ".join(sorted(skus))

        while True:
            data = self.graphql(
                """
                query ProductsBySku($cursor: String, $query: String!) {
                  products(first: 100, after: $cursor, query: $query) {
                    pageInfo {
                      hasNextPage
                      endCursor
                    }
                    nodes {
                      id
                      title
                      handle
                      status
                      vendor
                      productType
                      descriptionHtml
                      category {
                        id
                        name
                      }
                      media(first: 250, sortKey: POSITION) {
                        nodes {
                          id
                          alt
                          ... on MediaImage {
                            image {
                              url
                            }
                          }
                        }
                      }
                      metafields(first: 50) {
                        nodes {
                          namespace
                          key
                          value
                        }
                      }
                      variants(first: 50) {
                        nodes {
                          id
                          title
                          price
                          sku
                          image {
                            id
                            url
                            altText
                          }
                          media(first: 10) {
                            nodes {
                              id
                              alt
                            }
                          }
                          inventoryItem {
                            tracked
                          }
                        }
                      }
                    }
                  }
                }
                """,
                {"cursor": cursor, "query": query},
            )
            page = data["products"]
            for product in page["nodes"]:
                product_skus = {(variant.get("sku") or "").strip().upper() for variant in product["variants"]["nodes"]}
                if product_skus & skus:
                    products.append(product)
            if not page["pageInfo"]["hasNextPage"]:
                break
            cursor = page["pageInfo"]["endCursor"]
        return products


def detail_image_paths(path: Path, sku: str) -> list[Path]:
    if not path.exists():
        return []

    try:
        from PIL import Image
    except ImportError:
        return [path]

    with Image.open(path) as image:
        width, height = image.size

        if height <= DETAIL_SLICE_MAX_HEIGHT and path.stat().st_size <= DETAIL_SLICE_MIN_BYTES:
            return [path]

        output_dir = OUT_DIR / "detail-slices" / sku / path.stem
        output_dir.mkdir(parents=True, exist_ok=True)
        output_paths: list[Path] = []
        part_count = max(1, (height + DETAIL_SLICE_MAX_HEIGHT - 1) // DETAIL_SLICE_MAX_HEIGHT)

        for index in range(part_count):
            top = round(index * height / part_count)
            bottom = round((index + 1) * height / part_count)
            output_path = output_dir / f"{path.stem}-part-{index + 1:02d}.jpg"
            crop = image.crop((0, top, width, bottom))

            if crop.mode != "RGB":
                crop = crop.convert("RGB")

            crop.save(output_path, format="JPEG", quality=DETAIL_SLICE_QUALITY, optimize=True)
            output_paths.append(output_path)

    return output_paths


def upload_detail_images_for_item(admin: ShopifyAdmin, item: dict[str, Any]) -> list[str]:
    urls: list[str] = []
    for detail_index, source in enumerate(item["detail_images"], start=1):
        path = Path(source)
        for part_index, upload_path in enumerate(detail_image_paths(path, item["base"]), start=1):
            part = f" part {part_index}" if part_index > 1 else ""
            urls.append(admin.file_create(upload_path, f"{item['title']} details {detail_index}{part}"))
    return urls


def description_html(item: dict[str, Any], detail_urls: list[str]) -> str:
    return "\n".join(
        f'<p><img src="{html.escape(url)}" alt="{html.escape(item["title"])} details part {index}" /></p>'
        for index, url in enumerate(detail_urls, start=1)
    )


def filter_existing(admin: ShopifyAdmin, manifest: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    existing_handles, existing_skus = admin.products_index()
    todo: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for item in manifest:
        item_skus = {variant["sku"].upper() for variant in item["variants"]}
        matched_skus = sorted(item_skus & existing_skus)
        if item["handle"] in existing_handles:
            skipped.append({"sku": item["base"], "handle": item["handle"], "title": item["title"], "issues": ["existing_handle"]})
        elif matched_skus:
            skipped.append(
                {
                    "sku": item["base"],
                    "handle": item["handle"],
                    "title": item["title"],
                    "issues": [f"existing_sku:{', '.join(matched_skus)}"],
                }
            )
        else:
            todo.append(item)

    return todo, skipped


def check_shopify_conflicts(manifest: list[dict[str, Any]]) -> list[dict[str, Any]]:
    admin = ShopifyAdmin()
    existing_handles, existing_skus = admin.products_index()
    conflicts: list[dict[str, Any]] = []

    for item in manifest:
        item_skus = {variant["sku"].upper() for variant in item["variants"]}
        matched_skus = sorted(item_skus & existing_skus)
        issues: list[str] = []
        if item["handle"] in existing_handles:
            issues.append("existing_handle")
        if matched_skus:
            issues.append(f"existing_sku:{', '.join(matched_skus)}")
        if issues:
            conflicts.append(
                {
                    "sku": item["base"],
                    "handle": item["handle"],
                    "title": item["title"],
                    "issues": issues,
                }
            )

    write_json(OUT_DIR / "guly-shopify-conflicts.json", conflicts)
    write_csv(
        OUT_DIR / "guly-shopify-conflicts.csv",
        [
            {
                "sku": row["sku"],
                "handle": row["handle"],
                "title": row["title"],
                "issues": "; ".join(row["issues"]),
            }
            for row in conflicts
        ],
        ["sku", "handle", "title", "issues"],
    )
    return conflicts


def apply_create(manifest: list[dict[str, Any]], report_name: str = "guly-created-products") -> list[dict[str, Any]]:
    admin = ShopifyAdmin()
    scopes = admin.access_scopes()
    missing_scopes = sorted(REQUIRED_SCOPES - scopes)
    if missing_scopes:
        raise RuntimeError(f"Missing Shopify scopes: {', '.join(missing_scopes)}")

    publication_ids: list[str] = []
    if PUBLICATION_SCOPES <= scopes:
        publication_ids = [publication["id"] for publication in admin.publications()]

    todo, existing_skipped = filter_existing(admin, manifest)
    results: list[dict[str, Any]] = [{"manifest": row, "ok": False, "skipped": True, "error": "; ".join(row["issues"])} for row in existing_skipped]
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for index, item in enumerate(todo, start=1):
        result: dict[str, Any] = {"manifest": item, "ok": False}
        try:
            detail_urls = upload_detail_images_for_item(admin, item)
            product = admin.product_set(item, description_html(item, detail_urls))
            product_id = product["id"]
            base_import.sync_product_assets(admin, item, product_id, update_description=False)
            admin.update_status_and_category(product_id)
            admin.publish_to_publications(product_id, publication_ids)
            result.update({"ok": True, "product": admin.fetch_product(product_id)})
        except Exception as error:  # noqa: BLE001 - batch should continue and report every SKU.
            result["error"] = str(error)

        results.append(result)
        write_json(OUT_DIR / f"{report_name}.json", results)
        print(f"Processed {index}/{len(todo)} {item['base']} ok={result['ok']}", flush=True)
        time.sleep(0.5)

    return results


def verify_products(manifest: list[dict[str, Any]]) -> list[dict[str, Any]]:
    admin = ShopifyAdmin()
    expected_by_sku = {item["base"]: item for item in manifest}
    products = admin.products_by_sku(set(expected_by_sku))
    rows: list[dict[str, Any]] = []

    for sku, item in expected_by_sku.items():
        product = next(
            (
                candidate
                for candidate in products
                if any((variant.get("sku") or "").strip().upper() == sku for variant in candidate["variants"]["nodes"])
            ),
            None,
        )
        if not product:
            rows.append({"sku": sku, "ok": False, "issues": ["missing_product"]})
            continue

        variant = next(
            variant for variant in product["variants"]["nodes"] if (variant.get("sku") or "").strip().upper() == sku
        )
        metafields = {f"{node['namespace']}.{node['key']}": node["value"] for node in product["metafields"]["nodes"]}
        first_alt = (product["media"]["nodes"][0].get("alt") or "") if product["media"]["nodes"] else ""
        issues = []
        checks = {
            "title": product["title"] == item["title"],
            "vendor": product["vendor"] == VENDOR,
            "status": product["status"] == STATUS,
            "price": same_money(variant.get("price"), PRICE),
            "product_type": product["productType"] == item["product_type"],
            "category": (product.get("category") or {}).get("id") == CATEGORY_ID,
            "description_has_images": "<img" in (product.get("descriptionHtml") or ""),
            "first_media_white": "白底" in first_alt,
            "variant_has_media": bool((variant.get("media") or {}).get("nodes") or variant.get("image")),
        }
        checks["metafields"] = all(str(metafields.get(key, "")) == str(value) for key, value in item["metafields"].items())
        for key, ok in checks.items():
            if not ok:
                issues.append(key)
        rows.append(
            {
                "sku": sku,
                "product_id": product["id"],
                "handle": product["handle"],
                "title": product["title"],
                "ok": not issues,
                "issues": issues,
                "checks": checks,
                "media_count": len(product["media"]["nodes"]),
                "detail_image_count": len(item["detail_images"]),
            }
        )

    write_json(OUT_DIR / "guly-verify.json", rows)
    write_csv(
        OUT_DIR / "guly-verify.csv",
        [
            {
                "sku": row["sku"],
                "title": row.get("title", ""),
                "handle": row.get("handle", ""),
                "ok": row["ok"],
                "issues": "; ".join(row.get("issues", [])),
                "media_count": row.get("media_count", ""),
                "detail_image_count": row.get("detail_image_count", ""),
            }
            for row in rows
        ],
        ["sku", "title", "handle", "ok", "issues", "media_count", "detail_image_count"],
    )
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare and upload GULY products to Shopify.")
    parser.add_argument("--dry-run", action="store_true", help="Generate manifest, skipped report, and metadata workbook only.")
    parser.add_argument("--check-shopify", action="store_true", help="Read-only Shopify SKU/handle conflict check.")
    parser.add_argument("--apply", action="store_true", help="Create ready GULY products in Shopify.")
    parser.add_argument("--verify", action="store_true", help="Verify created GULY products in Shopify.")
    parser.add_argument(
        "--include-skipped-exit-zero",
        action="store_true",
        help="Return 0 even if skipped rows exist. Useful when missing assets are expected.",
    )
    args = parser.parse_args()

    if not (args.dry_run or args.check_shopify or args.apply or args.verify):
        parser.error("Choose at least one of --dry-run, --check-shopify, --apply, or --verify")

    manifest, skipped, quote_rows = build_manifest()
    write_reports(manifest, skipped, quote_rows)

    skipped_with_assets = [
        row
        for row in skipped
        if row.get("has_any_asset") and "asset_without_quote_row" not in row.get("issues", [])
    ]
    skipped_without_assets = [
        row
        for row in skipped
        if not row.get("has_any_asset") and "asset_without_quote_row" not in row.get("issues", [])
    ]
    asset_without_quote = [row for row in skipped if "asset_without_quote_row" in row.get("issues", [])]

    summary: dict[str, Any] = {
        "quote_xlsx": str(QUOTE_XLSX),
        "image_root": str(IMAGE_ROOT),
        "metadata_xlsx": str(METADATA_XLSX),
        "quote_rows": len(quote_rows),
        "manifest_count": len(manifest),
        "skipped_count": len(skipped),
        "skipped_with_assets": len(skipped_with_assets),
        "skipped_without_assets": len(skipped_without_assets),
        "asset_without_quote_count": len(asset_without_quote),
        "out_dir": str(OUT_DIR),
    }

    if args.apply:
        results = apply_create(manifest)
        summary["created_ok"] = sum(1 for result in results if result.get("ok") and not result.get("skipped"))
        summary["created_failed"] = sum(1 for result in results if not result.get("ok") and not result.get("skipped"))
        summary["existing_skipped"] = sum(1 for result in results if result.get("skipped"))

    if args.check_shopify:
        conflicts = check_shopify_conflicts(manifest)
        summary["shopify_conflicts"] = len(conflicts)

    if args.verify:
        verify_rows = verify_products(manifest)
        summary["verify_ok"] = sum(1 for row in verify_rows if row.get("ok"))
        summary["verify_failed"] = sum(1 for row in verify_rows if not row.get("ok"))

    write_json(OUT_DIR / "guly-summary.json", summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if skipped and not args.include_skipped_exit_zero:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
