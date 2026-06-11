#!/usr/bin/env python3
"""
JIESTAR 上架产品「素材缺失 + 详情图重复」审查。

覆盖范围（仅 status:active 产品）：
  缺失检查
    - 主图缺失（无 IMAGE 媒体 / 无 featuredMedia）
    - 描述详情图缺失（descriptionHtml 无可用 <img src>）
    - 关键 specs 元字段缺失（piece_count / recommended_age / finished_model_size / package_size）
    - 变体 SKU 图缺失（多变体产品才算问题）
    - 描述里存在空 <img>（无 src）标签 —— 上传残留 / 损坏
  重复检查（重点：描述中的详情图）
    - 同一产品内详情图重复（同一图被插入多次）
    - 跨产品复用同一详情图（极可能是贴错素材）
    - 详情图与该产品主图重复
  可选：--hash 下载图片算 MD5，识别「换了文件名但内容相同」的重复（更慢，更严格）

输出：
    <out>/detail-image-audit.csv         每个产品一行的合并报告（可直接用 Excel 打开）
    <out>/detail-image-duplicates.csv     每个跨产品重复图片一组的明细
    <out>/detail-image-audit.xlsx         若本机装了 openpyxl，则额外产出多工作表 Excel
    <out>/detail-image-audit-summary.json 汇总

用法：
    python3 scripts/shopify_detail_image_audit.py
    python3 scripts/shopify_detail_image_audit.py --hash            # 额外做内容哈希查重
    python3 scripts/shopify_detail_image_audit.py --out ./reports   # 自定义输出目录
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

# 复用已验证的 Admin API 认证 + 重试逻辑
from shopify_active_product_health_audit import ShopifyAdmin

REQUIRED_SPECS_METAFIELDS = [
    "piece_count",
    "recommended_age",
    "finished_model_size",
    "package_size",
]

IMG_SRC_RE = re.compile(r"<img\b[^>]*?\bsrc\s*=\s*[\"']([^\"']+)[\"']", re.I)
IMG_TAG_RE = re.compile(r"<img\b[^>]*>", re.I)


def normalize_url(url: str) -> str:
    """去掉 ?v=... 版本参数，得到稳定的文件标识。"""
    url = (url or "").strip()
    return url.split("?", 1)[0]


def basename(url: str) -> str:
    return normalize_url(url).rsplit("/", 1)[-1]


def product_admin_url(domain: str, product_id: str) -> str:
    numeric_id = product_id.rsplit("/", 1)[-1]
    return f"https://admin.shopify.com/store/{domain.split('.')[0]}/products/{numeric_id}"


def fetch_active_products(admin: ShopifyAdmin, vendor: str | None = None) -> list[dict[str, Any]]:
    search = "status:active"
    if vendor:
        search += f' vendor:{vendor}'
    query = (
        """
    query DetailImageAudit($cursor: String) {
      products(first: 250, after: $cursor, query: "%s", sortKey: TITLE) {""" % search
    ) + """
        pageInfo { hasNextPage endCursor }
        nodes {
          id
          handle
          title
          descriptionHtml
          featuredMedia { mediaContentType }
          media(first: 100) {
            nodes {
              mediaContentType
              ... on MediaImage { image { url } }
            }
          }
          variants(first: 100) {
            nodes { sku image { url } }
          }
          metafields(first: 100) { nodes { namespace key value } }
        }
      }
    }
    """
    products: list[dict[str, Any]] = []
    cursor = None
    while True:
        data = admin.graphql(query, {"cursor": cursor})
        page = data["products"]
        products.extend(page["nodes"])
        if not page["pageInfo"]["hasNextPage"]:
            break
        cursor = page["pageInfo"]["endCursor"]
    return products


def detail_image_urls(description_html: str) -> tuple[list[str], int]:
    """返回 (按出现顺序的详情图 URL 列表, 空 <img> 标签数量)。"""
    urls = [m.strip() for m in IMG_SRC_RE.findall(description_html or "")]
    all_tags = IMG_TAG_RE.findall(description_html or "")
    empty_tags = sum(1 for tag in all_tags if not re.search(r"\bsrc\s*=", tag, re.I))
    return urls, empty_tags


def main_image_urls(product: dict[str, Any]) -> list[str]:
    out = []
    for node in product.get("media", {}).get("nodes", []):
        if node.get("mediaContentType") == "IMAGE":
            url = (node.get("image") or {}).get("url")
            if url:
                out.append(url)
    return out


def specs_metafields(product: dict[str, Any]) -> dict[str, str]:
    fields: dict[str, str] = {}
    for item in product.get("metafields", {}).get("nodes", []):
        if (item.get("namespace") or "") == "specs":
            fields[item.get("key") or ""] = item.get("value") or ""
    return fields


def md5_of_url(url: str, cache: dict[str, str]) -> str | None:
    key = normalize_url(url)
    if key in cache:
        return cache[key]
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "jiestar-audit/1.0"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            digest = hashlib.md5(resp.read()).hexdigest()
        cache[key] = digest
        return digest
    except Exception:
        cache[key] = ""
        return None


def audit(products: list[dict[str, Any]], domain: str, use_hash: bool) -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, Any]]:
    # 第一遍：抽取每个产品的详情图，建立跨产品索引
    per_product: list[dict[str, Any]] = []
    url_to_handles: dict[str, set[str]] = defaultdict(set)
    hash_cache: dict[str, str] = {}
    hash_to_urls: dict[str, set[str]] = defaultdict(set)

    for product in products:
        handle = product.get("handle") or ""
        detail_urls, empty_tags = detail_image_urls(product.get("descriptionHtml") or "")
        norm_detail = [normalize_url(u) for u in detail_urls]
        for u in set(norm_detail):
            url_to_handles[u].add(handle)
        if use_hash:
            for u in set(detail_urls):
                digest = md5_of_url(u, hash_cache)
                if digest:
                    hash_to_urls[digest].add(normalize_url(u))
        per_product.append(
            {
                "product": product,
                "handle": handle,
                "detail_urls_raw": detail_urls,
                "detail_urls": norm_detail,
                "empty_tags": empty_tags,
            }
        )

    # 跨产品复用：同一详情图 URL 被多于 1 个产品引用
    shared_url_to_handles = {u: hs for u, hs in url_to_handles.items() if len(hs) > 1}

    # 内容哈希复用（换名同图）
    shared_hash_groups = {h: us for h, us in hash_to_urls.items() if len({normalize_url(x) for x in us}) > 1}
    url_to_dup_hash_partners: dict[str, set[str]] = defaultdict(set)
    if use_hash:
        for _h, us in shared_hash_groups.items():
            for u in us:
                url_to_dup_hash_partners[u] |= (us - {u})

    rows: list[dict[str, str]] = []
    dup_detail_rows: list[dict[str, str]] = []

    for entry in per_product:
        product = entry["product"]
        handle = entry["handle"]
        title = product.get("title") or ""
        skus = [(v.get("sku") or "").strip() for v in product.get("variants", {}).get("nodes", []) if (v.get("sku") or "").strip()]
        variants = product.get("variants", {}).get("nodes", [])
        admin_url = product_admin_url(domain, product.get("id") or "")

        main_urls = main_image_urls(product)
        main_norm = {normalize_url(u) for u in main_urls}
        detail_norm = entry["detail_urls"]
        empty_tags = entry["empty_tags"]

        specs = specs_metafields(product)
        missing_mf = [f"specs.{k}" for k in REQUIRED_SPECS_METAFIELDS if not specs.get(k)]

        critical: list[str] = []
        warning: list[str] = []

        # --- 缺失检查 ---
        if not main_urls:
            critical.append("missing_main_image")
        if not product.get("featuredMedia"):
            critical.append("missing_featured_media")
        if not detail_norm:
            critical.append("missing_detail_images")
        if empty_tags:
            warning.append(f"empty_img_tags({empty_tags})")
        if missing_mf:
            warning.append("missing_metafields(" + ",".join(m.replace("specs.", "") for m in missing_mf) + ")")
        if len(variants) > 1:
            no_img = [(v.get("sku") or "").strip() or "?" for v in variants if not (v.get("image") or {}).get("url")]
            if no_img:
                warning.append("variant_missing_sku_image(" + ",".join(no_img) + ")")

        # --- 重复检查（详情图） ---
        # 1) 产品内重复
        counts = Counter(detail_norm)
        intra_dups = [u for u, c in counts.items() if c > 1]
        if intra_dups:
            warning.append(f"detail_dup_in_product({len(intra_dups)})")
            for u in intra_dups:
                dup_detail_rows.append({
                    "type": "intra_product",
                    "image": basename(u),
                    "image_url": u,
                    "handle": handle,
                    "title": title,
                    "skus": "|".join(skus),
                    "detail": f"在本产品描述中出现 {counts[u]} 次",
                    "admin_url": admin_url,
                })

        # 2) 跨产品复用
        cross_partners: set[str] = set()
        for u in set(detail_norm):
            others = shared_url_to_handles.get(u, set()) - {handle}
            if others:
                cross_partners |= others
                dup_detail_rows.append({
                    "type": "cross_product_same_url",
                    "image": basename(u),
                    "image_url": u,
                    "handle": handle,
                    "title": title,
                    "skus": "|".join(skus),
                    "detail": "同一详情图也用于: " + ", ".join(sorted(others)[:12]),
                    "admin_url": admin_url,
                })
        if cross_partners:
            warning.append(f"detail_shared_across_products({len(cross_partners)})")

        # 2b) 内容哈希复用（换名同图）
        hash_partners: set[str] = set()
        if use_hash:
            for u in set(detail_norm):
                partners = url_to_dup_hash_partners.get(u, set())
                # 仅保留不同文件名的伙伴
                partners = {p for p in partners if basename(p) != basename(u)}
                if partners:
                    hash_partners |= partners
            if hash_partners:
                warning.append(f"detail_dup_by_content({len(hash_partners)})")

        # 3) 详情图与主图重复
        detail_eq_main = sorted(set(detail_norm) & main_norm)
        if detail_eq_main:
            warning.append(f"detail_equals_main({len(detail_eq_main)})")

        severity = "critical" if critical else ("warning" if warning else "ok")
        action = (
            "修复缺失(主图/详情图/featured)" if critical
            else ("复核重复/补元字段" if warning else "ok")
        )

        rows.append({
            "sku": "|".join(skus),
            "title": title,
            "handle": handle,
            "admin_url": admin_url,
            "main_image_count": str(len(main_urls)),
            "detail_image_count": str(len(detail_norm)),
            "detail_unique_count": str(len(set(detail_norm))),
            "empty_img_tags": str(empty_tags),
            "missing_required_metafields": "|".join(missing_mf),
            "detail_dup_in_product": "|".join(basename(u) for u in intra_dups),
            "detail_shared_with_handles": "|".join(sorted(cross_partners)[:20]),
            "detail_dup_by_content": "|".join(basename(u) for u in sorted(hash_partners))[:500] if use_hash else "",
            "detail_equals_main": "|".join(basename(u) for u in detail_eq_main),
            "severity": severity,
            "issues": " ; ".join(critical + warning),
            "recommended_action": action,
        })

    summary = {
        "active_products_checked": len(products),
        "with_critical": sum(1 for r in rows if r["severity"] == "critical"),
        "with_warning_only": sum(1 for r in rows if r["severity"] == "warning"),
        "clean": sum(1 for r in rows if r["severity"] == "ok"),
        "missing_main_image": sum(1 for r in rows if "missing_main_image" in r["issues"]),
        "missing_detail_images": sum(1 for r in rows if "missing_detail_images" in r["issues"]),
        "empty_img_tags_products": sum(1 for r in rows if int(r["empty_img_tags"]) > 0),
        "detail_dup_in_product_products": sum(1 for r in rows if r["detail_dup_in_product"]),
        "detail_shared_across_products_products": sum(1 for r in rows if r["detail_shared_with_handles"]),
        "detail_equals_main_products": sum(1 for r in rows if r["detail_equals_main"]),
        "products_missing_any_metafield": sum(1 for r in rows if r["missing_required_metafields"]),
        "cross_product_shared_image_groups": len(shared_url_to_handles),
        "content_hash_dup_groups": len(shared_hash_groups) if use_hash else None,
    }
    return rows, dup_detail_rows, summary


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as f:  # utf-8-sig 便于 Excel 直接打开中文
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def try_write_xlsx(path: Path, sheets: dict[str, list[dict[str, str]]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return False
    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(bold=True, color="FFFFFF")
    for name, rows in sheets.items():
        ws = wb.create_sheet(name[:31])
        if not rows:
            ws.append(["(无)"])
            continue
        cols = list(rows[0].keys())
        ws.append(cols)
        for c in ws[1]:
            c.fill = head_fill
            c.font = head_font
        for r in rows:
            ws.append([r.get(c, "") for c in cols])
        ws.freeze_panes = "A2"
        for i, col in enumerate(cols, 1):
            width = min(60, max(12, max([len(col)] + [len(str(r.get(col, ""))) for r in rows[:200]]) + 2))
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = width
    wb.save(path)
    return True


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default="/private/tmp/jiestar-detail-image-audit")
    parser.add_argument("--vendor", default=None, help="只审查指定厂商，例如 --vendor iBlock")
    parser.add_argument("--hash", action="store_true", help="下载图片算内容哈希，识别换名同图（更慢）")
    args = parser.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    admin = ShopifyAdmin()
    products = fetch_active_products(admin, args.vendor)
    rows, dup_rows, summary = audit(products, admin.domain, args.hash)

    # 排序：critical 在前，其次有重复/缺失的
    sev_rank = {"critical": 0, "warning": 1, "ok": 2}
    rows.sort(key=lambda r: (sev_rank.get(r["severity"], 9), r["title"]))

    audit_csv = out_dir / "detail-image-audit.csv"
    dup_csv = out_dir / "detail-image-duplicates.csv"
    summary_json = out_dir / "detail-image-audit-summary.json"
    write_csv(audit_csv, rows)
    write_csv(dup_csv, dup_rows)
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    xlsx_path = out_dir / "detail-image-audit.xlsx"
    xlsx_ok = try_write_xlsx(xlsx_path, {"逐产品审查": rows, "重复详情图明细": dup_rows})

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"audit_csv={audit_csv}")
    print(f"duplicates_csv={dup_csv}")
    print(f"summary_json={summary_json}")
    print(f"xlsx={'(written) ' + str(xlsx_path) if xlsx_ok else '(skipped: pip install openpyxl 后可生成 .xlsx)'}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(str(error), file=sys.stderr)
        raise SystemExit(1)
