#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import shopify_sample_import as base_import
from shopify_cn_pending_import import ShopifyAdmin as BaseShopifyAdmin


QUOTE_XLSX = Path("/Volumes/ORICO/TK two/TK TWO积木报价表26.05.25(1).xlsx")
IMAGE_ROOT = Path("/Volumes/ORICO/TK two/TK2详情图")
METADATA_XLSX = Path("/Volumes/ORICO/TK two/TK Two产品元字段资料表.xlsx")
OUT_DIR = Path("/private/tmp/jiestar-shopify-tktwo-import")

VENDOR = "TK Two"
STATUS = "ACTIVE"
PRICE = "999"
PRODUCT_TYPE = "Tank"
CUSTOM_SERIES = "Tank"
CATEGORY_ID = "gid://shopify/TaxonomyCategory/tg-5-7-12"
CATEGORY_NAME = "Interlocking Blocks"
OPTION_NAME = "Model"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
REQUIRED_SCOPES = {"read_products", "write_products", "read_files", "write_files"}
PUBLICATION_SCOPES = {"read_publications", "write_publications"}
DETAIL_SLICE_MAX_HEIGHT = 4000
DETAIL_SLICE_MIN_BYTES = 5 * 1024 * 1024
DETAIL_SLICE_QUALITY = 86

TITLE_BY_SKU = {
    "TK8001": "TK Two Tiger I Heavy Tank Model Kit",
    "TK8002": "TK Two Panther Tank Model Kit",
    "TK8003": "TK Two Lucky Tiger Tank Model Kit",
    "TK8004": "TK Two King Tiger Tank Model Kit",
    "TK8005": "TK Two Cheetah Tank Model Kit",
    "TK8006": "TK Two Jagdtiger Tank Model Kit",
    "TK9001": "TK Two Panzer IV Tank Model Kit",
    "TK9002": "TK Two Panzer III Tank Model Kit",
}

TARGET_SKUS = tuple(TITLE_BY_SKU)


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_header(value: Any) -> str:
    return clean(value).replace(" ", "")


def contains_cjk(value: str) -> bool:
    return bool(re.search(r"[\u3400-\u9fff]", value))


def same_money(left: Any, right: Any) -> bool:
    try:
        return Decimal(str(left)) == Decimal(str(right))
    except (InvalidOperation, ValueError):
        return False


def natural_key(path: Path) -> tuple[Any, ...]:
    parts = re.split(r"(\d+)", path.name.lower())
    return tuple(int(part) if part.isdigit() else part for part in parts)


def slugify(value: str) -> str:
    return base_import.slugify(value)


def read_quote_rows() -> list[dict[str, str]]:
    workbook = load_workbook(QUOTE_XLSX, read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    headers = [normalize_header(cell.value) for cell in sheet[2]]
    rows: list[dict[str, str]] = []

    for values in sheet.iter_rows(min_row=3, values_only=True):
        row = {headers[index]: clean(value) for index, value in enumerate(values) if index < len(headers)}
        sku = clean(row.get("产品货号")).upper()
        if sku in TARGET_SKUS:
            row["sku"] = sku
            rows.append(row)

    rows.sort(key=lambda row: TARGET_SKUS.index(row["sku"]))
    return rows


def image_files_for_sku(sku: str) -> list[Path]:
    code = sku.removeprefix("TK")
    files: list[Path] = []
    for path in IMAGE_ROOT.rglob("*"):
        if not path.is_file():
            continue
        if path.name.startswith("._") or path.suffix.lower() not in IMAGE_EXTS:
            continue
        if any(part in {"源文件", "__MACOSX"} for part in path.parts):
            continue
        text = str(path)
        if re.search(rf"(?i)TK\s*[-_]?\s*{re.escape(code)}|(?<!\d){re.escape(code)}(?!\d)", text):
            files.append(path)
    return sorted(files, key=natural_key)


def detail_rank(path: Path) -> tuple[int, tuple[Any, ...]]:
    text = str(path)
    if "详情页-切片" in text:
        bucket = 0
    elif "详情-切片" in text:
        bucket = 1
    elif "长图" in path.name:
        bucket = 2
    elif "短图" in path.name:
        bucket = 3
    else:
        bucket = 9
    return (bucket, natural_key(path))


def media_for_sku(sku: str) -> dict[str, list[Path]]:
    files = image_files_for_sku(sku)

    white = [
        path
        for path in files
        if "白底" in str(path)
        and "尺寸" not in path.name
        and "透明" not in path.name
    ]
    white = sorted(white, key=lambda path: (0 if path.suffix.lower() in {".jpg", ".jpeg"} else 1, len(str(path)), natural_key(path)))

    sku_images = sorted([path for path in files if "sku" in path.name.lower()], key=natural_key)

    details = [
        path
        for path in files
        if (
            "详情页-切片" in str(path)
            or "详情-切片" in str(path)
            or (sku == "TK8005" and ("长图" in path.name or "短图" in path.name))
        )
    ]
    details = sorted(details, key=detail_rank)

    main = [
        path
        for path in files
        if path not in white
        and path not in sku_images
        and path not in details
        and "尺寸" not in path.name
        and "透明" not in path.name
        and ("主图" in str(path.parent) or "主图" in path.name or re.search(r"(?:^|[-_])\d{1,2}\.", path.name))
    ]
    main = sorted(main, key=natural_key)

    fallback_sku = sku_images[:1] or white[:1]
    return {
        "white": white[:1],
        "main": main,
        "sku": fallback_sku,
        "detail": details,
        "all": files,
    }


def metafields_for_row(row: dict[str, str]) -> dict[str, str]:
    metafields = {
        "specs.piece_count": clean(row.get("颗粒数/PCS")),
        "specs.finished_model_size": clean(row.get("产品尺寸/CM")),
        "specs.package_size": clean(row.get("彩盒尺寸/CM")),
        "specs.difficulty_level": "See product package",
        "custom.series": CUSTOM_SERIES,
    }
    return {key: value for key, value in metafields.items() if value}


def manifest_item(row: dict[str, str]) -> tuple[dict[str, Any], list[str]]:
    sku = row["sku"]
    media = media_for_sku(sku)
    title = TITLE_BY_SKU[sku]
    handle = slugify(title)
    issues: list[str] = []

    if contains_cjk(title):
        issues.append("title_contains_chinese")
    if not media["white"]:
        issues.append("missing_white_image")
    if not media["detail"]:
        issues.append("missing_detail_image")
    if VENDOR != "TK Two":
        issues.append("vendor_mismatch")
    if PRICE != "999":
        issues.append("price_not_999")
    if STATUS != "ACTIVE":
        issues.append("status_not_active")

    main_media = media["white"] + media["main"]
    if not main_media:
        issues.append("missing_main_media")

    item = {
        "folder": sku,
        "folder_path": str(IMAGE_ROOT),
        "base": sku,
        "handle": handle,
        "title": title,
        "vendor": VENDOR,
        "status": STATUS,
        "product_type": PRODUCT_TYPE,
        "category": CATEGORY_ID,
        "category_name": CATEGORY_NAME,
        "price": PRICE,
        "option_name": OPTION_NAME,
        "variants": [
            {
                "sku": sku,
                "option_name": f"{sku} - {title.removeprefix(VENDOR).strip()}",
                "title_source": clean(row.get("产品名称")),
                "series": CUSTOM_SERIES,
                "age": "",
                "piece_count": clean(row.get("颗粒数/PCS")),
                "package_size": clean(row.get("彩盒尺寸/CM")),
                "finished_size": clean(row.get("产品尺寸/CM")),
            }
        ],
        "metafields": metafields_for_row(row),
        "main_media": [str(path) for path in main_media],
        "sku_images": [str(path) for path in media["sku"]],
        "detail_images": [str(path) for path in media["detail"]],
        "transparent_images": [],
        "source_row": row,
        "media_status": {
            "all_count": len(media["all"]),
            "white_count": len(media["white"]),
            "main_count": len(media["main"]),
            "sku_count": len(media["sku"]),
            "detail_count": len(media["detail"]),
        },
        "missing": {
            "white": not bool(media["white"]),
            "detail": not bool(media["detail"]),
            "sku_image_fallback_to_white": not any("sku" in Path(path).name.lower() for path in item_sku_images(media)),
        },
    }
    return item, issues


def item_sku_images(media: dict[str, list[Path]]) -> list[str]:
    return [str(path) for path in media["sku"]]


def build_manifest() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    manifest: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for row in read_quote_rows():
        item, issues = manifest_item(row)
        if issues:
            skipped.append(
                {
                    "sku": row["sku"],
                    "handle": item["handle"],
                    "title": item["title"],
                    "issues": issues,
                    "media_status": item["media_status"],
                }
            )
        else:
            manifest.append(item)

    return manifest, skipped


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def write_metadata_workbook(manifest: list[dict[str, Any]], skipped: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TK Two Metafields"
    headers = [
        "sku",
        "vendor",
        "shopify_title",
        "handle",
        "status",
        "price",
        "product_type",
        "category",
        "specs.piece_count",
        "specs.finished_model_size",
        "specs.package_size",
        "specs.difficulty_level",
        "custom.series",
        "original_name_cn",
        "carton_qty",
        "domestic_control_price_cny",
        "1688_control_price_cny",
        "cross_border_control_price_usd",
        "product_weight_with_package",
        "gross_net_weight_kg",
        "media_status",
        "upload_status",
        "issues",
    ]
    sheet.append(headers)

    all_rows = [(item, "READY", "") for item in manifest] + [
        (
            {
                "base": row["sku"],
                "vendor": VENDOR,
                "title": row["title"],
                "handle": row["handle"],
                "status": STATUS,
                "price": PRICE,
                "product_type": PRODUCT_TYPE,
                "category_name": CATEGORY_NAME,
                "metafields": {},
                "source_row": next((source for source in read_quote_rows() if source["sku"] == row["sku"]), {}),
                "media_status": row.get("media_status", {}),
            },
            "SKIPPED",
            "; ".join(row.get("issues", [])),
        )
        for row in skipped
    ]

    for item, upload_status, issues in all_rows:
        source = item.get("source_row", {})
        metafields = item.get("metafields", {})
        media_status = item.get("media_status", {})
        sheet.append(
            [
                item.get("base", ""),
                item.get("vendor", ""),
                item.get("title", ""),
                item.get("handle", ""),
                item.get("status", ""),
                item.get("price", ""),
                item.get("product_type", ""),
                item.get("category_name", ""),
                metafields.get("specs.piece_count", ""),
                metafields.get("specs.finished_model_size", ""),
                metafields.get("specs.package_size", ""),
                metafields.get("specs.difficulty_level", ""),
                metafields.get("custom.series", ""),
                source.get("产品名称", ""),
                source.get("装箱量", ""),
                source.get("国内电商控价/元", ""),
                source.get("1688控价/元", ""),
                source.get("跨境控价/$", ""),
                source.get("产品重量/（含包装）") or source.get("产品重量/ （含包装）") or "",
                source.get("毛净重/kg", ""),
                json.dumps(media_status, ensure_ascii=False),
                upload_status,
                issues,
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 12,
        "B": 12,
        "C": 44,
        "D": 40,
        "E": 12,
        "F": 10,
        "G": 16,
        "H": 22,
        "I": 16,
        "J": 20,
        "K": 18,
        "L": 24,
        "M": 14,
        "N": 18,
        "O": 12,
        "P": 18,
        "Q": 24,
        "R": 18,
        "S": 22,
        "T": 16,
        "U": 42,
        "V": 14,
        "W": 30,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

    METADATA_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(METADATA_XLSX)


def write_reports(manifest: list[dict[str, Any]], skipped: list[dict[str, Any]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    write_json(OUT_DIR / "tktwo-manifest.json", manifest)
    write_json(OUT_DIR / "tktwo-skipped.json", skipped)
    write_metadata_workbook(manifest, skipped)
    write_csv(
        OUT_DIR / "tktwo-manifest.csv",
        [
            {
                "sku": item["base"],
                "handle": item["handle"],
                "title": item["title"],
                "vendor": item["vendor"],
                "status": item["status"],
                "price": item["price"],
                "product_type": item["product_type"],
                "main_media_count": len(item["main_media"]),
                "detail_count": len(item["detail_images"]),
                "sku_image_count": len(item["sku_images"]),
                "first_media": Path(item["main_media"][0]).name if item["main_media"] else "",
            }
            for item in manifest
        ],
        [
            "sku",
            "handle",
            "title",
            "vendor",
            "status",
            "price",
            "product_type",
            "main_media_count",
            "detail_count",
            "sku_image_count",
            "first_media",
        ],
    )
    write_csv(
        OUT_DIR / "tktwo-skipped.csv",
        [
            {
                "sku": row["sku"],
                "handle": row["handle"],
                "title": row["title"],
                "issues": "; ".join(row["issues"]),
                "media_status": json.dumps(row.get("media_status", {}), ensure_ascii=False),
            }
            for row in skipped
        ],
        ["sku", "handle", "title", "issues", "media_status"],
    )


class ShopifyAdmin(BaseShopifyAdmin):
    def product_set(self, item: dict[str, Any], description_html: str) -> dict[str, Any]:
        variants = [
            {
                "optionValues": [
                    {
                        "optionName": item.get("option_name") or OPTION_NAME,
                        "name": variant["option_name"],
                    }
                ],
                "price": item["price"],
                "inventoryItem": {
                    "sku": variant["sku"],
                    "tracked": False,
                },
            }
            for variant in item["variants"]
        ]
        data = self.graphql(
            """
            mutation ProductSet($input: ProductSetInput!, $synchronous: Boolean!) {
              productSet(input: $input, synchronous: $synchronous) {
                product {
                  id
                  title
                  handle
                  status
                  vendor
                  productType
                  variants(first: 50) {
                    nodes {
                      id
                      title
                      price
                      sku
                      inventoryItem {
                        id
                        tracked
                      }
                    }
                  }
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {
                "synchronous": True,
                "input": {
                    "title": item["title"],
                    "handle": item["handle"],
                    "vendor": item["vendor"],
                    "status": item["status"],
                    "productType": item["product_type"],
                    "descriptionHtml": description_html,
                    "productOptions": [
                        {
                            "name": item.get("option_name") or OPTION_NAME,
                            "values": [{"name": variant["option_name"]} for variant in item["variants"]],
                        }
                    ],
                    "variants": variants,
                    "metafields": base_import.product_metafields(item),
                },
            },
        )
        result = data["productSet"]
        base_import.assert_no_user_errors("productSet", result["userErrors"])
        return result["product"]

    def products_by_sku(self, skus: set[str]) -> list[dict[str, Any]]:
        products: list[dict[str, Any]] = []
        cursor: str | None = None
        query = " OR ".join(sorted(skus))

        while True:
            data = self.graphql(
                """
                query ProductsBySku($cursor: String, $query: String!) {
                  products(first: 100, after: $cursor, query: $query) {
                    pageInfo {
                      hasNextPage
                      endCursor
                    }
                    nodes {
                      id
                      title
                      handle
                      status
                      vendor
                      productType
                      descriptionHtml
                      category {
                        id
                        name
                      }
                      media(first: 250, sortKey: POSITION) {
                        nodes {
                          id
                          alt
                          ... on MediaImage {
                            image {
                              url
                            }
                          }
                        }
                      }
                      metafields(first: 50) {
                        nodes {
                          namespace
                          key
                          value
                        }
                      }
                      variants(first: 50) {
                        nodes {
                          id
                          title
                          price
                          sku
                          image {
                            id
                            url
                            altText
                          }
                          media(first: 10) {
                            nodes {
                              id
                              alt
                            }
                          }
                          inventoryItem {
                            tracked
                          }
                        }
                      }
                    }
                  }
                }
                """,
                {"cursor": cursor, "query": query},
            )
            page = data["products"]
            for product in page["nodes"]:
                product_skus = {(variant.get("sku") or "").strip().upper() for variant in product["variants"]["nodes"]}
                if product_skus & skus:
                    products.append(product)
            if not page["pageInfo"]["hasNextPage"]:
                break
            cursor = page["pageInfo"]["endCursor"]
        return products


def upload_detail_images_for_item(admin: ShopifyAdmin, item: dict[str, Any]) -> list[str]:
    urls: list[str] = []
    for detail_index, source in enumerate(item["detail_images"], start=1):
        path = Path(source)
        for part_index, upload_path in enumerate(detail_image_paths(path, item["base"]), start=1):
            part = f" part {part_index}" if part_index > 1 else ""
            urls.append(admin.file_create(upload_path, f"{item['title']} details {detail_index}{part}"))
    return urls


def detail_image_paths(path: Path, sku: str) -> list[Path]:
    if not path.exists():
        return []

    try:
        from PIL import Image
    except ImportError:
        return [path]

    with Image.open(path) as image:
        width, height = image.size

        if height <= DETAIL_SLICE_MAX_HEIGHT and path.stat().st_size <= DETAIL_SLICE_MIN_BYTES:
            return [path]

        output_dir = OUT_DIR / "detail-slices" / sku / path.stem
        output_dir.mkdir(parents=True, exist_ok=True)
        output_paths: list[Path] = []
        part_count = max(1, (height + DETAIL_SLICE_MAX_HEIGHT - 1) // DETAIL_SLICE_MAX_HEIGHT)

        for index in range(part_count):
            top = round(index * height / part_count)
            bottom = round((index + 1) * height / part_count)
            output_path = output_dir / f"{path.stem}-part-{index + 1:02d}.jpg"
            crop = image.crop((0, top, width, bottom))

            if crop.mode != "RGB":
                crop = crop.convert("RGB")

            crop.save(output_path, format="JPEG", quality=DETAIL_SLICE_QUALITY, optimize=True)
            output_paths.append(output_path)

    return output_paths


def description_html(item: dict[str, Any], detail_urls: list[str]) -> str:
    return "\n".join(
        f'<p><img src="{html.escape(url)}" alt="{html.escape(item["title"])} details part {index}" /></p>'
        for index, url in enumerate(detail_urls, start=1)
    )


def filter_existing(admin: ShopifyAdmin, manifest: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    existing_handles, existing_skus = admin.products_index()
    todo: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for item in manifest:
        item_skus = {variant["sku"].upper() for variant in item["variants"]}
        matched_skus = sorted(item_skus & existing_skus)
        if item["handle"] in existing_handles:
            skipped.append({"sku": item["base"], "handle": item["handle"], "title": item["title"], "issues": ["existing_handle"]})
        elif matched_skus:
            skipped.append(
                {
                    "sku": item["base"],
                    "handle": item["handle"],
                    "title": item["title"],
                    "issues": [f"existing_sku:{', '.join(matched_skus)}"],
                }
            )
        else:
            todo.append(item)

    return todo, skipped


def apply_create(manifest: list[dict[str, Any]], report_name: str = "tktwo-created-products") -> list[dict[str, Any]]:
    admin = ShopifyAdmin()
    scopes = admin.access_scopes()
    missing_scopes = sorted(REQUIRED_SCOPES - scopes)
    if missing_scopes:
        raise RuntimeError(f"Missing Shopify scopes: {', '.join(missing_scopes)}")

    publication_ids: list[str] = []
    if PUBLICATION_SCOPES <= scopes:
        publication_ids = [publication["id"] for publication in admin.publications()]

    todo, existing_skipped = filter_existing(admin, manifest)
    results: list[dict[str, Any]] = [{"manifest": row, "ok": False, "skipped": True, "error": "; ".join(row["issues"])} for row in existing_skipped]
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for index, item in enumerate(todo, start=1):
        result: dict[str, Any] = {"manifest": item, "ok": False}
        try:
            detail_urls = upload_detail_images_for_item(admin, item)
            product = admin.product_set(item, description_html(item, detail_urls))
            product_id = product["id"]
            base_import.sync_product_assets(admin, item, product_id, update_description=False)
            admin.update_status_and_category(product_id)
            admin.publish_to_publications(product_id, publication_ids)
            result.update({"ok": True, "product": admin.fetch_product(product_id)})
        except Exception as error:  # noqa: BLE001 - batch should continue and report every SKU.
            result["error"] = str(error)

        results.append(result)
        write_json(OUT_DIR / f"{report_name}.json", results)
        print(f"Processed {index}/{len(todo)} {item['base']} ok={result['ok']}", flush=True)
        time.sleep(0.5)

    return results


def verify_products(manifest: list[dict[str, Any]]) -> list[dict[str, Any]]:
    admin = ShopifyAdmin()
    expected_by_sku = {item["base"]: item for item in manifest}
    products = admin.products_by_sku(set(expected_by_sku))
    rows: list[dict[str, Any]] = []

    for sku, item in expected_by_sku.items():
        product = next(
            (
                candidate
                for candidate in products
                if any((variant.get("sku") or "").strip().upper() == sku for variant in candidate["variants"]["nodes"])
            ),
            None,
        )
        if not product:
            rows.append({"sku": sku, "ok": False, "issues": ["missing_product"]})
            continue

        variant = next(
            variant for variant in product["variants"]["nodes"] if (variant.get("sku") or "").strip().upper() == sku
        )
        metafields = {f"{node['namespace']}.{node['key']}": node["value"] for node in product["metafields"]["nodes"]}
        first_alt = (product["media"]["nodes"][0].get("alt") or "") if product["media"]["nodes"] else ""
        issues = []
        checks = {
            "title": product["title"] == item["title"],
            "vendor": product["vendor"] == VENDOR,
            "status": product["status"] == STATUS,
            "price": same_money(variant.get("price"), PRICE),
            "product_type": product["productType"] == PRODUCT_TYPE,
            "category": (product.get("category") or {}).get("id") == CATEGORY_ID,
            "description_has_images": "<img" in (product.get("descriptionHtml") or ""),
            "first_media_white": "白底" in first_alt,
            "variant_has_media": bool((variant.get("media") or {}).get("nodes") or variant.get("image")),
        }
        checks["metafields"] = all(str(metafields.get(key, "")) == str(value) for key, value in item["metafields"].items())
        for key, ok in checks.items():
            if not ok:
                issues.append(key)
        rows.append(
            {
                "sku": sku,
                "product_id": product["id"],
                "handle": product["handle"],
                "title": product["title"],
                "ok": not issues,
                "issues": issues,
                "checks": checks,
                "media_count": len(product["media"]["nodes"]),
                "detail_image_count": len(item["detail_images"]),
            }
        )

    write_json(OUT_DIR / "tktwo-verify.json", rows)
    write_csv(
        OUT_DIR / "tktwo-verify.csv",
        [
            {
                "sku": row["sku"],
                "title": row.get("title", ""),
                "handle": row.get("handle", ""),
                "ok": row["ok"],
                "issues": "; ".join(row.get("issues", [])),
                "media_count": row.get("media_count", ""),
                "detail_image_count": row.get("detail_image_count", ""),
            }
            for row in rows
        ],
        ["sku", "title", "handle", "ok", "issues", "media_count", "detail_image_count"],
    )
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare and upload TK Two products to Shopify.")
    parser.add_argument("--dry-run", action="store_true", help="Generate manifest, skipped report, and metadata workbook only.")
    parser.add_argument("--apply", action="store_true", help="Create ready TK Two products in Shopify.")
    parser.add_argument("--verify", action="store_true", help="Verify created TK Two products in Shopify.")
    args = parser.parse_args()

    if not (args.dry_run or args.apply or args.verify):
        parser.error("Choose at least one of --dry-run, --apply, or --verify")

    manifest, skipped = build_manifest()
    write_reports(manifest, skipped)

    summary: dict[str, Any] = {
        "quote_xlsx": str(QUOTE_XLSX),
        "image_root": str(IMAGE_ROOT),
        "metadata_xlsx": str(METADATA_XLSX),
        "manifest_count": len(manifest),
        "skipped_count": len(skipped),
        "out_dir": str(OUT_DIR),
    }

    if args.apply:
        results = apply_create(manifest)
        summary["created_ok"] = sum(1 for result in results if result.get("ok") and not result.get("skipped"))
        summary["created_failed"] = sum(1 for result in results if not result.get("ok") and not result.get("skipped"))
        summary["existing_skipped"] = sum(1 for result in results if result.get("skipped"))

    if args.verify:
        verify_rows = verify_products(manifest)
        summary["verify_ok"] = sum(1 for row in verify_rows if row.get("ok"))
        summary["verify_failed"] = sum(1 for row in verify_rows if not row.get("ok"))

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if not skipped else 2


if __name__ == "__main__":
    raise SystemExit(main())
