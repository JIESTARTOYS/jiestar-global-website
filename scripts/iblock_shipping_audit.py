#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from dataclasses import asdict, dataclass
from decimal import Decimal, ROUND_CEILING
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from shopify_shipping_update_from_template import ShopifyAdmin, clean, grams_from_weight, sku_key


SOURCE_WORKBOOK = Path("/Volumes/ORICO/各品牌报价表/中文源报价表/积趣IBLOCK-全品报价单202607（含重量）.xlsx")
BASE_TEMPLATE = Path("/Users/chensen/jiestar/定价参考/Shopify运费模板_体积重_Shopify盒规补全_缺失SKU补全_20260701.xlsx")
PRIOR_READINESS = Path("/Volumes/ORICO/iblock/iblock-上架前整理/reports/iblock-shopify-readiness.csv")
OUT_JSON = Path("/private/tmp/jiestar-shopify-iblock-shipping/iblock-shipping-audit.json")

MANUAL_REVIEW_SKUS = {"IB1101获奖版"}
NON_LISTABLE_SKUS = {"IB1101-5", "IB1102-5", "IB2202"}
DEFAULT_PROFILE = "Standard goods"


@dataclass(frozen=True)
class SourceProduct:
    sku: str
    series: str
    name: str
    box_size_raw: str
    box_length_cm: float
    box_width_cm: float
    box_height_cm: float
    actual_weight_g: float
    volumetric_weight_g: float
    target_weight_g: int
    old_template_weight_g: int | None
    delta_from_old_template_g: int | None
    shipping_profile: str
    disposition: str


def parse_dimension_triplets(value: Any) -> list[tuple[float, float, float]]:
    text = clean(value).replace("×", "*").replace("X", "*").replace("x", "*")
    triplets = []
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)", text):
        triplets.append(tuple(float(match.group(index)) for index in range(1, 4)))
    return triplets


def sellable_box_dimensions(sku: str, value: Any) -> tuple[float, float, float]:
    triplets = parse_dimension_triplets(value)
    if not triplets:
        raise ValueError(f"Missing parseable box dimensions for {sku}: {value!r}")
    # Parent/display-pack rows explicitly sell the end box. Child variants use
    # their own one-box dimensions and normally contain only one triplet.
    if sku_key(sku) in {"IB2201", "IB2202"}:
        return triplets[0]
    return triplets[-1]


def chargeable_weight_g(actual_weight_g: float, dimensions: tuple[float, float, float]) -> tuple[float, int]:
    length, width, height = dimensions
    volumetric = length * width * height / 5000 * 1000
    target = int(Decimal(str(max(actual_weight_g, volumetric))).quantize(Decimal("1"), rounding=ROUND_CEILING))
    return round(volumetric, 4), target


def load_base_targets(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    sheet = workbook["Shopify商品重量导入"]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    output: dict[str, dict[str, Any]] = {}
    for row in rows:
        record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers)) if headers[index]}
        sku = sku_key(record.get("Variant SKU"))
        if not sku.startswith("IB"):
            continue
        output[sku] = {
            "weight_g": int(record["Variant Weight"]) if record.get("Variant Weight") not in (None, "") else None,
            "profile": clean(record.get("Shipping Profile Suggestion")) or DEFAULT_PROFILE,
        }
    return output


def load_source_products(source_path: Path, base_template: Path) -> list[SourceProduct]:
    base_targets = load_base_targets(base_template)
    workbook = load_workbook(source_path, read_only=True, data_only=True, keep_links=False)
    sheet = workbook["全品系列"]
    products: list[SourceProduct] = []
    seen: set[str] = set()
    for row in sheet.iter_rows(min_row=5, values_only=True):
        sku = sku_key(row[4] if len(row) > 4 else None)
        if not sku:
            continue
        if sku in seen:
            raise ValueError(f"Duplicate source SKU: {sku}")
        seen.add(sku)
        box_raw = clean(row[18] if len(row) > 18 else None)
        actual_weight = row[23] if len(row) > 23 else None
        if actual_weight in (None, ""):
            raise ValueError(f"Missing single-box weight for {sku}")
        dimensions = sellable_box_dimensions(sku, box_raw)
        volumetric, target = chargeable_weight_g(float(actual_weight), dimensions)
        old = base_targets.get(sku, {}).get("weight_g")
        if sku in MANUAL_REVIEW_SKUS:
            disposition = "manual_review"
        elif sku in NON_LISTABLE_SKUS:
            disposition = "audit_only"
        else:
            disposition = "eligible_if_active"
        products.append(
            SourceProduct(
                sku=sku,
                series=clean(row[3] if len(row) > 3 else None),
                name=clean(row[5] if len(row) > 5 else None),
                box_size_raw=box_raw,
                box_length_cm=dimensions[0],
                box_width_cm=dimensions[1],
                box_height_cm=dimensions[2],
                actual_weight_g=float(actual_weight),
                volumetric_weight_g=volumetric,
                target_weight_g=target,
                old_template_weight_g=old,
                delta_from_old_template_g=None if old is None else target - old,
                shipping_profile=base_targets.get(sku, {}).get("profile") or DEFAULT_PROFILE,
                disposition=disposition,
            )
        )
    return sorted(products, key=lambda item: item.sku)


def fetch_shopify_iblock_variants(admin: ShopifyAdmin) -> list[dict[str, Any]]:
    query = """
    query IblockVariantsForShippingAudit($cursor: String) {
      products(first: 100, after: $cursor, query: "vendor:iBlock", sortKey: TITLE) {
        pageInfo { hasNextPage endCursor }
        nodes {
          id handle title status vendor
          variants(first: 250) {
            nodes {
              id title sku
              inventoryItem {
                id requiresShipping
                measurement { weight { value unit } }
              }
            }
          }
        }
      }
    }
    """
    output: list[dict[str, Any]] = []
    cursor = None
    while True:
        page = admin.graphql(query, {"cursor": cursor})["products"]
        for product in page["nodes"]:
            for variant in product.get("variants", {}).get("nodes", []):
                inventory_item = variant.get("inventoryItem") or {}
                output.append(
                    {
                        "product_id": product["id"],
                        "product_title": clean(product.get("title")),
                        "handle": clean(product.get("handle")),
                        "status": clean(product.get("status")),
                        "vendor": clean(product.get("vendor")),
                        "variant_id": variant["id"],
                        "variant_title": clean(variant.get("title")),
                        "sku": sku_key(variant.get("sku")),
                        "current_weight_g": grams_from_weight((inventory_item.get("measurement") or {}).get("weight")),
                        "requires_shipping": inventory_item.get("requiresShipping"),
                    }
                )
        if not page["pageInfo"]["hasNextPage"]:
            break
        cursor = page["pageInfo"]["endCursor"]
    return output


def build_audit(products: list[SourceProduct], shopify_variants: list[dict[str, Any]] | None, snapshot_error: str = "") -> dict[str, Any]:
    eligible = {product.sku: product for product in products if product.disposition == "eligible_if_active"}
    manual_rows = [asdict(product) | {"reason": product.disposition} for product in products if product.disposition != "eligible_if_active"]
    shopify_variants = shopify_variants or []
    shopify_by_sku: dict[str, list[dict[str, Any]]] = {}
    for row in shopify_variants:
        shopify_by_sku.setdefault(sku_key(row.get("sku")), []).append(row)

    weight_rows: list[dict[str, Any]] = []
    for sku, product in eligible.items():
        matches = shopify_by_sku.get(sku, [])
        active = [row for row in matches if clean(row.get("status")).upper() == "ACTIVE"]
        if len(active) == 1:
            current = active[0].get("current_weight_g")
            action = "noop" if current == product.target_weight_g and active[0].get("requires_shipping") is True else "update"
            weight_rows.append(asdict(product) | active[0] | {"action": action})

    missing = []
    if shopify_variants:
        missing = [asdict(product) | {"reason": "missing_in_shopify"} for sku, product in eligible.items() if sku not in shopify_by_sku]
    else:
        old_template_skus = {product.sku for product in products if product.old_template_weight_g is not None}
        missing = [
            asdict(product) | {"reason": "candidate_only_snapshot_unavailable"}
            for sku, product in eligible.items()
            if sku not in old_template_skus
        ]

    extra = [row | {"reason": "shopify_sku_not_in_current_source"} for sku, rows in shopify_by_sku.items() if sku and sku not in eligible for row in rows]
    for sku, rows in shopify_by_sku.items():
        if not sku:
            manual_rows.extend(row | {"reason": "blank_shopify_sku"} for row in rows)
        elif len(rows) > 1:
            manual_rows.extend(row | {"reason": "duplicate_shopify_sku"} for row in rows)

    return {
        "summary": {
            "source_sku_count": len(products),
            "eligible_source_sku_count": len(eligible),
            "source_duplicate_count": len(products) - len({product.sku for product in products}),
            "source_missing_weight_count": 0,
            "source_missing_box_size_count": 0,
            "shopify_snapshot_ok": bool(shopify_variants),
            "shopify_snapshot_error": snapshot_error,
            "shopify_variant_count": len(shopify_variants),
            "weight_update_count": sum(1 for row in weight_rows if row["action"] == "update"),
            "weight_noop_count": sum(1 for row in weight_rows if row["action"] == "noop"),
            "missing_in_shopify_count": len(missing) if shopify_variants else None,
            "candidate_missing_count": len(missing) if not shopify_variants else 0,
            "extra_in_shopify_count": len(extra),
            "manual_review_count": len(manual_rows),
            "disposition_counts": dict(Counter(product.disposition for product in products)),
        },
        "source_products": [asdict(product) for product in products],
        "weight_updates": weight_rows,
        "missing_in_shopify": missing,
        "extra_in_shopify": extra,
        "manual_review": manual_rows,
        "shopify_variants": shopify_variants,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the iBlock shipping-weight and Shopify SKU audit input JSON.")
    parser.add_argument("--source-workbook", type=Path, default=SOURCE_WORKBOOK)
    parser.add_argument("--base-template", type=Path, default=BASE_TEMPLATE)
    parser.add_argument("--out-json", type=Path, default=OUT_JSON)
    parser.add_argument("--skip-shopify", action="store_true", help="Build a source-only audit when Shopify Admin access is unavailable.")
    args = parser.parse_args()

    products = load_source_products(args.source_workbook, args.base_template)
    variants: list[dict[str, Any]] | None = None
    snapshot_error = ""
    if not args.skip_shopify:
        try:
            variants = fetch_shopify_iblock_variants(ShopifyAdmin())
        except Exception as error:
            snapshot_error = str(error)
    audit = build_audit(products, variants, snapshot_error)
    audit["sources"] = {
        "source_workbook": str(args.source_workbook),
        "base_template": str(args.base_template),
    }
    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit["summary"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
