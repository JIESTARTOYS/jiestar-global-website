#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import json
import re
import subprocess
import sys
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import shopify_sample_import as base_import
import shopify_title_cleanup as title_cleanup
from shopify_cn_pending_import import ShopifyAdmin
from shopify_cn_pending_import import recover_partial_product
from shopify_cn_pending_import import repair_existing_product


ROOT = Path("/Volumes/ORICO/砖悦英文电商图")
WORKBOOK = ROOT / "砖悦积木报价2026.1.20.xlsx"
OUT_DIR = Path("/private/tmp/jiestar-shopify-xbert-import")
PRICE = "999"
VENDOR = "Xbert"
STATUS = "ACTIVE"
CATEGORY_ID = "gid://shopify/TaxonomyCategory/tg-5-7-12"
PRODUCT_TYPE_DEFAULT = "Building Block Sets"
PRODUCT_TITLE_SUFFIX = "Building Block Set"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
IGNORED_FILE_NAMES = {"thumbs.db", ".ds_store"}
REQUIRED_SCOPES = {"read_products", "write_products", "read_files", "write_files"}
PUBLICATION_SCOPES = {"read_publications", "write_publications"}


SERIES_EN_BY_CN = {
    "昆虫": "Insect",
    "花": "Botanical",
    "植物": "Botanical",
    "动物": "Animal",
    "建筑": "Architecture",
    "街景": "Street View",
    "街景系列": "Street View",
    "皇家海盗系列": "Pirates",
    "海盗": "Pirates",
    "摆件系列": "Ornament",
    "飞机": "Aircraft",
    "机械": "Mechanical",
    "恐龙": "Dinosaur",
    "武器": "Display Model",
}

NAME_EN_BY_CN = {
    "蜜蜂": "Bee",
    "瓢虫": "Ladybug",
    "螳螂": "Mantis",
    "蝎子": "Scorpion",
    "虎甲虫": "Tiger Beetle",
    "蜻蜓": "Dragonfly",
    "蝴蝶": "Butterfly",
    "鹿": "Deer",
    "老虎": "Tiger",
    "狮子": "Lion",
    "老鹰": "Eagle",
    "大白鲨": "Great White Shark",
    "海龟": "Sea Turtle",
    "霸王龙": "Tyrannosaurus",
    "花束": "Flower Bouquet",
    "向日葵房": "Sunflower House",
    "植物园": "Botanical Garden",
    "热带雨林": "Tropical Rainforest",
    "博物馆": "Museum",
    "消防站": "Fire Station",
    "警车": "Police Car",
    "救护车": "Ambulance",
    "酒馆": "Tavern",
    "打字机": "Typewriter",
    "房子": "House",
    "船": "Ship",
    "重型拖车": "Heavy-Duty Tow Truck",
    "重型自卸车": "Heavy Dump Truck",
    "航天飞机": "Space Shuttle",
    "航班追踪": "Flight Tracker",
}

LOCAL_NAME_BY_SKU = {
    "66095": "Sierra Madre Fortress",
    "66139": "Red Rocket",
    "66141": "Aerospace Museum",
    "66154": "Watch Charging Stand",
    "66171": "Sandworm Strike",
    "66192": "Common Room",
    "66194": "European Train Station",
    "66202": "Post Apocalyptic Survivors Base",
    "66205": "Sky Bison",
    "66226": "Deer Head",
    "66255": "White Rabbit",
}

SENSITIVE_REPLACEMENTS = [
    (re.compile(r"\bvault\s*33\b", re.I), "Vault Door"),
    (re.compile(r"\bthe\s+botanical\s+garden\s+with\s+van\s+gogh\s+art\s+exhibition\b", re.I), "Botanical Garden Art Gallery"),
    (re.compile(r"\bvan\s+gogh\b", re.I), "Art Exhibition"),
    (re.compile(r"\bpiranha\s+plant\b", re.I), "Carnivorous Plant"),
    (re.compile(r"\bgt[-\s]?r\b", re.I), "Performance Sports Car"),
    (re.compile(r"\bwednesdays?\b", re.I), "Gothic"),
    (re.compile(r"\bforbidden\s+forest\b", re.I), "Dark Forest"),
    (re.compile(r"\bdangers?\s*&\s*dragons\b", re.I), "Fantasy"),
    (re.compile(r"\bdangers?\s+and\s+dragons\b", re.I), "Fantasy"),
    (re.compile(r"\bf[-\s]?150\s+raptor\b", re.I), "Off-Road Vehicle"),
    (re.compile(r"\bf[-\s]?150\b", re.I), "Off-Road Vehicle"),
    (re.compile(r"\b(chevrolet|corvette)\b", re.I), "Classic Sports Car"),
    (re.compile(r"\b(ferrari)\b", re.I), "Italian Sports Car"),
    (re.compile(r"\b(ford|raptor|bronco)\b", re.I), "Off-Road Vehicle"),
    (re.compile(r"\b(lamborghini|urus)\b", re.I), "Luxury SUV"),
    (re.compile(r"\b(land rover|defender)\b", re.I), "Off-Road SUV"),
    (re.compile(r"\b(mclaren)\b", re.I), "Super Sports Car"),
    (re.compile(r"雪佛兰|科尔维特", re.I), "Classic Sports Car"),
    (re.compile(r"法拉利", re.I), "Italian Sports Car"),
    (re.compile(r"福特|猛禽|烈马", re.I), "Off-Road Vehicle"),
    (re.compile(r"兰博基尼", re.I), "Luxury SUV"),
    (re.compile(r"路虎|卫士", re.I), "Off-Road SUV"),
]


@dataclass(frozen=True)
class XbertWorkbookRow:
    sku: str
    series: str
    name: str
    package_size: str
    finished_size: str
    age: str
    notes: str
    factory_price: str


def normalize_sku(value: Any) -> str:
    sku = str(value).strip().upper() if value is not None else ""
    return sku[:-2] if sku.endswith(".0") else sku


def clean_cell(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value).strip()) if value is not None else ""


def strip_language_suffix(value: str) -> str:
    return re.sub(r"(英文|中文)$", "", value).strip()


def extract_sku_from_path(path: Path) -> str:
    for part in reversed(path.parts):
        normalized = unicodedata.normalize("NFKC", part).upper()
        match = re.search(r"(?<!\d)([A-Z]{0,3}66\d{3,4})(?![A-Z0-9])", normalized)
        if match:
            return match.group(1)
    return ""


def load_workbook_rows(path: Path = WORKBOOK) -> dict[str, XbertWorkbookRow]:
    if not path.exists():
        return {}

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: dict[str, XbertWorkbookRow] = {}

    for sheet in workbook.worksheets:
        header_index: dict[str, int] | None = None

        for row in sheet.iter_rows(values_only=True):
            values = [clean_cell(value) for value in row]

            if header_index is None:
                if "货号" in values:
                    header_index = {header: index for index, header in enumerate(values) if header}
                continue

            sku_index = header_index.get("货号")
            if sku_index is None or sku_index >= len(row):
                continue

            sku = normalize_sku(row[sku_index])
            if not sku:
                continue

            def value(*names: str) -> str:
                for name in names:
                    index = header_index.get(name)
                    if index is not None and index < len(row):
                        return clean_cell(row[index])
                return ""

            rows[sku] = XbertWorkbookRow(
                sku=sku,
                series=value("系列"),
                name=value("品名", "名称"),
                package_size=value("包装规格"),
                finished_size=value("造型规格", "组装尺寸", "产品规格"),
                age=value("年龄段", "适合年龄"),
                notes=value("备注"),
                factory_price=value("出厂价"),
            )

    return rows


def parse_piece_count(notes: str) -> str:
    match = re.search(r"(\d+)\s*块", notes)
    return match.group(1) if match else ""


def safe_ascii_words(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "")
    value = re.sub(r"[/|].*$", "", value).strip()

    for pattern, replacement in [*title_cleanup.SENSITIVE_REPLACEMENTS, *SENSITIVE_REPLACEMENTS]:
        value = pattern.sub(replacement, value)

    value = re.sub(r"[\u3400-\u9fff]+", " ", value)
    value = re.sub(r"[^A-Za-z0-9+&' -]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip(" -.")
    return value


def title_case(value: str) -> str:
    small_words = {"and", "or", "with", "for", "the", "a", "an"}
    words = []

    for index, word in enumerate(value.split()):
        if word.isupper() and len(word) <= 4:
            words.append(word)
        elif index > 0 and word.lower() in small_words:
            words.append(word.lower())
        elif "-" in word:
            words.append("-".join(part[:1].upper() + part[1:].lower() for part in word.split("-") if part))
        else:
            words.append(word[:1].upper() + word[1:].lower())

    return " ".join(words)


def product_name_from_sources(row: XbertWorkbookRow | None, folder_name: str) -> str:
    if row:
        english = safe_ascii_words(row.name)
        if english:
            return english

        for chinese, english_name in NAME_EN_BY_CN.items():
            if chinese in row.name:
                return english_name

        series_name = SERIES_EN_BY_CN.get(row.series, safe_ascii_words(row.series))
        if series_name:
            return f"{series_name} Model"

    sku = extract_sku_from_path(Path(folder_name))
    if sku in LOCAL_NAME_BY_SKU:
        return LOCAL_NAME_BY_SKU[sku]

    descriptor = safe_ascii_words(strip_language_suffix(folder_name))
    descriptor = re.sub(r"^[A-Z]{0,3}\d{5,6}\s*", "", descriptor, flags=re.I).strip()
    return descriptor or "Display Model"


def title_for_product(base: str, row: XbertWorkbookRow | None, folder_name: str) -> str:
    raw_name = product_name_from_sources(row, folder_name)
    product_name = title_case(raw_name)
    title = f"{VENDOR} {product_name} {PRODUCT_TITLE_SUFFIX}"
    title = re.sub(r"\s+", " ", title).strip()

    if re.search(r"[\u3400-\u9fff]", title):
        title = f"{VENDOR} Display Model {PRODUCT_TITLE_SUFFIX} {base}"

    return title


def variant_option_name(sku: str, row: XbertWorkbookRow | None) -> str:
    name = product_name_from_sources(row, sku)
    return f"{sku} - {title_case(name)}"


def product_type_for_row(row: XbertWorkbookRow | None) -> str:
    if not row:
        return ""

    if row.series == "摆件系列" and re.search(r"\bvault\s*33\b|避难所", row.name, re.I):
        return "Movie & Game"

    return SERIES_EN_BY_CN.get(row.series, safe_ascii_words(row.series))


def image_number(path: Path) -> int:
    match = re.search(r"(?:^|-)(\d+)(?:-[^.]+)?\.\w+$", path.name)
    return int(match.group(1)) if match else 9999


def image_dimensions(path: Path) -> tuple[int, int]:
    try:
        output = subprocess.check_output(
            ["sips", "-g", "pixelWidth", "-g", "pixelHeight", str(path)],
            text=True,
            stderr=subprocess.DEVNULL,
        )
    except Exception:  # noqa: BLE001 - dimensions are a best-effort classification aid.
        return 0, 0

    width = re.search(r"pixelWidth:\s*(\d+)", output)
    height = re.search(r"pixelHeight:\s*(\d+)", output)
    return (int(width.group(1)), int(height.group(1))) if width and height else (0, 0)


def is_detail_like_image(path: Path) -> bool:
    width, height = image_dimensions(path)
    return bool(width and height and height / width >= 2.4)


def is_ignored_file(path: Path) -> bool:
    return path.name.startswith("._") or path.name.lower() in IGNORED_FILE_NAMES


def image_files(folder: Path) -> list[Path]:
    return sorted(
        [
            path
            for path in folder.iterdir()
            if path.is_file() and path.suffix.lower() in IMAGE_EXTS and not is_ignored_file(path)
        ],
        key=lambda path: path.name.lower(),
    )


def image_buckets(folder: Path, base: str) -> dict[str, list[Path]]:
    files = image_files(folder)
    white = [path for path in files if "白底" in path.stem]

    def is_numbered_product_image(path: Path) -> bool:
        name = path.name
        return bool(
            re.search(rf"^{re.escape(base)}-\d+(?:-[^.]+)?\.", name, re.I)
            or re.search(r"^\d+(?:-[^.]+)?\.", name, re.I)
        )

    numbered_candidates = [
        path
        for path in files
        if is_numbered_product_image(path)
        and "详情" not in path.stem
        and "尺寸" not in path.stem
        and "包装" not in path.stem
        and "白底" not in path.stem
        and "透明" not in path.stem
        and "sku" not in path.stem.lower()
    ]
    detail_like_numbered = [path for path in numbered_candidates if is_detail_like_image(path)]
    numbered = [path for path in numbered_candidates if path not in detail_like_numbered]
    sku = [path for path in files if re.search(r"(?:^|-)sku(?:-[^.]+)?\.", path.name, re.I) or "尺寸" in path.stem]
    detail = [*detail_like_numbered, *[path for path in files if "详情" in path.stem]]
    transparent = [path for path in files if "透明" in path.stem]

    return {
        "white": sorted(white, key=lambda path: path.name.lower()),
        "numbered": sorted(numbered, key=lambda path: (image_number(path), path.name.lower())),
        "sku": sorted(sku, key=lambda path: path.name.lower()),
        "detail": sorted(detail, key=lambda path: path.name.lower()),
        "transparent": sorted(transparent, key=lambda path: path.name.lower()),
    }


def image_dir_score(folder: Path) -> tuple[int, str]:
    files = image_files(folder)
    parts = folder.parts
    part_text = "/".join(parts)
    score = len(files)

    if folder.name.endswith("英文"):
        score += 50
    if folder.name.endswith("中文"):
        score -= 50
    if "中文" in part_text and "英文" not in folder.name:
        score -= 20
    if "3比4" in part_text or "切片" in part_text:
        score -= 30

    score += 8 if any("-白底." in path.name for path in files) else 0
    score += 5 if any("-详情" in path.name for path in files) else 0
    score += 5 if any("-sku" in path.name.lower() for path in files) else 0
    return score, folder.as_posix()


def product_image_dirs(root: Path = ROOT) -> dict[str, Path]:
    candidates: dict[str, list[Path]] = {}

    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in IMAGE_EXTS or is_ignored_file(path):
            continue

        sku = extract_sku_from_path(path.parent)
        if not sku:
            continue

        candidates.setdefault(sku, []).append(path.parent)

    selected = {}
    for sku, folders in candidates.items():
        unique_folders = sorted(set(folders), key=lambda folder: image_dir_score(folder), reverse=True)
        selected[sku] = unique_folders[0]

    return dict(sorted(selected.items()))


def build_supplement_rows(workbook_rows: dict[str, XbertWorkbookRow], local_skus: set[str]) -> list[dict[str, str]]:
    rows = []

    for sku in sorted(set(workbook_rows) - local_skus):
        row = workbook_rows[sku]
        rows.append(
            {
                "sku": sku,
                "reason": "workbook_row_without_local_images",
                "brick4_search_url": f"https://brick4.com/search/set/?s={sku}",
                "series": row.series,
                "name": row.name,
                "title": title_for_product(sku, row, sku),
            }
        )

    return rows


def build_manifest(
    root: Path = ROOT,
    workbook_rows: dict[str, XbertWorkbookRow] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, str]]]:
    rows = workbook_rows if workbook_rows is not None else load_workbook_rows()
    folders_by_sku = product_image_dirs(root)
    manifest: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for sku, folder in folders_by_sku.items():
        base = sku
        row = rows.get(sku)
        images = image_buckets(folder, base)
        main_media = images["white"][:1] + images["numbered"]

        if not main_media:
            skipped.append({"folder": str(folder), "sku": sku, "reason": "missing_main_media"})
            continue

        title = title_for_product(base, row, folder.name)
        handle = base_import.slugify(f"xbert-{base}-{title}")
        piece_count = parse_piece_count(row.notes) if row else ""
        product_type = product_type_for_row(row)

        manifest.append(
            {
                "folder": folder.name,
                "folder_path": str(folder),
                "base": base,
                "handle": handle,
                "title": title,
                "vendor": VENDOR,
                "status": STATUS,
                "product_type": product_type or PRODUCT_TYPE_DEFAULT,
                "price": PRICE,
                "variants": [
                    {
                        "sku": sku,
                        "option_name": variant_option_name(sku, row),
                        "title_source": row.name if row else "",
                        "series": row.series if row else "",
                        "age": row.age if row else "",
                        "piece_count": piece_count,
                        "package_size": row.package_size if row else "",
                        "finished_size": row.finished_size if row else "",
                    }
                ],
                "metafields": {
                    "specs.piece_count": piece_count,
                    "specs.recommended_age": row.age if row else "",
                    "specs.finished_model_size": row.finished_size if row else "",
                    "specs.package_size": row.package_size if row else "",
                    "specs.difficulty_level": "See product package",
                    "custom.series": product_type or "",
                },
                "main_media": [str(path) for path in main_media],
                "sku_images": [str(path) for path in images["sku"]],
                "detail_images": [str(path) for path in images["detail"]],
                "transparent_images": [str(path) for path in images["transparent"]],
                "missing": {
                    "white": not bool(images["white"]),
                    "detail": not bool(images["detail"]),
                    "sku_images": not bool(images["sku"]),
                    "workbook_rows": row is None,
                    "needs_brick4_info": row is None,
                },
            }
        )

    supplements = build_supplement_rows(rows, set(folders_by_sku))
    return manifest, skipped, supplements


def write_manifest(
    manifest: list[dict[str, Any]],
    skipped: list[dict[str, Any]],
    supplements: list[dict[str, str]],
    name: str = "xbert-pending",
) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / f"{name}-manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / f"{name}-skipped.json").write_text(json.dumps(skipped, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / f"{name}-brick4-supplement.json").write_text(
        json.dumps(supplements, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    with (OUT_DIR / f"{name}-manifest.csv").open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(
            [
                "folder",
                "handle",
                "title",
                "vendor",
                "status",
                "price",
                "variant_skus",
                "main_media_count",
                "sku_image_count",
                "detail_count",
                "missing",
            ]
        )

        for item in manifest:
            writer.writerow(
                [
                    item["folder"],
                    item["handle"],
                    item["title"],
                    item["vendor"],
                    item["status"],
                    item["price"],
                    ", ".join(variant["sku"] for variant in item["variants"]),
                    len(item["main_media"]),
                    len(item["sku_images"]),
                    len(item["detail_images"]),
                    json.dumps(item["missing"], ensure_ascii=False),
                ]
            )

    with (OUT_DIR / f"{name}-brick4-supplement.csv").open("w", encoding="utf-8", newline="") as file:
        fieldnames = ["sku", "reason", "brick4_search_url", "series", "name", "title"]
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(supplements)


def filter_existing(admin: ShopifyAdmin, manifest: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    existing_handles, existing_skus = admin.products_index()
    todo = []
    skipped = []

    for item in manifest:
        item_skus = {variant["sku"].upper() for variant in item["variants"]}
        matched_skus = sorted(item_skus & existing_skus)

        if item["handle"] in existing_handles:
            skipped.append({"folder": item["folder"], "handle": item["handle"], "reason": "existing_handle"})
        elif matched_skus:
            skipped.append({"folder": item["folder"], "handle": item["handle"], "reason": "existing_sku", "skus": matched_skus})
        else:
            todo.append(item)

    return todo, skipped


def upload_detail_images_for_item(admin: ShopifyAdmin, item: dict[str, Any]) -> list[str]:
    urls = []

    for detail_index, source in enumerate(item.get("detail_images", []), start=1):
        for part_index, path in enumerate(base_import.detail_image_paths(Path(source)), start=1):
            part = f" part {part_index}" if part_index > 1 else ""
            urls.append(admin.file_create(path, f"{item['title']} details {detail_index}{part}"))

    return urls


def description_html(item: dict[str, Any], detail_urls: list[str]) -> str:
    return "\n".join(
        f'<p><img src="{html.escape(url)}" alt="{html.escape(item["title"])} details part {index}" /></p>'
        for index, url in enumerate(detail_urls, start=1)
    )


def create_batch(manifest: list[dict[str, Any]], offset: int, batch_size: int, report_name: str) -> list[dict[str, Any]]:
    admin = ShopifyAdmin()
    scopes = admin.access_scopes()
    missing_scopes = sorted(REQUIRED_SCOPES - scopes)

    if missing_scopes:
        raise RuntimeError(f"Missing Shopify scopes: {', '.join(missing_scopes)}")

    publication_ids: list[str] = []
    if PUBLICATION_SCOPES <= scopes:
        publication_ids = [publication["id"] for publication in admin.publications()]

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    batch = manifest[offset : offset + batch_size]
    results = []

    for index, item in enumerate(batch, start=1):
        result = {"manifest": item, "ok": False}

        try:
            existing = admin.product_by_handle(item["handle"])
            if existing:
                verified = repair_existing_product(admin, item, existing["id"], publication_ids)
                result.update({"ok": True, "existing": True, "product": verified})
                results.append(result)
                continue

            detail_urls = upload_detail_images_for_item(admin, item)
            product = admin.product_set(item, description_html(item, detail_urls))
            product_id = product["id"]
            verified = base_import.sync_product_assets(admin, item, product_id, update_description=False)
            admin.update_status_and_category(product_id)
            admin.publish_to_publications(product_id, publication_ids)
            verified = admin.fetch_product(product_id)
            result.update({"ok": True, "existing": False, "product": verified})
        except Exception as error:  # noqa: BLE001 - batch uploads should continue and report failures.
            result["error"] = str(error)
            recovered = recover_partial_product(admin, item, publication_ids)

            if recovered:
                result.update({"ok": True, "recovered": True, "product": recovered})

        results.append(result)
        (OUT_DIR / f"{report_name}.json").write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Processed {offset + index}: {item['folder']} ok={result['ok']}", flush=True)

    return results


def run_auto(root: Path, workbook_rows: dict[str, XbertWorkbookRow], batch_size: int) -> dict[str, Any]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    runs = []
    started_at = time.strftime("%Y%m%d-%H%M%S")

    while True:
        manifest, local_skipped, supplements = build_manifest(root=root, workbook_rows=workbook_rows)
        write_manifest(manifest, local_skipped, supplements, name="xbert-source")
        admin = ShopifyAdmin()
        todo, existing_skipped = filter_existing(admin, manifest)
        write_manifest(todo, local_skipped + existing_skipped, supplements)

        remaining = len(todo)
        print(
            json.dumps(
                {
                    "remaining": remaining,
                    "skipped": len(local_skipped) + len(existing_skipped),
                    "brick4_supplements": len(supplements),
                    "missing_detail": sum(1 for item in todo if item["missing"]["detail"]),
                    "missing_workbook_rows": sum(1 for item in todo if item["missing"]["workbook_rows"]),
                },
                ensure_ascii=False,
            ),
            flush=True,
        )

        if remaining == 0:
            break

        report_name = f"xbert-auto-{started_at}-batch-{len(runs) + 1:03d}-size-{batch_size}"
        created = create_batch(todo, 0, batch_size, report_name)
        run = {
            "batch": len(runs) + 1,
            "report": str(OUT_DIR / f"{report_name}.json"),
            "processed": len(created),
            "ok": sum(1 for row in created if row.get("ok")),
            "failed": sum(1 for row in created if not row.get("ok")),
            "failed_folders": [row["manifest"]["folder"] for row in created if not row.get("ok")],
        }
        runs.append(run)
        (OUT_DIR / f"auto-{started_at}-summary.json").write_text(
            json.dumps({"runs": runs}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(json.dumps(run, ensure_ascii=False), flush=True)

        if run["failed"] == len(created):
            print("All products in this batch failed; stopping to avoid a tight retry loop.", flush=True)
            break

    summary = {"started_at": started_at, "runs": runs}
    (OUT_DIR / f"auto-{started_at}-summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def audit_xbert_products(
    root: Path = ROOT,
    workbook_rows: dict[str, XbertWorkbookRow] | None = None,
    admin: ShopifyAdmin | None = None,
) -> dict[str, Any]:
    admin = admin or ShopifyAdmin()
    _handles, skus = admin.products_index()
    manifest, local_skipped, supplements = build_manifest(root=root, workbook_rows=workbook_rows)
    todo, existing_skipped = filter_existing(admin, manifest)
    existing_xbert = [item for item in manifest if item not in todo]
    issues = []

    for item in manifest:
        if item["vendor"] != VENDOR:
            issues.append({"folder": item["folder"], "issue": "manifest_vendor_not_xbert", "value": item["vendor"]})
        if not item["title"].startswith(f"{VENDOR} "):
            issues.append({"folder": item["folder"], "issue": "manifest_title_not_xbert", "value": item["title"]})
        if item["price"] != PRICE:
            issues.append({"folder": item["folder"], "issue": "manifest_price_not_999", "value": item["price"]})

    return {
        "source_products": len(manifest),
        "shopify_skus_seen": len(skus),
        "todo_products": len(todo),
        "existing_or_skipped": len(existing_skipped),
        "local_skipped": len(local_skipped),
        "brick4_supplements": len(supplements),
        "manifest_issues": issues,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Xbert/Zhuanyue product folders to Shopify.")
    parser.add_argument("--dry-run", action="store_true", help="Generate manifest and skip existing Shopify products.")
    parser.add_argument("--create-batch", action="store_true", help="Create one batch of products.")
    parser.add_argument("--auto", action="store_true", help="Create repeated batches until no pending products remain.")
    parser.add_argument("--audit", action="store_true", help="Run a read-only Xbert manifest/Shopify overlap audit.")
    parser.add_argument("--source-root", type=Path, default=ROOT, help="Only scan this Xbert source directory.")
    parser.add_argument("--workbook", type=Path, default=WORKBOOK, help="Xbert workbook to use for titles/specs.")
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--batch-size", type=int, default=10)
    args = parser.parse_args()

    if not (args.dry_run or args.create_batch or args.auto or args.audit):
        parser.error("Choose --dry-run, --create-batch, --auto, or --audit")

    if args.batch_size < 1 or args.batch_size > 25:
        parser.error("--batch-size must be between 1 and 25")

    workbook_rows = load_workbook_rows(args.workbook)

    if args.auto:
        print(json.dumps(run_auto(args.source_root, workbook_rows, args.batch_size), ensure_ascii=False, indent=2))
        return 0

    manifest, local_skipped, supplements = build_manifest(root=args.source_root, workbook_rows=workbook_rows)
    write_manifest(manifest, local_skipped, supplements, name="xbert-source")
    admin = ShopifyAdmin()
    todo, existing_skipped = filter_existing(admin, manifest)
    write_manifest(todo, local_skipped + existing_skipped, supplements)

    summary = {
        "source_manifest_json": str(OUT_DIR / "xbert-source-manifest.json"),
        "manifest_json": str(OUT_DIR / "xbert-pending-manifest.json"),
        "skipped_json": str(OUT_DIR / "xbert-pending-skipped.json"),
        "brick4_supplement_json": str(OUT_DIR / "xbert-pending-brick4-supplement.json"),
        "source_products": len(manifest),
        "todo_products": len(todo),
        "skipped": len(local_skipped) + len(existing_skipped),
        "brick4_supplements": len(supplements),
        "missing_detail": sum(1 for item in todo if item["missing"]["detail"]),
        "missing_workbook_rows": sum(1 for item in todo if item["missing"]["workbook_rows"]),
        "vendor_not_xbert": sum(1 for item in todo if item["vendor"] != VENDOR),
        "titles_not_xbert": sum(1 for item in todo if not item["title"].startswith(f"{VENDOR} ")),
        "price_not_999": sum(1 for item in todo if item["price"] != PRICE),
    }

    if args.audit:
        summary["audit"] = audit_xbert_products(root=args.source_root, workbook_rows=workbook_rows, admin=admin)

    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.create_batch:
        report_name = f"xbert-batch-{time.strftime('%Y%m%d-%H%M%S')}-offset-{args.offset}-size-{args.batch_size}"
        created = create_batch(todo, args.offset, args.batch_size, report_name)
        print(
            json.dumps(
                {
                    "processed": len(created),
                    "ok": sum(1 for row in created if row.get("ok")),
                    "failed": sum(1 for row in created if not row.get("ok")),
                    "report": str(OUT_DIR / f"{report_name}.json"),
                    "next_offset": args.offset + args.batch_size,
                },
                ensure_ascii=False,
                indent=2,
            )
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
