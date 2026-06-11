#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image, UnidentifiedImageError


TARGET_ROOT = Path("/Volumes/ORICO/iblock/iblock-上架前整理")
SOURCE_PRODUCT_ROOT = TARGET_ROOT / "shopify-products"
UPLOAD_READY_ROOT = TARGET_ROOT / "shopify-products-upload-ready"
REPORTS_ROOT = TARGET_ROOT / "reports"
WORKBOOK_PATH = TARGET_ROOT / "iblock-shopify-readiness.xlsx"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
DETAIL_MIN_HEIGHT = 80


def classify_role(filename: str) -> str:
    lower = filename.lower()
    if "-白底" in filename:
        return "white"
    if "-sku" in lower:
        return "sku"
    if "-详情" in filename:
        return "detail"
    return "main"


def sha1_file(path: Path) -> str:
    digest = hashlib.sha1()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def image_size(path: Path) -> tuple[int, int]:
    try:
        with Image.open(path) as image:
            return image.width, image.height
    except UnidentifiedImageError:
        return 0, 0


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def remove_appledouble_files(root: Path) -> int:
    removed = 0
    if not root.exists():
        return removed
    for path in root.rglob("._*"):
        if path.is_file():
            path.unlink(missing_ok=True)
            removed += 1
    return removed


def skip_reason(path: Path, role: str) -> str:
    if path.stat().st_size == 0:
        return "empty_or_unreadable_image"
    width, height = image_size(path)
    if not width or not height:
        return "empty_or_unreadable_image"
    if role == "detail" and height < DETAIL_MIN_HEIGHT:
        return f"extremely_short_detail_image_{width}x{height}"
    return ""


def group_image_files(group_folder: Path) -> list[Path]:
    image_folder = group_folder / "images"
    if not image_folder.exists():
        return []
    return sorted(
        [
            path
            for path in image_folder.iterdir()
            if path.is_file() and path.suffix.lower() in IMAGE_EXTS and not path.name.startswith("._")
        ],
        key=lambda path: path.name,
    )


def build_upload_ready_pack(
    source_root: Path = SOURCE_PRODUCT_ROOT, target_root: Path = UPLOAD_READY_ROOT
) -> tuple[list[dict[str, Any]], list[dict[str, str]], list[dict[str, str]]]:
    if target_root.exists():
        shutil.rmtree(target_root, ignore_errors=True)
    target_root.mkdir(parents=True, exist_ok=True)

    summaries: list[dict[str, Any]] = []
    copied_assets: list[dict[str, str]] = []
    skipped_assets: list[dict[str, str]] = []

    for group_folder in sorted([path for path in source_root.iterdir() if path.is_dir()]):
        group = group_folder.name
        target_group = target_root / group
        target_images = target_group / "images"
        if (group_folder / "variants.csv").exists():
            target_group.mkdir(parents=True, exist_ok=True)
            shutil.copy2(group_folder / "variants.csv", target_group / "variants.csv")

        seen_role_hashes: set[tuple[str, str]] = set()
        role_counts = Counter()
        kept_counts = Counter()
        duplicate_skipped = 0
        invalid_skipped = 0
        input_images = group_image_files(group_folder)
        seen_any_hashes: set[str] = set()

        for source in input_images:
            role = classify_role(source.name)
            reason = skip_reason(source, role)
            if reason:
                skipped_assets.append(
                    {
                        "upload_group": group,
                        "role": role,
                        "source": source.as_posix(),
                        "skip_reason": reason,
                    }
                )
                invalid_skipped += 1
                continue

            digest = sha1_file(source)
            if role == "detail" and digest in seen_any_hashes:
                skipped_assets.append(
                    {
                        "upload_group": group,
                        "role": role,
                        "source": source.as_posix(),
                        "skip_reason": "exact_duplicate_detail_matches_existing_image",
                    }
                )
                duplicate_skipped += 1
                continue
            if role in {"main", "detail"}:
                role_hash = (role, digest)
                if role_hash in seen_role_hashes:
                    skipped_assets.append(
                        {
                            "upload_group": group,
                            "role": role,
                            "source": source.as_posix(),
                            "skip_reason": "exact_duplicate_main_or_detail",
                        }
                    )
                    duplicate_skipped += 1
                    continue
                seen_role_hashes.add(role_hash)

            target = target_images / source.name
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
            copied_assets.append(
                {
                    "upload_group": group,
                    "role": role,
                    "source": source.as_posix(),
                    "target": target.as_posix(),
                }
            )
            seen_any_hashes.add(digest)
            role_counts[role] += 1
            kept_counts[role] += 1

        summaries.append(
            {
                "upload_group": group,
                "input_image_count": len(input_images),
                "kept_image_count": sum(kept_counts.values()),
                "kept_white_count": kept_counts["white"],
                "kept_main_count": kept_counts["main"],
                "kept_sku_count": kept_counts["sku"],
                "kept_detail_count": kept_counts["detail"],
                "duplicate_skipped_count": duplicate_skipped,
                "invalid_skipped_count": invalid_skipped,
                "upload_ready_folder": target_group.as_posix(),
            }
        )
    return summaries, copied_assets, skipped_assets


def update_workbook(summary_rows: list[dict[str, Any]], asset_rows: list[dict[str, str]], skipped_rows: list[dict[str, str]]) -> None:
    if not WORKBOOK_PATH.exists():
        return
    workbook = load_workbook(WORKBOOK_PATH)
    for name in ("UploadReadySummary", "UploadReadyAssets", "UploadReadySkipped"):
        if name in workbook.sheetnames:
            del workbook[name]
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for name, rows in (
        ("UploadReadySummary", summary_rows),
        ("UploadReadyAssets", asset_rows),
        ("UploadReadySkipped", skipped_rows),
    ):
        sheet = workbook.create_sheet(name)
        if not rows:
            continue
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 70)
    workbook.save(WORKBOOK_PATH)


def run(update_xlsx: bool = True) -> dict[str, Any]:
    summary_rows, asset_rows, skipped_rows = build_upload_ready_pack()
    summary_csv = REPORTS_ROOT / "iblock-upload-ready-summary.csv"
    asset_csv = REPORTS_ROOT / "iblock-upload-ready-assets.csv"
    skipped_csv = REPORTS_ROOT / "iblock-upload-ready-skipped.csv"
    summary_json = REPORTS_ROOT / "iblock-upload-ready-summary.json"
    write_csv(summary_csv, summary_rows)
    write_csv(asset_csv, asset_rows)
    write_csv(skipped_csv, skipped_rows)
    if update_xlsx:
        update_workbook(summary_rows, asset_rows, skipped_rows)
    appledouble_removed = remove_appledouble_files(TARGET_ROOT)
    result = {
        "product_group_count": len(summary_rows),
        "input_image_count": sum(int(row["input_image_count"]) for row in summary_rows),
        "kept_image_count": sum(int(row["kept_image_count"]) for row in summary_rows),
        "duplicate_skipped_count": sum(int(row["duplicate_skipped_count"]) for row in summary_rows),
        "invalid_skipped_count": sum(int(row["invalid_skipped_count"]) for row in summary_rows),
        "appledouble_removed_count": appledouble_removed,
        "upload_ready_root": UPLOAD_READY_ROOT.as_posix(),
        "summary_csv": summary_csv.as_posix(),
        "asset_csv": asset_csv.as_posix(),
        "skipped_csv": skipped_csv.as_posix(),
    }
    summary_json.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    remove_appledouble_files(TARGET_ROOT)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Build de-duplicated iBlock Shopify upload image folders.")
    parser.add_argument("--no-xlsx", action="store_true", help="Do not update the readiness workbook.")
    args = parser.parse_args()
    run(update_xlsx=not args.no_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
