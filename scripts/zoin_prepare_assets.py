#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


QUOTE_FILE = Path("/Volumes/ORICO/积域资料/集域产品报价2026.3.17(1).xlsx")
SOURCE_ROOT = Path("/Volumes/ORICO/积域资料/积域-产品资料.rar/积域-产品资料/集域产品图")
TARGET_ROOT = Path("/Volumes/ORICO/积域资料/Zoin-上架前整理")
REPORT_DIR = Path("/private/tmp/zoin-product-prep")
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
IGNORE_NAMES = {".DS_Store", "Thumbs.db", ".WeDrive"}
SKU_RE = re.compile(r"(?:GT|GK|FT)\s*[-_ ]?\s*\d{3}", re.I)


@dataclass(frozen=True)
class ProductRow:
    sku: str
    series: str
    name_cn: str
    package_type: str
    package_size: str
    model_size: str
    carton_size: str
    carton_qty: str
    volume: str
    gross_net_weight: str
    recommended_age: str
    factory_price_cny: str
    minifigure: str
    certifications: str
    notes: str


@dataclass(frozen=True)
class ImageEntry:
    source: Path
    role: str
    skus: tuple[str, ...]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_sku(value: str) -> str:
    value = unicodedata.normalize("NFKC", value).upper()
    value = re.sub(r"[^A-Z0-9]", "", value)
    return value


def parse_quote_rows(path: Path) -> list[ProductRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    products: list[ProductRow] = []

    for row in sheet.iter_rows(min_row=3, values_only=True):
        sku = normalize_sku(normalize_text(row[4] if len(row) > 4 else ""))
        if not sku:
            continue
        products.append(
            ProductRow(
                sku=sku,
                series=normalize_text(row[3] if len(row) > 3 else ""),
                name_cn=normalize_text(row[5] if len(row) > 5 else ""),
                package_type=normalize_text(row[6] if len(row) > 6 else ""),
                package_size=normalize_text(row[7] if len(row) > 7 else ""),
                model_size=normalize_text(row[8] if len(row) > 8 else ""),
                carton_size=normalize_text(row[9] if len(row) > 9 else ""),
                carton_qty=normalize_text(row[10] if len(row) > 10 else ""),
                volume=normalize_text(row[11] if len(row) > 11 else ""),
                gross_net_weight=normalize_text(row[12] if len(row) > 12 else ""),
                recommended_age=normalize_text(row[13] if len(row) > 13 else ""),
                factory_price_cny=normalize_text(row[14] if len(row) > 14 else ""),
                minifigure=normalize_text(row[15] if len(row) > 15 else ""),
                certifications=normalize_text(row[16] if len(row) > 16 else ""),
                notes=normalize_text(row[17] if len(row) > 17 else ""),
            )
        )
    return products


def expand_sku_segment(text: str) -> list[str]:
    normalized = unicodedata.normalize("NFKC", text).upper()
    output: list[str] = []

    for hit in SKU_RE.finditer(normalized):
        first = normalize_sku(hit.group(0))
        if first not in output:
            output.append(first)

        tail = normalized[hit.end() : hit.end() + 40]
        prefix_match = re.match(r"([A-Z]+)(\d+)", first)
        if not prefix_match:
            continue
        prefix, digits = prefix_match.groups()
        range_match = re.match(r"\s*[-~至]\s*(?:(GT|GK|FT)\s*)?(\d{3})", tail, re.I)
        if range_match:
            end_prefix = (range_match.group(1) or prefix).upper()
            end_number = int(range_match.group(2))
            start_number = int(digits)
            if end_prefix == prefix and start_number <= end_number and end_number - start_number <= 100:
                for number in range(start_number + 1, end_number + 1):
                    code = f"{prefix}{number:03d}"
                    if code not in output:
                        output.append(code)

        for bare in re.finditer(r"[、,，/]\s*(\d{3})", tail):
            code = f"{prefix}{bare.group(1)}"
            if code not in output:
                output.append(code)

    return output


def extract_skus_from_path(path: Path, known_skus: set[str]) -> list[str]:
    text = path.as_posix()
    skus: list[str] = []
    for code in expand_sku_segment(text):
        if code in known_skus and code not in skus:
            skus.append(code)
    return skus


def classify_role(path: Path) -> str:
    text = path.as_posix()
    if "白底" in text:
        return "white"
    if "详情" in text or "切片" in text:
        return "detail"
    if "商品主图" in text or "主图" in path.name:
        return "main"
    if "收单图" in text:
        return "main"
    if "透明" in text:
        return "transparent"
    if "彩盒" in text or "包装" in text or "说明书" in text:
        return "package"
    if "/产品/" in text or "\\产品\\" in text:
        return "product"
    return "other"


def image_files(root: Path) -> list[Path]:
    return sorted(
        [
            path
            for path in root.rglob("*")
            if path.is_file()
            and path.suffix.lower() in IMAGE_EXTS
            and not path.name.startswith("._")
            and path.name not in IGNORE_NAMES
            and ".Temp" not in path.parts
        ],
        key=lambda path: path.as_posix().lower(),
    )


def collect_images(root: Path, known_skus: set[str]) -> tuple[list[ImageEntry], list[Path]]:
    entries: list[ImageEntry] = []
    unmatched: list[Path] = []

    for path in image_files(root):
        skus = extract_skus_from_path(path.relative_to(root), known_skus)
        if not skus:
            unmatched.append(path)
            continue
        entries.append(ImageEntry(source=path, role=classify_role(path.relative_to(root)), skus=tuple(skus)))

    return entries, unmatched


def natural_key(path: Path) -> list[Any]:
    parts = re.split(r"(\d+)", path.as_posix().lower())
    return [int(part) if part.isdigit() else part for part in parts]


def useful_media(entries: list[ImageEntry], sku: str, role: str) -> list[Path]:
    paths = [entry.source for entry in entries if entry.role == role and sku in entry.skus]
    return sorted(dict.fromkeys(paths), key=natural_key)


def fallback_main_media(entries: list[ImageEntry], sku: str) -> list[Path]:
    paths: list[Path] = []
    for role in ("product", "transparent", "package", "other"):
        paths.extend(entry.source for entry in entries if entry.role == role and sku in entry.skus)
    return sorted(dict.fromkeys(paths), key=natural_key)


def image_target_names(sku: str, white: list[Path], main: list[Path], detail: list[Path]) -> list[tuple[Path, str, str]]:
    copies: list[tuple[Path, str, str]] = []

    for index, path in enumerate(white, start=1):
        suffix = path.suffix.lower()
        name = f"{sku}-白底{suffix}" if index == 1 else f"{sku}-白底-{index:02d}{suffix}"
        copies.append((path, name, "white"))

    for index, path in enumerate(main, start=1):
        copies.append((path, f"{sku}-{index}{path.suffix.lower()}", "main"))

    if len(detail) == 1:
        copies.append((detail[0], f"{sku}-详情{detail[0].suffix.lower()}", "detail"))
    else:
        for index, path in enumerate(detail, start=1):
            copies.append((path, f"{sku}-详情-{index:02d}{path.suffix.lower()}", "detail"))

    return copies


def copy_assets(rows: list[ProductRow], entries: list[ImageEntry], target_root: Path, apply: bool) -> list[dict[str, str]]:
    copy_rows: list[dict[str, str]] = []
    for product in rows:
        sku = product.sku
        white = useful_media(entries, sku, "white")
        main = useful_media(entries, sku, "main")
        if not main:
            main = fallback_main_media(entries, sku)
        detail = useful_media(entries, sku, "detail")
        sku_dir = target_root / "images" / sku

        for source, target_name, role in image_target_names(sku, white, main, detail):
            target = sku_dir / target_name
            copy_rows.append(
                {
                    "sku": sku,
                    "role": role,
                    "source": source.as_posix(),
                    "target": target.as_posix(),
                    "applied": "yes" if apply else "no",
                }
            )
            if apply:
                target.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source, target)

    return copy_rows


def product_manifest(rows: list[ProductRow], entries: list[ImageEntry]) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for product in rows:
        sku = product.sku
        white = useful_media(entries, sku, "white")
        main = useful_media(entries, sku, "main")
        fallback_main = fallback_main_media(entries, sku)
        detail = useful_media(entries, sku, "detail")
        image_count = sum(1 for entry in entries if sku in entry.skus)
        output.append(
            {
                "sku": sku,
                "vendor": "Zoin",
                "series": product.series,
                "name_cn": product.name_cn,
                "shopify_title_draft": f"Zoin {product.name_cn} Building Block Set",
                "product_type": product.series or "Building Block Sets",
                "category": "Interlocking Blocks",
                "price_rule": "999",
                "package_type": product.package_type,
                "package_size": product.package_size,
                "finished_model_size": "" if product.model_size == "/" else product.model_size,
                "carton_size": product.carton_size,
                "carton_qty": product.carton_qty,
                "volume": product.volume,
                "gross_net_weight": product.gross_net_weight,
                "recommended_age": product.recommended_age,
                "factory_price_cny": product.factory_price_cny,
                "has_minifigure": product.minifigure,
                "certifications": product.certifications,
                "notes": product.notes,
                "piece_count": "",
                "piece_count_source_needed": "Brick4",
                "local_image_count": str(image_count),
                "white_image_count": str(len(white)),
                "main_image_count": str(len(main) or len(fallback_main)),
                "detail_image_count": str(len(detail)),
                "missing_white_image": "yes" if not white else "no",
                "missing_main_image": "yes" if not main and not fallback_main else "no",
                "missing_detail_image": "yes" if not detail else "no",
                "needs_brick4_image": "yes" if not white or not detail else "no",
                "needs_english_title_review": "yes",
            }
        )
    return output


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, sheets: dict[str, list[dict[str, str]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, rows in sheets.items():
        sheet = workbook.create_sheet(sheet_name[:31])
        if not rows:
            continue
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)

    workbook.save(path)


def mirror_reports_to_target(report_paths: list[Path], target_root: Path) -> dict[str, str]:
    target_reports = target_root / "reports"
    target_reports.mkdir(parents=True, exist_ok=True)
    copied: dict[str, str] = {}
    for source in report_paths:
        target = target_reports / source.name
        shutil.copy2(source, target)
        copied[source.name] = target.as_posix()
    return copied


def build_reports(
    quote_file: Path,
    source_root: Path,
    target_root: Path,
    report_dir: Path,
    apply: bool,
) -> dict[str, Any]:
    rows = parse_quote_rows(quote_file)
    known_skus = {row.sku for row in rows}
    entries, unmatched = collect_images(source_root, known_skus)
    manifest = product_manifest(rows, entries)
    copy_rows = copy_assets(rows, entries, target_root, apply=apply)

    missing_rows = [
        row
        for row in manifest
        if row["missing_white_image"] == "yes"
        or row["missing_main_image"] == "yes"
        or row["missing_detail_image"] == "yes"
        or row["piece_count_source_needed"] == "Brick4"
    ]
    unmatched_rows = [{"source": path.as_posix()} for path in unmatched]

    report_dir.mkdir(parents=True, exist_ok=True)
    if apply:
        target_root.mkdir(parents=True, exist_ok=True)

    manifest_csv = report_dir / "zoin-products-manifest.csv"
    missing_csv = report_dir / "zoin-missing-assets.csv"
    copy_plan_csv = report_dir / "zoin-image-copy-plan.csv"
    unmatched_csv = report_dir / "zoin-unmatched-images.csv"
    workbook_xlsx = report_dir / "zoin-products-metafields-ready.xlsx"
    summary_json = report_dir / "zoin-prep-summary.json"

    write_csv(manifest_csv, manifest)
    write_csv(missing_csv, missing_rows)
    write_csv(copy_plan_csv, copy_rows)
    write_csv(unmatched_csv, unmatched_rows)
    write_xlsx(
        workbook_xlsx,
        {
            "Products": manifest,
            "MissingAssets": missing_rows,
            "ImageCopyPlan": copy_rows,
        },
    )

    role_counts = Counter(entry.role for entry in entries)
    copied_role_counts = Counter(row["role"] for row in copy_rows)
    summary = {
        "quote_file": quote_file.as_posix(),
        "source_root": source_root.as_posix(),
        "target_root": target_root.as_posix(),
        "apply": apply,
        "quote_sku_count": len(rows),
        "matched_image_sku_count": len({sku for entry in entries for sku in entry.skus}),
        "image_entry_count": len(entries),
        "unmatched_image_count": len(unmatched),
        "source_role_counts": dict(sorted(role_counts.items())),
        "planned_copy_count": len(copy_rows),
        "planned_copy_role_counts": dict(sorted(copied_role_counts.items())),
        "missing_white_sku_count": sum(1 for row in manifest if row["missing_white_image"] == "yes"),
        "missing_main_sku_count": sum(1 for row in manifest if row["missing_main_image"] == "yes"),
        "missing_detail_sku_count": sum(1 for row in manifest if row["missing_detail_image"] == "yes"),
        "missing_piece_count_sku_count": len(rows),
        "reports": {
            "manifest_csv": manifest_csv.as_posix(),
            "missing_csv": missing_csv.as_posix(),
            "copy_plan_csv": copy_plan_csv.as_posix(),
            "unmatched_csv": unmatched_csv.as_posix(),
            "workbook_xlsx": workbook_xlsx.as_posix(),
            "summary_json": summary_json.as_posix(),
        },
    }
    if apply:
        summary["target_reports"] = mirror_reports_to_target(
            [manifest_csv, missing_csv, copy_plan_csv, unmatched_csv, workbook_xlsx],
            target_root,
        )
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    if apply:
        target_summary = target_root / "reports" / summary_json.name
        target_summary.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare Zoin quote data and staged Shopify image assets.")
    parser.add_argument("--quote-file", type=Path, default=QUOTE_FILE)
    parser.add_argument("--source-root", type=Path, default=SOURCE_ROOT)
    parser.add_argument("--target-root", type=Path, default=TARGET_ROOT)
    parser.add_argument("--report-dir", type=Path, default=REPORT_DIR)
    parser.add_argument("--apply", action="store_true", help="Copy normalized assets into the target root.")
    args = parser.parse_args()

    summary = build_reports(
        quote_file=args.quote_file,
        source_root=args.source_root,
        target_root=args.target_root,
        report_dir=args.report_dir,
        apply=args.apply,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
