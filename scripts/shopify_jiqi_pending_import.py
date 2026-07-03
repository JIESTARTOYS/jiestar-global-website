#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import shopify_sample_import as base_import
from shopify_cn_pending_import import ShopifyAdmin as BaseShopifyAdmin


QUOTE_XLSX = Path("/Volumes/ORICO/积琪积木/积琪报价表-26.6.3 32款(3).xlsx")
IMAGE_ROOT = Path("/Volumes/ORICO/积琪积木/积琪电商图整理")
METADATA_XLSX = Path("/Volumes/ORICO/积琪积木/JIQI产品元字段资料表.xlsx")
OUT_DIR = Path("/private/tmp/jiestar-shopify-jiqi-import")
DEFAULT_OUT_DIR = OUT_DIR

VENDOR = "JIQI"
STATUS = "ACTIVE"
PRICE = "999"
CATEGORY_ID = "gid://shopify/TaxonomyCategory/tg-5-7-12"
CATEGORY_NAME = "Interlocking Blocks"
OPTION_NAME = "Model"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
SOURCE_IMAGE_EXTS = IMAGE_EXTS | {".gif"}
REQUIRED_SCOPES = {"read_products", "write_products", "read_files", "write_files"}
PUBLICATION_SCOPES = {"read_publications", "write_publications"}
DETAIL_SLICE_MAX_HEIGHT = 4000
DETAIL_SLICE_MIN_BYTES = 5 * 1024 * 1024
DETAIL_SLICE_QUALITY = 86
MAX_STAGED_UPLOAD_BYTES = 15 * 1024 * 1024
UPLOAD_IMAGE_QUALITY_STEPS = (86, 80, 74, 68)

IP_RISK_SKUS = {
    "JQ1103": "Colt firearm brand reference",
    "JQ1105": "anime character reference",
    "JQ1108": "game IP reference",
    "JQ1109": "game IP reference",
    "JQ1111": "Mario game IP reference",
    "JQ1112": "Mario game IP reference",
    "JQ1113": "film IP reference",
    "JQ1122": "film IP reference",
    "JQ1144": "anime IP reference",
    "JQ1145": "anime IP reference",
    "JQ1146": "anime IP reference",
    "JQ1147": "anime IP reference",
}

SOURCE_FIELDS = [
    "sheet",
    "sku",
    "original_sku_cell",
    "original_name_cn",
    "factory_price",
    "carton_qty",
    "package_size",
    "outer_carton_size",
    "gross_net_weight",
    "piece_count",
    "finished_model_size",
    "recommended_age",
    "power",
]

JULY_2_TITLE_OVERRIDES = {
    "JQ1167": {
        "title": "JIQI Mechanical Horse Display Model Kit JQ1167",
        "descriptor": "Mechanical Horse Display",
        "product_type": "Animal",
        "series": "Mechanical Animals",
    },
    "JQ1168": {
        "title": "JIQI Mechanical Snail Display Model Kit JQ1168",
        "descriptor": "Mechanical Snail Display",
        "product_type": "Animal",
        "series": "Mechanical Animals",
    },
    "JQ1150": {
        "title": "JIQI Moon Base Space Building Set JQ1150",
        "descriptor": "Moon Base Space",
        "product_type": "Space",
        "series": "Space",
    },
    "JQ1152": {
        "title": "JIQI Mechanical Phoenix Display Model Kit JQ1152",
        "descriptor": "Mechanical Phoenix Display",
        "product_type": "Animal",
        "series": "Mechanical Animals",
    },
    "JQ1153": {
        "title": "JIQI Deep Space Starry Sky Wall Art Building Set JQ1153",
        "descriptor": "Deep Space Starry Sky Wall Art",
        "product_type": "Wall Art",
        "series": "Space Wall Art",
    },
}


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_header(value: Any) -> str:
    return clean(value).replace(" ", "").replace("\n", "")


def contains_cjk(value: str) -> bool:
    return bool(re.search(r"[\u3400-\u9fff]", value))


def same_money(left: Any, right: Any) -> bool:
    try:
        return Decimal(str(left)) == Decimal(str(right))
    except (InvalidOperation, ValueError):
        return False


def money(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def money_str(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value:.2f}"


def natural_key(value: str | Path) -> tuple[Any, ...]:
    text = str(value).lower()
    parts = re.split(r"(\d+)", text)
    return tuple(int(part) if part.isdigit() else part for part in parts)


def slugify(value: str) -> str:
    return base_import.slugify(value)


def piece_count_value(value: str) -> str:
    match = re.search(r"\d+", clean(value))
    return match.group(0) if match else ""


def sku_from_cell(value: Any) -> str:
    match = re.search(r"\bJQ\s*[-_]?\s*(\d{4})\b", str(value or ""), flags=re.I)
    if match:
        return f"JQ{match.group(1)}"
    return ""


def read_quote_rows(path: Path = QUOTE_XLSX) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    seen: set[str] = set()

    for sheet in workbook.worksheets:
        header_row: int | None = None
        headers: list[str] = []
        for index, cells in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            header_candidates = [normalize_header(cell) for cell in cells]
            if "货号" in header_candidates:
                header_row = index
                headers = header_candidates
                break

        if not header_row:
            continue

        sku_index = headers.index("货号")
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if sku_index >= len(values):
                continue

            sku = sku_from_cell(values[sku_index]).upper()
            if not sku or sku in seen:
                continue
            seen.add(sku)

            row_by_header = {
                header: clean(values[index]) if index < len(values) else ""
                for index, header in enumerate(headers)
                if header
            }
            rows.append(
                {
                    "sheet": sheet.title,
                    "sku": sku,
                    "original_sku_cell": clean(values[sku_index]),
                    "original_name_cn": row_by_header.get("品名", ""),
                    "factory_price": row_by_header.get("出厂价", ""),
                    "carton_qty": row_by_header.get("装箱数", ""),
                    "package_size": row_by_header.get("彩盒规格", ""),
                    "outer_carton_size": row_by_header.get("外箱规格", ""),
                    "gross_net_weight": row_by_header.get("整件重量/KG", ""),
                    "piece_count": row_by_header.get("颗粒数/pcs", "") or row_by_header.get("颗粒数/PCS", ""),
                    "finished_model_size": row_by_header.get("实物尺寸/CM", ""),
                    "recommended_age": row_by_header.get("年龄", ""),
                    "power": row_by_header.get("是否带电", ""),
                }
            )

    return rows


def row_dicts(sheet: Any) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [clean(cell) for cell in next(rows)]
    except StopIteration:
        return []
    output: list[dict[str, Any]] = []
    for row in rows:
        if not row or not any(cell not in (None, "") for cell in row):
            continue
        output.append({headers[index]: row[index] if index < len(row) else None for index in range(len(headers)) if headers[index]})
    return output


def load_initial_pricing_rows(path: Path | None) -> dict[str, dict[str, str]]:
    if not path:
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Shopify导入价格" not in workbook.sheetnames:
        raise RuntimeError(f"Missing Shopify导入价格 sheet: {path}")

    c_rows = row_dicts(workbook["C端_公开售价"]) if "C端_公开售价" in workbook.sheetnames else []
    c_by_sku = {clean(row.get("SKU")).upper(): row for row in c_rows if clean(row.get("SKU"))}
    rows: dict[str, dict[str, str]] = {}

    for row in row_dicts(workbook["Shopify导入价格"]):
        sku = clean(row.get("Variant SKU")).upper()
        if not sku:
            continue
        price = money(row.get("Variant Price"))
        compare_at = money(row.get("Variant Compare At Price"))
        if price is None:
            continue
        if compare_at is not None and compare_at <= price:
            compare_at = None
        c_row = c_by_sku.get(sku, {})
        rows[sku] = {
            "price": money_str(price),
            "compare_at_price": money_str(compare_at),
            "pricing_status": clean(c_row.get("控价状态")),
            "pricing_source_file": path.name,
        }
    return rows


def all_image_files(root: Path = IMAGE_ROOT) -> list[Path]:
    files: list[Path] = []
    if not root.exists():
        return files
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if path.name.startswith("._") or path.name.lower() in {"thumbs.db", ".ds_store"}:
            continue
        if any(part in {"__MACOSX", "源文件"} for part in path.parts):
            continue
        if path.suffix.lower() not in SOURCE_IMAGE_EXTS:
            continue
        files.append(path)
    return sorted(files, key=natural_key)


def folder_skus(folder_name: str) -> list[str]:
    matches = [match.upper() for match in re.findall(r"JQ\d{4}", folder_name, flags=re.I)]
    if len(matches) == 2 and "-" in folder_name:
        start = int(matches[0][2:])
        end = int(matches[1][2:])
        if start <= end and end - start < 20:
            return [f"JQ{number}" for number in range(start, end + 1)]
    return matches


def folder_for_sku(root: Path, sku: str) -> Path | None:
    if not root.exists():
        return None
    for folder in sorted((path for path in root.iterdir() if path.is_dir()), key=natural_key):
        if sku in folder_skus(folder.name):
            return folder
    return None


def files_for_sku(root: Path, sku: str) -> list[Path]:
    folder = folder_for_sku(root, sku)
    if not folder:
        return []
    return sorted(
        [
            path
            for path in folder.rglob("*")
            if path.is_file()
            and not path.name.startswith("._")
            and path.suffix.lower() in SOURCE_IMAGE_EXTS
            and not any(part in {"__MACOSX", "源文件"} for part in path.parts)
        ],
        key=natural_key,
    )


def path_text(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def has_part(path: Path, root: Path, term: str) -> bool:
    try:
        parts = path.relative_to(root).parts
    except ValueError:
        parts = path.parts
    return any(term in part for part in parts)


def has_dir_part(path: Path, root: Path, term: str) -> bool:
    try:
        parts = path.relative_to(root).parts[:-1]
    except ValueError:
        parts = path.parts[:-1]
    return any(term in part for part in parts)


def media_for_sku(root: Path, sku: str) -> dict[str, list[Path]]:
    files = files_for_sku(root, sku)
    uploadable = [path for path in files if path.suffix.lower() in IMAGE_EXTS]

    def in_main_folder(path: Path) -> bool:
        return has_dir_part(path, root, "主图")

    def in_white_folder(path: Path) -> bool:
        return has_dir_part(path, root, "白底")

    def in_detail_folder(path: Path) -> bool:
        return has_dir_part(path, root, "详情")

    details = [
        path
        for path in uploadable
        if in_detail_folder(path)
        or (("详情" in path.name or "长图" in path.name) and not in_main_folder(path) and not in_white_folder(path))
    ]
    details = sorted(details, key=natural_key)

    whites = [
        path
        for path in uploadable
        if in_white_folder(path)
        or "白底" in path_text(path, root)
        or "透明底" in path.name
        or "透明" in path_text(path, root)
        or "尺寸" in path_text(path, root)
    ]
    whites = sorted(
        whites,
        key=lambda path: (
            0 if "白底" in path.name else 1,
            0 if path.suffix.lower() in {".jpg", ".jpeg"} else 1,
            natural_key(path),
        ),
    )

    sku_images = sorted([path for path in uploadable if "sku" in path.name.lower()], key=natural_key)
    size_images = sorted([path for path in uploadable if "尺寸" in path_text(path, root)], key=natural_key)

    main = [
        path
        for path in uploadable
        if path not in details
        and path not in whites
        and path not in sku_images
        and path not in size_images
        and (in_main_folder(path) or re.search(r"(?:^|[-_ ])\(?\d+\)?\.", path.name))
    ]
    main = sorted(main, key=natural_key)

    fallback_sku = sku_images[:1] or whites[:1]
    gif_details = [
        path
        for path in files
        if path.suffix.lower() == ".gif" and (has_dir_part(path, root, "详情") or "详情" in path.name)
    ]
    return {
        "white": whites[:1],
        "main": main,
        "sku": fallback_sku,
        "detail": details,
        "gif_detail": sorted(gif_details, key=natural_key),
        "size": size_images,
        "all": files,
    }


def upload_ready_image_path(path: Path, sku: str, role: str) -> Path:
    if path.stat().st_size <= MAX_STAGED_UPLOAD_BYTES:
        return path

    try:
        from PIL import Image
    except ImportError:
        return path

    output_dir = OUT_DIR / "upload-images" / sku
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{path.stem}-{role}-upload.jpg"

    with Image.open(path) as image:
        if image.mode != "RGB":
            image = image.convert("RGB")

        width, height = image.size
        max_dimension = max(width, height)
        if max_dimension > 2400:
            scale = 2400 / max_dimension
            image = image.resize((round(width * scale), round(height * scale)))

        for quality in UPLOAD_IMAGE_QUALITY_STEPS:
            image.save(output_path, format="JPEG", quality=quality, optimize=True)
            if output_path.stat().st_size <= MAX_STAGED_UPLOAD_BYTES:
                return output_path

        width, height = image.size
        image = image.resize((round(width * 0.85), round(height * 0.85)))
        image.save(output_path, format="JPEG", quality=68, optimize=True)

    return output_path


def product_type_and_series(name: str) -> tuple[str, str]:
    override = JULY_2_TITLE_OVERRIDES.get(sku_from_cell(name).upper())
    if override:
        return override["product_type"], override["series"]
    if any(term in name for term in ["独角兽", "大熊", "鱼缸", "机械鹿"]):
        return "Animal", "Display Animals"
    if any(term in name for term in ["机械马", "机械蜗牛", "机械凤凰"]):
        return "Animal", "Mechanical Animals"
    if any(term in name for term in ["星空画"]):
        return "Wall Art", "Space Wall Art"
    if any(term in name for term in ["月球基地"]):
        return "Space", "Space"
    if any(term in name for term in ["宇航员", "星舰", "钛战斗机"]):
        return "Space", "Space"
    if any(term in name for term in ["龙墟"]):
        return "Legendary Dragon", "Fantasy Display"
    if any(term in name for term in ["巴黎铁塔", "悉尼歌剧院"]):
        return "Street View", "Landmarks"
    if any(term in name for term in ["相机"]):
        return "Home Appliance", "Retro Camera"
    if any(term in name for term in ["机甲"]):
        return "Mecha", "Mecha"
    if any(term in name for term in ["宝剑"]):
        return "Weapon", "Display Weapon"
    return "Ornament", "Display Model"


def descriptor_for_name(name: str) -> str:
    if "独角兽" in name:
        return "Mirror Unicorn Display"
    if "宇航员" in name:
        return "Astronaut Display"
    if "半机械大熊" in name:
        return "Cyber Bear Display"
    if "龙墟" in name:
        return "Dragon Ruins Fantasy"
    if "梦幻鱼缸" in name:
        return "Fantasy Aquarium"
    if "黑色礼花" in name:
        return "Black Firework Display"
    if "星舰火箭" in name:
        return "Starship Rocket"
    if "巴黎铁塔" in name:
        return "Landmark Tower"
    if "悉尼歌剧院" in name:
        return "Opera House Landmark"
    if "星核强袭" in name:
        return "Star Core Mecha"
    if "机械鹿" in name:
        return "Mechanical Deer"
    if "悬浮大宝剑" in name:
        return "Floating Sword Display"
    if "悬浮钛战斗机" in name:
        return "Floating Space Fighter"
    if "宝丽来" in name:
        return "Retro Instant Camera"
    if "宝莱克斯" in name:
        return "Retro Film Camera"
    if "奥林巴斯" in name:
        return "Retro SLR Camera"
    if "禄来" in name:
        return "Retro Twin-Lens Camera"
    if "粉扎机甲" in name:
        return "Pink Mecha"
    if "天狼座机甲" in name:
        return "Sirius Mecha"
    if "机械马" in name:
        return "Mechanical Horse Display"
    if "机械蜗牛" in name:
        return "Mechanical Snail Display"
    if "月球基地" in name:
        return "Moon Base Space"
    if "机械凤凰" in name:
        return "Mechanical Phoenix Display"
    if "深邃星空画" in name:
        return "Deep Space Starry Sky Wall Art"
    return "Display Model"


def title_for_row(row: dict[str, str]) -> str:
    sku = row["sku"]
    override = JULY_2_TITLE_OVERRIDES.get(sku.upper())
    if override:
        return override["title"]
    descriptor = descriptor_for_name(row.get("original_name_cn", ""))
    product_type, _series = product_type_and_series(row.get("original_name_cn", ""))
    suffix = "Model Kit" if product_type in {"Mecha", "Weapon", "Space"} else "Building Set"
    return re.sub(r"\s+", " ", f"{VENDOR} {descriptor} {suffix} {sku}").strip()


def metafields_for_row(row: dict[str, str], product_type: str, series: str) -> dict[str, str]:
    metafields = {
        "specs.piece_count": piece_count_value(row.get("piece_count", "")),
        "specs.recommended_age": clean(row.get("recommended_age", "")),
        "specs.finished_model_size": clean(row.get("finished_model_size", "")),
        "specs.package_size": clean(row.get("package_size", "")),
        "specs.difficulty_level": "See product package",
        "custom.series": series or product_type,
    }
    return {key: value for key, value in metafields.items() if value}


def manifest_item(row: dict[str, str], image_root: Path, pricing: dict[str, str] | None = None) -> tuple[dict[str, Any], list[str]]:
    sku = row["sku"].upper()
    pricing = pricing or {}
    media = media_for_sku(image_root, sku)
    override = JULY_2_TITLE_OVERRIDES.get(sku)
    if override:
        product_type, series = override["product_type"], override["series"]
    else:
        product_type, series = product_type_and_series(row.get("original_name_cn", ""))
    title = title_for_row(row)
    handle = slugify(title)
    item_price = pricing.get("price") or PRICE
    compare_at_price = pricing.get("compare_at_price", "")
    issues: list[str] = []

    if sku in IP_RISK_SKUS:
        issues.append("ip_risk_skip")
    if contains_cjk(title):
        issues.append("title_contains_chinese")
    if not media["white"]:
        issues.append("missing_white_image")
    if not media["main"]:
        issues.append("missing_main_image")
    if not media["detail"]:
        issues.append("missing_detail_image")
    if media["gif_detail"]:
        issues.append("gif_detail_requires_conversion")
    if VENDOR != "JIQI":
        issues.append("vendor_mismatch")
    if money(item_price) is None:
        issues.append("invalid_price")
    if compare_at_price and money(compare_at_price) is None:
        issues.append("invalid_compare_at_price")
    if STATUS != "ACTIVE":
        issues.append("status_not_active")

    main_media = media["white"] + media["main"]
    upload_main_media = [
        str(upload_ready_image_path(path, sku, f"main-{index:02d}"))
        for index, path in enumerate(main_media, start=1)
    ]
    upload_sku_images = [
        str(upload_ready_image_path(path, sku, f"sku-{index:02d}"))
        for index, path in enumerate(media["sku"], start=1)
    ]
    item = {
        "folder": folder_for_sku(image_root, sku).name if folder_for_sku(image_root, sku) else sku,
        "folder_path": str(folder_for_sku(image_root, sku) or image_root),
        "base": sku,
        "handle": handle,
        "title": title,
        "vendor": VENDOR,
        "status": STATUS,
        "product_type": product_type,
        "category": CATEGORY_ID,
        "category_name": CATEGORY_NAME,
        "price": item_price,
        "compare_at_price": compare_at_price,
        "pricing_status": pricing.get("pricing_status", ""),
        "pricing_source_file": pricing.get("pricing_source_file", ""),
        "option_name": OPTION_NAME,
        "variants": [
            {
                "sku": sku,
                "option_name": f"{sku} - {descriptor_for_name(row.get('original_name_cn', ''))}",
                "title_source": clean(row.get("original_name_cn")),
                "series": series,
                "age": clean(row.get("recommended_age")),
                "piece_count": piece_count_value(row.get("piece_count", "")),
                "package_size": clean(row.get("package_size")),
                "finished_size": clean(row.get("finished_model_size")),
            }
        ],
        "metafields": metafields_for_row(row, product_type, series),
        "main_media": upload_main_media,
        "sku_images": upload_sku_images,
        "detail_images": [str(path) for path in media["detail"]],
        "transparent_images": [],
        "source_row": row,
        "media_status": {
            "all_count": len(media["all"]),
            "white_count": len(media["white"]),
            "main_count": len(media["main"]),
            "sku_count": len(media["sku"]),
            "detail_count": len(media["detail"]),
            "gif_detail_count": len(media["gif_detail"]),
            "size_count": len(media["size"]),
            "first_media": Path(main_media[0]).name if main_media else "",
        },
        "missing": {
            "white": not bool(media["white"]),
            "main": not bool(media["main"]),
            "detail": not bool(media["detail"]),
            "sku_image_fallback_to_white": not any("sku" in Path(path).name.lower() for path in media["sku"]),
        },
        "ip_risk_reason": IP_RISK_SKUS.get(sku, ""),
    }
    return item, issues


def build_manifest(
    image_root: Path = IMAGE_ROOT,
    quote_rows: list[dict[str, str]] | None = None,
    pricing_rows: dict[str, dict[str, str]] | None = None,
    sku_filter: set[str] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, str]]]:
    rows = quote_rows if quote_rows is not None else read_quote_rows()
    pricing_rows = pricing_rows or {}
    normalized_filter = {sku.upper() for sku in sku_filter} if sku_filter else set()
    if normalized_filter:
        rows = [row for row in rows if row["sku"].upper() in normalized_filter]
    manifest: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    rows_by_sku = {row["sku"].upper(): row for row in rows}

    for row in rows:
        item, issues = manifest_item(row, image_root, pricing_rows.get(row["sku"].upper()))
        has_any_asset = bool(item["media_status"]["all_count"])
        if issues:
            skipped.append(
                {
                    "sku": row["sku"],
                    "handle": item["handle"],
                    "title": item["title"],
                    "sheet": row.get("sheet", ""),
                    "original_name_cn": row.get("original_name_cn", ""),
                    "issues": issues,
                    "media_status": item["media_status"],
                    "has_any_asset": has_any_asset,
                    "ip_risk_reason": item.get("ip_risk_reason", ""),
                }
            )
        else:
            manifest.append(item)

    asset_skus: set[str] = set()
    if image_root.exists():
        for folder in image_root.iterdir():
            if folder.is_dir():
                asset_skus.update(folder_skus(folder.name))
    if normalized_filter:
        asset_skus &= normalized_filter

    for sku in sorted(asset_skus - set(rows_by_sku), key=natural_key):
        skipped.append(
            {
                "sku": sku,
                "handle": "",
                "title": "",
                "sheet": "",
                "original_name_cn": "",
                "issues": ["asset_without_quote_row"],
                "media_status": {"all_count": len(files_for_sku(image_root, sku))},
                "has_any_asset": True,
                "ip_risk_reason": "",
            }
        )

    return manifest, skipped, rows


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def source_row_for_skipped(row: dict[str, Any], quote_rows: list[dict[str, str]]) -> dict[str, str]:
    return next((source for source in quote_rows if source["sku"] == row.get("sku")), {})


def metadata_row(item: dict[str, Any], upload_status: str, issues: str, source: dict[str, str]) -> list[Any]:
    metafields = item.get("metafields", {})
    media_status = item.get("media_status", {})
    return [
        item.get("base", ""),
        item.get("vendor", ""),
        item.get("title", ""),
        item.get("handle", ""),
        item.get("status", ""),
        item.get("price", ""),
        item.get("compare_at_price", ""),
        item.get("pricing_status", ""),
        item.get("product_type", ""),
        item.get("category_name", ""),
        metafields.get("specs.piece_count", ""),
        metafields.get("specs.recommended_age", ""),
        metafields.get("specs.finished_model_size", ""),
        metafields.get("specs.package_size", ""),
        metafields.get("specs.difficulty_level", ""),
        metafields.get("custom.series", ""),
        source.get("sheet", ""),
        source.get("original_sku_cell", ""),
        source.get("original_name_cn", ""),
        source.get("factory_price", ""),
        source.get("carton_qty", ""),
        source.get("outer_carton_size", ""),
        source.get("gross_net_weight", ""),
        source.get("power", ""),
        json.dumps(media_status, ensure_ascii=False),
        upload_status,
        issues,
    ]


def write_metadata_workbook(manifest: list[dict[str, Any]], skipped: list[dict[str, Any]], quote_rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "JIQI Metafields"
    headers = [
        "sku",
        "vendor",
        "shopify_title",
        "handle",
        "status",
        "price",
        "compare_at_price",
        "pricing_status",
        "product_type",
        "category",
        "specs.piece_count",
        "specs.recommended_age",
        "specs.finished_model_size",
        "specs.package_size",
        "specs.difficulty_level",
        "custom.series",
        "source_sheet",
        "original_sku_cell",
        "original_name_cn",
        "factory_price",
        "carton_qty",
        "outer_carton_size",
        "gross_net_weight",
        "power",
        "media_status",
        "upload_status",
        "issues",
    ]
    sheet.append(headers)

    for item in manifest:
        sheet.append(metadata_row(item, "READY", "", item.get("source_row", {})))

    for row in skipped:
        source = source_row_for_skipped(row, quote_rows)
        if source:
            product_type, series = product_type_and_series(source.get("original_name_cn", ""))
            pseudo = {
                "base": row["sku"],
                "vendor": VENDOR,
                "title": row["title"],
                "handle": row["handle"],
                "status": STATUS,
                "price": PRICE,
                "product_type": product_type,
                "category_name": CATEGORY_NAME,
                "metafields": metafields_for_row(source, product_type, series),
                "media_status": row.get("media_status", {}),
            }
        else:
            pseudo = {
                "base": row.get("sku", ""),
                "vendor": VENDOR,
                "title": row.get("title", ""),
                "handle": row.get("handle", ""),
                "status": STATUS,
                "price": PRICE,
                "product_type": "",
                "category_name": CATEGORY_NAME,
                "metafields": {},
                "media_status": row.get("media_status", {}),
            }
        sheet.append(metadata_row(pseudo, "SKIPPED", "; ".join(row.get("issues", [])), source))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_fill = {
        "READY": PatternFill("solid", fgColor="E2F0D9"),
        "SKIPPED": PatternFill("solid", fgColor="FCE4D6"),
    }
    status_col = headers.index("upload_status") + 1
    for row in sheet.iter_rows(min_row=2):
        fill = status_fill.get(str(row[status_col - 1].value or ""))
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    widths = {
        "A": 12,
        "B": 12,
        "C": 48,
        "D": 48,
        "E": 12,
        "F": 10,
        "G": 16,
        "H": 18,
        "I": 20,
        "J": 24,
        "K": 16,
        "L": 18,
        "M": 22,
        "N": 18,
        "O": 24,
        "P": 22,
        "Q": 16,
        "R": 18,
        "S": 32,
        "T": 12,
        "U": 12,
        "V": 20,
        "W": 18,
        "X": 12,
        "Y": 42,
        "Z": 14,
        "AA": 34,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

    METADATA_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(METADATA_XLSX)


def write_reports(manifest: list[dict[str, Any]], skipped: list[dict[str, Any]], quote_rows: list[dict[str, str]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    write_json(OUT_DIR / "jiqi-manifest.json", manifest)
    write_json(OUT_DIR / "jiqi-skipped.json", skipped)
    write_metadata_workbook(manifest, skipped, quote_rows)
    write_csv(
        OUT_DIR / "jiqi-manifest.csv",
        [
            {
                "sku": item["base"],
                "handle": item["handle"],
                "title": item["title"],
                "vendor": item["vendor"],
                "status": item["status"],
                "price": item["price"],
                "compare_at_price": item.get("compare_at_price", ""),
                "pricing_status": item.get("pricing_status", ""),
                "product_type": item["product_type"],
                "main_media_count": len(item["main_media"]),
                "detail_count": len(item["detail_images"]),
                "sku_image_count": len(item["sku_images"]),
                "first_media": Path(item["main_media"][0]).name if item["main_media"] else "",
                "source_sheet": item.get("source_row", {}).get("sheet", ""),
                "original_name_cn": item.get("source_row", {}).get("original_name_cn", ""),
            }
            for item in manifest
        ],
        [
            "sku",
            "handle",
            "title",
            "vendor",
            "status",
            "price",
            "compare_at_price",
            "pricing_status",
            "product_type",
            "main_media_count",
            "detail_count",
            "sku_image_count",
            "first_media",
            "source_sheet",
            "original_name_cn",
        ],
    )
    write_csv(
        OUT_DIR / "jiqi-skipped.csv",
        [
            {
                "sku": row["sku"],
                "handle": row.get("handle", ""),
                "title": row.get("title", ""),
                "source_sheet": row.get("sheet", ""),
                "original_name_cn": row.get("original_name_cn", ""),
                "issues": "; ".join(row.get("issues", [])),
                "ip_risk_reason": row.get("ip_risk_reason", ""),
                "has_any_asset": row.get("has_any_asset", ""),
                "media_status": json.dumps(row.get("media_status", {}), ensure_ascii=False),
            }
            for row in skipped
        ],
        [
            "sku",
            "handle",
            "title",
            "source_sheet",
            "original_name_cn",
            "issues",
            "ip_risk_reason",
            "has_any_asset",
            "media_status",
        ],
    )


def detail_image_parts(path: Path, sku: str) -> list[Path]:
    if path.stat().st_size < DETAIL_SLICE_MIN_BYTES:
        return [path]

    try:
        from PIL import Image
    except ImportError:
        return [path]

    with Image.open(path) as image:
        width, height = image.size
        if height <= DETAIL_SLICE_MAX_HEIGHT:
            return [path]

        output_dir = OUT_DIR / "detail-slices" / sku / path.stem
        output_dir.mkdir(parents=True, exist_ok=True)
        parts: list[Path] = []
        for top in range(0, height, DETAIL_SLICE_MAX_HEIGHT):
            bottom = min(height, top + DETAIL_SLICE_MAX_HEIGHT)
            crop = image.crop((0, top, width, bottom))
            if crop.mode != "RGB":
                crop = crop.convert("RGB")
            output = output_dir / f"{path.stem}-{len(parts) + 1:02d}.jpg"
            crop.save(output, format="JPEG", quality=DETAIL_SLICE_QUALITY, optimize=True)
            parts.append(output)
        return parts


def upload_detail_images_for_item(admin: BaseShopifyAdmin, item: dict[str, Any]) -> list[str]:
    urls: list[str] = []
    for detail_index, detail_path in enumerate(item["detail_images"], start=1):
        path = Path(detail_path)
        for part_index, upload_path in enumerate(detail_image_parts(path, item["base"]), start=1):
            part = f"-{part_index:02d}" if part_index > 1 else ""
            try:
                urls.append(admin.file_create(upload_path, f"{item['title']} detail {detail_index}{part}"))
            except Exception as error:  # noqa: BLE001 - preserve Shopify context and the local file path.
                raise RuntimeError(f"{item['base']} detail upload failed for {upload_path}: {error}") from error
    return urls


def description_html(item: dict[str, Any], detail_urls: list[str]) -> str:
    html_parts: list[str] = []
    for url in detail_urls:
        html_parts.append(f'<p><img src="{html.escape(url)}" alt="{html.escape(item["title"])} detail image"></p>')
    return "\n".join(html_parts)


def remove_generated_intro(description_html_value: str) -> str:
    cleaned = re.sub(r"^\s*<h2\b[^>]*>.*?</h2>\s*", "", description_html_value, count=1, flags=re.I | re.S)
    cleaned = re.sub(r"^\s*<ul\b[^>]*>.*?</ul>\s*", "", cleaned, count=1, flags=re.I | re.S)
    cleaned = re.sub(
        r"\s*<p>\s*Interested in wholesale or custom versions of this product\?\s*Contact us for business cooperation\.\s*</p>\s*",
        "\n",
        cleaned,
        flags=re.I,
    )
    return re.sub(r"\n{3,}", "\n\n", cleaned).strip()


def expected_first_media_alt(item: dict[str, Any]) -> str:
    if not item.get("main_media"):
        return ""
    return f"{item['title']} - {Path(item['main_media'][0]).name}"


class ShopifyAdmin(BaseShopifyAdmin):
    def product_set(self, item: dict[str, Any], description_html_value: str) -> dict[str, Any]:
        variants = [
            dict(
                {
                    "optionValues": [
                        {
                            "optionName": "SKU",
                            "name": variant["option_name"],
                        }
                    ],
                    "price": item["price"],
                    "inventoryItem": {
                        "sku": variant["sku"],
                        "tracked": False,
                    },
                },
                **({"compareAtPrice": item["compare_at_price"]} if item.get("compare_at_price") else {}),
            )
            for variant in item["variants"]
        ]
        data = self.graphql(
            """
            mutation ProductSet($input: ProductSetInput!, $synchronous: Boolean!) {
              productSet(input: $input, synchronous: $synchronous) {
                product {
                  id
                  title
                  handle
                  status
                  vendor
                  variants(first: 250) {
                    nodes {
                      id
                      title
                      price
                      sku
                      inventoryItem {
                        id
                        tracked
                      }
                    }
                  }
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {
                "synchronous": True,
                "input": {
                    "title": item["title"],
                    "handle": item["handle"],
                    "vendor": item["vendor"],
                    "status": item["status"],
                    "productType": item["product_type"],
                    "descriptionHtml": description_html_value,
                    "productOptions": [
                        {
                            "name": "SKU",
                            "values": [{"name": variant["option_name"]} for variant in item["variants"]],
                        }
                    ],
                    "variants": variants,
                    "metafields": base_import.product_metafields(item),
                },
            },
        )
        result = data["productSet"]
        base_import.assert_no_user_errors("productSet", result["userErrors"])
        return result["product"]

    def products_by_sku(self, skus: set[str]) -> list[dict[str, Any]]:
        products: list[dict[str, Any]] = []
        cursor: str | None = None
        query = " OR ".join(f"sku:{sku}" for sku in sorted(skus))

        while True:
            data = self.graphql(
                """
                query ProductsBySku($first: Int!, $after: String, $query: String!) {
                  products(first: $first, after: $after, query: $query) {
                    pageInfo {
                      hasNextPage
                      endCursor
                    }
                    nodes {
                      id
                      handle
                      title
                      vendor
                      status
                      productType
                      descriptionHtml
                      category {
                        id
                        name
                      }
                      media(first: 250, sortKey: POSITION) {
                        nodes {
                          id
                          alt
                          mediaContentType
                        }
                      }
                      metafields(first: 50) {
                        nodes {
                          namespace
                          key
                          value
                        }
                      }
                      variants(first: 250) {
                        nodes {
                          id
                          title
                          price
                          compareAtPrice
                          sku
                          image {
                            id
                            url
                            altText
                          }
                          media(first: 10) {
                            nodes {
                              id
                              alt
                              mediaContentType
                            }
                          }
                        }
                      }
                    }
                  }
                }
                """,
                {"first": 40, "after": cursor, "query": query},
            )
            page = data["products"]
            products.extend(page["nodes"])
            if not page["pageInfo"]["hasNextPage"]:
                break
            cursor = page["pageInfo"]["endCursor"]

        return products


def filter_existing(admin: ShopifyAdmin, manifest: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    existing_handles, existing_skus = admin.products_index()
    todo: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for item in manifest:
        item_skus = {variant["sku"].upper() for variant in item["variants"]}
        matched_skus = sorted(item_skus & existing_skus)
        issues: list[str] = []
        if item["handle"] in existing_handles:
            issues.append("existing_handle")
        if matched_skus:
            issues.append(f"existing_sku:{', '.join(matched_skus)}")
        if issues:
            skipped.append({**item, "issues": issues})
        else:
            todo.append(item)

    return todo, skipped


def check_shopify_conflicts(manifest: list[dict[str, Any]]) -> list[dict[str, Any]]:
    admin = ShopifyAdmin()
    existing_handles, existing_skus = admin.products_index()
    conflicts: list[dict[str, Any]] = []

    for item in manifest:
        item_skus = {variant["sku"].upper() for variant in item["variants"]}
        matched_skus = sorted(item_skus & existing_skus)
        issues: list[str] = []
        if item["handle"] in existing_handles:
            issues.append("existing_handle")
        if matched_skus:
            issues.append(f"existing_sku:{', '.join(matched_skus)}")
        if issues:
            conflicts.append({"sku": item["base"], "handle": item["handle"], "title": item["title"], "issues": issues})

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    write_json(OUT_DIR / "jiqi-shopify-conflicts.json", conflicts)
    write_csv(
        OUT_DIR / "jiqi-shopify-conflicts.csv",
        [{"sku": row["sku"], "handle": row["handle"], "title": row["title"], "issues": "; ".join(row["issues"])} for row in conflicts],
        ["sku", "handle", "title", "issues"],
    )
    return conflicts


def apply_create(manifest: list[dict[str, Any]], report_name: str = "jiqi-created-products") -> list[dict[str, Any]]:
    admin = ShopifyAdmin()
    scopes = admin.access_scopes()
    missing_scopes = sorted(REQUIRED_SCOPES - scopes)
    if missing_scopes:
        raise RuntimeError(f"Missing Shopify scopes: {', '.join(missing_scopes)}")

    publication_ids: list[str] = []
    if PUBLICATION_SCOPES <= scopes:
        publication_ids = [publication["id"] for publication in admin.publications()]

    todo, existing_skipped = filter_existing(admin, manifest)
    results: list[dict[str, Any]] = [
        {"manifest": row, "ok": False, "skipped": True, "error": "; ".join(row["issues"])} for row in existing_skipped
    ]
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for index, item in enumerate(todo, start=1):
        result: dict[str, Any] = {"manifest": item, "ok": False}
        try:
            detail_urls = upload_detail_images_for_item(admin, item)
            product = admin.product_set(item, description_html(item, detail_urls))
            product_id = product["id"]
            base_import.sync_product_assets(admin, item, product_id, update_description=False)
            admin.update_status_and_category(product_id)
            admin.publish_to_publications(product_id, publication_ids)
            result.update({"ok": True, "product": admin.fetch_product(product_id)})
        except Exception as error:  # noqa: BLE001 - keep batch moving and report every SKU.
            result["error"] = str(error)

        results.append(result)
        write_json(OUT_DIR / f"{report_name}.json", results)
        print(f"Processed {index}/{len(todo)} {item['base']} ok={result['ok']}", flush=True)
        time.sleep(0.5)

    return results


def verify_products(manifest: list[dict[str, Any]]) -> list[dict[str, Any]]:
    admin = ShopifyAdmin()
    expected_by_sku = {item["base"]: item for item in manifest}
    products = admin.products_by_sku(set(expected_by_sku))
    rows: list[dict[str, Any]] = []

    for sku, item in expected_by_sku.items():
        product = next(
            (
                candidate
                for candidate in products
                if any((variant.get("sku") or "").strip().upper() == sku for variant in candidate["variants"]["nodes"])
            ),
            None,
        )
        if not product:
            rows.append({"sku": sku, "ok": False, "issues": ["missing_product"]})
            continue

        variant = next(
            variant for variant in product["variants"]["nodes"] if (variant.get("sku") or "").strip().upper() == sku
        )
        metafields = {f"{node['namespace']}.{node['key']}": node["value"] for node in product["metafields"]["nodes"]}
        first_alt = (product["media"]["nodes"][0].get("alt") or "") if product["media"]["nodes"] else ""
        issues: list[str] = []
        checks = {
            "title": product["title"] == item["title"],
            "vendor": product["vendor"] == VENDOR,
            "status": product["status"] == STATUS,
            "price": same_money(variant.get("price"), item.get("price") or PRICE),
            "compare_at_price": (
                same_money(variant.get("compareAtPrice"), item["compare_at_price"])
                if item.get("compare_at_price")
                else not variant.get("compareAtPrice")
            ),
            "product_type": product["productType"] == item["product_type"],
            "category": (product.get("category") or {}).get("id") == CATEGORY_ID,
            "description_has_images": "<img" in (product.get("descriptionHtml") or ""),
            "first_media_matches_manifest": first_alt == expected_first_media_alt(item),
            "variant_has_media": bool((variant.get("media") or {}).get("nodes") or variant.get("image")),
        }
        checks["metafields"] = all(str(metafields.get(key, "")) == str(value) for key, value in item["metafields"].items())
        for key, ok in checks.items():
            if not ok:
                issues.append(key)
        rows.append(
            {
                "sku": sku,
                "product_id": product["id"],
                "handle": product["handle"],
                "title": product["title"],
                "ok": not issues,
                "issues": issues,
                "checks": checks,
                "media_count": len(product["media"]["nodes"]),
                "detail_image_count": len(item["detail_images"]),
            }
        )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    write_json(OUT_DIR / "jiqi-verify.json", rows)
    write_csv(
        OUT_DIR / "jiqi-verify.csv",
        [
            {
                "sku": row["sku"],
                "title": row.get("title", ""),
                "handle": row.get("handle", ""),
                "ok": row["ok"],
                "issues": "; ".join(row.get("issues", [])),
                "media_count": row.get("media_count", ""),
                "detail_image_count": row.get("detail_image_count", ""),
            }
            for row in rows
        ],
        ["sku", "title", "handle", "ok", "issues", "media_count", "detail_image_count"],
    )
    return rows


def main() -> int:
    global METADATA_XLSX, OUT_DIR

    parser = argparse.ArgumentParser(description="Prepare and upload JIQI products to Shopify.")
    parser.add_argument("--quote-xlsx", type=Path, default=QUOTE_XLSX, help="Quote workbook to read. Defaults to the historical JIQI quote workbook.")
    parser.add_argument("--source-root", type=Path, default=IMAGE_ROOT, help="Image root to scan. Defaults to the historical JIQI image root.")
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR, help="Directory for generated JSON/CSV reports and temporary upload assets.")
    parser.add_argument("--metadata-xlsx-output", type=Path, help="Optional metadata workbook output path.")
    parser.add_argument("--pricing-file", type=Path, help="Optional pricing workbook used for initial product price and compare-at price.")
    parser.add_argument("--sku", action="append", default=[], help="Limit manifest generation to this SKU. Can be passed more than once.")
    parser.add_argument("--dry-run", action="store_true", help="Generate manifest, skipped report, and metadata workbook only.")
    parser.add_argument("--check-shopify", action="store_true", help="Read-only Shopify SKU/handle conflict check.")
    parser.add_argument("--apply", action="store_true", help="Create ready JIQI products in Shopify.")
    parser.add_argument("--verify", action="store_true", help="Verify created JIQI products in Shopify.")
    parser.add_argument(
        "--include-skipped-exit-zero",
        action="store_true",
        help="Return 0 even if skipped rows exist. Useful when skipped IP-risk rows are expected.",
    )
    args = parser.parse_args()

    if not (args.dry_run or args.check_shopify or args.apply or args.verify):
        parser.error("Choose at least one of --dry-run, --check-shopify, --apply, or --verify")

    OUT_DIR = args.out_dir
    if args.metadata_xlsx_output:
        METADATA_XLSX = args.metadata_xlsx_output
    elif args.out_dir != DEFAULT_OUT_DIR:
        METADATA_XLSX = args.out_dir / "JIQI产品元字段资料表.xlsx"

    quote_rows = read_quote_rows(args.quote_xlsx)
    pricing_rows = load_initial_pricing_rows(args.pricing_file)
    sku_filter = {clean(sku).upper() for sku in args.sku if clean(sku)}
    manifest, skipped, quote_rows = build_manifest(args.source_root, quote_rows, pricing_rows, sku_filter or None)
    write_reports(manifest, skipped, quote_rows)

    skipped_with_assets = [
        row
        for row in skipped
        if row.get("has_any_asset") and "asset_without_quote_row" not in row.get("issues", [])
    ]
    skipped_without_assets = [
        row
        for row in skipped
        if not row.get("has_any_asset") and "asset_without_quote_row" not in row.get("issues", [])
    ]
    asset_without_quote = [row for row in skipped if "asset_without_quote_row" in row.get("issues", [])]
    ip_risk_skipped = [row for row in skipped if "ip_risk_skip" in row.get("issues", [])]

    summary: dict[str, Any] = {
        "quote_xlsx": str(QUOTE_XLSX),
        "effective_quote_xlsx": str(args.quote_xlsx),
        "image_root": str(args.source_root),
        "pricing_file": str(args.pricing_file) if args.pricing_file else "",
        "sku_filter": sorted(sku_filter),
        "metadata_xlsx": str(METADATA_XLSX),
        "quote_rows": len(quote_rows),
        "manifest_count": len(manifest),
        "skipped_count": len(skipped),
        "ip_risk_skipped": len(ip_risk_skipped),
        "skipped_with_assets": len(skipped_with_assets),
        "skipped_without_assets": len(skipped_without_assets),
        "asset_without_quote_count": len(asset_without_quote),
        "out_dir": str(OUT_DIR),
    }

    if args.apply:
        results = apply_create(manifest)
        summary["created_ok"] = sum(1 for result in results if result.get("ok") and not result.get("skipped"))
        summary["created_failed"] = sum(1 for result in results if not result.get("ok") and not result.get("skipped"))
        summary["existing_skipped"] = sum(1 for result in results if result.get("skipped"))

    if args.check_shopify:
        conflicts = check_shopify_conflicts(manifest)
        summary["shopify_conflicts"] = len(conflicts)

    if args.verify:
        verify_rows = verify_products(manifest)
        summary["verify_ok"] = sum(1 for row in verify_rows if row.get("ok"))
        summary["verify_failed"] = sum(1 for row in verify_rows if not row.get("ok"))

    write_json(OUT_DIR / "jiqi-summary.json", summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if skipped and not args.include_skipped_exit_zero:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
