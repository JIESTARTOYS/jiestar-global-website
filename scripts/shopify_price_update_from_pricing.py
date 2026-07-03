#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


OUT_DIR = Path("/private/tmp/jiestar-shopify-price-update")
API_VERSION_FALLBACK = "2026-01"
DEFAULT_PRICING_DIR = Path("/Users/chensen/jiestar/定价参考")
PRICING_GLOB = "*_保守费率_控价保护_C端B端代理_不含运费_20260630.xlsx"

FULL_REPORT = "price-update-full-report.csv"
ELIGIBLE_REPORT = "price-update-eligible.csv"
BLOCKED_REPORT = "price-update-blocked-review.csv"
UNMATCHED_REPORT = "price-update-unmatched-active.csv"
DUPLICATE_REPORT = "price-update-duplicate-pricing.csv"
DUPLICATE_SOURCE_REPORT = "price-update-duplicate-pricing-source.csv"
SUMMARY_JSON = "price-update-summary.json"
APPLY_RESULT_JSON = "price-update-apply-result.json"
PASS_STATUS = "PASS"
MISSING_MAP_REVIEW_STATUS = "REVIEW: 缺品牌/平台控价"
PASS_UPDATE_REASON = "eligible_to_update"
MISSING_MAP_UPDATE_REASON = "eligible_missing_map_original_profit"


@dataclass(frozen=True)
class PricingRow:
    sku: str
    sku_key: str
    brand: str
    title: str
    status: str
    target_price: Decimal
    target_compare_at: Decimal | None
    source_file: str


@dataclass(frozen=True)
class ShopifyVariant:
    product_id: str
    product_handle: str
    product_title: str
    product_status: str
    vendor: str
    online_store_url: str
    variant_id: str
    variant_title: str
    sku: str
    sku_key: str
    current_price: Decimal | None
    current_compare_at: Decimal | None


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def money(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def money_str(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value:.2f}"


def sku_key(value: str) -> str:
    return (value or "").strip().upper()


def row_dicts(sheet: Any) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        return []
    output: list[dict[str, Any]] = []
    for row in rows:
        if not row or not any(cell not in (None, "") for cell in row):
            continue
        output.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers)) if headers[idx]})
    return output


def load_pricing_rows_from_files(files: list[Path]) -> dict[str, list[PricingRow]]:
    pricing_by_sku: dict[str, list[PricingRow]] = defaultdict(list)
    if not files:
        raise RuntimeError("No pricing workbooks provided")

    for path in sorted(files):
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "C端_公开售价" not in workbook.sheetnames:
            raise RuntimeError(f"Missing C端_公开售价 sheet: {path}")
        if "Shopify导入价格" not in workbook.sheetnames:
            raise RuntimeError(f"Missing Shopify导入价格 sheet: {path}")

        c_rows = row_dicts(workbook["C端_公开售价"])
        shopify_rows = row_dicts(workbook["Shopify导入价格"])
        shopify_by_sku = {
            sku_key(str(row.get("Variant SKU") or "")): row
            for row in shopify_rows
            if sku_key(str(row.get("Variant SKU") or ""))
        }

        for row in c_rows:
            sku = str(row.get("SKU") or "").strip()
            key = sku_key(sku)
            if not key:
                continue
            shopify_row = shopify_by_sku.get(key, {})
            target_price = money(shopify_row.get("Variant Price")) or money(row.get("最终公开价USD_不含运"))
            target_compare_at = money(shopify_row.get("Variant Compare At Price")) or money(row.get("最终CompareAtPriceUSD"))
            if target_price is None:
                continue
            if target_compare_at is not None and target_compare_at <= target_price:
                target_compare_at = None
            pricing_by_sku[key].append(
                PricingRow(
                    sku=sku,
                    sku_key=key,
                    brand=str(row.get("品牌") or ""),
                    title=str(shopify_row.get("Title") or row.get("Shopify标题") or row.get("品名") or ""),
                    status=str(row.get("控价状态") or ""),
                    target_price=target_price,
                    target_compare_at=target_compare_at,
                    source_file=path.name,
                )
            )

    return pricing_by_sku


def load_pricing_rows(pricing_dir: Path) -> dict[str, list[PricingRow]]:
    files = sorted(pricing_dir.glob(PRICING_GLOB))
    if not files:
        raise RuntimeError(f"No pricing workbooks found in {pricing_dir} with glob {PRICING_GLOB}")
    return load_pricing_rows_from_files(files)


class ShopifyAdmin:
    def __init__(self) -> None:
        load_dotenv(Path(".env.local"))
        self.domain = os.environ.get("SHOPIFY_STORE_DOMAIN", "").strip()
        self.version = os.environ.get("SHOPIFY_API_VERSION", API_VERSION_FALLBACK).strip() or API_VERSION_FALLBACK
        self.token = os.environ.get("SHOPIFY_ADMIN_ACCESS_TOKEN", "").strip()
        if not self.domain:
            raise RuntimeError("Missing SHOPIFY_STORE_DOMAIN in .env.local")
        if not self.token:
            raise RuntimeError("Missing SHOPIFY_ADMIN_ACCESS_TOKEN in .env.local")
        self.endpoint = f"https://{self.domain}/admin/api/{self.version}/graphql.json"

    def graphql(self, query: str, variables: dict[str, Any] | None = None) -> dict[str, Any]:
        body = json.dumps({"query": query, "variables": variables or {}}).encode("utf-8")
        for attempt in range(6):
            request = urllib.request.Request(
                self.endpoint,
                data=body,
                method="POST",
                headers={"Content-Type": "application/json", "X-Shopify-Access-Token": self.token},
            )
            try:
                with urllib.request.urlopen(request, timeout=120) as response:
                    payload = json.loads(response.read().decode("utf-8"))
            except urllib.error.HTTPError as error:
                error_body = error.read().decode("utf-8", errors="ignore")
                if error.code in {429, 500, 502, 503, 504} and attempt < 5:
                    time.sleep(min(30, 2**attempt))
                    continue
                raise RuntimeError(f"Shopify HTTP {error.code}: {error_body[:1200]}") from error

            errors = payload.get("errors")
            if errors:
                error_text = json.dumps(errors, ensure_ascii=False)
                retryable = any(
                    str(error.get("extensions", {}).get("code", "")).upper() == "THROTTLED"
                    or "throttl" in str(error).lower()
                    or "timeout" in str(error).lower()
                    for error in errors
                    if isinstance(error, dict)
                )
                if retryable and attempt < 5:
                    time.sleep(min(30, 2**attempt))
                    continue
                raise RuntimeError(f"Shopify GraphQL errors: {error_text}")
            return payload["data"]
        raise RuntimeError("Shopify GraphQL retry limit exceeded")

    def active_variants(self) -> list[ShopifyVariant]:
        query = """
        query ActiveProductsForPriceUpdate($cursor: String) {
          products(first: 250, after: $cursor, query: "status:active", sortKey: TITLE) {
            pageInfo {
              hasNextPage
              endCursor
            }
            nodes {
              id
              handle
              title
              status
              vendor
              onlineStoreUrl
              variants(first: 100) {
                nodes {
                  id
                  title
                  sku
                  price
                  compareAtPrice
                }
              }
            }
          }
        }
        """
        variants: list[ShopifyVariant] = []
        cursor = None
        while True:
            data = self.graphql(query, {"cursor": cursor})
            page = data["products"]
            for product in page["nodes"]:
                for variant in product.get("variants", {}).get("nodes", []):
                    sku = (variant.get("sku") or "").strip()
                    variants.append(
                        ShopifyVariant(
                            product_id=product["id"],
                            product_handle=product.get("handle") or "",
                            product_title=product.get("title") or "",
                            product_status=product.get("status") or "",
                            vendor=product.get("vendor") or "",
                            online_store_url=product.get("onlineStoreUrl") or "",
                            variant_id=variant["id"],
                            variant_title=variant.get("title") or "",
                            sku=sku,
                            sku_key=sku_key(sku),
                            current_price=money(variant.get("price")),
                            current_compare_at=money(variant.get("compareAtPrice")),
                        )
                    )
            if not page["pageInfo"]["hasNextPage"]:
                break
            cursor = page["pageInfo"]["endCursor"]
        return variants

    def variants_bulk_update(self, product_id: str, variants: list[dict[str, Any]]) -> list[dict[str, Any]]:
        data = self.graphql(
            """
            mutation ProductVariantsBulkUpdate($productId: ID!, $variants: [ProductVariantsBulkInput!]!) {
              productVariantsBulkUpdate(productId: $productId, variants: $variants, allowPartialUpdates: true) {
                product {
                  id
                }
                productVariants {
                  id
                  sku
                  price
                  compareAtPrice
                }
                userErrors {
                  field
                  message
                }
              }
            }
            """,
            {"productId": product_id, "variants": variants},
        )
        return data["productVariantsBulkUpdate"]["userErrors"]


REPORT_FIELDS = [
    "action",
    "skip_reason",
    "product_id",
    "variant_id",
    "handle",
    "product_title",
    "vendor",
    "variant_title",
    "sku",
    "current_price",
    "current_compare_at_price",
    "target_price",
    "target_compare_at_price",
    "pricing_status",
    "pricing_brand",
    "pricing_source_file",
    "online_store_url",
]


def report_row(variant: ShopifyVariant, pricing: PricingRow | None, action: str, skip_reason: str) -> dict[str, str]:
    return {
        "action": action,
        "skip_reason": skip_reason,
        "product_id": variant.product_id,
        "variant_id": variant.variant_id,
        "handle": variant.product_handle,
        "product_title": variant.product_title,
        "vendor": variant.vendor,
        "variant_title": variant.variant_title,
        "sku": variant.sku,
        "current_price": money_str(variant.current_price),
        "current_compare_at_price": money_str(variant.current_compare_at),
        "target_price": money_str(pricing.target_price if pricing else None),
        "target_compare_at_price": money_str(pricing.target_compare_at if pricing else None),
        "pricing_status": pricing.status if pricing else "",
        "pricing_brand": pricing.brand if pricing else "",
        "pricing_source_file": pricing.source_file if pricing else "",
        "online_store_url": variant.online_store_url,
    }


def pricing_status_skip_reason(status: str, allow_missing_map: bool) -> str | None:
    if status == PASS_STATUS:
        return None
    if allow_missing_map and status == MISSING_MAP_REVIEW_STATUS:
        return None
    if status == MISSING_MAP_REVIEW_STATUS:
        return "pricing_status_missing_map"
    if status.startswith("FAIL"):
        return "pricing_status_fail"
    return "pricing_status_not_pass"


def update_reason(status: str) -> str:
    if status == MISSING_MAP_REVIEW_STATUS:
        return MISSING_MAP_UPDATE_REASON
    return PASS_UPDATE_REASON


def build_report(
    variants: list[ShopifyVariant],
    pricing_by_sku: dict[str, list[PricingRow]],
    allow_missing_map: bool,
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for variant in variants:
        if not variant.sku_key:
            rows.append(report_row(variant, None, "skip", "missing_variant_sku"))
            continue
        pricing_matches = pricing_by_sku.get(variant.sku_key, [])
        if not pricing_matches:
            rows.append(report_row(variant, None, "skip", "unmatched_active_sku"))
            continue
        if len(pricing_matches) > 1:
            rows.append(report_row(variant, None, "skip", "duplicate_pricing_rows"))
            continue

        pricing = pricing_matches[0]
        status_skip_reason = pricing_status_skip_reason(pricing.status, allow_missing_map)
        if status_skip_reason:
            rows.append(report_row(variant, pricing, "skip", status_skip_reason))
            continue
        if pricing.target_price <= Decimal("0.00"):
            rows.append(report_row(variant, pricing, "skip", "invalid_target_price"))
            continue
        if pricing.target_compare_at is not None and pricing.target_compare_at <= pricing.target_price:
            rows.append(report_row(variant, pricing, "skip", "compare_at_not_above_price"))
            continue
        if variant.current_price == pricing.target_price and variant.current_compare_at == pricing.target_compare_at:
            rows.append(report_row(variant, pricing, "noop", "already_current"))
            continue
        rows.append(report_row(variant, pricing, "update", update_reason(pricing.status)))
    return rows


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=REPORT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def summarize(
    rows: list[dict[str, str]],
    variants: list[ShopifyVariant],
    pricing_by_sku: dict[str, list[PricingRow]],
    applied: bool,
    allow_missing_map: bool,
) -> dict[str, Any]:
    action_counts = Counter(row["action"] for row in rows)
    reason_counts = Counter(row["skip_reason"] for row in rows)
    vendor_counts = Counter(row["vendor"] for row in rows)
    current_price_counts = Counter(money_str(variant.current_price) for variant in variants)
    pricing_status_counts = Counter()
    for matches in pricing_by_sku.values():
        for row in matches:
            pricing_status_counts[row.status or "missing"] += 1

    return {
        "pricing_workbook_count": len({row.source_file for matches in pricing_by_sku.values() for row in matches}),
        "pricing_row_count": sum(len(matches) for matches in pricing_by_sku.values()),
        "pricing_sku_count": len(pricing_by_sku),
        "pricing_duplicate_sku_count": sum(1 for matches in pricing_by_sku.values() if len(matches) > 1),
        "pricing_duplicate_row_count": sum(len(matches) for matches in pricing_by_sku.values() if len(matches) > 1),
        "shopify_active_variant_count": len(variants),
        "shopify_active_vendor_counts": dict(sorted(vendor_counts.items())),
        "current_price_counts": dict(current_price_counts.most_common(20)),
        "action_counts": dict(sorted(action_counts.items())),
        "skip_reason_counts": dict(sorted(reason_counts.items())),
        "pricing_status_counts": dict(sorted(pricing_status_counts.items())),
        "eligible_update_count": action_counts.get("update", 0),
        "eligible_pass_count": reason_counts.get(PASS_UPDATE_REASON, 0),
        "eligible_missing_map_count": reason_counts.get(MISSING_MAP_UPDATE_REASON, 0),
        "allowed_missing_map_count": sum(
            1
            for row in rows
            if row["pricing_status"] == MISSING_MAP_REVIEW_STATUS and row["action"] in {"update", "noop"}
        ),
        "blocked_review_count": sum(count for reason, count in reason_counts.items() if reason.startswith("pricing_status")),
        "blocked_fail_count": reason_counts.get("pricing_status_fail", 0),
        "blocked_missing_map_count": reason_counts.get("pricing_status_missing_map", 0),
        "unmatched_active_count": reason_counts.get("unmatched_active_sku", 0),
        "duplicate_pricing_count": reason_counts.get("duplicate_pricing_rows", 0),
        "allow_missing_map": allow_missing_map,
        "applied": applied,
    }


def duplicate_source_rows(pricing_by_sku: dict[str, list[PricingRow]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for matches in pricing_by_sku.values():
        if len(matches) <= 1:
            continue
        for row in matches:
            rows.append(
                {
                    "sku": row.sku,
                    "pricing_brand": row.brand,
                    "pricing_title": row.title,
                    "target_price": money_str(row.target_price),
                    "target_compare_at_price": money_str(row.target_compare_at),
                    "pricing_status": row.status,
                    "pricing_source_file": row.source_file,
                }
            )
    return rows


def write_duplicate_source_csv(path: Path, rows: list[dict[str, str]]) -> None:
    fields = [
        "sku",
        "pricing_brand",
        "pricing_title",
        "target_price",
        "target_compare_at_price",
        "pricing_status",
        "pricing_source_file",
    ]
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def approved_update_keys(path: Path) -> tuple[set[tuple[str, str, str]], bool]:
    with path.open(newline="", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        required = {"variant_id", "target_price", "target_compare_at_price", "action", "skip_reason"}
        if not required.issubset(reader.fieldnames or []):
            raise RuntimeError(f"Approved report missing required columns: {sorted(required)}")
        keys = set()
        contains_missing_map_updates = False
        for row in reader:
            if row.get("action") != "update":
                continue
            if row.get("skip_reason") == MISSING_MAP_UPDATE_REASON:
                contains_missing_map_updates = True
            keys.add((row["variant_id"], row["target_price"], row.get("target_compare_at_price") or ""))
        return keys, contains_missing_map_updates


def apply_updates(
    admin: ShopifyAdmin,
    report_rows: list[dict[str, str]],
    approved_path: Path,
    allow_missing_map: bool,
) -> list[dict[str, Any]]:
    approved, approved_contains_missing_map_updates = approved_update_keys(approved_path)
    if approved_contains_missing_map_updates and not allow_missing_map:
        raise RuntimeError(f"Approved report contains {MISSING_MAP_UPDATE_REASON} rows; rerun apply with --allow-missing-map")
    current_updates = {
        (row["variant_id"], row["target_price"], row["target_compare_at_price"]): row
        for row in report_rows
        if row["action"] == "update"
    }
    extra = approved - set(current_updates)
    if extra:
        raise RuntimeError(f"Approved report contains rows that are not currently eligible: {len(extra)}")

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for key in sorted(approved):
        row = current_updates[key]
        variant_input: dict[str, Any] = {"id": row["variant_id"], "price": row["target_price"]}
        if row["target_compare_at_price"]:
            variant_input["compareAtPrice"] = row["target_compare_at_price"]
        else:
            variant_input["compareAtPrice"] = None
        grouped[row["product_id"]].append(variant_input)

    results = []
    for product_id, variants in grouped.items():
        errors = admin.variants_bulk_update(product_id, variants)
        results.append(
            {
                "product_id": product_id,
                "variant_update_count": len(variants),
                "errors": errors,
                "ok": not errors,
            }
        )
    return results


def main() -> int:
    parser = argparse.ArgumentParser(description="Update Shopify active variant prices from approved MAP-protected pricing workbooks.")
    parser.add_argument("--pricing-dir", type=Path, default=DEFAULT_PRICING_DIR)
    parser.add_argument(
        "--pricing-file",
        type=Path,
        action="append",
        default=[],
        help="Explicit pricing workbook to load. Can be passed more than once; overrides --pricing-dir glob.",
    )
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    parser.add_argument("--apply", action="store_true", help="Apply approved eligible price updates to Shopify.")
    parser.add_argument("--yes", action="store_true", help="Required with --apply.")
    parser.add_argument("--input-approved-report", type=Path, help="Required with --apply. Use a reviewed price-update-eligible.csv.")
    parser.add_argument(
        "--allow-missing-map",
        action="store_true",
        help=f"Allow rows with pricing status '{MISSING_MAP_REVIEW_STATUS}' to update at the original profit-model price.",
    )
    args = parser.parse_args()

    if args.apply and (not args.yes or not args.input_approved_report):
        raise RuntimeError("--apply requires --yes and --input-approved-report")

    args.out_dir.mkdir(parents=True, exist_ok=True)
    pricing_by_sku = load_pricing_rows_from_files(args.pricing_file) if args.pricing_file else load_pricing_rows(args.pricing_dir)
    admin = ShopifyAdmin()
    variants = admin.active_variants()
    report_rows = build_report(variants, pricing_by_sku, allow_missing_map=args.allow_missing_map)

    write_csv(args.out_dir / FULL_REPORT, report_rows)
    write_csv(args.out_dir / ELIGIBLE_REPORT, [row for row in report_rows if row["action"] == "update"])
    write_csv(args.out_dir / BLOCKED_REPORT, [row for row in report_rows if row["skip_reason"].startswith("pricing_status")])
    write_csv(args.out_dir / UNMATCHED_REPORT, [row for row in report_rows if row["skip_reason"] == "unmatched_active_sku"])
    write_csv(args.out_dir / DUPLICATE_REPORT, [row for row in report_rows if row["skip_reason"] == "duplicate_pricing_rows"])
    write_duplicate_source_csv(args.out_dir / DUPLICATE_SOURCE_REPORT, duplicate_source_rows(pricing_by_sku))

    apply_results: list[dict[str, Any]] = []
    if args.apply:
        apply_results = apply_updates(admin, report_rows, args.input_approved_report, allow_missing_map=args.allow_missing_map)
        (args.out_dir / APPLY_RESULT_JSON).write_text(json.dumps(apply_results, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = summarize(report_rows, variants, pricing_by_sku, applied=args.apply, allow_missing_map=args.allow_missing_map)
    if apply_results:
        summary["apply_ok_count"] = sum(1 for row in apply_results if row["ok"])
        summary["apply_error_count"] = sum(1 for row in apply_results if not row["ok"])
        summary["applied_variant_update_count"] = sum(row["variant_update_count"] for row in apply_results)
    (args.out_dir / SUMMARY_JSON).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(str(error), file=sys.stderr)
        raise SystemExit(1)
