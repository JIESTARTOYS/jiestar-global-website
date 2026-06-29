#!/usr/bin/env python3
"""Generate a unified English B2B quotation workbook from ORICO brand sheets."""

from __future__ import annotations

import argparse
import hashlib
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps

import generate_b2b_product_catalogs as catalog


OUTPUT_DIR = Path("output/spreadsheets/english-quotations")
OUTPUT_FILE = OUTPUT_DIR / "all-brands-english-quotation.xlsx"
TMP_ROOT = Path("tmp/spreadsheets/english-quotations")

BRAND_ORDER = ["JIESTAR", "Xbert", "TK Two", "Small Angle", "GULY", "Zoin", "iBlock", "JIQI"]
FORBIDDEN_TERMS = ("控价", "零售价", "1688", "跨境控价", "Control Price", "Retail Price")
FORBIDDEN_TEXT_TERMS = ("控价", "零售价", "跨境控价", "Control Price", "Retail Price")

HEADERS = [
    "Product Image",
    "Brand",
    "Series",
    "Product Name (EN)",
    "Item No. / SKU",
    "RMB Factory Price",
    "USD Price",
    "Pieces",
    "Carton Qty",
    "Color Box Size (CM)",
    "Outer Carton Size (CM)",
    "Gross / Net Weight (KG)",
    "Model Size (CM)",
    "Recommended Age",
    "Remarks",
    "Source Sheet",
    "Source Row",
]


@dataclass
class QuotationProduct:
    product: catalog.Product
    rmb_price: float | int | None
    remarks: str
    price_source: str

    @property
    def brand(self) -> str:
        return self.product.brand

    @property
    def sku(self) -> str:
        return self.product.sku

    @property
    def image_path(self) -> str:
        return self.product.image_path

    def row_values(self) -> list[Any]:
        p = self.product
        return [
            "",
            p.brand,
            p.series,
            p.name_en,
            p.sku,
            self.rmb_price,
            None,
            dash(p.piece_count),
            dash(p.carton_qty),
            dash(p.color_box_size),
            dash(p.outer_carton_size),
            dash(p.gross_net_weight_kg),
            dash(p.model_size_cm),
            dash(p.recommended_age),
            dash(self.remarks),
            p.source_sheet,
            p.source_row,
        ]


def dash(value: Any) -> str:
    text = catalog.clean_text(value)
    return text if text != "" else "-"


def normalize_price(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value) if float(value).is_integer() else float(value)
    text = unicodedata.normalize("NFKC", str(value)).strip()
    if not text or text in {"-", "/", "--"}:
        return None
    text = text.replace("￥", "").replace("¥", "").replace("元", "").replace("RMB", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group(0))
    return int(number) if number.is_integer() else number


def clean_remarks(value: Any) -> str:
    text = catalog.pdf_safe_text(value)
    if text == "-":
        return "-"
    text = re.sub(r"\s+", " ", text).strip()
    return text if text else "-"


def electric_remark(value: Any) -> str:
    text = catalog.clean_text(value)
    if text == "-":
        return "-"
    if "电" in text:
        return "Electric"
    if "静" in text:
        return "Static"
    return clean_remarks(text)


def make_quote(
    *,
    product: catalog.Product | None,
    rmb_price: Any,
    remarks: Any = None,
    price_source: str,
) -> QuotationProduct | None:
    if not product:
        return None
    return QuotationProduct(
        product=product,
        rmb_price=normalize_price(rmb_price),
        remarks=clean_remarks(remarks),
        price_source=price_source,
    )


def extract_jiestar_and_xbert(title_maps: dict[str, dict[str, str]]) -> list[QuotationProduct]:
    wb = catalog.load_workbook(catalog.JIESTAR_BOOK)
    image_base = TMP_ROOT / "source-images" / "jiestar-source"
    products: list[QuotationProduct] = []
    configs = {
        "杰星积木": ("JIESTAR", 3, 4, 5, 7, 8, 9, 10, 12, 13, 15, 14),
        "杰星X系列": ("JIESTAR", 3, 4, 5, 7, 8, 9, 10, 12, 13, 15, 14),
        "杰星FF系列": ("JIESTAR", 3, 4, 5, 7, 8, 9, 10, 12, 13, 15, 14),
        "杰星JJ【积木】": ("JIESTAR", 3, 4, 5, 7, 8, 9, 10, 12, 13, 15, 14),
        "砖悦": ("Xbert", 3, 4, 5, 7, 8, 9, 10, 12, 13, 15, 14),
    }
    for sheet_name, config in configs.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        image_map = catalog.row_image_map(ws, image_base / sheet_name, preferred_cols={2})
        for row in range(4, ws.max_row + 1):
            product = catalog.make_product(
                brand=config[0],
                series=ws.cell(row, config[1]).value,
                sku=ws.cell(row, config[2]).value,
                source_name=ws.cell(row, config[3]).value,
                title_maps=title_maps,
                source_file=catalog.JIESTAR_BOOK,
                source_sheet=sheet_name,
                source_row=row,
                color_box_size=ws.cell(row, config[4]).value,
                model_size_cm=ws.cell(row, config[5]).value,
                outer_carton_size=ws.cell(row, config[6]).value,
                carton_qty=ws.cell(row, config[7]).value,
                gross_net_weight_kg=ws.cell(row, config[8]).value,
                recommended_age=ws.cell(row, config[9]).value,
                piece_count_value=ws.cell(row, config[10]).value,
                image_path=image_map.get(row),
                image_source="workbook",
            )
            quote = make_quote(
                product=product,
                rmb_price=ws.cell(row, config[11]).value,
                remarks=ws.cell(row, config[10]).value,
                price_source="出厂价",
            )
            if quote:
                products.append(quote)
    return products


def count_jiestar_baby_rows() -> int:
    wb = catalog.load_workbook(catalog.JIESTAR_BOOK)
    if "杰星婴童" not in wb.sheetnames:
        return 0
    ws = wb["杰星婴童"]
    count = 0
    for row in range(4, ws.max_row + 1):
        if catalog.clean_sku(ws.cell(row, 4).value):
            count += 1
    return count


def count_iblock_sheet1_no_dealer_price_rows() -> int:
    wb = catalog.load_workbook(catalog.IBLOCK_BOOK, data_only=False)
    if "Sheet1" not in wb.sheetnames:
        return 0
    ws = wb["Sheet1"]
    count = 0
    for row in range(5, ws.max_row + 1):
        if catalog.clean_sku(ws.cell(row, 4).value):
            count += 1
    return count


def extract_tk_two(title_maps: dict[str, dict[str, str]]) -> list[QuotationProduct]:
    wb = catalog.load_workbook(catalog.TK_BOOK)
    ws = wb["Sheet1"]
    image_map = catalog.row_image_map(ws, TMP_ROOT / "source-images" / "tk-two", preferred_cols={2})
    products: list[QuotationProduct] = []
    for row in range(3, ws.max_row + 1):
        product = catalog.make_product(
            brand="TK Two",
            series="Tank",
            sku=ws.cell(row, 1).value,
            source_name=ws.cell(row, 3).value,
            title_maps=title_maps,
            source_file=catalog.TK_BOOK,
            source_sheet=ws.title,
            source_row=row,
            carton_qty=ws.cell(row, 4).value,
            model_size_cm=ws.cell(row, 9).value,
            color_box_size=ws.cell(row, 10).value,
            outer_carton_size=ws.cell(row, 11).value,
            gross_net_weight_kg=ws.cell(row, 13).value,
            piece_count_value=ws.cell(row, 14).value,
            image_path=image_map.get(row),
            image_source="workbook",
        )
        quote = make_quote(product=product, rmb_price=ws.cell(row, 5).value, price_source="单价/元")
        if quote:
            products.append(quote)
    return products


def extract_small_angle(title_maps: dict[str, dict[str, str]]) -> list[QuotationProduct]:
    wb = catalog.load_workbook(catalog.SMALL_ANGLE_BOOK)
    ws = wb["Sheet1"]
    image_map = catalog.row_image_map(
        ws, TMP_ROOT / "source-images" / "small-angle", preferred_cols={1}
    )
    products: list[QuotationProduct] = []
    for row in range(5, ws.max_row + 1):
        product = catalog.make_product(
            brand="Small Angle",
            series=ws.cell(row, 2).value,
            sku=ws.cell(row, 3).value,
            source_name=ws.cell(row, 5).value,
            title_maps=title_maps,
            source_file=catalog.SMALL_ANGLE_BOOK,
            source_sheet=ws.title,
            source_row=row,
            piece_count_value=ws.cell(row, 4).value,
            carton_qty=ws.cell(row, 7).value,
            color_box_size=ws.cell(row, 8).value,
            outer_carton_size=ws.cell(row, 9).value,
            gross_net_weight_kg=catalog.format_gross_net(ws.cell(row, 10).value, ws.cell(row, 11).value),
            image_path=image_map.get(row),
            image_source="workbook",
        )
        quote = make_quote(
            product=product,
            rmb_price=ws.cell(row, 6).value,
            price_source="单价【单位：元】",
        )
        if quote:
            products.append(quote)
    return products


def extract_guly(title_maps: dict[str, dict[str, str]]) -> list[QuotationProduct]:
    wb = catalog.load_workbook(catalog.GULY_BOOK)
    products: list[QuotationProduct] = []
    configs = {
        "古励积木报价": (1, 2, 6, 7, 8, 9, 10, 4, 12),
        "双高积木报价": (1, 2, 6, 7, 8, 9, 10, 4, 12),
        "展鹏": (1, 2, 5, 6, 7, 8, 9, 4, 10),
    }
    for sheet_name, config in configs.items():
        ws = wb[sheet_name]
        image_map = catalog.row_image_map(
            ws, TMP_ROOT / "source-images" / "guly" / sheet_name, preferred_cols={3}
        )
        start_row = 5 if sheet_name != "双高积木报价" else 3
        for row in range(start_row, ws.max_row + 1):
            product = catalog.make_product(
                brand="GULY",
                series=sheet_name,
                sku=ws.cell(row, config[0]).value,
                source_name=ws.cell(row, config[1]).value,
                title_maps=title_maps,
                source_file=catalog.GULY_BOOK,
                source_sheet=sheet_name,
                source_row=row,
                carton_qty=ws.cell(row, config[2]).value,
                color_box_size=ws.cell(row, config[3]).value,
                outer_carton_size=ws.cell(row, config[4]).value,
                gross_net_weight_kg=ws.cell(row, config[5]).value,
                piece_count_value=ws.cell(row, config[6]).value,
                image_path=image_map.get(row),
                image_source="workbook",
            )
            quote = make_quote(
                product=product,
                rmb_price=ws.cell(row, config[7]).value,
                remarks=ws.cell(row, config[8]).value,
                price_source="出厂价",
            )
            if quote:
                products.append(quote)
    return products


def extract_zoin(title_maps: dict[str, dict[str, str]]) -> list[QuotationProduct]:
    wb = catalog.load_workbook(catalog.ZOIN_BOOK)
    ws = wb["Sheet1"]
    image_map = catalog.row_image_map(ws, TMP_ROOT / "source-images" / "zoin", preferred_cols={2})
    products: list[QuotationProduct] = []
    for row in range(3, ws.max_row + 1):
        product = catalog.make_product(
            brand="Zoin",
            series=ws.cell(row, 4).value,
            sku=ws.cell(row, 5).value,
            source_name=ws.cell(row, 6).value,
            title_maps=title_maps,
            source_file=catalog.ZOIN_BOOK,
            source_sheet=ws.title,
            source_row=row,
            color_box_size=ws.cell(row, 8).value,
            model_size_cm=ws.cell(row, 9).value,
            outer_carton_size=ws.cell(row, 10).value,
            carton_qty=ws.cell(row, 11).value,
            gross_net_weight_kg=ws.cell(row, 13).value,
            recommended_age=ws.cell(row, 14).value,
            image_path=image_map.get(row),
            image_source="workbook",
        )
        quote = make_quote(
            product=product,
            rmb_price=ws.cell(row, 15).value,
            remarks=ws.cell(row, 18).value,
            price_source="出厂价",
        )
        if quote:
            products.append(quote)
    return products


def extract_iblock(title_maps: dict[str, dict[str, str]]) -> list[QuotationProduct]:
    wb = catalog.load_workbook(catalog.IBLOCK_BOOK, data_only=False)
    image_index = catalog.build_iblock_image_index()
    products: list[QuotationProduct] = []

    for sheet_name in ("常规品", "女生品", "男生品"):
        ws = wb[sheet_name]
        for row in range(5, ws.max_row + 1):
            sku = catalog.clean_sku(ws.cell(row, 5).value)
            raw_series = catalog.clean_text(ws.cell(row, 4).value)
            product = catalog.make_product(
                brand="iBlock",
                series=sheet_name if raw_series == "-" else raw_series,
                sku=sku,
                source_name=ws.cell(row, 6).value,
                title_maps=title_maps,
                source_file=catalog.IBLOCK_BOOK,
                source_sheet=sheet_name,
                source_row=row,
                carton_qty=ws.cell(row, 11).value,
                model_size_cm=ws.cell(row, 17).value,
                color_box_size=ws.cell(row, 18).value,
                recommended_age=ws.cell(row, 10).value,
                image_path=image_index.get(sku),
                image_source="iblock-prepared-folder" if image_index.get(sku) else "",
            )
            quote = make_quote(
                product=product,
                rmb_price=ws.cell(row, 13).value,
                price_source="一级经销价（元）",
            )
            if quote:
                products.append(quote)

    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        for row in range(5, ws.max_row + 1):
            sku = catalog.clean_sku(ws.cell(row, 4).value)
            product = catalog.make_product(
                brand="iBlock",
                series=ws.cell(row, 3).value or "Zodiac",
                sku=sku,
                source_name=ws.cell(row, 5).value,
                title_maps=title_maps,
                source_file=catalog.IBLOCK_BOOK,
                source_sheet=ws.title,
                source_row=row,
                image_path=image_index.get(sku),
                image_source="iblock-prepared-folder" if image_index.get(sku) else "",
            )
            quote = make_quote(product=product, rmb_price=None, price_source="missing一级经销价")
            if quote:
                products.append(quote)
    return products


def extract_jiqi(title_maps: dict[str, dict[str, str]]) -> list[QuotationProduct]:
    wb = catalog.load_workbook(catalog.JIQI_BOOK)
    ws = wb["Sheet1"]
    image_map = catalog.row_image_map(ws, TMP_ROOT / "source-images" / "jiqi", preferred_cols={2})
    products: list[QuotationProduct] = []
    for row in range(2, ws.max_row + 1):
        product = catalog.make_product(
            brand="JIQI",
            series="Display Model",
            sku=ws.cell(row, 3).value,
            source_name=ws.cell(row, 5).value,
            title_maps=title_maps,
            source_file=catalog.JIQI_BOOK,
            source_sheet=ws.title,
            source_row=row,
            piece_count_value=ws.cell(row, 4).value,
            carton_qty=ws.cell(row, 7).value,
            color_box_size=ws.cell(row, 8).value,
            outer_carton_size=ws.cell(row, 9).value,
            gross_net_weight_kg=ws.cell(row, 10).value,
            model_size_cm=ws.cell(row, 11).value,
            recommended_age=ws.cell(row, 12).value,
            image_path=image_map.get(row),
            image_source="workbook",
        )
        quote = make_quote(
            product=product,
            rmb_price=ws.cell(row, 6).value,
            remarks=electric_remark(ws.cell(row, 13).value),
            price_source="出厂价",
        )
        if quote:
            products.append(quote)
    return products


def build_products() -> list[QuotationProduct]:
    catalog.TMP_ROOT = TMP_ROOT / "catalog-extract"
    title_maps = catalog.load_title_maps()
    products: list[QuotationProduct] = []
    products.extend(extract_jiestar_and_xbert(title_maps))
    products.extend(extract_tk_two(title_maps))
    products.extend(extract_small_angle(title_maps))
    products.extend(extract_guly(title_maps))
    products.extend(extract_zoin(title_maps))
    products.extend(extract_iblock(title_maps))
    products.extend(extract_jiqi(title_maps))
    return sort_products(dedupe_products(products))


def dedupe_products(products: Iterable[QuotationProduct]) -> list[QuotationProduct]:
    best: dict[tuple[str, str], QuotationProduct] = {}
    order: list[tuple[str, str]] = []
    for quote in products:
        key = (quote.brand, quote.sku)
        if key not in best:
            best[key] = quote
            order.append(key)
            continue
        if quote_quality_score(quote) > quote_quality_score(best[key]):
            best[key] = quote
    return [best[key] for key in order]


def quote_quality_score(quote: QuotationProduct) -> int:
    score = catalog.product_quality_score(quote.product)
    if quote.rmb_price is not None:
        score += 2
    if quote.remarks != "-":
        score += 1
    return score


def sort_products(products: list[QuotationProduct]) -> list[QuotationProduct]:
    brand_rank = {brand: index for index, brand in enumerate(BRAND_ORDER)}
    return sorted(
        products,
        key=lambda q: (
            brand_rank.get(q.brand, 999),
            q.product.series,
            sku_sort_key(q.sku),
        ),
    )


def sku_sort_key(sku: str) -> tuple[str, int, str]:
    match = re.match(r"([A-Z]+)[._/-]?(\d+)", sku.upper())
    if not match:
        return (sku.upper(), -1, sku.upper())
    return (match.group(1), int(match.group(2)), sku.upper())


def make_thumbnail(source: str, *, sku: str) -> Path | None:
    if not source:
        return None
    source_path = Path(source)
    if not source_path.exists():
        return None
    thumb_dir = TMP_ROOT / "workbook-thumbnails"
    thumb_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha1(f"{sku}:{source_path}:{source_path.stat().st_mtime_ns}".encode()).hexdigest()[:14]
    target = thumb_dir / f"{sku_safe(sku)}-{digest}.jpg"
    if target.exists():
        return target
    try:
        with Image.open(source_path) as image:
            image = ImageOps.exif_transpose(image)
            image.thumbnail((180, 180), Image.Resampling.LANCZOS)
            canvas = Image.new("RGB", image.size, "white")
            if image.mode == "RGBA":
                canvas.paste(image, mask=image.split()[-1])
            else:
                canvas.paste(image.convert("RGB"))
            canvas.save(target, "JPEG", quality=78, optimize=True)
        return target
    except Exception:
        return None


def sku_safe(sku: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "-", sku)[:60] or "image"


def grouped_by_brand(products: list[QuotationProduct]) -> dict[str, list[QuotationProduct]]:
    result: dict[str, list[QuotationProduct]] = {brand: [] for brand in BRAND_ORDER}
    for product in products:
        result.setdefault(product.brand, []).append(product)
    return result


def write_workbook(products: list[QuotationProduct], output: Path = OUTPUT_FILE) -> dict[str, Any]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    by_brand = grouped_by_brand(products)
    all_ws = wb.create_sheet("All Products")
    write_product_sheet(all_ws, products)

    for brand in BRAND_ORDER:
        write_product_sheet(wb.create_sheet(safe_sheet_name(brand)), by_brand.get(brand, []))

    write_qa_sheet(wb.create_sheet("QA Summary"), products, by_brand)
    wb.save(output)

    return {
        "output": str(output),
        "total_products": len(products),
        "brand_counts": {brand: len(by_brand.get(brand, [])) for brand in BRAND_ORDER},
    }


def write_product_sheet(ws: Any, products: list[QuotationProduct]) -> None:
    ws.append(HEADERS)
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(products) + 1}"
    set_column_widths(ws)

    for index, quote in enumerate(products, start=2):
        ws.append(quote.row_values())
        ws.row_dimensions[index].height = 64
        style_data_row(ws, index)
        thumbnail = make_thumbnail(quote.image_path, sku=quote.sku)
        if thumbnail:
            try:
                image = XLImage(str(thumbnail))
                image.width = 76
                image.height = 76
                ws.add_image(image, f"A{index}")
            except Exception:
                pass


def style_header(ws: Any) -> None:
    fill = PatternFill("solid", fgColor="111827")
    font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28


def style_data_row(ws: Any, row: int) -> None:
    thin = Side(style="thin", color="E5E7EB")
    border = Border(bottom=thin)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row, col)
        cell.border = border
        cell.font = Font(name="Arial", size=9, color="111827")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in (6, 7):
        ws.cell(row, col).number_format = '#,##0.00'
    ws.cell(row, 7).fill = PatternFill("solid", fgColor="FFF7D6")
    ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")


def set_column_widths(ws: Any) -> None:
    widths = {
        "A": 13,
        "B": 14,
        "C": 21,
        "D": 42,
        "E": 16,
        "F": 15,
        "G": 12,
        "H": 11,
        "I": 14,
        "J": 19,
        "K": 19,
        "L": 20,
        "M": 19,
        "N": 13,
        "O": 28,
        "P": 22,
        "Q": 10,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def write_qa_sheet(
    ws: Any,
    products: list[QuotationProduct],
    by_brand: dict[str, list[QuotationProduct]],
) -> None:
    headers = [
        "Brand",
        "Products",
        "Missing RMB Price",
        "Missing Image",
        "Low Confidence English Names",
        "iBlock Sheet1 Missing Price",
        "Notes",
    ]
    ws.append(headers)
    style_header(ws)
    ws.freeze_panes = "A2"
    baby_excluded = count_jiestar_baby_rows()
    iblock_sheet1_no_dealer_price = count_iblock_sheet1_no_dealer_price_rows()
    for brand in BRAND_ORDER:
        rows = by_brand.get(brand, [])
        note = ""
        if brand == "JIESTAR":
            note = f"JIESTAR domestic quotation excluded; Baby and Toddler excluded rows: {baby_excluded}"
        elif brand == "iBlock":
            note = (
                "RMB price uses 一级经销价（元） only; "
                f"Sheet1 source rows without 一级经销价: {iblock_sheet1_no_dealer_price}. "
                "Those SKUs are duplicates of priced category-sheet rows after consolidation."
            )
        ws.append(
            [
                brand,
                len(rows),
                sum(1 for q in rows if q.rmb_price is None),
                sum(1 for q in rows if not q.image_path or not Path(q.image_path).exists()),
                sum(1 for q in rows if q.product.name_confidence == "low"),
                iblock_sheet1_no_dealer_price if brand == "iBlock" else 0,
                note,
            ]
        )
    ws.append(
        [
            "TOTAL",
            len(products),
            sum(1 for q in products if q.rmb_price is None),
            sum(1 for q in products if not q.image_path or not Path(q.image_path).exists()),
            sum(1 for q in products if q.product.name_confidence == "low"),
            iblock_sheet1_no_dealer_price,
            "Price-control columns, consumer retail columns, barcode, certificate text, and internal cost fields excluded.",
        ]
    )
    for row in range(2, ws.max_row + 1):
        style_data_row(ws, row)
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    widths = {"A": 18, "B": 12, "C": 18, "D": 14, "E": 26, "F": 25, "G": 72}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def safe_sheet_name(name: str) -> str:
    return re.sub(r"[\[\]:*?/\\]", "-", name)[:31]


def validate_workbook(path: Path, products: list[QuotationProduct]) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    sheet_names = wb.sheetnames
    all_rows = wb["All Products"].max_row - 1
    brand_rows = {
        brand: wb[safe_sheet_name(brand)].max_row - 1 for brand in BRAND_ORDER if safe_sheet_name(brand) in wb.sheetnames
    }
    forbidden_hits: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    terms = FORBIDDEN_TERMS if cell.row == 1 else FORBIDDEN_TEXT_TERMS
                    for term in terms:
                        if term.lower() in value.lower():
                            forbidden_hits.append(f"{ws.title}!{cell.coordinate}:{term}")
    usd_nonblank = 0
    if "All Products" in wb.sheetnames:
        ws = wb["All Products"]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 7).value not in (None, ""):
                usd_nonblank += 1
    qa_ws = wb["QA Summary"]
    qa_values = list(qa_ws.iter_rows(min_row=2, values_only=True))
    return {
        "sheet_names": sheet_names,
        "all_rows": all_rows,
        "brand_rows": brand_rows,
        "brand_total": sum(brand_rows.values()),
        "expected_total": len(products),
        "forbidden_hits": forbidden_hits[:20],
        "forbidden_hit_count": len(forbidden_hits),
        "usd_nonblank": usd_nonblank,
        "qa_rows": qa_values,
        "image_counts": {ws.title: len(getattr(ws, "_images", [])) for ws in wb.worksheets},
    }


def print_summary(summary: dict[str, Any], validation: dict[str, Any]) -> None:
    print(f"Output: {summary['output']}")
    print(f"Total products: {summary['total_products']}")
    print(f"Brand counts: {summary['brand_counts']}")
    print(f"All Products rows: {validation['all_rows']}")
    print(f"Brand sheet total rows: {validation['brand_total']}")
    print(f"Forbidden term hits: {validation['forbidden_hit_count']}")
    print(f"USD nonblank cells: {validation['usd_nonblank']}")
    print(f"Image counts: {validation['image_counts']}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate the unified English quotation workbook.")
    parser.add_argument("--output", type=Path, default=OUTPUT_FILE)
    parser.add_argument("--validate-only", action="store_true")
    args = parser.parse_args()

    if args.validate_only:
        products = build_products()
        validation = validate_workbook(args.output, products)
        print_summary({"output": str(args.output), "total_products": len(products), "brand_counts": {}}, validation)
        return

    products = build_products()
    summary = write_workbook(products, args.output)
    validation = validate_workbook(args.output, products)
    print_summary(summary, validation)
    if validation["all_rows"] != validation["expected_total"]:
        raise SystemExit("All Products row count does not match extracted product count.")
    if validation["brand_total"] != validation["expected_total"]:
        raise SystemExit("Brand sheet rows do not sum to extracted product count.")
    if validation["forbidden_hit_count"]:
        raise SystemExit(f"Forbidden terms found: {validation['forbidden_hits']}")
    if validation["usd_nonblank"]:
        raise SystemExit("USD Price column should be blank.")


if __name__ == "__main__":
    main()
