from __future__ import annotations

import importlib.util
import sys
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "iblock_prepare_assets.py"


def load_module():
    spec = importlib.util.spec_from_file_location("iblock_prepare_assets", SCRIPT_PATH)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def make_quote(path: Path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    common_headers = [
        "序号",
        "上市日期",
        "品牌系列",
        "产品系列",
        "产品货号",
        "产品名称",
        "产品图片",
        "",
        "条码",
        "年龄段",
        "装箱数\npcs",
        "装箱\n规格",
        "一级\n经销价\n（元）",
        "二级\n经销价\n（元）",
        "建议\n零售价（元）",
        "电商\n活动控价\n（元）",
        "产品尺寸\n长*宽*高（CM）",
        "彩盒尺寸\n长*宽*高\n（CM）",
    ]

    regular = workbook.create_sheet("常规品")
    for _ in range(3):
        regular.append([])
    regular.append(common_headers)
    regular.append(
        [
            1,
            "2025-01-03",
            "iBlock\n潮玩系列",
            "十二生肖",
            "ib1001-4",
            "兔(领航机师)",
            "",
            "",
            "6977228050001",
            "8+",
            72,
            "1*72",
            13,
            14.5,
            29,
            29,
            "7.8*8*7.8",
            "9*9*14",
        ]
    )
    regular.append(
        [
            2,
            "2025-01-03",
            "iBlock\n潮玩系列",
            "瓶中童话",
            "ib1065",
            "瓶中童话-灰姑娘",
            "",
            "",
            "6977228050002",
            "8+",
            48,
            "1*48",
            9.2,
            12,
            25,
            25,
            "10*7.5*28",
            "13*8.5*17",
        ]
    )

    sheet1 = workbook.create_sheet("Sheet1")
    sheet1.append(["广东积趣文化科技有限公司"])
    sheet1.append(["瑞诺祥优商贸有限公司的报价单（含税）"])
    sheet1.append(["序号", "品牌系列", "产品系列", "产品货号", "产品名称", "含税单价（元）", "建议\n零售价\n（元）"])
    sheet1.append(["十二生肖系列"])
    sheet1.append([1, "iBlock潮玩系列", "十二生肖", "ib1001-4", "兔(领航机师)", 40.05, 89])

    workbook.save(path)


def touch(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"fake-image")


class IblockPrepareAssetsTest(unittest.TestCase):
    def test_expands_iblock_range_codes(self) -> None:
        module = load_module()

        self.assertEqual(module.expand_sku_codes("ib1065-8_瓶中童话"), ["IB1065", "IB1066", "IB1067", "IB1068"])
        self.assertEqual(module.expand_sku_codes("ib1204_99A坦克机甲"), ["IB1204"])
        self.assertEqual(
            module.expand_sku_codes("ib1076_ib1081_极速方程"),
            [
                "IB1076",
                "IB1077",
                "IB1078",
                "IB1079",
                "IB1080",
                "IB1081",
            ],
        )

    def test_matches_mini_series_aliases(self) -> None:
        module = load_module()

        rows = [
            module.ProductRow(
                sku="IB1301-4",
                source_sheets=("男生品",),
                brand_series="iBlock 潮玩系列",
                product_series="MINI战线",
                name_cn="99A主战坦克",
                barcode="",
                recommended_age="",
                carton_qty="",
                carton_spec="",
                dealer_price_level_1_cny="",
                dealer_price_level_2_cny="",
                retail_price_cny="",
                ecommerce_control_price_cny="",
                product_size_cm="",
                box_size_cm="",
                tax_included_price_cny="",
            )
        ]

        self.assertEqual(
            module.candidate_skus_for_path(Path("iblock_ib1301-6_Mini战线_99A_电商_头图_800x800.jpg"), rows),
            ["IB1301-4"],
        )

    def test_mini_series_combo_white_is_not_single_sku_white(self) -> None:
        module = load_module()

        self.assertEqual(module.classify_role(Path("ib1301_Mini战线_六合一白底图_800x800.png")), ("other", ""))
        self.assertEqual(module.classify_role(Path("ib1301_Mini战线_详情页_六合一.jpg")), ("detail", ""))

    def test_shared_city_dream_assets_are_split_by_model_terms_and_numbers(self) -> None:
        module = load_module()

        rows = [
            module.ProductRow(
                sku=sku,
                source_sheets=("常规品",),
                brand_series="iBlock 潮玩系列",
                product_series="城市梦英雄",
                name_cn=name,
                barcode="",
                recommended_age="",
                carton_qty="",
                carton_spec="",
                dealer_price_level_1_cny="",
                dealer_price_level_2_cny="",
                retail_price_cny="",
                ecommerce_control_price_cny="",
                product_size_cm="",
                box_size_cm="",
                tax_included_price_cny="",
            )
            for sku, name in [
                ("IB1106", "望宇小队-宇航员"),
                ("IB1107", "望宇小队-火箭"),
                ("IB1108", "望宇小队-探测车"),
                ("IB1109", "望宇小队-卫星"),
                ("IB1116", "治愈小队-医护员"),
                ("IB1117", "治愈小队-救护车"),
                ("IB1118", "治愈小队-献血车"),
                ("IB1119", "治愈小队-救护站"),
            ]
        ]

        self.assertEqual(
            module.candidate_skus_for_path(
                Path("ib1106-ib1107-ib1108-ib1109_望宇小队_电商_主图_800x800_jpg_out (4).jpg"),
                rows,
            ),
            ["IB1107"],
        )
        self.assertEqual(
            module.candidate_skus_for_path(Path("ib1106-ib1107-ib1108-ib1109-望宇小队_卫星_白底图_800x800.png"), rows),
            ["IB1109"],
        )
        self.assertEqual(
            module.candidate_skus_for_path(Path("ib1116_治愈小队_头图800x800_SKU/救护车.jpg"), rows),
            ["IB1117"],
        )
        self.assertEqual(
            module.candidate_skus_for_path(Path("ib1116_治愈小队_头图800x800_SKU/画板 1 副本 12-80.jpg"), rows),
            ["IB1117"],
        )
        self.assertEqual(
            module.candidate_skus_for_path(Path("ib1116_治愈小队_头图800x800/ib1116_治愈小队_电商_头图1_800x800_out.jpg"), rows),
            [],
        )
        self.assertEqual(
            module.candidate_skus_for_path(
                Path("ib1116_治愈小队_详情页800/ib1116_治愈小队_电商_详情页_PSD-out_01.jpg"),
                rows,
            ),
            ["IB1116", "IB1117", "IB1118", "IB1119"],
        )

    def test_child_sku_matching_prefers_sibling_specific_keywords(self) -> None:
        module = load_module()

        rows = [
            module.ProductRow(
                sku=sku,
                source_sheets=("常规品",),
                brand_series="iBlock 潮玩系列",
                product_series="城市梦英雄",
                name_cn=name,
                barcode="",
                recommended_age="",
                carton_qty="",
                carton_spec="",
                dealer_price_level_1_cny="",
                dealer_price_level_2_cny="",
                retail_price_cny="",
                ecommerce_control_price_cny="",
                product_size_cm="",
                box_size_cm="",
                tax_included_price_cny="",
            )
            for sku, name in [
                ("IB1104", "重工崛起"),
                ("IB1104-1", "重工小队-工程师"),
                ("IB1104-2", "重工小队-搅拌车"),
                ("IB1104-3", "重工小队-推土机"),
                ("IB1104-4", "重工小队-挖掘机"),
            ]
        ]

        self.assertEqual(
            module.candidate_skus_for_path(
                Path("iB1104_城市梦英雄_重工小队_(1-4)/iblock_ib1104_重工小队_电商_工程师_白底图800x800.png"),
                rows,
            ),
            ["IB1104-1"],
        )
        self.assertEqual(
            module.candidate_skus_for_path(
                Path("iB1104_城市梦英雄_重工小队_(1-4)/iblock_ib1104_重工小队_电商_搅拌车_白底图800x800.png"),
                rows,
            ),
            ["IB1104-2"],
        )

    def test_range_sku_matching_prefers_specific_model_keywords(self) -> None:
        module = load_module()

        rows = [
            module.ProductRow(
                sku=sku,
                source_sheets=("女生品",),
                brand_series="iBlock 潮玩系列",
                product_series="花漾玲珑",
                name_cn=name,
                barcode="",
                recommended_age="",
                carton_qty="",
                carton_spec="",
                dealer_price_level_1_cny="",
                dealer_price_level_2_cny="",
                retail_price_cny="",
                ecommerce_control_price_cny="",
                product_size_cm="",
                box_size_cm="",
                tax_included_price_cny="",
            )
            for sku, name in [
                ("IB1085", "花漾玲珑-弗朗花"),
                ("IB1086", "花漾玲珑-樱花"),
                ("IB1087", "花漾玲珑-绣球花"),
                ("IB1088", "花漾玲珑-向日葵"),
                ("IB1089", "花漾玲珑-四叶草"),
                ("IB1090", "花漾玲珑-蓝风铃"),
            ]
        ]

        self.assertEqual(
            module.candidate_skus_for_path(
                Path("ib1085-1090_花漾玲珑_电商素材/ib1085-1090_花漾玲珑_电商_SKU图800x800/ib1086_花漾玲珑-樱花_电商_SKU图800x800.jpg"),
                rows,
            ),
            ["IB1086"],
        )
        self.assertEqual(
            module.candidate_skus_for_path(
                Path("ib1085-1090_花漾玲珑_电商素材/ib1085-1090_花漾玲珑_电商_SKU图800x800/ib1087_花漾玲珑-绣球花_电商_SKU图800x800.jpg"),
                rows,
            ),
            ["IB1087"],
        )

    def test_build_reports_matches_subsku_and_range_assets(self) -> None:
        from tempfile import TemporaryDirectory

        with TemporaryDirectory() as temp_dir:
            module = load_module()
            tmp_path = Path(temp_dir)
            quote = tmp_path / "quote.xlsx"
            source = tmp_path / "source"
            target = tmp_path / "target"
            make_quote(quote)

            touch(source / "iB1001_玩酷份子十二生肖_(1-12)" / "iblock_ib1001-十二生肖_兔_电商_JPG" / "兔_白底图_800x800.jpg")
            touch(source / "iB1001_玩酷份子十二生肖_(1-12)" / "iblock_ib1001-十二生肖_兔_电商_JPG" / "头图800x800" / "兔_头图_1.jpg")
            touch(source / "iB1001_玩酷份子十二生肖_(1-12)" / "iblock_ib1001-十二生肖_兔_电商_JPG" / "详情页分切" / "详情_01.png")
            touch(source / "ib1065-8_瓶中童话_电商素材" / "ib1065-8_瓶中童话_SKU头图800x800" / "ib1065_SKU图.png")
            touch(source / "ib1065-8_瓶中童话_电商素材" / "ib1065-8_瓶中童话_头图800x800" / "ib1065_main.jpg")
            touch(source / "ib1065-8_瓶中童话_电商素材" / "ib1065-8_瓶中童话_详情页800" / "ib1065_detail.jpg")

            summary = module.build_reports(
                quote_file=quote,
                source_root=source,
                target_root=target,
                dry_run=False,
            )

            self.assertEqual(summary["unique_quote_sku_count"], 2)
            self.assertEqual(summary["duplicate_quote_row_count"], 1)
            self.assertEqual(summary["ready_to_create_count"], 2)
            self.assertTrue((target / "images" / "IB1001-4" / "IB1001-4-白底.jpg").exists())
            self.assertTrue((target / "images" / "IB1001-4" / "IB1001-4-1.jpg").exists())
            self.assertTrue((target / "images" / "IB1001-4" / "IB1001-4-详情-01.png").exists())
            self.assertTrue((target / "images" / "IB1065" / "IB1065-sku.png").exists())

            workbook = load_workbook(target / "iblock-catalog-ready.xlsx", read_only=True, data_only=True)
            rows = list(workbook["Catalog"].iter_rows(values_only=True))
            self.assertEqual(rows[0][:6], ("sku", "vendor", "category", "brand_series", "product_series", "name_cn"))
            catalog = {row[0]: row for row in rows[1:]}
            self.assertEqual(catalog["IB1001-4"][20], "READY_TO_CREATE")
            self.assertEqual(catalog["IB1065"][16:20], (0, 1, 1, 1))

    def test_dry_run_writes_reports_without_copying_images(self) -> None:
        from tempfile import TemporaryDirectory

        with TemporaryDirectory() as temp_dir:
            module = load_module()
            tmp_path = Path(temp_dir)
            quote = tmp_path / "quote.xlsx"
            source = tmp_path / "source"
            target = tmp_path / "target"
            make_quote(quote)
            touch(source / "ib1065-8_瓶中童话_电商素材" / "ib1065-8_瓶中童话_头图800x800" / "ib1065_main.jpg")

            summary = module.build_reports(
                quote_file=quote,
                source_root=source,
                target_root=target,
                dry_run=True,
            )

            self.assertIs(summary["dry_run"], True)
            self.assertTrue((target / "reports" / "iblock-asset-copy-plan.csv").exists())
            self.assertFalse((target / "images").exists())


if __name__ == "__main__":
    unittest.main()
