from __future__ import annotations

import importlib.util
import sys
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook


SCRIPTS_DIR = Path(__file__).resolve().parents[1] / "scripts"


def load_shipping_script():
    path = SCRIPTS_DIR / "shopify_shipping_update_from_template.py"
    spec = importlib.util.spec_from_file_location("shopify_shipping_update_from_template_v2", path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


class ShopifyShippingUpdateFromTemplateTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.shipping = load_shipping_script()
        cls.air_rates, cls.sea_rates, cls.targets = cls.shipping.load_shipping_workbooks(
            cls.shipping.DEFAULT_AIR_RATE_WORKBOOK,
            cls.shipping.DEFAULT_SEA_RATE_WORKBOOK,
            cls.shipping.DEFAULT_WEIGHT_WORKBOOK,
        )

    def target(self, sku: str, weight: int | None, status: str = "Yes"):
        profile = (
            self.shipping.STANDARD_GOODS_PROFILE
            if status == "Yes"
            else self.shipping.MANUAL_REVIEW_PROFILE if status == "Review" else ""
        )
        return self.shipping.WeightTarget(
            sku=sku,
            sku_key=sku.upper(),
            handle=sku.lower(),
            title=f"Product {sku}",
            vendor="JIESTAR",
            target_weight_g=weight,
            source_profile="Standard goods" if status == "Yes" else "Manual review" if status == "Review" else "",
            target_profile=profile,
            weight_import_status=status,
            dimension_verification="Verified color-box dimensions; no +2cm buffer" if status != "No" else "Missing reliable color-box dimensions",
            shipping_status="",
            listing_status="",
            notes="",
        )

    def variant(self, product_id: str, variant_id: str, sku: str, current_weight: int = 100):
        return self.shipping.ShopifyVariant(
            product_id=product_id,
            product_handle=product_id.lower(),
            product_title=f"Product {product_id}",
            product_status="ACTIVE",
            vendor="JIESTAR",
            variant_id=variant_id,
            variant_title="Default",
            sku=sku,
            sku_key=sku.upper(),
            inventory_item_id=f"inventory-{variant_id}",
            inventory_tracked=False,
            requires_shipping=True,
            current_weight_g=current_weight,
        )

    def test_generated_workbooks_have_exact_approved_rate_counts(self) -> None:
        self.assertEqual(len(self.air_rates), 240)
        self.assertEqual(len(self.sea_rates), 31)
        self.assertEqual({row.rate_name for row in self.air_rates}, {"Air Shipping"})
        self.assertEqual({row.rate_name for row in self.sea_rates}, {"Sea Shipping"})
        self.assertEqual({row.target_profile for row in [*self.air_rates, *self.sea_rates]}, {"JIESTAR Standard goods"})
        self.assertEqual(
            {row.country_code for row in self.sea_rates},
            {"US", "AU"},
        )

    def test_x88058_acceptance_weight_and_us_prices(self) -> None:
        target = self.targets["X88058"]
        self.assertEqual(target.target_weight_g, 1058)
        self.assertEqual(target.target_profile, "JIESTAR Standard goods")
        air_price = next(
            row.price_usd
            for row in self.air_rates
            if row.country_code == "US" and row.max_weight_kg == Decimal("1.5")
        )
        sea_price = next(
            row.price_usd
            for row in self.sea_rates
            if row.country_code == "US" and row.max_weight_kg == Decimal("1.5")
        )
        self.assertEqual(air_price, Decimal("31.99"))
        self.assertEqual(sea_price, Decimal("26.99"))

    def test_unverified_sibling_drafts_the_whole_shopify_product(self) -> None:
        variants = [
            self.variant("product-1", "variant-1", "GOOD"),
            self.variant("product-1", "variant-2", "MISSING"),
        ]
        targets = {
            "GOOD": self.target("GOOD", 1000, "Yes"),
            "MISSING": self.target("MISSING", None, "No"),
        }
        weights, assignments, heavy, draft_actions, unmatched = self.shipping.build_shipping_report(variants, targets)
        self.assertEqual(weights, [])
        self.assertEqual(assignments, [])
        self.assertEqual(heavy, [])
        self.assertEqual(len(draft_actions), 1)
        self.assertEqual(draft_actions[0]["product_id"], "product-1")
        self.assertEqual(len(unmatched), 2)
        self.assertEqual({row["planned_action"] for row in unmatched}, {"change_product_status_to_draft"})

    def test_eligible_and_heavy_products_use_only_standard_or_manual_profiles(self) -> None:
        variants = [
            self.variant("product-1", "variant-1", "NORMAL"),
            self.variant("product-2", "variant-2", "HEAVY"),
        ]
        targets = {
            "NORMAL": self.target("NORMAL", 1058, "Yes"),
            "HEAVY": self.target("HEAVY", 10001, "Review"),
        }
        weights, assignments, heavy, draft_actions, unmatched = self.shipping.build_shipping_report(variants, targets)
        self.assertEqual(len(weights), 2)
        self.assertEqual(
            {row["target_profile"] for row in assignments},
            {"JIESTAR Standard goods", "JIESTAR Manual Shipping Review"},
        )
        self.assertEqual(len(heavy), 1)
        self.assertEqual(draft_actions, [])
        self.assertEqual(unmatched, [])

    def test_scoped_weight_update_never_drafts_products_outside_workbook(self) -> None:
        variants = [
            self.variant("product-in-scope", "variant-in-scope", "MATCHED"),
            self.variant("product-outside", "variant-outside", "UNRELATED"),
        ]
        targets = {"MATCHED": self.target("MATCHED", 1058, "Yes")}

        weights, assignments, heavy, draft_actions, unmatched = self.shipping.build_scoped_shipping_report(
            variants,
            targets,
            scope_to_targets=True,
        )

        self.assertEqual([row["sku"] for row in weights], ["MATCHED"])
        self.assertEqual([row["sku"] for row in assignments], ["MATCHED"])
        self.assertEqual(heavy, [])
        self.assertEqual(draft_actions, [])
        self.assertEqual(len(unmatched), 1)
        self.assertEqual(unmatched[0]["sku"], "UNRELATED")
        self.assertEqual(unmatched[0]["planned_action"], "report_only_no_write")

    def test_battery_rate_text_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bad-air.xlsx"
            workbook = Workbook()
            workbook.active.title = "Inputs"
            workbook.create_sheet("Rate_Source")
            sheet = workbook.create_sheet("Shopify_Rates")
            workbook.create_sheet("QA")
            sheet.append(
                [
                    "Shipping Profile", "Zone Country", "Country Code", "Service Type", "Rate Name",
                    "Transit Time", "Min Weight kg", "Max Weight kg", "Price USD", "Freight Cost RMB",
                    "Active", "Notes",
                ]
            )
            sheet.append(
                [
                    "Standard goods", "United States", "US", "Battery/electric", "Air Shipping",
                    "8-12 days", 0, 0.5, 14.99, 78, "Active", "",
                ]
            )
            workbook.save(path)
            with self.assertRaisesRegex(RuntimeError, "Battery/electric rate text is prohibited"):
                self.shipping.load_rate_workbook(path, "air")

    def test_live_rate_fingerprint_preserves_friendly_name_and_weight_tier(self) -> None:
        profile = {
            "profileLocationGroups": [
                {
                    "locationGroupZones": {
                        "nodes": [
                            {
                                "zone": {"countries": [{"code": {"countryCode": "US"}}]},
                                "methodDefinitions": {
                                    "nodes": [
                                        {
                                            "name": "Air Shipping",
                                            "rateProvider": {"price": {"amount": "31.99"}},
                                            "methodConditions": [
                                                {
                                                    "operator": "GREATER_THAN_OR_EQUAL_TO",
                                                    "conditionCriteria": {"__typename": "Weight", "value": 1.001},
                                                },
                                                {
                                                    "operator": "LESS_THAN_OR_EQUAL_TO",
                                                    "conditionCriteria": {"__typename": "Weight", "value": 1.5},
                                                },
                                            ],
                                        }
                                    ]
                                },
                            }
                        ]
                    }
                }
            ]
        }
        self.assertEqual(
            self.shipping.profile_rate_keys(profile),
            {("US", "Air Shipping", "1.001", "1.5", "31.99")},
        )

    def test_rate_fingerprint_can_be_scoped_to_chenghai_location_group(self) -> None:
        def location_group(group_id: str, location_id: str, location_name: str, rate_name: str):
            return {
                "locationGroup": {
                    "id": group_id,
                    "locations": {"nodes": [{"id": location_id, "name": location_name}]},
                },
                "locationGroupZones": {
                    "nodes": [
                        {
                            "zone": {"id": f"zone-{group_id}", "countries": [{"code": {"countryCode": "US"}}]},
                            "methodDefinitions": {
                                "nodes": [
                                    {
                                        "id": f"method-{group_id}",
                                        "name": rate_name,
                                        "rateProvider": {"price": {"amount": "19.99"}},
                                        "methodConditions": [
                                            {
                                                "operator": "GREATER_THAN_OR_EQUAL_TO",
                                                "conditionCriteria": {"__typename": "Weight", "value": 0},
                                            },
                                            {
                                                "operator": "LESS_THAN_OR_EQUAL_TO",
                                                "conditionCriteria": {"__typename": "Weight", "value": 0.5},
                                            },
                                        ],
                                    }
                                ]
                            },
                        }
                    ]
                },
            }

        profile = {
            "profileLocationGroups": [
                location_group("chenghai-group", "chenghai-location", self.shipping.CHENGHAI_LOCATION_NAME, "Air Shipping"),
                location_group("us-group", "us-location", "Amazon U.S. Warehouse", "U.S. Warehouse Shipping"),
            ]
        }

        self.assertEqual(
            self.shipping.profile_rate_keys(profile, location_name=self.shipping.CHENGHAI_LOCATION_NAME),
            {("US", "Air Shipping", "0", "0.5", "19.99")},
        )
        zone_ids, method_ids, group_id = self.shipping.profile_existing_ids(profile, "chenghai-location")
        self.assertEqual(zone_ids, ["zone-chenghai-group"])
        self.assertEqual(method_ids, ["method-chenghai-group"])
        self.assertEqual(group_id, "chenghai-group")

    def test_chenghai_location_selection_never_falls_back_to_another_location(self) -> None:
        with self.assertRaisesRegex(RuntimeError, "Expected exactly one Shopify location"):
            self.shipping.target_location_id(
                [{"id": "us-location", "name": "Amazon U.S. Warehouse", "isActive": True}]
            )


if __name__ == "__main__":
    unittest.main()
